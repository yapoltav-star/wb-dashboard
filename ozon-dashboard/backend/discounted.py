"""Уценённые товары Ozon.

Что даёт Seller API:
- /v3/product/list и /v3/product/info/list — флаг is_discounted (карточку уценил
  сам продавец) и discounted_fbo_stocks (остатки, которые уценил Ozon на FBO);
- /v1/product/info/discounted — состояние и дефекты уценённого товара + SKU
  основного товара;
- /v1/product/update/discount — размер скидки (3–99%) на уценённый товар FBS.

Создать уценённую карточку через API нельзя — она появляется из кабинета
(уценка своими руками по FBS) или её создаёт Ozon при повреждении на FBO.
"""

from __future__ import annotations

import io
import logging
import threading
from datetime import datetime, timezone
from typing import Callable

logger = logging.getLogger("ozon-dashboard.discounted")

DISC_MIN = 3
DISC_MAX = 99

DISC_CACHE: dict = {
    "rows": [],
    "updated_at": None,
    "syncing": False,
    "error": None,
    "counts": {},
}

_lock = threading.Lock()

CONDITION_LABELS = {
    "1": "удовлетворительное",
    "2": "хорошее",
    "3": "очень хорошее",
    "4": "отличное",
    "5": "как новое",
    "6": "как новое",
    "7": "как новое",
}

DEFECT_FIELDS = (
    ("defects", "дефекты"),
    ("mechanical_damage", "механические повреждения"),
    ("package_damage", "повреждение упаковки"),
    ("packaging_violation", "вскрытая упаковка"),
    ("shortage", "некомплект"),
    ("repair", "был в ремонте"),
    ("reason_damaged", "причина уценки"),
    ("comment_reason_damaged", "комментарий"),
)


def _to_num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _int(v) -> int:
    try:
        return int(v or 0)
    except (TypeError, ValueError):
        return 0


def _fetch_product_list(ozon_post: Callable, visibility: str = "ALL") -> list[dict]:
    items: list[dict] = []
    last_id = ""
    for _ in range(100):
        payload = ozon_post(
            "/v3/product/list",
            {"filter": {"visibility": visibility}, "last_id": last_id, "limit": 1000},
        )
        result = payload.get("result") or payload
        batch = result.get("items") or []
        items.extend(batch)
        last_id = result.get("last_id") or ""
        if not batch or not last_id:
            break
        total = result.get("total")
        if total is not None and len(items) >= _int(total):
            break
    return items


def _fetch_product_info(ozon_post: Callable, product_ids: list[int]) -> list[dict]:
    out: list[dict] = []
    for i in range(0, len(product_ids), 100):
        payload = ozon_post("/v3/product/info/list", {"product_id": product_ids[i : i + 100]})
        out.extend(payload.get("items") or (payload.get("result") or {}).get("items") or [])
    return out


def _fetch_discount_info(ozon_post: Callable, skus: list[int]) -> dict[str, dict]:
    """/v1/product/info/discounted → {discounted_sku: детали уценки}."""
    out: dict[str, dict] = {}
    for i in range(0, len(skus), 100):
        part = [str(s) for s in skus[i : i + 100] if s]
        if not part:
            continue
        try:
            payload = ozon_post("/v1/product/info/discounted", {"discounted_skus": part})
        except Exception as e:
            logger.warning("info/discounted: %s", e)
            continue
        for it in payload.get("items") or []:
            key = str(it.get("discounted_sku") or "")
            if key:
                out[key] = it
    return out


def _stocks(info: dict) -> tuple[int, int]:
    """(FBS, FBO) из /v3/product/info/list."""
    fbs = fbo = 0
    st = info.get("stocks")
    if isinstance(st, dict):
        rows = st.get("stocks")
        if isinstance(rows, list):
            for r in rows:
                src = str((r or {}).get("source") or "").lower()
                present = _int((r or {}).get("present"))
                if src == "fbs":
                    fbs += present
                elif src == "fbo":
                    fbo += present
        else:
            fbs = _int(st.get("fbs"))
            fbo = _int(st.get("fbo"))
    return fbs, fbo


def _sku_of(info: dict):
    sources = info.get("sources")
    if isinstance(sources, list):
        for s in sources:
            if isinstance(s, dict) and s.get("sku"):
                return s["sku"]
    return info.get("sku")


def _defects_text(det: dict) -> str:
    parts = []
    for field, label in DEFECT_FIELDS:
        val = str((det or {}).get(field) or "").strip()
        if val and val.lower() not in ("нет", "no", "false"):
            parts.append(f"{label}: {val}" if len(val) > 3 else label)
    return "; ".join(parts)


def _condition_text(det: dict) -> str:
    est = str((det or {}).get("condition_estimation") or "").strip()
    cond = str((det or {}).get("condition") or "").strip()
    label = CONDITION_LABELS.get(est, "")
    out = []
    if est:
        out.append(f"{est}/7" + (f" — {label}" if label else ""))
    if cond:
        out.append(cond)
    return " · ".join(out)


def sync_discounted(ozon_post: Callable) -> dict:
    if not _lock.acquire(blocking=False):
        DISC_CACHE["syncing"] = True
        return {"ok": False, "syncing": True}
    DISC_CACHE["syncing"] = True
    DISC_CACHE["error"] = None
    try:
        listed: dict[int, dict] = {}
        for visibility in ("ALL", "ARCHIVED"):
            for row in _fetch_product_list(ozon_post, visibility):
                pid = _int(row.get("product_id"))
                if not pid:
                    continue
                row = dict(row)
                row["archived"] = visibility == "ARCHIVED" or bool(row.get("archived"))
                listed[pid] = row

        infos = _fetch_product_info(ozon_post, list(listed.keys())) if listed else []
        info_by_id: dict[int, dict] = {}
        for it in infos:
            pid = _int(it.get("id") or it.get("product_id"))
            if pid:
                info_by_id[pid] = it

        # основные товары для связки «уценка → оригинал»
        main_by_sku: dict[str, dict] = {}
        for pid, info in info_by_id.items():
            sku = _sku_of(info)
            if sku:
                main_by_sku[str(sku)] = {
                    "product_id": pid,
                    "offer_id": info.get("offer_id") or "",
                    "name": info.get("name") or "",
                    "price": _to_num(info.get("price")),
                }

        own_skus = []
        for pid, info in info_by_id.items():
            if bool(info.get("is_discounted") or listed.get(pid, {}).get("is_discounted")):
                sku = _sku_of(info)
                if sku:
                    own_skus.append(_int(sku))
        details = _fetch_discount_info(ozon_post, own_skus) if own_skus else {}

        rows: list[dict] = []
        for pid, list_row in listed.items():
            info = info_by_id.get(pid) or {}
            is_own = bool(info.get("is_discounted") or list_row.get("is_discounted"))
            fbo_discounted = _int(info.get("discounted_fbo_stocks"))
            has_ozon_disc = bool(info.get("has_discounted_fbo_item")) or fbo_discounted > 0
            if not is_own and not has_ozon_disc:
                continue

            sku = _sku_of(info)
            det = details.get(str(sku)) if sku else None
            main_sku = (det or {}).get("sku")
            main = main_by_sku.get(str(main_sku)) if main_sku else None
            fbs, fbo = _stocks(info)
            price = _to_num(info.get("price"))
            old_price = _to_num(info.get("old_price"))
            discount_pct = None
            if price and old_price and old_price > 0 and price < old_price:
                discount_pct = round((old_price - price) / old_price * 100, 1)

            rows.append({
                "product_id": pid,
                "offer_id": info.get("offer_id") or list_row.get("offer_id") or "",
                "sku": sku,
                "name": info.get("name") or "",
                "kind": "own" if is_own else "ozon",
                "archived": bool(list_row.get("archived")),
                "price": price,
                "old_price": old_price,
                "min_price": _to_num(info.get("min_price")),
                "marketing_price": _to_num(info.get("marketing_price")),
                "discount_pct": discount_pct,
                "stock_fbs": fbs,
                "stock_fbo": fbo,
                "stock_total": fbs + fbo,
                "discounted_fbo_stocks": fbo_discounted,
                "condition": _condition_text(det) if det else "",
                "condition_estimation": (det or {}).get("condition_estimation") if det else None,
                "defects": _defects_text(det) if det else "",
                "warranty": str((det or {}).get("warranty_type") or "") if det else "",
                "main_sku": main_sku,
                "main_offer_id": (main or {}).get("offer_id") or "",
                "main_name": (main or {}).get("name") or "",
                "main_price": (main or {}).get("price"),
                "can_set_discount": is_own,
            })

        rows.sort(key=lambda r: (
            0 if r["kind"] == "own" else 1,
            -(r["stock_total"] or 0),
            str(r["offer_id"]),
        ))

        counts = {
            "own": sum(1 for r in rows if r["kind"] == "own"),
            "ozon": sum(1 for r in rows if r["kind"] == "ozon"),
            "with_stock": sum(1 for r in rows if r["stock_total"] > 0),
            "products_scanned": len(listed),
        }
        DISC_CACHE.update({
            "rows": rows,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "syncing": False,
            "error": None,
            "counts": counts,
        })
        logger.info("discounted sync: %s карточек (%s)", len(rows), counts)
        return {"ok": True, "count": len(rows)}
    except Exception as e:
        logger.exception("discounted sync")
        DISC_CACHE["error"] = str(e)
        DISC_CACHE["syncing"] = False
        return {"ok": False, "error": str(e)}
    finally:
        _lock.release()


def get_cached() -> dict:
    rows = DISC_CACHE.get("rows") or []
    own = [r for r in rows if r.get("kind") == "own"]
    discounts = [r["discount_pct"] for r in rows if r.get("discount_pct")]
    stock_value = sum(
        (r.get("price") or 0) * (r.get("stock_total") or 0) for r in rows
    )
    return {
        "rows": rows,
        "updated_at": DISC_CACHE.get("updated_at"),
        "syncing": bool(DISC_CACHE.get("syncing")),
        "error": DISC_CACHE.get("error"),
        "counts": DISC_CACHE.get("counts") or {},
        "summary": {
            "total": len(rows),
            "own": len(own),
            "ozon": len(rows) - len(own),
            "with_stock": sum(1 for r in rows if (r.get("stock_total") or 0) > 0),
            "stock_qty": sum(r.get("stock_total") or 0 for r in rows),
            "ozon_fbo_qty": sum(r.get("discounted_fbo_stocks") or 0 for r in rows),
            "avg_discount": round(sum(discounts) / len(discounts), 1) if discounts else None,
            "stock_value": round(stock_value, 2),
        },
        "limits": {"min": DISC_MIN, "max": DISC_MAX},
    }


def set_discount(ozon_post: Callable, product_id: int, discount: int) -> dict:
    """POST /v1/product/update/discount — скидка на уценённый товар FBS."""
    pid = _int(product_id)
    disc = _int(discount)
    if not pid:
        raise ValueError("нужен product_id")
    if disc < DISC_MIN or disc > DISC_MAX:
        raise ValueError(f"скидка должна быть от {DISC_MIN} до {DISC_MAX}%")
    payload = ozon_post("/v1/product/update/discount", {"product_id": pid, "discount": disc})
    ok = bool(payload.get("result", True))
    for r in DISC_CACHE.get("rows") or []:
        if r.get("product_id") == pid:
            r["discount_pct"] = disc
            r["discount_set_at"] = datetime.now(timezone.utc).isoformat()
            break
    return {"ok": ok, "product_id": pid, "discount": disc}


def build_xlsx_bytes(rows: list[dict], summary: dict | None = None) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError("На сервере нет openpyxl — добавь в requirements.txt и задеплой") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Discounted"
    ws.append([
        "Артикул", "Название", "SKU уценки", "Кто уценил", "Состояние", "Дефекты",
        "Цена", "Без скидки", "Скидка, %", "FBS", "FBO", "Уценено Ozon (FBO)",
        "Основной артикул", "Основной SKU", "Гарантия", "product_id",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows or []:
        ws.append([
            r.get("offer_id"), r.get("name"), r.get("sku"),
            "продавец" if r.get("kind") == "own" else "Ozon",
            r.get("condition"), r.get("defects"),
            r.get("price"), r.get("old_price"), r.get("discount_pct"),
            r.get("stock_fbs"), r.get("stock_fbo"), r.get("discounted_fbo_stocks"),
            r.get("main_offer_id"), r.get("main_sku"), r.get("warranty"), r.get("product_id"),
        ])
    for col, w in zip("ABCDEFGHIJKLMNOP", [22, 46, 14, 12, 26, 40, 10, 10, 10, 8, 8, 16, 22, 14, 14, 12]):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("Info")
    s = summary or {}
    ws2.append(["Показатель", "Значение"])
    ws2.append(["Уценённых карточек", s.get("total")])
    ws2.append(["Уценили сами (FBS)", s.get("own")])
    ws2.append(["Уценил Ozon (FBO)", s.get("ozon")])
    ws2.append(["С остатком", s.get("with_stock")])
    ws2.append(["Штук в наличии", s.get("stock_qty")])
    ws2.append(["Средняя скидка, %", s.get("avg_discount")])
    ws2.append(["Стоимость остатка, ₽", s.get("stock_value")])
    ws2.append(["Сформировано", datetime.now(timezone.utc).isoformat()])
    ws2["A1"].font = Font(bold=True)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
