"""География заказов FBO/FBS: откуда отгружаем и куда уезжают заказы.

Аналог раздела «География FBS/FBW» на WB. Источник — отправления Ozon
(/v3/posting/fbo/list и /v4/posting/fbs/list) с analytics_data: там регион,
город и склад отгрузки. Кэш в памяти, синк по кнопке.
"""

from __future__ import annotations

import io
import logging
import threading
from datetime import datetime, timedelta, timezone
from typing import Callable

from sales_pace import MSK, fetch_fbo_postings, fetch_fbs_postings

logger = logging.getLogger("ozon-dashboard.orders_geo")

UNKNOWN = "Не указан"

GEO_CACHE: dict = {
    "rows": [],
    "updated_at": None,
    "syncing": False,
    "error": None,
    "days": 30,
    "since": None,
    "until": None,
    "count_fbo": 0,
    "count_fbs": 0,
}

_lock = threading.Lock()


def _analytics(p: dict) -> dict:
    a = p.get("analytics_data")
    return a if isinstance(a, dict) else {}


def _posting_dt(p: dict) -> datetime | None:
    raw = p.get("in_process_at") or p.get("created_at") or p.get("shipment_date")
    if not raw:
        return None
    try:
        d = datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
    except ValueError:
        return None
    if d.tzinfo is None:
        d = d.replace(tzinfo=timezone.utc)
    return d.astimezone(MSK)


def _warehouse_name(p: dict, channel: str) -> str:
    a = _analytics(p)
    dm = p.get("delivery_method")
    dm = dm if isinstance(dm, dict) else {}
    name = (
        a.get("warehouse_name")
        or a.get("warehouse")
        or dm.get("warehouse")
        or ""
    )
    name = str(name).strip()
    if not name:
        return f"Склад {channel}" if channel == "FBO" else UNKNOWN
    return name


def _to_float(v) -> float:
    try:
        return float(str(v).replace(",", ".") or 0)
    except (TypeError, ValueError):
        return 0.0


def _norm_rows(p: dict, channel: str) -> list[dict]:
    """Одна строка на товар отправления: заказы считаем по posting_number."""
    d = _posting_dt(p)
    if d is None:
        return []
    a = _analytics(p)
    region = str(a.get("region") or "").strip() or UNKNOWN
    city = str(a.get("city") or "").strip() or UNKNOWN
    status = str(p.get("status") or "").strip()
    base = {
        "posting": str(p.get("posting_number") or ""),
        "date": d.strftime("%Y-%m-%d"),
        "dt": d.strftime("%Y-%m-%d %H:%M:%S"),
        "channel": channel,
        "warehouse": _warehouse_name(p, channel),
        "region": region,
        "city": city,
        "delivery_type": str(a.get("delivery_type") or "").strip(),
        "status": status,
        "cancelled": status.lower() in ("cancelled", "canceled"),
    }
    out = []
    for prod in p.get("products") or []:
        qty = int(prod.get("quantity") or 1)
        price = _to_float(prod.get("price"))
        out.append({
            **base,
            "offer_id": str(prod.get("offer_id") or "").strip(),
            "sku": prod.get("sku"),
            "name": str(prod.get("name") or "").strip(),
            "qty": qty,
            "revenue": round(price * qty, 2),
        })
    if not out:
        out.append({**base, "offer_id": "", "sku": None, "name": "", "qty": 0, "revenue": 0.0})
    return out


def sync_orders_geo(ozon_post: Callable, days: int = 30) -> dict:
    if not _lock.acquire(blocking=False):
        GEO_CACHE["syncing"] = True
        return {"ok": False, "syncing": True}
    GEO_CACHE["syncing"] = True
    GEO_CACHE["error"] = None
    try:
        days = max(1, min(int(days or 30), 90))
        now = datetime.now(MSK)
        since = (now - timedelta(days=days)).replace(hour=0, minute=0, second=0, microsecond=0)
        fbo, fbs = [], []
        try:
            fbo = fetch_fbo_postings(ozon_post, since, now, with_analytics=True)
        except Exception as e:
            logger.warning("orders-geo FBO: %s", e)
        try:
            fbs = fetch_fbs_postings(ozon_post, since, now, with_analytics=True)
        except Exception as e:
            logger.warning("orders-geo FBS: %s", e)

        rows: list[dict] = []
        for p in fbo:
            rows.extend(_norm_rows(p, "FBO"))
        for p in fbs:
            rows.extend(_norm_rows(p, "FBS"))
        rows.sort(key=lambda r: r.get("dt") or "", reverse=True)

        GEO_CACHE.update({
            "rows": rows,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "syncing": False,
            "error": None,
            "days": days,
            "since": since.isoformat(),
            "until": now.isoformat(),
            "count_fbo": len(fbo),
            "count_fbs": len(fbs),
        })
        logger.info("orders-geo: %s строк (FBO %s / FBS %s)", len(rows), len(fbo), len(fbs))
        return {"ok": True, "count": len(rows)}
    except Exception as e:
        logger.exception("orders-geo sync")
        GEO_CACHE["error"] = str(e)
        GEO_CACHE["syncing"] = False
        return {"ok": False, "error": str(e)}
    finally:
        _lock.release()


def _match(row: dict, channel: str, warehouse: str, region: str, city: str, search: str) -> bool:
    if channel not in ("all", "") and (row.get("channel") or "") != channel:
        return False
    if warehouse not in ("all", "") and (row.get("warehouse") or "") != warehouse:
        return False
    if region not in ("all", "") and (row.get("region") or "") != region:
        return False
    if city not in ("all", "") and (row.get("city") or "") != city:
        return False
    if search:
        blob = " ".join([
            str(row.get("offer_id") or ""),
            str(row.get("sku") or ""),
            str(row.get("name") or ""),
            str(row.get("warehouse") or ""),
            str(row.get("region") or ""),
            str(row.get("city") or ""),
            str(row.get("posting") or ""),
        ]).lower()
        if search not in blob:
            return False
    return True


def aggregate(
    date_from: str | None = None,
    date_to: str | None = None,
    channel: str = "all",
    warehouse: str = "all",
    region: str = "all",
    city: str = "all",
    search: str = "",
    include_cancelled: bool = False,
) -> dict:
    rows = GEO_CACHE.get("rows") or []
    search = (search or "").strip().lower()
    channel = (channel or "all").upper()
    if channel not in ("FBO", "FBS"):
        channel = "all"
    warehouse = warehouse or "all"
    region = region or "all"
    city = city or "all"

    # опции селектов считаем по периоду, чтобы они не схлопывались при фильтрах
    scoped = []
    for r in rows:
        d = r.get("date") or ""
        if date_from and d < date_from:
            continue
        if date_to and d > date_to:
            continue
        if not include_cancelled and r.get("cancelled"):
            continue
        scoped.append(r)

    filtered = [r for r in scoped if _match(r, channel, warehouse, region, city, search)]

    by_day: dict[str, dict] = {}
    by_wh: dict[str, dict] = {}
    by_reg: dict[str, dict] = {}
    by_city: dict[str, dict] = {}
    by_art: dict[str, dict] = {}
    postings_all: set[str] = set()
    postings_fbo: set[str] = set()
    postings_fbs: set[str] = set()
    total_qty = 0
    total_rev = 0.0
    qty_fbo = qty_fbs = 0
    rev_fbo = rev_fbs = 0.0

    for r in filtered:
        ch = r.get("channel") or "FBO"
        qty = int(r.get("qty") or 0)
        rev = float(r.get("revenue") or 0)
        posting = r.get("posting") or ""
        day = r.get("date") or ""
        wh = r.get("warehouse") or UNKNOWN
        reg = r.get("region") or UNKNOWN
        cty = r.get("city") or UNKNOWN
        art = (r.get("offer_id") or "").strip() or (f"sku:{r.get('sku')}" if r.get("sku") else "—")

        total_qty += qty
        total_rev += rev
        if posting:
            postings_all.add(posting)
            (postings_fbo if ch == "FBO" else postings_fbs).add(posting)
        if ch == "FBO":
            qty_fbo += qty
            rev_fbo += rev
        else:
            qty_fbs += qty
            rev_fbs += rev

        dslot = by_day.setdefault(day, {
            "date": day, "postings": set(), "fbo_postings": set(), "fbs_postings": set(),
            "qty": 0, "fbo_qty": 0, "fbs_qty": 0, "revenue": 0.0,
        })
        dslot["qty"] += qty
        dslot["revenue"] += rev
        if posting:
            dslot["postings"].add(posting)
            (dslot["fbo_postings"] if ch == "FBO" else dslot["fbs_postings"]).add(posting)
        if ch == "FBO":
            dslot["fbo_qty"] += qty
        else:
            dslot["fbs_qty"] += qty

        wslot = by_wh.setdefault(wh, {
            "warehouse": wh, "channels": set(), "postings": set(),
            "qty": 0, "revenue": 0.0, "regions": {},
        })
        wslot["channels"].add(ch)
        wslot["qty"] += qty
        wslot["revenue"] += rev
        if posting:
            wslot["postings"].add(posting)
        wslot["regions"][reg] = wslot["regions"].get(reg, 0) + qty

        rslot = by_reg.setdefault(reg, {
            "region": reg, "postings": set(), "qty": 0, "fbo_qty": 0, "fbs_qty": 0,
            "revenue": 0.0, "cities": {},
        })
        rslot["qty"] += qty
        rslot["revenue"] += rev
        if posting:
            rslot["postings"].add(posting)
        if ch == "FBO":
            rslot["fbo_qty"] += qty
        else:
            rslot["fbs_qty"] += qty
        rslot["cities"][cty] = rslot["cities"].get(cty, 0) + qty

        cslot = by_city.setdefault(f"{cty}|{reg}", {
            "city": cty, "region": reg, "postings": set(), "qty": 0,
            "fbo_qty": 0, "fbs_qty": 0, "revenue": 0.0,
        })
        cslot["qty"] += qty
        cslot["revenue"] += rev
        if posting:
            cslot["postings"].add(posting)
        if ch == "FBO":
            cslot["fbo_qty"] += qty
        else:
            cslot["fbs_qty"] += qty

        aslot = by_art.setdefault(art, {
            "offer_id": art, "sku": r.get("sku"), "name": r.get("name") or "",
            "qty": 0, "fbo_qty": 0, "fbs_qty": 0, "revenue": 0.0, "cities": {}, "regions": {},
        })
        aslot["qty"] += qty
        aslot["revenue"] += rev
        if ch == "FBO":
            aslot["fbo_qty"] += qty
        else:
            aslot["fbs_qty"] += qty
        if not aslot["name"] and r.get("name"):
            aslot["name"] = r["name"]
        aslot["cities"][cty] = aslot["cities"].get(cty, 0) + qty
        aslot["regions"][reg] = aslot["regions"].get(reg, 0) + qty

    total_orders = len(postings_all)

    days_out = []
    for s in sorted(by_day.values(), key=lambda x: x["date"]):
        days_out.append({
            "date": s["date"],
            "orders": len(s["postings"]),
            "fbo_orders": len(s["fbo_postings"]),
            "fbs_orders": len(s["fbs_postings"]),
            "qty": s["qty"],
            "fbo_qty": s["fbo_qty"],
            "fbs_qty": s["fbs_qty"],
            "revenue": round(s["revenue"], 2),
        })

    wh_out = []
    for s in by_wh.values():
        top_regions = [x[0] for x in sorted(s["regions"].items(), key=lambda i: -i[1])[:3]]
        wh_out.append({
            "warehouse": s["warehouse"],
            "channel": "/".join(sorted(s["channels"])),
            "orders": len(s["postings"]),
            "qty": s["qty"],
            "revenue": round(s["revenue"], 2),
            "share_pct": round(s["qty"] / max(1, total_qty) * 100, 1),
            "top_regions": top_regions,
        })
    wh_out.sort(key=lambda x: -x["qty"])

    reg_out = []
    for s in by_reg.values():
        top_cities = [x[0] for x in sorted(s["cities"].items(), key=lambda i: -i[1])[:4]]
        reg_out.append({
            "region": s["region"],
            "orders": len(s["postings"]),
            "qty": s["qty"],
            "fbo_qty": s["fbo_qty"],
            "fbs_qty": s["fbs_qty"],
            "revenue": round(s["revenue"], 2),
            "share_pct": round(s["qty"] / max(1, total_qty) * 100, 1),
            "top_cities": top_cities,
        })
    reg_out.sort(key=lambda x: -x["qty"])

    city_out = []
    for s in by_city.values():
        city_out.append({
            "city": s["city"],
            "region": s["region"],
            "orders": len(s["postings"]),
            "qty": s["qty"],
            "fbo_qty": s["fbo_qty"],
            "fbs_qty": s["fbs_qty"],
            "revenue": round(s["revenue"], 2),
            "share_pct": round(s["qty"] / max(1, total_qty) * 100, 1),
        })
    city_out.sort(key=lambda x: -x["qty"])

    art_out = []
    for s in by_art.values():
        top_city = sorted(s["cities"].items(), key=lambda i: -i[1])[0][0] if s["cities"] else ""
        top_region = sorted(s["regions"].items(), key=lambda i: -i[1])[0][0] if s["regions"] else ""
        art_out.append({
            "offer_id": s["offer_id"],
            "sku": s["sku"],
            "name": s["name"],
            "qty": s["qty"],
            "fbo_qty": s["fbo_qty"],
            "fbs_qty": s["fbs_qty"],
            "revenue": round(s["revenue"], 2),
            "avg_price": round(s["revenue"] / max(1, s["qty"]), 2),
            "top_city": top_city,
            "top_region": top_region,
            "cities_count": len(s["cities"]),
        })
    art_out.sort(key=lambda x: -x["qty"])

    all_dates = sorted({r.get("date") for r in rows if r.get("date")})
    return {
        "summary": {
            "orders": total_orders,
            "qty": total_qty,
            "revenue": round(total_rev, 2),
            "avg_order_price": round(total_rev / max(1, total_orders), 2),
            "fbo_orders": len(postings_fbo),
            "fbo_qty": qty_fbo,
            "fbo_revenue": round(rev_fbo, 2),
            "fbo_share_pct": round(qty_fbo / max(1, total_qty) * 100, 1),
            "fbs_orders": len(postings_fbs),
            "fbs_qty": qty_fbs,
            "fbs_revenue": round(rev_fbs, 2),
            "fbs_share_pct": round(qty_fbs / max(1, total_qty) * 100, 1),
            "regions": len(reg_out),
            "cities": len(city_out),
            "warehouses": len(wh_out),
        },
        "by_day": days_out,
        "by_warehouse": wh_out[:60],
        "by_region": reg_out[:60],
        "by_city": city_out[:150],
        "by_article": art_out[:150],
        "filters": {
            "warehouses": sorted({r.get("warehouse") or UNKNOWN for r in scoped}),
            "regions": sorted({r.get("region") or UNKNOWN for r in scoped}),
            "cities": sorted({r.get("city") or UNKNOWN for r in scoped})[:400],
            "date_min": all_dates[0] if all_dates else None,
            "date_max": all_dates[-1] if all_dates else None,
            "date_from": date_from,
            "date_to": date_to,
            "channel": channel,
            "total_cached": len(rows),
        },
        "meta": {
            "updated_at": GEO_CACHE.get("updated_at"),
            "syncing": bool(GEO_CACHE.get("syncing")),
            "error": GEO_CACHE.get("error"),
            "days": GEO_CACHE.get("days"),
            "since": GEO_CACHE.get("since"),
            "until": GEO_CACHE.get("until"),
            "count_fbo": GEO_CACHE.get("count_fbo") or 0,
            "count_fbs": GEO_CACHE.get("count_fbs") or 0,
            "include_cancelled": bool(include_cancelled),
            "rows_filtered": len(filtered),
        },
    }


def build_xlsx_bytes(agg: dict) -> bytes:
    """Выгрузка географии в Excel: сводка + регионы, города, склады, артикулы, дни."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError("На сервере нет openpyxl — добавь в requirements.txt и задеплой") from e

    wb = Workbook()
    s = agg.get("summary") or {}
    f = agg.get("filters") or {}
    m = agg.get("meta") or {}

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Показатель", "Значение"])
    ws.append(["Период", f"{f.get('date_from') or '—'} — {f.get('date_to') or '—'}"])
    ws.append(["Канал", f.get("channel") or "all"])
    ws.append(["Заказов (отправлений)", s.get("orders")])
    ws.append(["Товаров, шт", s.get("qty")])
    ws.append(["Сумма, ₽", s.get("revenue")])
    ws.append(["Средний чек, ₽", s.get("avg_order_price")])
    ws.append(["FBO: заказов / шт / ₽", f"{s.get('fbo_orders')} / {s.get('fbo_qty')} / {s.get('fbo_revenue')}"])
    ws.append(["FBS: заказов / шт / ₽", f"{s.get('fbs_orders')} / {s.get('fbs_qty')} / {s.get('fbs_revenue')}"])
    ws.append(["Регионов / городов / складов", f"{s.get('regions')} / {s.get('cities')} / {s.get('warehouses')}"])
    ws.append(["Синк", m.get("updated_at")])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 42

    ws2 = wb.create_sheet("Regions")
    ws2.append(["Регион", "Заказов", "Штук", "FBO шт", "FBS шт", "Сумма, ₽", "Доля, %", "Топ города"])
    for r in agg.get("by_region") or []:
        ws2.append([
            r.get("region"), r.get("orders"), r.get("qty"), r.get("fbo_qty"), r.get("fbs_qty"),
            r.get("revenue"), r.get("share_pct"), ", ".join(r.get("top_cities") or []),
        ])

    ws3 = wb.create_sheet("Cities")
    ws3.append(["Город", "Регион", "Заказов", "Штук", "FBO шт", "FBS шт", "Сумма, ₽", "Доля, %"])
    for r in agg.get("by_city") or []:
        ws3.append([
            r.get("city"), r.get("region"), r.get("orders"), r.get("qty"),
            r.get("fbo_qty"), r.get("fbs_qty"), r.get("revenue"), r.get("share_pct"),
        ])

    ws4 = wb.create_sheet("Warehouses")
    ws4.append(["Склад отгрузки", "Канал", "Заказов", "Штук", "Сумма, ₽", "Доля, %", "Топ регионы"])
    for r in agg.get("by_warehouse") or []:
        ws4.append([
            r.get("warehouse"), r.get("channel"), r.get("orders"), r.get("qty"),
            r.get("revenue"), r.get("share_pct"), ", ".join(r.get("top_regions") or []),
        ])

    ws5 = wb.create_sheet("Articles")
    ws5.append(["Артикул", "SKU", "Название", "Штук", "FBO шт", "FBS шт", "Сумма, ₽", "Ср. цена", "Топ город", "Топ регион", "Городов"])
    for r in agg.get("by_article") or []:
        ws5.append([
            r.get("offer_id"), r.get("sku"), r.get("name"), r.get("qty"), r.get("fbo_qty"),
            r.get("fbs_qty"), r.get("revenue"), r.get("avg_price"), r.get("top_city"),
            r.get("top_region"), r.get("cities_count"),
        ])

    ws6 = wb.create_sheet("By_days")
    ws6.append(["Дата", "Заказов", "FBO заказов", "FBS заказов", "Штук", "FBO шт", "FBS шт", "Сумма, ₽"])
    for r in agg.get("by_day") or []:
        ws6.append([
            r.get("date"), r.get("orders"), r.get("fbo_orders"), r.get("fbs_orders"),
            r.get("qty"), r.get("fbo_qty"), r.get("fbs_qty"), r.get("revenue"),
        ])

    for sheet, widths in (
        (ws2, [30, 10, 10, 10, 10, 14, 10, 46]),
        (ws3, [26, 28, 10, 10, 10, 10, 14, 10]),
        (ws4, [34, 10, 10, 10, 14, 10, 46]),
        (ws5, [24, 14, 46, 10, 10, 10, 14, 12, 22, 26, 10]),
        (ws6, [12, 10, 12, 12, 10, 10, 10, 14]),
    ):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for col, w in zip("ABCDEFGHIJK", widths):
            sheet.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
