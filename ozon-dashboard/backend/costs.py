"""Себестоимость из Excel (формат cost_price) — матчинг по артикулу и SKU Ozon."""

from __future__ import annotations

import io
import logging
from datetime import date, datetime, timedelta, timezone
logger = logging.getLogger("ozon-dashboard.costs")

COST_PRICES_KEY = "ozon_cost_prices"
COST_META_KEY = "ozon_cost_prices_meta"


def _parse_cost_number(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v) if float(v) >= 0 else None
    s = str(v).strip().replace("\xa0", " ").replace(" ", "").replace(",", ".")
    s = s.replace("₽", "").replace("руб.", "").replace("руб", "")
    if not s or s.lower() in ("nan", "none", "-", "—"):
        return None
    try:
        n = float(s)
        return n if n >= 0 else None
    except Exception:
        return None


def _norm_offer(v) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    return s.replace("\u041e", "O").replace("\u043e", "o")


def _parse_header_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or s.lower() in ("по умолчанию", "default", "sku", "артикул", "наименование", "размер"):
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            pass
    try:
        n = float(s)
        if 30000 < n < 60000:
            return date(1899, 12, 30) + timedelta(days=int(n))
    except Exception:
        pass
    return None


def _effective_cost(default_cost, dated_costs, as_of=None):
    as_of = as_of or datetime.now(timezone.utc).date()
    applicable = [(d, c) for d, c in dated_costs if d is not None and d <= as_of and c is not None]
    if applicable:
        d, c = max(applicable, key=lambda x: x[0])
        return c, d.isoformat()
    if default_cost is not None:
        return default_cost, None
    return None, None


def parse_cost_price_workbook(contents: bytes, as_of=None) -> tuple[dict, dict]:
    """
    Лист «Себестоимость»:
      row1: SKU | Артикул | … | По умолчанию | даты…
      row2:          …        | Себестоимость | …
    Возвращает by_offer / by_sku + meta.
    """
    from openpyxl import load_workbook

    as_of = as_of or datetime.now(timezone.utc).date()
    wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    ws = None
    for name in wb.sheetnames:
        if "себестоим" in name.lower() or "cost" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        h1 = next(rows_iter)
        h2 = next(rows_iter)
    except StopIteration:
        wb.close()
        return {"by_offer": {}, "by_sku": {}}, {"error": "Пустой файл"}

    cost_cols = []
    for i, (top, sub) in enumerate(zip(h1, h2)):
        sub_l = str(sub or "").strip().lower()
        top_s = str(top or "").strip().lower()
        if "фулфил" in sub_l or "fulfill" in sub_l:
            continue
        has_cost = "себестоим" in sub_l or "себестоим" in top_s or "cost" in sub_l
        is_default_top = "умолчан" in top_s
        d = _parse_header_date(top)
        if has_cost or (is_default_top and ("себестоим" in sub_l or sub_l in ("", "none", "nan"))):
            if d is None and not is_default_top and not has_cost:
                continue
            if d is None and not is_default_top and "умолчан" not in top_s:
                if not has_cost:
                    continue
            cost_cols.append((i, d if not is_default_top else None))

    if not cost_cols:
        for i, top in enumerate(h1):
            top_s = str(top or "").strip().lower()
            d = _parse_header_date(top)
            if "умолчан" in top_s:
                prev_same = i > 0 and str(h1[i - 1] or "").strip().lower() == top_s
                if prev_same:
                    continue
                cost_cols.append((i, None))
            elif d is not None:
                prev_d = _parse_header_date(h1[i - 1]) if i > 0 else None
                if prev_d == d:
                    continue
                cost_cols.append((i, d))

    offer_col = 1
    sku_col = 0
    for i, top in enumerate(h1):
        t = str(top or "").strip().lower()
        if t in ("артикул", "артикул продавца", "vendorcode", "vendor_code", "offer_id"):
            offer_col = i
        if t in ("sku", "ozon sku", "sku ozon", "product_id"):
            sku_col = i

    rows_list = list(rows_iter)
    default_idxs = [i for i, d in cost_cols if d is None]
    dated_idxs = [(i, d) for i, d in cost_cols if d is not None]

    by_offer: dict[str, dict] = {}
    by_sku: dict[str, dict] = {}
    for row in rows_list:
        if not row or offer_col >= len(row):
            continue
        offer = _norm_offer(row[offer_col])
        if not offer or offer.lower() in ("артикул", "nan", "none"):
            continue
        sku = None
        if sku_col is not None and sku_col < len(row) and row[sku_col] not in (None, ""):
            try:
                raw = str(row[sku_col]).strip()
                if raw.replace(".", "", 1).isdigit():
                    sku_i = int(float(raw))
                    if sku_i >= 10000:
                        sku = str(sku_i)
            except Exception:
                sku = None

        default_cost = None
        for i in default_idxs:
            if i < len(row):
                c = _parse_cost_number(row[i])
                if c is not None:
                    default_cost = c
                    break
        dated = []
        for i, d in dated_idxs:
            if i < len(row) and row[i] not in (None, ""):
                c = _parse_cost_number(row[i])
                if c is not None:
                    dated.append((d, c))
        eff, as_of_used = _effective_cost(default_cost, dated, as_of)
        if eff is None:
            continue
        entry = {
            "cost": round(eff, 4),
            "default": round(default_cost, 4) if default_cost is not None else None,
            "as_of": as_of_used,
            "offer_id": offer,
            "sku": sku,
        }
        by_offer[offer] = entry
        if sku:
            by_sku[sku] = entry

    wb.close()
    return {"by_offer": by_offer, "by_sku": by_sku}, {
        "format": "dated_cost_matrix",
        "default_cols": len(default_idxs),
        "date_cols": len(dated_idxs),
        "as_of": as_of.isoformat(),
        "offers": len(by_offer),
        "skus": len(by_sku),
    }


def load_cost_indexes(get_setting) -> tuple[dict[str, dict], dict[str, dict], dict]:
    raw = get_setting(COST_PRICES_KEY, "{}") or "{}"
    meta = get_setting(COST_META_KEY, "{}") or "{}"
    try:
        data = raw if isinstance(raw, dict) else __import__("json").loads(raw)
    except Exception:
        data = {}
    try:
        meta_d = meta if isinstance(meta, dict) else __import__("json").loads(meta)
    except Exception:
        meta_d = {}
    by_offer = {}
    by_sku = {}
    for k, e in (data.get("by_offer") or {}).items():
        offer = _norm_offer(k)
        if offer and isinstance(e, dict):
            by_offer[offer] = e
    for k, e in (data.get("by_sku") or {}).items():
        sku = str(k).strip()
        if sku and isinstance(e, dict):
            by_sku[sku] = e
    # плоский формат {offer: cost}
    if not by_offer and isinstance(data, dict) and "by_offer" not in data:
        for k, v in data.items():
            offer = _norm_offer(k)
            if not offer:
                continue
            if isinstance(v, dict):
                by_offer[offer] = v
            else:
                c = _parse_cost_number(v)
                if c is not None:
                    by_offer[offer] = {"cost": c, "offer_id": offer}
    return by_offer, by_sku, meta_d


def resolve_cost(by_offer: dict, by_sku: dict, offer_id=None, sku=None) -> dict:
    offer = _norm_offer(offer_id)
    if offer and offer in by_offer:
        return by_offer[offer]
    if sku is not None:
        s = str(sku).strip()
        if s in by_sku:
            return by_sku[s]
    if offer:
        low = offer.lower()
        for k, e in by_offer.items():
            if k.lower() == low:
                return e
    return {}


def save_costs(save_setting, by_offer: dict, by_sku: dict, meta: dict | None = None) -> dict:
    import json

    payload = {"by_offer": by_offer, "by_sku": by_sku}
    save_setting(COST_PRICES_KEY, json.dumps(payload, ensure_ascii=False))
    meta = dict(meta or {})
    meta["updated_at"] = datetime.now(timezone.utc).isoformat()
    meta["offers"] = len(by_offer)
    meta["skus"] = len(by_sku)
    save_setting(COST_META_KEY, json.dumps(meta, ensure_ascii=False))
    return meta
