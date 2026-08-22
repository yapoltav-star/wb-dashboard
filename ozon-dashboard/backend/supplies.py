"""FBO поставки Ozon — список по статусам + Excel состава «на точке отгрузки»."""

from __future__ import annotations

import io
import logging
import threading
import time
from datetime import datetime, timezone
from typing import Callable

logger = logging.getLogger("ozon-dashboard.supplies")

# статусы из /v3/supply-order/get (без префикса ORDER_STATE_)
STATE_LABELS: dict[str, str] = {
    "DATA_FILLING": "Заполнение данных",
    "READY_TO_SUPPLY": "Готова к отгрузке",
    "ACCEPTED_AT_SUPPLY_WAREHOUSE": "На точке отгрузки",
    "IN_TRANSIT": "В пути",
    "ACCEPTANCE_AT_STORAGE_WAREHOUSE": "Приёмка на складе",
    "REPORTS_CONFIRMATION_AWAITING": "Согласование актов",
    "REPORT_REJECTED": "Спор",
    "COMPLETED": "Завершена",
    "REJECTED_AT_SUPPLY_WAREHOUSE": "Отказ на точке",
    "CANCELLED": "Отменена",
    "OVERDUE": "Просрочена",
    "UNSPECIFIED": "Не указан",
}

# активные / «отгруженные» — то, что обычно нужно видеть
ACTIVE_STATES = [
    "DATA_FILLING",
    "READY_TO_SUPPLY",
    "ACCEPTED_AT_SUPPLY_WAREHOUSE",
    "IN_TRANSIT",
    "ACCEPTANCE_AT_STORAGE_WAREHOUSE",
    "REPORTS_CONFIRMATION_AWAITING",
    "REPORT_REJECTED",
]

AT_DROPOFF = "ACCEPTED_AT_SUPPLY_WAREHOUSE"

SUPPLIES_CACHE: dict = {
    "orders": [],
    "by_state": {},
    "counters": [],
    "updated_at": None,
    "syncing": False,
    "error": None,
}

_lock = threading.Lock()


def _norm_state(raw) -> str:
    s = str(raw or "").strip().upper()
    if s.startswith("ORDER_STATE_"):
        s = s[len("ORDER_STATE_") :]
    return s or "UNSPECIFIED"


def state_label(state: str) -> str:
    return STATE_LABELS.get(_norm_state(state), state or "—")


def fetch_status_counters(ozon_post: Callable) -> list[dict]:
    payload = ozon_post("/v1/supply-order/status/counter", {})
    items = payload.get("items") or []
    out = []
    for it in items:
        if not isinstance(it, dict):
            continue
        st = _norm_state(it.get("order_state") or it.get("state"))
        out.append({
            "state": st,
            "label": state_label(st),
            "count": int(it.get("count") or 0),
        })
    out.sort(key=lambda x: (-x["count"], x["label"]))
    return out


def list_order_ids(ozon_post: Callable, states: list[str] | None = None, max_pages: int = 40) -> list[str]:
    """POST /v3/supply-order/list → order_ids."""
    states = [_norm_state(s) for s in (states or ACTIVE_STATES)]
    # API иногда ждёт с префиксом — пробуем короткие имена как в доке
    last_id = ""
    ids: list[str] = []
    seen = set()
    for _ in range(max_pages):
        body: dict = {
            "filter": {"states": states},
            "limit": 100,
            "sort_by": "ORDER_STATE_UPDATED_AT",
            "sort_dir": "DESC",
        }
        if last_id and last_id != "null":
            body["last_id"] = last_id
        try:
            payload = ozon_post("/v3/supply-order/list", body)
        except Exception as e:
            # fallback: с префиксом ORDER_STATE_
            logger.warning("supply-order/list short states failed: %s — retry prefixed", e)
            body["filter"]["states"] = [f"ORDER_STATE_{s}" if not s.startswith("ORDER_STATE_") else s for s in states]
            payload = ozon_post("/v3/supply-order/list", body)
        batch = payload.get("order_ids") or []
        for oid in batch:
            key = str(oid)
            if key not in seen:
                seen.add(key)
                ids.append(key)
        new_last = payload.get("last_id") or ""
        if not batch or not new_last or new_last == last_id:
            break
        last_id = str(new_last)
        time.sleep(0.12)
    return ids


def get_orders(ozon_post: Callable, order_ids: list[str]) -> list[dict]:
    """POST /v3/supply-order/get — пачками до 50."""
    out: list[dict] = []
    cleaned = [str(x) for x in order_ids if x is not None and str(x).strip()]
    for i in range(0, len(cleaned), 50):
        chunk = cleaned[i : i + 50]
        if not chunk:
            continue
        # API принимает и int и string — шлём как есть
        ids_payload: list = []
        for x in chunk:
            try:
                ids_payload.append(int(x))
            except (TypeError, ValueError):
                ids_payload.append(x)
        payload = ozon_post("/v3/supply-order/get", {"order_ids": ids_payload})
        orders = payload.get("orders") or []
        out.extend(orders)
        time.sleep(0.1)
    return out


def fetch_bundle_items(ozon_post: Callable, bundle_id: str, max_pages: int = 50) -> list[dict]:
    if not bundle_id:
        return []
    items: list[dict] = []
    last_id = ""
    for _ in range(max_pages):
        body: dict = {
            "bundle_ids": [str(bundle_id)],
            "limit": 100,
            "sort_field": "SKU",
            "is_asc": True,
        }
        if last_id:
            body["last_id"] = last_id
        payload = ozon_post("/v1/supply-order/bundle", body)
        batch = payload.get("items")
        if batch is None and isinstance(payload.get("result"), dict):
            batch = payload["result"].get("items")
        if batch is None and isinstance(payload.get("result"), list):
            batch = payload["result"]
        batch = batch or []
        items.extend(batch)
        has_next = bool(payload.get("has_next"))
        if isinstance(payload.get("result"), dict):
            has_next = has_next or bool(payload["result"].get("has_next"))
        if not has_next or not batch:
            break
        last_id = payload.get("last_id") or ""
        if isinstance(payload.get("result"), dict) and not last_id:
            last_id = payload["result"].get("last_id") or ""
        if not last_id:
            break
        time.sleep(0.08)
    return items


def fetch_order_details_items(ozon_post: Callable, order_id) -> list[dict]:
    """Fallback: /v1/supply-order/details — если bundle пустой."""
    try:
        oid = int(order_id) if str(order_id).isdigit() else order_id
        payload = ozon_post("/v1/supply-order/details", {"order_ids": [oid]})
    except Exception as e:
        logger.warning("supply-order/details %s: %s", order_id, e)
        return []
    orders = payload.get("orders") or payload.get("result") or []
    if isinstance(orders, dict):
        orders = [orders]
    out = []
    for order in orders:
        if not isinstance(order, dict):
            continue
        for supply in order.get("supplies") or []:
            if not isinstance(supply, dict):
                continue
            for it in supply.get("items") or supply.get("products") or []:
                if isinstance(it, dict):
                    out.append(it)
            bundle = supply.get("bundle") or {}
            if isinstance(bundle, dict):
                for it in bundle.get("items") or []:
                    if isinstance(it, dict):
                        out.append(it)
        for it in order.get("items") or order.get("products") or []:
            if isinstance(it, dict):
                out.append(it)
    return out


def _summarize_order(order: dict) -> dict:
    state = _norm_state(order.get("state"))
    drop = order.get("dropoff_warehouse") or {}
    timeslot = ((order.get("timeslot") or {}).get("timeslot") or {})
    supplies_raw = order.get("supplies") or []
    supplies = []
    bundle_ids = []
    for s in supplies_raw:
        if not isinstance(s, dict):
            continue
        wh = s.get("storage_warehouse") or {}
        bid = s.get("bundle_id")
        if bid:
            bundle_ids.append(str(bid))
        supplies.append({
            "supply_id": s.get("supply_id"),
            "bundle_id": bid,
            "state": _norm_state(s.get("state")),
            "is_crossdock": bool(s.get("is_crossdock")),
            "warehouse_id": wh.get("warehouse_id"),
            "warehouse_name": wh.get("name") or "",
            "warehouse_address": wh.get("address") or "",
            "arrival_date": wh.get("arrival_date"),
        })
    return {
        "order_id": order.get("order_id"),
        "order_number": order.get("order_number") or str(order.get("order_id") or ""),
        "state": state,
        "state_label": state_label(state),
        "created_date": order.get("created_date"),
        "state_updated_date": order.get("state_updated_date"),
        "dropoff_warehouse_id": drop.get("warehouse_id"),
        "dropoff_name": drop.get("name") or "",
        "dropoff_address": drop.get("address") or "",
        "timeslot_from": timeslot.get("from"),
        "timeslot_to": timeslot.get("to"),
        "supplies": supplies,
        "bundle_ids": bundle_ids,
        "supplies_count": len(supplies),
    }


def sync_supplies(ozon_post: Callable, states: list[str] | None = None) -> dict:
    if not _lock.acquire(blocking=False):
        SUPPLIES_CACHE["syncing"] = True
        return {"ok": False, "syncing": True}
    SUPPLIES_CACHE["syncing"] = True
    SUPPLIES_CACHE["error"] = None
    try:
        counters = []
        try:
            counters = fetch_status_counters(ozon_post)
        except Exception as e:
            logger.warning("status/counter: %s", e)

        want = states or (ACTIVE_STATES + ["COMPLETED"])
        # не тащим отменённые по умолчанию — слишком много шума
        order_ids = list_order_ids(ozon_post, want)
        orders_raw = get_orders(ozon_post, order_ids) if order_ids else []
        orders = [_summarize_order(o) for o in orders_raw if isinstance(o, dict)]
        orders.sort(
            key=lambda o: (
                0 if o.get("state") == AT_DROPOFF else 1,
                str(o.get("state_updated_date") or ""),
            ),
            reverse=True,
        )

        by_state: dict[str, list] = {}
        for o in orders:
            by_state.setdefault(o["state"], []).append(o)

        SUPPLIES_CACHE.update({
            "orders": orders,
            "by_state": {k: len(v) for k, v in by_state.items()},
            "counters": counters,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "syncing": False,
            "error": None,
        })
        return {"ok": True, "orders": len(orders), "counters": len(counters)}
    except Exception as e:
        logger.exception("supplies sync")
        SUPPLIES_CACHE["error"] = str(e)
        SUPPLIES_CACHE["syncing"] = False
        return {"ok": False, "error": str(e)}
    finally:
        _lock.release()


def get_cached() -> dict:
    return {
        "orders": SUPPLIES_CACHE.get("orders") or [],
        "by_state": SUPPLIES_CACHE.get("by_state") or {},
        "counters": SUPPLIES_CACHE.get("counters") or [],
        "updated_at": SUPPLIES_CACHE.get("updated_at"),
        "syncing": bool(SUPPLIES_CACHE.get("syncing")),
        "error": SUPPLIES_CACHE.get("error"),
        "states": [{"state": k, "label": v} for k, v in STATE_LABELS.items()],
        "at_dropoff_state": AT_DROPOFF,
    }


def collect_products_for_state(ozon_post: Callable, state: str = AT_DROPOFF) -> dict:
    """Товары из всех поставок в статусе state — для Excel одним файлом."""
    state = _norm_state(state)
    orders = [
        o for o in (SUPPLIES_CACHE.get("orders") or [])
        if _norm_state(o.get("state")) == state
    ]
    if not orders and not SUPPLIES_CACHE.get("syncing"):
        try:
            sync_supplies(ozon_post, states=[state])
        except Exception as e:
            logger.warning("sync for export: %s", e)
        orders = [
            o for o in (SUPPLIES_CACHE.get("orders") or [])
            if _norm_state(o.get("state")) == state
        ]

    rows: list[dict] = []
    errors: list[str] = []
    for order in orders:
        supplies = order.get("supplies") or [{}]
        got_any = False
        for supply in supplies:
            if not isinstance(supply, dict):
                continue
            bid = supply.get("bundle_id")
            items: list[dict] = []
            if bid:
                try:
                    items = fetch_bundle_items(ozon_post, str(bid))
                except Exception as e:
                    msg = f"bundle {bid}: {e}"
                    logger.warning(msg)
                    errors.append(msg[:300])
                    items = []
            if not items:
                try:
                    items = fetch_order_details_items(ozon_post, order.get("order_id"))
                except Exception as e:
                    errors.append(f"details {order.get('order_id')}: {e}"[:300])
                    items = []
            if not items:
                continue
            got_any = True
            for it in items:
                if not isinstance(it, dict):
                    continue
                try:
                    qty = int(float(it.get("quantity") or it.get("qty") or 0))
                except (TypeError, ValueError):
                    qty = 0
                rows.append({
                    "order_id": order.get("order_id"),
                    "order_number": order.get("order_number"),
                    "order_state": order.get("state"),
                    "order_state_label": order.get("state_label") or state_label(state),
                    "supply_id": supply.get("supply_id"),
                    "bundle_id": bid,
                    "dropoff_name": order.get("dropoff_name") or "",
                    "warehouse_name": supply.get("warehouse_name") or "",
                    "offer_id": it.get("offer_id") or it.get("article") or "",
                    "sku": it.get("sku") or it.get("sku_id"),
                    "name": it.get("name") or it.get("product_name") or "",
                    "quantity": qty,
                    "barcode": it.get("barcode") or "",
                    "product_id": it.get("product_id"),
                    "volume_in_litres": it.get("volume_in_litres"),
                    "total_volume_in_litres": it.get("total_volume_in_litres"),
                    "created_date": order.get("created_date"),
                    "state_updated_date": order.get("state_updated_date"),
                    "timeslot_from": order.get("timeslot_from"),
                })
            time.sleep(0.05)
        if not got_any:
            # ещё одна попытка details на всю заявку
            try:
                items = fetch_order_details_items(ozon_post, order.get("order_id"))
            except Exception:
                items = []
            for it in items:
                if not isinstance(it, dict):
                    continue
                try:
                    qty = int(float(it.get("quantity") or 0))
                except (TypeError, ValueError):
                    qty = 0
                rows.append({
                    "order_id": order.get("order_id"),
                    "order_number": order.get("order_number"),
                    "order_state": order.get("state"),
                    "order_state_label": order.get("state_label") or state_label(state),
                    "supply_id": None,
                    "bundle_id": None,
                    "dropoff_name": order.get("dropoff_name") or "",
                    "warehouse_name": "",
                    "offer_id": it.get("offer_id") or "",
                    "sku": it.get("sku"),
                    "name": it.get("name") or "",
                    "quantity": qty,
                    "barcode": it.get("barcode") or "",
                    "product_id": it.get("product_id"),
                    "volume_in_litres": it.get("volume_in_litres"),
                    "total_volume_in_litres": it.get("total_volume_in_litres"),
                    "created_date": order.get("created_date"),
                    "state_updated_date": order.get("state_updated_date"),
                    "timeslot_from": order.get("timeslot_from"),
                })

    agg: dict[str, dict] = {}
    for r in rows:
        key = str(r.get("offer_id") or r.get("sku") or "")
        if not key:
            continue
        a = agg.setdefault(key, {
            "offer_id": r.get("offer_id") or "",
            "sku": r.get("sku"),
            "name": r.get("name") or "",
            "barcode": r.get("barcode") or "",
            "quantity": 0,
            "orders": set(),
        })
        a["quantity"] += int(r.get("quantity") or 0)
        if r.get("order_number"):
            a["orders"].add(str(r["order_number"]))
        if not a.get("name") and r.get("name"):
            a["name"] = r["name"]
        if not a.get("sku") and r.get("sku"):
            a["sku"] = r["sku"]

    summary = []
    for a in agg.values():
        summary.append({
            "offer_id": a["offer_id"],
            "sku": a["sku"],
            "name": a["name"],
            "barcode": a["barcode"],
            "quantity": a["quantity"],
            "orders_count": len(a["orders"]),
            "order_numbers": ", ".join(sorted(a["orders"])),
        })
    summary.sort(key=lambda x: (-int(x.get("quantity") or 0), str(x.get("offer_id") or "")))

    return {
        "state": state,
        "state_label": state_label(state),
        "orders_count": len(orders),
        "rows": rows,
        "summary": summary,
        "errors": errors[:20],
    }


def build_xlsx_bytes(payload: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError("На сервере нет openpyxl — добавь в requirements.txt и задеплой") from e

    wb = Workbook()
    # ASCII sheet titles — надёжнее в некоторых окружениях
    ws1 = wb.active
    ws1.title = "Summary"
    # «Количество» — чтобы WB «Наш склад» принял файл как shk-отгрузку
    ws1.append(["Артикул", "SKU", "Название", "Штрихкод", "Количество", "Поставок", "Номера поставок"])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for r in payload.get("summary") or []:
        ws1.append([
            r.get("offer_id"),
            r.get("sku"),
            r.get("name"),
            r.get("barcode"),
            r.get("quantity"),
            r.get("orders_count"),
            r.get("order_numbers"),
        ])
    for col, w in zip("ABCDEFG", [22, 14, 48, 16, 10, 10, 36]):
        ws1.column_dimensions[col].width = w

    ws2 = wb.create_sheet("By_orders")
    ws2.append([
        "Номер заявки", "ID заявки", "Статус", "Supply ID",
        "Точка отгрузки", "Склад назначения",
        "Артикул", "SKU", "Название", "Количество", "Штрихкод",
        "Создана", "Статус обновлён", "Таймслот с",
    ])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in payload.get("rows") or []:
        ws2.append([
            r.get("order_number"),
            r.get("order_id"),
            r.get("order_state_label"),
            r.get("supply_id"),
            r.get("dropoff_name"),
            r.get("warehouse_name"),
            r.get("offer_id"),
            r.get("sku"),
            r.get("name"),
            r.get("quantity"),
            r.get("barcode"),
            r.get("created_date"),
            r.get("state_updated_date"),
            r.get("timeslot_from"),
        ])
    for col, w in zip("ABCDEFGHIJKLMN", [16, 12, 18, 12, 28, 28, 20, 14, 40, 10, 16, 20, 20, 20]):
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet("Info")
    ws3.append(["Status", payload.get("state_label"), payload.get("state")])
    ws3.append(["Orders", payload.get("orders_count")])
    ws3.append(["Detail rows", len(payload.get("rows") or [])])
    ws3.append(["Unique offers", len(payload.get("summary") or [])])
    ws3.append(["Generated", datetime.now(timezone.utc).isoformat()])
    if payload.get("errors"):
        ws3.append(["API errors", "; ".join(payload["errors"])[:2000]])
    if not (payload.get("rows") or []):
        ws3.append(["Note", "No items: check status / bundle_id / API rights"])
    ws3["A1"].font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
