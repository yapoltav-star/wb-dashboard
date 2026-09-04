import httpx
import os
import io
import json
import gzip
import base64
import time
import logging
import re
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone, date
from apscheduler.schedulers.background import BackgroundScheduler
from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
import pandas as pd
from pathlib import Path

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

WB_TOKEN = os.getenv("WB_TOKEN", "")
SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")
WB_FEEDBACKS_URL = "https://feedbacks-api.wildberries.ru"
WB_ANALYTICS_URL = "https://seller-analytics-api.wildberries.ru"
WB_STATISTICS_URL = "https://statistics-api.wildberries.ru"
WB_SUPPLIES_URL = "https://supplies-api.wildberries.ru"
WB_PROMOTION_URL = "https://advert-api.wildberries.ru"
WB_CALENDAR_URL = "https://dp-calendar-api.wildberries.ru"
WB_CONTENT_URL = "https://content-api.wildberries.ru"
WB_PRICES_URL = "https://discounts-prices-api.wildberries.ru"
WB_MARKETPLACE_URL = "https://marketplace-api.wildberries.ru"
WB_CHAT_URL = "https://buyer-chat-api.wildberries.ru"
WB_FINANCE_URL = "https://finance-api.wildberries.ru"
WB_CHAT_AUTOREPLY_KEY = "wb_chat_autoreply"
WB_CHAT_DEFAULT_TEXT = "Здравствуйте! Сообщение получено, ответим в ближайшее время."
WB_CHAT_REPLIED_KEEP = 2500
_WB_CHAT_LOCK = threading.Lock()
_WB_CHAT_RUNNING = False

# Team CRM — задачи «прокачать полки» менеджерам (Афина / Заира)
CRM_API_URL = (os.getenv("CRM_API_URL") or os.getenv("TEAM_CRM_URL") or "").rstrip("/")
CRM_PASSWORD = os.getenv("CRM_PASSWORD") or os.getenv("CRM_WEB_PASSWORD") or ""
CRM_MANAGER_ALIASES = {
    "afina": ("афина", "афине", "afina"),
    "zaira": ("заира", "заире", "zaira"),
}

# Спец-строки в ответе WB warehouse_remains, которые на самом деле не склады,
# а агрегаты — переносим их в отдельные поля stock_totals вместо списка складов.
STOCK_SPECIAL_FIELDS = {
    "В пути до получателей": "in_way_to_client",
    "В пути возвраты на склад WB": "in_way_from_client",
    "Всего находится на складах": "quantity_warehouses_full",
}

def sb_headers():
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates"
    }

def wb_headers():
    return {"Authorization": WB_TOKEN}

def upsert_feedbacks(feedbacks: list):
    if not feedbacks:
        return 0
    url = f"{SUPABASE_URL}/rest/v1/feedbacks"
    resp = httpx.post(url, json=feedbacks, headers=sb_headers(), timeout=30)
    if not resp.is_success:
        logger.error(f"Supabase upsert error: {resp.status_code} {resp.text[:200]}")
        return 0
    return len(feedbacks)

def fetch_feedbacks_page(is_answered: bool, skip: int):
    resp = httpx.get(
        f"{WB_FEEDBACKS_URL}/api/v1/feedbacks",
        headers=wb_headers(),
        params={"isAnswered": str(is_answered).lower(), "take": 5000, "skip": skip, "order": "dateDesc"},
        timeout=30
    )
    if not resp.is_success:
        logger.error(f"WB API error {resp.status_code}")
        return []
    return resp.json().get("data", {}).get("feedbacks", [])

def fetch_archive_page(skip: int):
    resp = httpx.get(
        f"{WB_FEEDBACKS_URL}/api/v1/feedbacks/archive",
        headers=wb_headers(),
        params={"isAnswered": "true", "take": 5000, "skip": skip, "order": "dateDesc"},
        timeout=30
    )
    if not resp.is_success:
        logger.error(f"WB archive error {resp.status_code}")
        return []
    return resp.json().get("data", {}).get("feedbacks", [])

def is_supplemented(f: dict) -> bool:
    return bool(f.get("isEdited") or f.get("supplementedFeedbackId"))

def process_feedback(f: dict, nm_to_vendor: dict = None) -> dict:
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(days=365)
    pd_data = f.get("productDetails", {})
    nm_id = pd_data.get("nmId")
    supplier_article = pd_data.get("supplierArticle", "")
    # Если WB не вернул supplierArticle — берём из нашей карты nm_id→vendor_code
    # чтобы не хранить отзыв как "208715116" вместо "000Braslet1"
    if not supplier_article and nm_id and nm_to_vendor:
        supplier_article = nm_to_vendor.get(nm_id, "")
    article = supplier_article or str(nm_id or "")
    date_str = f.get("createdDate") or f.get("updatedDate") or ""
    try:
        date = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
    except:
        date = now
    return {
        "id": f.get("id", ""),
        "article": article,
        "nm_id": nm_id,
        "stars": f.get("productValuation", 0),
        "created_date": date.isoformat(),
        "is_old": date < cutoff,
        "is_answered": bool(f.get("answer")),
        "text": (f.get("text") or "")[:500],
        "updated_at": now.isoformat()
    }

def sync_all(full: bool = False):
    """Подтягивает отзывы с WB.
    Обычный запуск (каждые 30 мин) — только свежие (~14 дней), без архива.
    full=True — глубокий проход (редко / вручную).
    """
    if not WB_TOKEN:
        logger.error("WB_TOKEN not set")
        return
    logger.info("Starting sync%s...", " FULL" if full else "")
    total = 0
    now = datetime.now(timezone.utc)
    # обычный синк не тащит всю историю — иначе Railway виснет на 80k+ отзывах
    cutoff = now - timedelta(days=400 if full else 14)
    max_skip = 199990 if full else 20000

    nm_to_vendor = {}
    try:
        st = httpx.get(
            f"{SUPABASE_URL}/rest/v1/stock_totals?select=nm_id,vendor_code",
            headers=sb_headers(), timeout=15
        )
        if st.is_success:
            nm_to_vendor = {r["nm_id"]: r["vendor_code"] for r in st.json() if r.get("nm_id") and r.get("vendor_code")}
        # stock_totals часто без vendor_code — добираем из ratings
        rt = httpx.get(
            f"{SUPABASE_URL}/rest/v1/ratings_official?select=nm_id,article&nm_id=not.is.null&article=not.is.null&limit=5000",
            headers=sb_headers(), timeout=20,
        )
        if rt.is_success:
            for r in rt.json() or []:
                nm, art = r.get("nm_id"), (r.get("article") or "").strip()
                if nm and art and nm not in nm_to_vendor:
                    nm_to_vendor[nm] = art
        logger.info(f"sync_all: nm_to_vendor map built: {len(nm_to_vendor)} entries")
    except Exception as e:
        logger.error(f"sync_all: failed to build nm_to_vendor: {e}")

    def _fb_date(f: dict):
        date_str = f.get("createdDate") or f.get("updatedDate") or ""
        try:
            return datetime.fromisoformat(date_str.replace("Z", "+00:00"))
        except Exception:
            return now

    for is_answered in [True, False]:
        skip = 0
        while skip <= max_skip:
            batch = fetch_feedbacks_page(is_answered, skip)
            if not batch:
                break
            processed = []
            hit_old = False
            for f in batch:
                if not f.get("id") or is_supplemented(f):
                    continue
                if _fb_date(f) < cutoff:
                    hit_old = True
                    continue
                processed.append(process_feedback(f, nm_to_vendor))
            total += upsert_feedbacks(processed)
            logger.info(f"  answered={is_answered} skip={skip} saved={len(processed)}")
            skip += len(batch)
            if hit_old or len(batch) < 5000:
                break
            time.sleep(0.3)

    # архив — только в полном синке (тяжёлый)
    if full:
        skip = 0
        while skip <= max_skip:
            batch = fetch_archive_page(skip)
            if not batch:
                break
            processed = []
            hit_old = False
            for f in batch:
                if not f.get("id"):
                    continue
                if _fb_date(f) < cutoff:
                    hit_old = True
                    continue
                processed.append(process_feedback(f, nm_to_vendor))
            total += upsert_feedbacks(processed)
            logger.info(f"  archive skip={skip} saved={len(processed)}")
            skip += len(batch)
            if hit_old or len(batch) < 5000:
                break
            time.sleep(0.3)

    now_str = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")
    httpx.post(
        f"{SUPABASE_URL}/rest/v1/settings",
        json={"key": "last_sync", "value": now_str, "updated_at": datetime.now(timezone.utc).isoformat()},
        headers=sb_headers(), timeout=10
    )
    logger.info(f"Sync complete. Total: {total}")
    # Рейтинги склейки считаем по всем отзывам из feedbacks (в т.ч. старше года).
    # xlsx «Оценка товара» — опционально через /api/upload-ratings.

def sync_ratings_official():
    """Не используем Analytics item-rating: там только прирост за период и нет старых карточек,
    из-за этого склейки давали бред вроде 5.06. Чистим ошибочные source=api записи."""
    try:
        httpx.delete(
            f"{SUPABASE_URL}/rest/v1/ratings_official?source=eq.api",
            headers={**sb_headers(), "Prefer": "return=minimal"}, timeout=30,
        )
        httpx.delete(
            f"{SUPABASE_URL}/rest/v1/settings?key=eq.last_ratings_sync",
            headers={**sb_headers(), "Prefer": "return=minimal"}, timeout=10,
        )
        httpx.delete(
            f"{SUPABASE_URL}/rest/v1/settings?key=eq.last_ratings_sync_error",
            headers={**sb_headers(), "Prefer": "return=minimal"}, timeout=10,
        )
        try:
            _DASH_CACHE["ts"] = 0.0
            _DASH_CACHE["data"] = None
        except NameError:
            pass
    except Exception as e:
        logger.warning(f"sync_ratings_official cleanup: {e}")
    logger.info("sync_ratings_official: skipped (use feedbacks + optional xlsx)")
    return {"status": "skipped"}

# ---------- Остатки на складах (WB Analytics: warehouse_remains report) ----------

def create_stock_report():
    resp = httpx.get(
        f"{WB_ANALYTICS_URL}/api/v1/warehouse_remains",
        headers=wb_headers(), params={"groupByNm": "true"}, timeout=30
    )
    if not resp.is_success:
        logger.error(f"WB stock report create error {resp.status_code} {resp.text[:200]}")
        return None
    return resp.json().get("data", {}).get("taskId")

def wait_stock_report(task_id: str, max_wait: int = 180) -> bool:
    elapsed = 0
    while elapsed < max_wait:
        time.sleep(5)
        elapsed += 5
        resp = httpx.get(
            f"{WB_ANALYTICS_URL}/api/v1/warehouse_remains/tasks/{task_id}/status",
            headers=wb_headers(), timeout=15
        )
        status = resp.json().get("data", {}).get("status") if resp.is_success else f"http_{resp.status_code}"
        logger.info(f"Stock report status (t+{elapsed}s): {status}")
        if status == "done":
            return True
    return False

def download_stock_report(task_id: str) -> list:
    resp = httpx.get(
        f"{WB_ANALYTICS_URL}/api/v1/warehouse_remains/tasks/{task_id}/download",
        headers=wb_headers(), timeout=60
    )
    if not resp.is_success:
        logger.error(f"WB stock report download error {resp.status_code} {resp.text[:200]}")
        return []
    data = resp.json()
    if not isinstance(data, list):
        logger.error(f"Unexpected stock report shape ({type(data).__name__}): {str(data)[:300]}")
        return []
    logger.info(f"Stock report download: {len(data)} items" + (f", raw snippet: {resp.text[:300]}" if not data else ""))
    if data:
        logger.info(f"Stock report sample item 0: {json.dumps(data[0], ensure_ascii=False)[:600]}")
        if len(data) > 1:
            logger.info(f"Stock report sample item 1: {json.dumps(data[1], ensure_ascii=False)[:600]}")
    return data

def process_stock_items(items: list):
    now = datetime.now(timezone.utc).isoformat()
    totals, warehouses = [], []
    for it in items:
        nm_id = it.get("nmId")
        if not nm_id:
            continue
        row = {
            "nm_id": nm_id,
            "vendor_code": it.get("vendorCode", ""),
            "subject_name": it.get("subjectName", ""),
            "brand": it.get("brand", ""),
            "volume": it.get("volume"),
            "in_way_to_client": 0,
            "in_way_from_client": 0,
            "quantity_warehouses_full": 0,
            "updated_at": now,
        }
        for w in it.get("warehouses", []):
            name, qty = w.get("warehouseName"), w.get("quantity", 0)
            field = STOCK_SPECIAL_FIELDS.get(name)
            if field:
                row[field] = qty
            else:
                warehouses.append({"nm_id": nm_id, "warehouse_name": name, "quantity": qty, "updated_at": now})
        totals.append(row)
    return totals, warehouses

def upsert_stock(totals: list, warehouses: list) -> int:
    saved = 0
    for i in range(0, len(totals), 200):
        batch = totals[i:i + 200]
        resp = httpx.post(f"{SUPABASE_URL}/rest/v1/stock_totals", json=batch, headers=sb_headers(), timeout=30)
        if resp.is_success:
            saved += len(batch)
        else:
            logger.error(f"stock_totals upsert error {resp.status_code} {resp.text[:200]}")
    # Полная перезаливка детализации по складам — проще, чем строить составной upsert-ключ
    httpx.delete(f"{SUPABASE_URL}/rest/v1/stock_warehouses?id=gte.0", headers=sb_headers(), timeout=15)
    for i in range(0, len(warehouses), 500):
        batch = warehouses[i:i + 500]
        resp = httpx.post(f"{SUPABASE_URL}/rest/v1/stock_warehouses", json=batch, headers=sb_headers(), timeout=30)
        if not resp.is_success:
            logger.error(f"stock_warehouses insert error {resp.status_code} {resp.text[:200]}")
    return saved

def fetch_seller_fbs_warehouses() -> tuple:
    """Склады продавца (FBS / Маркетплейс). → (list, error_or_None)."""
    if not WB_TOKEN:
        return [], "WB_TOKEN не задан"
    try:
        resp = httpx.get(
            f"{WB_MARKETPLACE_URL}/api/v3/warehouses",
            headers=wb_headers(),
            timeout=30,
        )
    except Exception as e:
        logger.error(f"FBS warehouses exception: {e}")
        return [], str(e)
    if not resp.is_success:
        logger.error(f"FBS warehouses {resp.status_code}: {resp.text[:240]}")
        hint = ""
        if resp.status_code in (401, 403):
            hint = " — добавь категорию «Маркетплейс» в WB_TOKEN и перевыпусти токен"
        return [], f"HTTP {resp.status_code}: {resp.text[:180]}{hint}"
    data = resp.json()
    if isinstance(data, list):
        whs = data
    elif isinstance(data, dict):
        whs = data.get("warehouses") or data.get("data") or data.get("result") or []
        # иногда один склад приходит объектом
        if not whs and (data.get("id") is not None or data.get("warehouseId") is not None):
            whs = [data]
    else:
        whs = []
    if not whs:
        return [], "API ответил ок, но складов FBS в кабинете нет (создай склад продавца в WB)"
    # нормализуем id/name
    norm = []
    for w in whs:
        if not isinstance(w, dict):
            continue
        wid = w.get("id") if w.get("id") is not None else w.get("warehouseId")
        name = w.get("name") or w.get("warehouseName") or w.get("officeName") or ""
        if wid is None:
            continue
        item = dict(w)
        item["id"] = wid
        item["name"] = str(name).strip() or f"склад {wid}"
        norm.append(item)
    if not norm:
        sample = str(whs[:2])[:300]
        return [], f"Склады пришли в неожиданном формате: {sample}"
    return norm, None


def fetch_all_card_skus() -> list:
    """Все баркоды карточек: [{sku, nm_id, vendor_code}]."""
    if not WB_TOKEN:
        return []
    out, seen = [], set()
    cursor = {"limit": 100}
    for _ in range(200):
        try:
            resp = httpx.post(
                f"{WB_CONTENT_URL}/content/v2/get/cards/list",
                headers=wb_headers(),
                json={
                    "settings": {
                        "sort": {"ascending": True},
                        "filter": {"withPhoto": -1},
                        "cursor": cursor,
                    }
                },
                timeout=40,
            )
        except Exception as e:
            logger.error(f"cards/list for FBS skus: {e}")
            break
        if not resp.is_success:
            logger.error(f"cards/list FBS {resp.status_code}: {resp.text[:200]}")
            break
        payload = resp.json() or {}
        cards = payload.get("cards") or []
        if not cards:
            break
        for c in cards:
            nm = c.get("nmID") or c.get("nmId")
            vc = (c.get("vendorCode") or "").strip()
            for sz in c.get("sizes") or []:
                for sku in sz.get("skus") or []:
                    sku_s = str(sku).strip()
                    if not sku_s or sku_s in seen:
                        continue
                    seen.add(sku_s)
                    out.append({"sku": sku_s, "nm_id": nm, "vendor_code": vc})
        curs = payload.get("cursor") or {}
        updated = curs.get("updatedAt")
        nm_cur = curs.get("nmID") or curs.get("nmId")
        if len(cards) < 100 or not updated or nm_cur is None:
            break
        cursor = {"limit": 100, "updatedAt": updated, "nmID": nm_cur}
        time.sleep(0.35)
    logger.info(f"FBS skus from cards: {len(out)}")
    return out


def fetch_fbs_stocks() -> dict:
    """
    Остатки FBS (система Маркетплейс) по складам продавца.
    → {
        warehouses: [{nm_id, warehouse_name, quantity, updated_at}],
        by_nm: {nm_id: qty},
        samples: [...],
        error: optional
      }
    """
    now = datetime.now(timezone.utc).isoformat()
    whs, wh_err = fetch_seller_fbs_warehouses()
    if not whs:
        return {
            "warehouses": [],
            "by_nm": {},
            "samples": [],
            "error": wh_err or "Нет складов FBS",
        }
    sku_rows = fetch_all_card_skus()
    if not sku_rows:
        return {
            "warehouses": [],
            "by_nm": {},
            "samples": [],
            "error": "Не удалось получить баркоды из Content API",
            "fbs_warehouses": [
                {"id": w.get("id"), "name": w.get("name")} for w in whs if isinstance(w, dict)
            ],
        }
    sku_meta = {r["sku"]: r for r in sku_rows}
    skus = list(sku_meta.keys())
    # qty по (nm_id, warehouse_label)
    qty_map = {}
    errors = []
    for w in whs:
        if not isinstance(w, dict):
            continue
        wid = w.get("id")
        wname = (w.get("name") or f"склад {wid}").strip()
        label = f"Маркетплейс (FBS) · {wname}"
        if wid is None:
            continue
        for i in range(0, len(skus), 1000):
            batch = skus[i:i + 1000]
            try:
                resp = httpx.post(
                    f"{WB_MARKETPLACE_URL}/api/v3/stocks/{wid}",
                    headers=wb_headers(),
                    json={"skus": batch},
                    timeout=60,
                )
            except Exception as e:
                errors.append(f"{wname}: {e}")
                break
            if resp.status_code == 429:
                time.sleep(2)
                try:
                    resp = httpx.post(
                        f"{WB_MARKETPLACE_URL}/api/v3/stocks/{wid}",
                        headers=wb_headers(),
                        json={"skus": batch},
                        timeout=60,
                    )
                except Exception as e:
                    errors.append(f"{wname}: {e}")
                    break
            if not resp.is_success:
                errors.append(f"{wname}: {resp.status_code} {resp.text[:160]}")
                break
            body = resp.json() or {}
            stocks = body.get("stocks") if isinstance(body, dict) else body
            # на всякий случай считаем и нули — в samples потом отфильтруем >0
            for s in stocks or []:
                if not isinstance(s, dict):
                    continue
                sku = str(s.get("sku") or s.get("barcode") or "").strip()
                try:
                    amount = int(s.get("amount") if s.get("amount") is not None else s.get("quantity") or 0)
                except Exception:
                    amount = 0
                if not sku:
                    continue
                meta = sku_meta.get(sku) or {}
                nm = meta.get("nm_id")
                if nm is None:
                    continue
                key = (int(nm), label)
                qty_map[key] = qty_map.get(key, 0) + amount
            if i + 1000 < len(skus):
                time.sleep(0.2)

    warehouses = []
    by_nm = {}
    samples_acc = {}
    for (nm, label), qty in qty_map.items():
        if qty <= 0:
            continue
        warehouses.append({
            "nm_id": nm,
            "warehouse_name": label,
            "quantity": qty,
            "updated_at": now,
        })
        by_nm[nm] = by_nm.get(nm, 0) + qty
        meta = next((r for r in sku_rows if r.get("nm_id") == nm), {})
        samples_acc[nm] = {
            "nm_id": nm,
            "vendor_code": meta.get("vendor_code") or "",
            "fbs_qty": by_nm[nm],
        }
    samples = sorted(samples_acc.values(), key=lambda x: -x["fbs_qty"])[:20]
    logger.info(
        f"FBS stocks: wh={len(whs)}, skus={len(skus)}, rows={len(warehouses)}, "
        f"nms_with_stock={len(by_nm)}, errors={len(errors)}"
    )
    err = None
    if errors and not warehouses:
        err = "; ".join(errors[:3])
    elif not warehouses and not errors:
        err = "Склады FBS есть, но остатки по всем баркодам = 0 (или SKU не совпали)"
    return {
        "warehouses": warehouses,
        "by_nm": by_nm,
        "samples": samples,
        "fbs_warehouses": [
            {"id": w.get("id"), "name": w.get("name")} for w in whs if isinstance(w, dict)
        ],
        "skus_count": len(skus),
        "errors": errors,
        "error": err,
    }


def sync_stock():
    if not WB_TOKEN:
        logger.error("WB_TOKEN not set")
        return
    logger.info("Starting stock sync...")
    task_id = create_stock_report()
    if not task_id:
        return
    if not wait_stock_report(task_id):
        logger.error("Stock report generation timed out")
        return
    time.sleep(5)  # небольшой буфер: статус иногда становится "done" чуть раньше, чем файл реально готов к скачиванию
    items = download_stock_report(task_id)
    if not items:
        logger.info("Stock report empty on first download, retrying once after 15s")
        time.sleep(15)
        items = download_stock_report(task_id)
    totals, warehouses = process_stock_items(items)

    # FBS / Маркетплейс — отдельные колонки складов (не входят в quantity_warehouses_full WB)
    fbs = {}
    try:
        fbs = fetch_fbs_stocks()
        fbs_rows = fbs.get("warehouses") or []
        if fbs_rows:
            warehouses.extend(fbs_rows)
            by_nm = fbs.get("by_nm") or {}
            # артикулы только на FBS (нет в FBW-отчёте) — строка в totals с нулём на WB
            have = {int(t["nm_id"]) for t in totals if t.get("nm_id") is not None}
            now = datetime.now(timezone.utc).isoformat()
            for nm, qty in by_nm.items():
                if int(nm) in have or int(qty) <= 0:
                    continue
                vc = next(
                    (s.get("vendor_code") for s in (fbs.get("samples") or []) if s.get("nm_id") == nm),
                    "",
                )
                totals.append({
                    "nm_id": int(nm),
                    "vendor_code": vc or "",
                    "subject_name": "",
                    "brand": "",
                    "volume": None,
                    "in_way_to_client": 0,
                    "in_way_from_client": 0,
                    "quantity_warehouses_full": 0,
                    "updated_at": now,
                })
        elif fbs.get("error"):
            logger.warning(f"FBS stocks skipped: {fbs.get('error')}")
    except Exception as e:
        logger.error(f"FBS stocks merge error: {e}")

    saved = upsert_stock(totals, warehouses)
    try:
        save_stock_warehouse_snapshot_from_rows(warehouses)
    except Exception as e:
        logger.error(f"stock warehouse snapshot error: {e}")
    httpx.post(
        f"{SUPABASE_URL}/rest/v1/settings",
        json={"key": "last_stock_sync", "value": datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M"),
              "updated_at": datetime.now(timezone.utc).isoformat()},
        headers=sb_headers(), timeout=10
    )
    logger.info(
        f"Stock sync complete. Articles: {saved}, warehouse rows: {len(warehouses)}, "
        f"fbs_nms: {len((fbs or {}).get('by_nm') or {})}"
    )


# ---------- Снимки остатков по складам (для сравнения географии) ----------
STOCK_WH_SNAPS_KEY = "stock_warehouse_snaps"
STOCK_WH_SNAPS_KEEP_DAYS = 7
# Если живых складов меньше — считаем географию узкой (срок доставки бьёт по конверсии)
STOCK_WH_NARROW_LIVE = 2
# Склады, отключённые в «Рекомендациях поставок» (общие для всех устройств)
SUPPLY_WH_DISABLED_KEY = "supply_wh_disabled"


def get_disabled_warehouses() -> set:
    """Имена складов, снятых галкой в матрице поставок — не участвуют в расчётах темпа."""
    raw = get_setting_raw(SUPPLY_WH_DISABLED_KEY, None)
    if raw is None:
        # fallback: json path
        raw = get_setting_json(SUPPLY_WH_DISABLED_KEY, None)
    return set(_normalize_disabled_warehouses(raw if raw is not None else []))


def _normalize_disabled_warehouses(value) -> list:
    """Приводит value из API/settings к чистому list[str]."""
    import json as _json
    raw = value
    if raw is None or raw == "":
        return []
    if isinstance(raw, str):
        try:
            raw = _json.loads(raw)
            if isinstance(raw, str):
                raw = _json.loads(raw)
        except Exception:
            return [raw.strip()] if raw.strip() else []
    if not isinstance(raw, list):
        return []
    out, seen = [], set()
    for x in raw:
        n = str(x).strip()
        if n and n not in seen:
            seen.add(n)
            out.append(n)
    return out


def _save_disabled_warehouses(names: list) -> bool:
    clean = _normalize_disabled_warehouses(names)
    return save_setting_value(SUPPLY_WH_DISABLED_KEY, clean)


def _invalidate_dash_cache():
    try:
        with _DASH_CACHE_LOCK:
            _DASH_CACHE["ts"] = 0.0
            _DASH_CACHE["data"] = None
    except Exception:
        pass


def _filter_wh_map(wh_map: dict, disabled: set) -> dict:
    """Оставляет только включённые склады (qty > 0)."""
    if not wh_map:
        return {}
    if not disabled:
        return {str(k): int(v or 0) for k, v in wh_map.items() if int(v or 0) > 0}
    out = {}
    for name, qty in wh_map.items():
        n = str(name).strip()
        q = int(qty or 0)
        if q > 0 and n not in disabled:
            out[n] = q
    return out


def _stock_wh_by_nm_from_rows(warehouses: list) -> dict:
    """nm_id(str) -> {t: total, w: {warehouse_name: qty}} — только qty > 0."""
    by_nm = {}
    for row in warehouses or []:
        nm = row.get("nm_id")
        if nm is None:
            continue
        name = (row.get("warehouse_name") or "").strip()
        qty = int(row.get("quantity") or 0)
        if not name or qty <= 0:
            continue
        key = str(int(nm))
        slot = by_nm.setdefault(key, {"t": 0, "w": {}})
        slot["w"][name] = slot["w"].get(name, 0) + qty
        slot["t"] = sum(slot["w"].values())
    return by_nm


def _fetch_stock_wh_by_nm() -> dict:
    try:
        r = httpx.get(
            f"{SUPABASE_URL}/rest/v1/stock_warehouses?select=nm_id,warehouse_name,quantity",
            headers=sb_headers(),
            timeout=30,
        )
        rows = r.json() if r.is_success else []
        return _stock_wh_by_nm_from_rows(rows if isinstance(rows, list) else [])
    except Exception as e:
        logger.error(f"fetch stock_warehouses for snapshot: {e}")
        return {}


def _save_stock_wh_snaps(snaps: list) -> bool:
    import json as _json
    body = {
        "key": STOCK_WH_SNAPS_KEY,
        "value": _json.dumps(snaps, ensure_ascii=False),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    try:
        resp = httpx.post(
            f"{SUPABASE_URL}/rest/v1/settings?on_conflict=key",
            json=body,
            headers=sb_headers(),
            timeout=60,
        )
        return resp.is_success
    except Exception as e:
        logger.error(f"save stock_warehouse_snaps: {e}")
        return False


def save_stock_warehouse_snapshot_from_rows(warehouses: list = None) -> dict:
    """Пишет дневной снимок остатков по складам (перезаписывает сегодняшний)."""
    by_nm = _stock_wh_by_nm_from_rows(warehouses) if warehouses is not None else _fetch_stock_wh_by_nm()
    return save_stock_warehouse_snapshot_by_nm(by_nm)


def save_stock_warehouse_snapshot_by_nm(by_nm: dict) -> dict:
    now = _msk_now()
    day = now.strftime("%Y-%m-%d")
    hour_key = now.strftime("%Y-%m-%dT%H")
    payload = {
        "day": day,
        "hour_key": hour_key,
        "as_of": now.strftime("%Y-%m-%d %H:%M"),
        "by_nm": by_nm or {},
        "nm_count": len(by_nm or {}),
    }
    snaps = get_setting_json(STOCK_WH_SNAPS_KEY, []) or []
    if not isinstance(snaps, list):
        snaps = []
    snaps = [s for s in snaps if isinstance(s, dict) and s.get("day") != day]
    snaps.append(payload)
    cutoff = (now - timedelta(days=STOCK_WH_SNAPS_KEEP_DAYS)).strftime("%Y-%m-%d")
    snaps = [s for s in snaps if (s.get("day") or "") >= cutoff]
    snaps.sort(key=lambda s: s.get("day") or "")
    ok = _save_stock_wh_snaps(snaps)
    logger.info(
        f"stock WH snapshot day={day}: nms={len(by_nm or {})}, kept_days={len(snaps)}, saved={ok}"
    )
    return payload


def get_stock_warehouse_snap_for_day(day: str):
    """Последний снимок за календарный день YYYY-MM-DD (или None)."""
    snaps = get_setting_json(STOCK_WH_SNAPS_KEY, []) or []
    if not isinstance(snaps, list):
        return None
    day = str(day or "")[:10]
    best = None
    for s in snaps:
        if not isinstance(s, dict):
            continue
        if (s.get("day") or "")[:10] != day:
            continue
        if best is None or (s.get("hour_key") or "") >= (best.get("hour_key") or ""):
            best = s
    return best


def stock_wh_geo_compare(nm_id, cur_by_nm: dict, prev_snap: dict, disabled: set = None) -> dict:
    """Сравнивает текущую географию складов с вчерашним снимком (без отключённых складов)."""
    disabled = disabled if disabled is not None else set()
    key = str(int(nm_id))
    cur = (cur_by_nm or {}).get(key) or {}
    cur_w = _filter_wh_map(cur.get("w") or {}, disabled)
    prev = ((prev_snap or {}).get("by_nm") or {}).get(key) or {}
    prev_w = _filter_wh_map(prev.get("w") or {}, disabled)
    cur_live = sorted([n for n, q in cur_w.items() if int(q or 0) > 0])
    prev_live = sorted([n for n, q in prev_w.items() if int(q or 0) > 0])
    emptied = [n for n in prev_live if n not in cur_w or int(cur_w.get(n) or 0) <= 0]
    added = [n for n in cur_live if n not in prev_w]
    live_now = len(cur_live)
    live_prev = len(prev_live)
    total = sum(cur_w.values())
    geo_flag = "ok"
    if live_now <= 0:
        geo_flag = "oos"
    elif emptied and live_now < live_prev:
        geo_flag = "emptied"
    elif live_now <= STOCK_WH_NARROW_LIVE and total > 0:
        geo_flag = "narrow"
    return {
        "wh_live": live_now,
        "wh_live_prev": live_prev if prev_w or prev.get("t") is not None else None,
        "wh_emptied": emptied[:8],
        "wh_added": added[:8],
        "wh_names": cur_live[:12],
        "stock_qty_enabled": total,
        "stock_geo_flag": geo_flag,
        "wh_snap_prev_as_of": (prev_snap or {}).get("as_of"),
    }


# ---------- Остатки нашего склада (Google Sheets) ----------
OWN_WAREHOUSE_SHEET_ID = os.getenv(
    "OWN_WAREHOUSE_SHEET_ID",
    "1Lhoy4s_KX0pWndsd3Y5oCOjTFCtfEfVUM4AgtBv4Crc",
)
# Вкладка «Остатки на складе» (если на Railway задан старый OWN_WAREHOUSE_GID — обнови)
OWN_WAREHOUSE_GID = os.getenv("OWN_WAREHOUSE_GID", "787686207")
OWN_WAREHOUSE_GID_FALLBACKS = ("787686207", "0")
OWN_WAREHOUSE_CACHE = {
    "title": None,
    "as_of": None,
    "rows": [],
    "updated_at": None,
    "error": None,
    "syncing": False,
}

OWN_WH_SHIPMENTS_KEY = "own_wh_shipments"
OWN_WH_RECEIPTS_KEY = "own_wh_receipts"
OWN_WH_ARCHIVE_KEY = "own_wh_archive"
OWN_WH_SKU_ALIASES_KEY = "own_wh_sku_aliases"  # {alias_sku: canonical_vendor_code}
OWN_WH_STOCK_SNAPSHOT_KEY = "own_wh_stock_snapshot"  # общий снимок для WB+Ozon
OWN_WH_DOCS_KEEP = 200
OWN_WH_ARCHIVE_KEEP = 60
OWN_WH_CHANNELS = ("fbw", "fbs", "ozon_fbo", "ozon_fbs")
# Жёсткие семьи склада: LK11 Pro Max = только 046; 038 = S11 middle.
# Ручной model_map в settings перекрывает эти дефолты.
OWN_WH_MODEL_DEFAULTS = {
    "046_LK11_Promax_black_O": "046_LK11_Promax_black_O",
    "046_LK11_Promax_grey_O": "046_LK11_Promax_grey_O",
    "046_LK11_Promax_gold_O": "046_LK11_Promax_gold_O",
    "046_LK11_Promax_black_0": "046_LK11_Promax_black_O",  # опечатка в приёмках
    "038_LK11_gold_O": "031_LK11_gold_O",              # S11 Pro золото
    "038_LK11_black_O": "038_LK11_black_O",            # S11 middle чёрный (корень)
    "038_LK11_orahge_O": "038_LK11_orahge_O",          # S11 middle оранжевый
    "038_S11grey_3bras_O": "031_LK11_grey_O",          # S11 middle серебро
}
OWN_WH_MODEL_NAME_OVERRIDES = {
    "046_LK11_Promax_black_O": "LK11 Pro Max (Черный)",
    "046_LK11_Promax_grey_O": "LK11 Pro Max (Серебро)",
    "046_LK11_Promax_gold_O": "LK11 Pro Max (Золото)",
    "038_LK11_black_O": "S11 middle (Черный)",
    "038_LK11_orahge_O": "S11 middle (Оранжевый)",
}
OWN_WH_SKU_ALIAS_DEFAULTS = {
    "046_LK11_Promax_black_0": "046_LK11_Promax_black_O",
}
_RU_MONTHS_SHORT = (
    "", "янв", "фев", "мар", "апр", "май", "июн",
    "июл", "авг", "сен", "окт", "ноя", "дек",
)


def _own_wh_shipments() -> list:
    raw = get_setting_json(OWN_WH_SHIPMENTS_KEY, []) or []
    return raw if isinstance(raw, list) else []


def _own_wh_receipts() -> list:
    raw = get_setting_json(OWN_WH_RECEIPTS_KEY, []) or []
    return raw if isinstance(raw, list) else []


def _own_wh_archives() -> list:
    raw = get_setting_json(OWN_WH_ARCHIVE_KEY, []) or []
    return raw if isinstance(raw, list) else []


def _own_wh_sku_aliases() -> dict:
    raw = get_setting_json(OWN_WH_SKU_ALIASES_KEY, {}) or {}
    if not isinstance(raw, dict):
        raw = {}
    out = {str(k): str(v) for k, v in OWN_WH_SKU_ALIAS_DEFAULTS.items() if k and v}
    for k, v in raw.items():
        ak = str(k or "").strip()
        cv = str(v or "").strip()
        if ak and cv:
            out[ak] = cv
    return out


def _own_wh_canonical_vc(vc: str, aliases: dict | None = None) -> str:
    """Приводит артикул Ozon/альяс к каноническому (WB / Sheets), чтобы семьи считались одинаково."""
    vc = str(vc or "").strip()
    if not vc:
        return ""
    aliases = aliases if aliases is not None else _own_wh_sku_aliases()
    if vc in aliases:
        return str(aliases[vc]).strip() or vc
    low = vc.lower()
    for ak, cv in aliases.items():
        if str(ak).lower() == low:
            return str(cv).strip() or vc
    return vc


def _own_wh_normalize_items(items: list) -> list:
    """Агрегирует qty по каноническому артикулу (алиасы Ozon → WB)."""
    aliases = _own_wh_sku_aliases()
    agg = {}
    for it in items or []:
        raw_vc = str((it or {}).get("vendor_code") or "").strip()
        if not raw_vc:
            continue
        try:
            qty = int((it or {}).get("qty") or 0)
        except Exception:
            qty = 0
        if qty <= 0:
            continue
        vc = _own_wh_canonical_vc(raw_vc, aliases)
        if vc not in agg:
            agg[vc] = {"vendor_code": vc, "qty": 0, "aliases": set()}
        agg[vc]["qty"] += qty
        if raw_vc != vc:
            agg[vc]["aliases"].add(raw_vc)
    out = []
    for vc, row in sorted(agg.items(), key=lambda x: (-x[1]["qty"], x[0])):
        item = {"vendor_code": vc, "qty": row["qty"]}
        if row["aliases"]:
            item["source_skus"] = sorted(row["aliases"])
        out.append(item)
    return out


def _own_wh_qty_map_from_docs(docs: list) -> dict:
    out = {}
    aliases = _own_wh_sku_aliases()
    for sh in docs or []:
        for it in sh.get("items") or []:
            vc = _own_wh_canonical_vc(str(it.get("vendor_code") or "").strip(), aliases)
            if not vc:
                continue
            try:
                qty = int(it.get("qty") or 0)
            except Exception:
                qty = 0
            if qty <= 0:
                continue
            out[vc] = out.get(vc, 0) + qty
    return out


def _persist_own_wh_snapshot():
    """Пишет общий снимок остатков (WB+Ozon читают одно и то же)."""
    by_vendor = OWN_WAREHOUSE_CACHE.get("by_vendor") or {}
    slim = {}
    for vc, meta in by_vendor.items():
        if not vc or not isinstance(meta, dict):
            continue
        slim[str(vc)] = {
            "stock": int(meta.get("stock") or 0),
            "family_stock": int(meta.get("family_stock") or 0),
            "family": list(meta.get("family") or [vc]),
            "root": meta.get("root") or vc,
            "model_name": meta.get("model_name") or "",
        }
    snap = {
        "as_of": OWN_WAREHOUSE_CACHE.get("as_of"),
        "updated_at": OWN_WAREHOUSE_CACHE.get("updated_at"),
        "title": OWN_WAREHOUSE_CACHE.get("title"),
        "by_vendor": slim,
        "shared": True,
        "marketplaces": ["wb", "ozon"],
    }
    save_setting_value(OWN_WH_STOCK_SNAPSHOT_KEY, snap)
    OWN_WAREHOUSE_CACHE["stock_snapshot"] = snap
    return snap


def _own_wh_archive_active(
    reason: str = "friday",
    note: str = "",
    include_shipments: bool = True,
    include_receipts: bool = True,
) -> dict | None:
    """Переносит активные поступления/отгрузки в архив по дате (для аудита)."""
    ships = _own_wh_shipments() if include_shipments else []
    receipts = _own_wh_receipts() if include_receipts else []
    if not ships and not receipts:
        return None
    created_at, created_iso = _own_wh_now_stamp()
    ship_qty = sum(int(s.get("total_qty") or 0) for s in ships)
    recv_qty = sum(int(s.get("total_qty") or 0) for s in receipts)
    entry = {
        "id": f"arch_{int(time.time())}",
        "archived_at": created_at,
        "archived_at_iso": created_iso,
        "reason": reason,
        "note": str(note or "").strip(),
        "sheet_as_of": OWN_WAREHOUSE_CACHE.get("as_of"),
        "shipments": ships,
        "receipts": receipts,
        "shipments_qty": ship_qty,
        "receipts_qty": recv_qty,
        "shipments_files": len(ships),
        "receipts_files": len(receipts),
    }
    archive = _own_wh_archives()
    archive.insert(0, entry)
    archive = archive[:OWN_WH_ARCHIVE_KEEP]
    if not save_setting_value(OWN_WH_ARCHIVE_KEY, archive):
        raise RuntimeError("Не удалось сохранить архив в settings")
    OWN_WAREHOUSE_CACHE["archives"] = archive
    return entry


def _own_wh_deduction_map() -> dict:
    """Суммарные списания по артикулу из загруженных отгрузок."""
    return _own_wh_qty_map_from_docs(_own_wh_shipments())


def _own_wh_receipt_map() -> dict:
    """Суммарные поступления по артикулу."""
    return _own_wh_qty_map_from_docs(_own_wh_receipts())


def _apply_own_wh_deductions(personal_sheet: dict) -> dict:
    """Остаток = Sheets + поступления − отгрузки (оба списка — оверлеи до правки таблицы)."""
    received = _own_wh_receipt_map()
    shipped = _own_wh_deduction_map()
    out = {str(k): int(v or 0) for k, v in (personal_sheet or {}).items()}
    all_vc = set(out) | set(received) | set(shipped)
    for vc in all_vc:
        base = int(out.get(vc, 0))
        out[vc] = max(0, base + int(received.get(vc, 0)) - int(shipped.get(vc, 0)))
    return out


def _own_wh_parse_doc_dt(doc: dict):
    """Парсит дату документа → datetime (MSK-naive, для недель)."""
    if not doc:
        return None
    iso = str(doc.get("created_at_iso") or "").strip()
    if iso:
        try:
            s = iso.replace("Z", "+00:00")
            dt = datetime.fromisoformat(s)
            if dt.tzinfo is not None:
                # к «настенному» Мск ±3 без zoneinfo
                dt = (dt.astimezone(timezone.utc) + timedelta(hours=3)).replace(tzinfo=None)
            return dt
        except Exception:
            pass
    raw = str(doc.get("created_at") or "").strip()
    if raw:
        for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(raw, fmt)
            except Exception:
                continue
    return None


def _own_wh_week_bounds(dt: datetime):
    d = dt.date() if hasattr(dt, "date") else dt
    start = d - timedelta(days=d.weekday())  # понедельник
    end = start + timedelta(days=6)
    iso_year, iso_week, _ = start.isocalendar()
    key = f"{iso_year}-W{iso_week:02d}"
    if start.year == end.year:
        label = f"{start.day} {_RU_MONTHS_SHORT[start.month]} – {end.day} {_RU_MONTHS_SHORT[end.month]} {end.year}"
    else:
        label = (
            f"{start.day} {_RU_MONTHS_SHORT[start.month]} {start.year} – "
            f"{end.day} {_RU_MONTHS_SHORT[end.month]} {end.year}"
        )
    return key, start.isoformat(), end.isoformat(), label


def _own_wh_weekly_ledger(shipments: list = None, receipts: list = None) -> list:
    """Учёт по календарным неделям: приход / WB / FBS."""
    shipments = shipments if shipments is not None else _own_wh_shipments()
    receipts = receipts if receipts is not None else _own_wh_receipts()
    weeks = {}

    def ensure(key, start, end, label):
        if key not in weeks:
            weeks[key] = {
                "week_key": key,
                "week_start": start,
                "week_end": end,
                "label": label,
                "receipts_qty": 0,
                "receipts_files": 0,
                "fbw_qty": 0,
                "fbw_files": 0,
                "fbs_qty": 0,
                "fbs_files": 0,
                "ozon_fbo_qty": 0,
                "ozon_fbo_files": 0,
                "ozon_fbs_qty": 0,
                "ozon_fbs_files": 0,
                "docs": [],
            }
        return weeks[key]

    for rec in receipts or []:
        dt = _own_wh_parse_doc_dt(rec) or datetime.now()
        key, start, end, label = _own_wh_week_bounds(dt)
        w = ensure(key, start, end, label)
        qty = int(rec.get("total_qty") or 0)
        w["receipts_qty"] += qty
        w["receipts_files"] += 1
        w["docs"].append({
            "id": rec.get("id"),
            "doc_type": "receipt",
            "channel": "in",
            "filename": rec.get("filename") or rec.get("note") or "поступление",
            "note": rec.get("note") or "",
            "total_qty": qty,
            "articles": rec.get("articles") or 0,
            "created_at": rec.get("created_at"),
            "created_at_iso": rec.get("created_at_iso"),
            "kind": rec.get("kind") or "receipt",
        })

    for sh in shipments or []:
        dt = _own_wh_parse_doc_dt(sh) or datetime.now()
        key, start, end, label = _own_wh_week_bounds(dt)
        w = ensure(key, start, end, label)
        ch = _own_wh_shipment_channel(sh)
        qty = int(sh.get("total_qty") or 0)
        if ch == "fbs":
            w["fbs_qty"] += qty
            w["fbs_files"] += 1
        elif ch == "ozon_fbo":
            w["ozon_fbo_qty"] += qty
            w["ozon_fbo_files"] += 1
        elif ch == "ozon_fbs":
            w["ozon_fbs_qty"] += qty
            w["ozon_fbs_files"] += 1
        else:
            w["fbw_qty"] += qty
            w["fbw_files"] += 1
        w["docs"].append({
            "id": sh.get("id"),
            "doc_type": "shipment",
            "channel": ch,
            "filename": sh.get("filename") or "файл",
            "note": sh.get("note") or "",
            "total_qty": qty,
            "articles": sh.get("articles") or 0,
            "created_at": sh.get("created_at"),
            "created_at_iso": sh.get("created_at_iso"),
            "kind": sh.get("kind"),
        })

    out = []
    for key in sorted(weeks.keys(), reverse=True):
        w = weeks[key]
        w["net_qty"] = (
            int(w["receipts_qty"])
            - int(w["fbw_qty"]) - int(w["fbs_qty"])
            - int(w["ozon_fbo_qty"]) - int(w["ozon_fbs_qty"])
        )
        w["docs"].sort(
            key=lambda d: str(d.get("created_at_iso") or d.get("created_at") or ""),
            reverse=True,
        )
        out.append(w)
    return out


def _parse_int_cell(v):
    s = str(v or "").strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s or s.lower() in ("nan", "none", "-"):
        return None
    try:
        return int(float(s))
    except Exception:
        return None

def _download_own_warehouse_csv() -> str:
    """Скачивает CSV вкладки остатков. export → gviz; при 400 пробует запасные gid."""
    if not OWN_WAREHOUSE_SHEET_ID:
        raise RuntimeError("OWN_WAREHOUSE_SHEET_ID не задан — вкладка «Наш склад» опциональна")
    headers = {"User-Agent": "Mozilla/5.0 (compatible; wb-dashboard/1.0)"}
    gids = []
    for g in (OWN_WAREHOUSE_GID, *OWN_WAREHOUSE_GID_FALLBACKS):
        g = str(g or "").strip()
        if g and g not in gids:
            gids.append(g)
    last_err = None
    for gid in gids:
        urls = [
            f"https://docs.google.com/spreadsheets/d/{OWN_WAREHOUSE_SHEET_ID}/export?format=csv&gid={gid}",
            f"https://docs.google.com/spreadsheets/d/{OWN_WAREHOUSE_SHEET_ID}/gviz/tq?tqx=out:csv&gid={gid}",
        ]
        for url in urls:
            try:
                resp = httpx.get(url, timeout=30, follow_redirects=True, headers=headers)
            except Exception as e:
                last_err = e
                continue
            if not resp.is_success:
                last_err = RuntimeError(f"Google Sheets HTTP {resp.status_code} (gid={gid})")
                continue
            text = resp.text or ""
            if not text.strip() or text.lstrip().startswith("<!"):
                last_err = RuntimeError("Таблица недоступна (нужен доступ «все у кого есть ссылка»)")
                continue
            # похоже на вкладку остатков
            low = text[:2000].lower()
            if "артикул" in low or "остатк" in low or "наименование" in low:
                return text
            last_err = RuntimeError(f"Не похоже на лист остатков (gid={gid})")
    raise RuntimeError(str(last_err) if last_err else "Не удалось скачать Google Sheets")


def fetch_own_warehouse_stock() -> dict:
    """Тянет CSV из Google Sheets «Остатки на складе».
    Берём только 1-ю таблицу (до ИТОГО / «Принято на склад»), без блоков принято/обмен.
    Строим семьи артикулов: пустые строки-артикулы под основным (044→037) делят остаток.
    Строки без артикула продавца (есть только наименование) — тоже в списке как «товар без продаж»."""
    import csv as _csv
    import re as _re

    text = _download_own_warehouse_csv()
    rows_raw = list(_csv.reader(io.StringIO(text)))
    if len(rows_raw) < 2:
        raise RuntimeError("Пустая таблица")

    # Иногда gviz склеивает заголовок в одну строку — ищем строку с «артикул»
    header_idx = 1
    for i, r in enumerate(rows_raw[:5]):
        joined = " ".join(str(c).lower() for c in r)
        if "артикул" in joined and ("наименован" in joined or "остатк" in joined or "на складе" in joined):
            header_idx = i
            break
    title_row = rows_raw[0] if header_idx > 0 else rows_raw[header_idx]
    title = (title_row[0] if title_row else "").strip()
    if "артикул" in title.lower():
        title = "Остатки на складе"
    as_of = None
    m = _re.search(r"(\d{1,2})[./](\d{1,2})[./](\d{2,4})", title)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        if len(y) == 2:
            y = "20" + y
        as_of = f"{int(d):02d}.{int(mo):02d}.{y}"
    if not as_of:
        m2 = _re.search(r"(\d{1,2})[./](\d{1,2})[./](\d{2,4})", " ".join(str(c) for c in rows_raw[0]))
        if m2:
            d, mo, y = m2.group(1), m2.group(2), m2.group(3)
            if len(y) == 2:
                y = "20" + y
            as_of = f"{int(d):02d}.{int(mo):02d}.{y}"

    header = [str(h).strip().lower() for h in rows_raw[header_idx]]

    def find_col(*needles):
        for i, h in enumerate(header):
            for n in needles:
                if n in h:
                    return i
        return None

    col_vc = find_col("артикул продавца", "артикул")
    col_name = find_col("наименование", "название")
    col_stock = find_col(
        "остататки на складе",
        "остатки на складе",
        "остаток на складе",
        "остатки",
    )
    col_note = find_col("примечание", "комплект")
    if col_stock is None and len(header) > 11:
        col_stock = 12 if len(header) > 12 else 11
    if col_vc is None:
        col_vc = 1
    if col_name is None:
        col_name = 2

    # ── Только 1-я таблица ──
    raw_rows = []
    for r in rows_raw[header_idx + 1:]:
        if not r or not any(str(c).strip() for c in r):
            continue
        pn = str(r[0]).strip() if r else ""
        joined = " ".join(str(c).lower() for c in r)
        if pn.upper().startswith("ИТОГО") or "принято на склад" in joined:
            break
        if pn.replace("\\", "") in ("П/Н", "ПН") and "артикул" not in joined:
            break

        def cell(i):
            return str(r[i]).strip() if i is not None and i < len(r) else ""

        vc = cell(col_vc)
        name = cell(col_name)
        note = cell(col_note)
        stock_raw = cell(col_stock)
        stock = _parse_int_cell(stock_raw)
        if not vc and not name:
            continue
        # служебная строка-заголовок второй таблицы
        if not vc and name.lower() in ("наименование", "название"):
            break
        raw_rows.append({
            "vendor_code": vc or None,
            "name": name or None,
            "stock": stock if stock is not None else 0,
            "note": note or None,
            "has_stock_cell": bool(stock_raw),
            "no_sales": not bool(vc),
        })

    # Личный остаток по артикулу (сумма, если vc повторяется)
    personal = {}
    for row in raw_rows:
        vc = row["vendor_code"]
        if not vc:
            continue
        personal[vc] = personal.get(vc, 0) + (row["stock"] or 0)

    # База из Google Sheets; отгрузки Excel списывают поверх (см. own_wh_shipments)
    personal_sheet = dict(personal)
    personal = _apply_own_wh_deductions(personal_sheet)

    # Семьи: основной (есть имя или ячейка остатка) + следующие «голые» артикулы без имени
    # Пример: 044_LK_GT5Pro_black_O (380) → 037_G7Pro_black_O (пусто) делят 380
    families = []  # [{root, members:[], name}]
    cur = None
    for row in raw_rows:
        vc = row["vendor_code"]
        if not vc:
            continue
        is_main = bool(row["name"]) or row["has_stock_cell"]
        if is_main:
            if cur:
                families.append(cur)
            cur = {"root": vc, "members": [vc], "name": row["name"]}
        else:
            if cur is None:
                cur = {"root": vc, "members": [vc], "name": None}
            elif vc not in cur["members"]:
                cur["members"].append(vc)
    if cur:
        families.append(cur)

    # Если артикул попал в несколько семей — объединяем
    parent = {}
    def find(x):
        if parent.get(x, x) != x:
            parent[x] = find(parent[x])
        return parent.get(x, x)
    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for fam in families:
        root = fam["root"]
        parent.setdefault(root, root)
        for m in fam["members"]:
            parent.setdefault(m, m)
            union(root, m)

    root_members = {}
    for vc in personal:
        parent.setdefault(vc, vc)
        r = find(vc)
        root_members.setdefault(r, [])
        if vc not in root_members[r]:
            root_members[r].append(vc)
    # также члены семей без личного остатка
    for fam in families:
        for m in fam["members"]:
            parent.setdefault(m, m)
            r = find(m)
            root_members.setdefault(r, [])
            if m not in root_members[r]:
                root_members[r].append(m)

    auto_by_vendor = {}
    for root, members in root_members.items():
        fam_stock = sum(personal.get(m, 0) for m in members)
        for m in members:
            auto_by_vendor[m] = {
                "stock": personal.get(m, 0),
                "family_stock": fam_stock,
                "family": list(members),
                "root": root,
            }

    # Имена моделей (наименование корня семьи)
    name_by_vc = {}
    for row in raw_rows:
        vc = row["vendor_code"]
        if vc and row.get("name"):
            name_by_vc[vc] = row["name"]

    model_map = get_setting_json("own_wh_model_map", {}) or {}
    by_vendor, models = _apply_own_wh_model_map(auto_by_vendor, personal, name_by_vc, model_map)

    out = []
    seen_vc = set()
    shipped_map = _own_wh_deduction_map()
    received_map = _own_wh_receipt_map()
    for row in raw_rows:
        vc = row["vendor_code"]
        if vc and vc in seen_vc and not row["name"] and not row["has_stock_cell"]:
            continue
        if vc:
            seen_vc.add(vc)
        meta = by_vendor.get(vc, {}) if vc else {}
        sheet_qty = personal_sheet.get(vc, row["stock"] or 0) if vc else (row["stock"] or 0)
        no_sales = bool(row.get("no_sales")) or not bool(vc)
        display_name = row["name"]
        if no_sales and not display_name:
            display_name = "товар без продаж"
        out.append({
            "vendor_code": vc,
            "name": display_name,
            "model_name": meta.get("model_name") or display_name,
            "model_root": meta.get("root"),
            "model_manual": bool(vc and vc in model_map),
            "stock": meta.get("stock", row["stock"] or 0) if vc else (row["stock"] or 0),
            "stock_sheet": sheet_qty,
            "shipped": shipped_map.get(vc, 0) if vc else 0,
            "received": received_map.get(vc, 0) if vc else 0,
            "family_stock": meta.get("family_stock", row["stock"] or 0) if vc else (row["stock"] or 0),
            "family": meta.get("family", [vc] if vc else []),
            "note": row["note"],
            "no_sales": no_sales,
        })

    return {
        "title": title or "Остатки на складе",
        "as_of": as_of,
        "rows": out,
        "by_vendor": by_vendor,
        "models": models,
        "model_map": model_map,
        "personal": personal,
        "personal_sheet": personal_sheet,
        "name_by_vc": name_by_vc,
        "auto_by_vendor": auto_by_vendor,
        "shipments": _own_wh_shipments(),
        "receipts": _own_wh_receipts(),
        "weekly_ledger": _own_wh_weekly_ledger(),
        "updated_at": datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M"),
        "error": None,
    }


def _own_wh_effective_model_map(model_map: dict = None) -> dict:
    """Дефолты Pro Max/038 + ручные привязки из settings (ручные важнее)."""
    saved = model_map if model_map is not None else (get_setting_json("own_wh_model_map", {}) or {})
    if not isinstance(saved, dict):
        saved = {}
    out = {str(k): str(v) for k, v in OWN_WH_MODEL_DEFAULTS.items() if k and v}
    for k, v in saved.items():
        if k and v:
            out[str(k)] = str(v)
    return out


def _apply_own_wh_model_map(auto_by_vendor: dict, personal: dict, name_by_vc: dict, model_map: dict):
    """Пересобирает семьи с учётом ручных привязок артикул → корень модели.
    model_map: {vendor_code: root_vendor_code}. Если root == vendor — отдельно."""
    saved_map = {str(k): str(v) for k, v in (model_map or {}).items() if k and v}
    model_map = _own_wh_effective_model_map(saved_map)

    all_vcs = set(auto_by_vendor.keys()) | set(personal.keys()) | set(model_map.keys())
    # эффективный корень
    root_of = {}
    for vc in all_vcs:
        if vc in model_map:
            root_of[vc] = model_map[vc]
        else:
            root_of[vc] = (auto_by_vendor.get(vc) or {}).get("root") or vc

    def resolve(vc, depth=0):
        if depth > 8:
            return vc
        r = root_of.get(vc, vc)
        if r == vc:
            return vc
        # если корень сам переназначен — идём дальше
        rr = root_of.get(r, r)
        if rr != r:
            return resolve(r, depth + 1)
        return r

    groups = {}
    for vc in all_vcs:
        r = resolve(vc)
        groups.setdefault(r, [])
        if vc not in groups[r]:
            groups[r].append(vc)

    by_vendor = {}
    models = []
    for root, members in sorted(groups.items(), key=lambda x: x[0]):
        members = sorted(members)
        fam_stock = sum(personal.get(m, 0) for m in members)
        model_name = (
            OWN_WH_MODEL_NAME_OVERRIDES.get(root)
            or name_by_vc.get(root)
            or (auto_by_vendor.get(root) or {}).get("model_name")
        )
        if not model_name:
            # любое имя из членов
            for m in members:
                if OWN_WH_MODEL_NAME_OVERRIDES.get(m):
                    model_name = OWN_WH_MODEL_NAME_OVERRIDES[m]
                    break
                if name_by_vc.get(m):
                    model_name = name_by_vc[m]
                    break
        if not model_name:
            model_name = root
        models.append({
            "root": root,
            "name": model_name,
            "members": members,
            "family_stock": fam_stock,
        })
        for m in members:
            by_vendor[m] = {
                "stock": personal.get(m, 0),
                "family_stock": fam_stock,
                "family": members,
                "root": root,
                "model_name": model_name,
                "manual": m in saved_map or m in OWN_WH_MODEL_DEFAULTS,
            }
    models.sort(key=lambda x: (x["name"] or "").lower())
    return by_vendor, models


def _rebuild_own_wh_from_cache():
    """Пересчитывает by_vendor/rows из кэша + актуальной model_map (без повторного Google Sheets)."""
    auto = OWN_WAREHOUSE_CACHE.get("auto_by_vendor") or {}
    personal_sheet = OWN_WAREHOUSE_CACHE.get("personal_sheet")
    if personal_sheet is None:
        personal_sheet = OWN_WAREHOUSE_CACHE.get("personal") or {}
    personal = _apply_own_wh_deductions(personal_sheet)
    name_by_vc = OWN_WAREHOUSE_CACHE.get("name_by_vc") or {}
    if not auto and not personal_sheet and not personal:
        return False
    model_map = get_setting_json("own_wh_model_map", {}) or {}
    by_vendor, models = _apply_own_wh_model_map(auto, personal, name_by_vc, model_map)
    OWN_WAREHOUSE_CACHE["personal"] = personal
    OWN_WAREHOUSE_CACHE["personal_sheet"] = personal_sheet
    OWN_WAREHOUSE_CACHE["by_vendor"] = by_vendor
    OWN_WAREHOUSE_CACHE["models"] = models
    OWN_WAREHOUSE_CACHE["model_map"] = model_map
    OWN_WAREHOUSE_CACHE["shipments"] = _own_wh_shipments()
    OWN_WAREHOUSE_CACHE["receipts"] = _own_wh_receipts()
    OWN_WAREHOUSE_CACHE["weekly_ledger"] = _own_wh_weekly_ledger(
        OWN_WAREHOUSE_CACHE["shipments"], OWN_WAREHOUSE_CACHE["receipts"]
    )
    shipped_map = _own_wh_deduction_map()
    received_map = _own_wh_receipt_map()
    # обновить поля в rows
    rows = OWN_WAREHOUSE_CACHE.get("rows") or []
    new_rows = []
    for row in rows:
        vc = row.get("vendor_code")
        meta = by_vendor.get(vc, {}) if vc else {}
        sheet_qty = personal_sheet.get(vc, row.get("stock_sheet", row.get("stock") or 0)) if vc else (row.get("stock") or 0)
        new_rows.append({
            **row,
            "model_name": meta.get("model_name") or row.get("name"),
            "model_root": meta.get("root"),
            "model_manual": bool(vc and vc in model_map),
            "stock": meta.get("stock", row.get("stock") or 0),
            "stock_sheet": sheet_qty,
            "shipped": shipped_map.get(vc, 0) if vc else 0,
            "received": received_map.get(vc, 0) if vc else 0,
            "family_stock": meta.get("family_stock", row.get("stock") or 0),
            "family": meta.get("family", [vc] if vc else []),
        })
    OWN_WAREHOUSE_CACHE["rows"] = new_rows
    try:
        _persist_own_wh_snapshot()
    except Exception as e:
        logger.warning(f"own-wh snapshot persist: {e}")
    return True


def refresh_own_warehouse_stock():
    OWN_WAREHOUSE_CACHE["syncing"] = True
    OWN_WAREHOUSE_CACHE["error"] = None
    try:
        data = fetch_own_warehouse_stock()
        OWN_WAREHOUSE_CACHE.update(data)
        OWN_WAREHOUSE_CACHE["archives"] = _own_wh_archives()
        OWN_WAREHOUSE_CACHE["sku_aliases"] = _own_wh_sku_aliases()
        try:
            _persist_own_wh_snapshot()
        except Exception as e:
            logger.warning(f"own-wh snapshot persist: {e}")
        OWN_WAREHOUSE_CACHE["syncing"] = False
        logger.info(f"own-warehouse: {len(data['rows'])} rows, as_of={data.get('as_of')}")
    except Exception as e:
        logger.error(f"own-warehouse refresh error: {e}")
        OWN_WAREHOUSE_CACHE["syncing"] = False
        OWN_WAREHOUSE_CACHE["error"] = str(e)

@app.get("/api/own-warehouse-stock")
def get_own_warehouse_stock(refresh: bool = False):
    """Остатки нашего склада (один физический склад для WB и Ozon)."""
    if refresh or not OWN_WAREHOUSE_CACHE.get("rows"):
        if OWN_WAREHOUSE_CACHE.get("syncing"):
            return {**OWN_WAREHOUSE_CACHE, "syncing": True}
        refresh_own_warehouse_stock()
    else:
        # подтянуть актуальные ручные привязки моделей
        if not _rebuild_own_wh_from_cache():
            refresh_own_warehouse_stock()
    archives = OWN_WAREHOUSE_CACHE.get("archives")
    if archives is None:
        archives = _own_wh_archives()
        OWN_WAREHOUSE_CACHE["archives"] = archives
    aliases = OWN_WAREHOUSE_CACHE.get("sku_aliases")
    if aliases is None:
        aliases = _own_wh_sku_aliases()
        OWN_WAREHOUSE_CACHE["sku_aliases"] = aliases
    return {
        "title": OWN_WAREHOUSE_CACHE.get("title"),
        "as_of": OWN_WAREHOUSE_CACHE.get("as_of"),
        "rows": OWN_WAREHOUSE_CACHE.get("rows") or [],
        "by_vendor": OWN_WAREHOUSE_CACHE.get("by_vendor") or {},
        "models": OWN_WAREHOUSE_CACHE.get("models") or [],
        "model_map": OWN_WAREHOUSE_CACHE.get("model_map") or {},
        "shipments": OWN_WAREHOUSE_CACHE.get("shipments") or _own_wh_shipments(),
        "receipts": OWN_WAREHOUSE_CACHE.get("receipts") or _own_wh_receipts(),
        "weekly_ledger": OWN_WAREHOUSE_CACHE.get("weekly_ledger") or _own_wh_weekly_ledger(),
        "channel_summaries": _own_wh_channel_summaries(
            OWN_WAREHOUSE_CACHE.get("shipments") or _own_wh_shipments()
        ),
        "archives": [
            {
                "id": a.get("id"),
                "archived_at": a.get("archived_at"),
                "archived_at_iso": a.get("archived_at_iso"),
                "reason": a.get("reason"),
                "note": a.get("note"),
                "sheet_as_of": a.get("sheet_as_of"),
                "shipments_qty": a.get("shipments_qty"),
                "receipts_qty": a.get("receipts_qty"),
                "shipments_files": a.get("shipments_files"),
                "receipts_files": a.get("receipts_files"),
            }
            for a in (archives or [])[:30]
        ],
        "sku_aliases": aliases,
        "shared_stock": True,
        "marketplaces": ["wb", "ozon"],
        "updated_at": OWN_WAREHOUSE_CACHE.get("updated_at"),
        "error": OWN_WAREHOUSE_CACHE.get("error"),
        "syncing": OWN_WAREHOUSE_CACHE.get("syncing", False),
    }

@app.post("/api/own-warehouse-set-model")
async def own_warehouse_set_model(request: dict):
    """Привязать артикул к модели (корню семьи) или сбросить на авто из таблицы.
    body: {vendor_code, root} — root=null|'' сброс на авто; root=vendor_code — отдельно;
    root=другой артикул — в его семью."""
    vc = (request.get("vendor_code") or "").strip()
    if not vc:
        return {"error": "vendor_code required"}
    root = request.get("root")
    model_map = get_setting_json("own_wh_model_map", {}) or {}
    reset = request.get("reset") or root is None or root == ""
    if reset:
        model_map.pop(vc, None)
    else:
        root = str(root).strip()
        if not root:
            model_map.pop(vc, None)
        else:
            model_map[vc] = root
    if not save_setting_value("own_wh_model_map", model_map):
        return {"error": "не удалось сохранить в settings"}
    # если кэш пуст — подтянем лист
    if not OWN_WAREHOUSE_CACHE.get("auto_by_vendor"):
        refresh_own_warehouse_stock()
    else:
        _rebuild_own_wh_from_cache()
    return {
        "status": "ok",
        "model_map": OWN_WAREHOUSE_CACHE.get("model_map") or model_map,
        "by_vendor": OWN_WAREHOUSE_CACHE.get("by_vendor") or {},
        "models": OWN_WAREHOUSE_CACHE.get("models") or [],
        "rows": OWN_WAREHOUSE_CACHE.get("rows") or [],
    }

@app.post("/api/sync-own-warehouse")
def sync_own_warehouse():
    import threading
    if OWN_WAREHOUSE_CACHE.get("syncing"):
        return {"status": "already_running"}
    threading.Thread(target=refresh_own_warehouse_stock, daemon=True).start()
    return {"status": "started"}


_BARCODE_VENDOR_CACHE = {"ts": 0.0, "map": {}}
_BARCODE_VENDOR_TTL = 3600


def _normalize_barcode(bc: str) -> str:
    s = str(bc or "").strip()
    if s.endswith(".0"):
        head = s[:-2]
        if head.isdigit():
            s = head
    return s


def _barcode_vendor_map(force: bool = False) -> dict[str, dict]:
    """Баркод → {vendor_code, nm_id} из Content API."""
    now = time.time()
    cached = _BARCODE_VENDOR_CACHE.get("map") or {}
    if not force and cached and now - float(_BARCODE_VENDOR_CACHE.get("ts") or 0) < _BARCODE_VENDOR_TTL:
        return cached
    mp: dict[str, dict] = {}
    for row in fetch_all_card_skus():
        bc = _normalize_barcode(row.get("sku"))
        vc = (row.get("vendor_code") or "").strip()
        if not bc or not vc:
            continue
        mp[bc] = {"vendor_code": vc, "nm_id": row.get("nm_id")}
    _BARCODE_VENDOR_CACHE["ts"] = now
    _BARCODE_VENDOR_CACHE["map"] = mp
    return mp


def _resolve_barcode_items(agg_by_bc: dict) -> tuple[list, list]:
    bc_map = _barcode_vendor_map()
    by_vc: dict[str, int] = {}
    unmapped = []
    for bc, qty in (agg_by_bc or {}).items():
        bc_n = _normalize_barcode(bc)
        try:
            q = int(qty or 0)
        except (TypeError, ValueError):
            q = 0
        if not bc_n or q <= 0:
            continue
        meta = bc_map.get(bc_n)
        vc = (meta or {}).get("vendor_code") if meta else None
        if not vc:
            unmapped.append({"barcode": bc_n, "qty": q})
            continue
        by_vc[vc] = by_vc.get(vc, 0) + q
    items = [{"vendor_code": vc, "qty": q} for vc, q in sorted(by_vc.items(), key=lambda x: -x[1])]
    return items, unmapped


def parse_own_wh_shipment_excel(content: bytes, filename: str = "") -> dict:
    """
    Форматы отгрузки / поступления со своего склада:
    1) shk-excel: «Артикул поставщика» + «Количество» (или «Количество, шт»)
    2) WB template: только «Баркод» + «Количество» → маппим в артикулы через Content API
    3) WB-GI лист подбора: «Артикул продавца» (1 строка = 1 шт)
    """
    import io as _io
    try:
        xl = pd.ExcelFile(_io.BytesIO(content))
    except Exception as e:
        return {"error": f"Не удалось прочитать Excel: {e}"}

    best = None
    for sheet in xl.sheet_names:
        try:
            df_raw = pd.read_excel(xl, sheet_name=sheet, header=None, dtype=object)
        except Exception:
            continue
        if df_raw is None or df_raw.empty:
            continue

        header_row = None
        col_vc = col_qty = None
        kind = None
        for i in range(min(12, len(df_raw))):
            vals = [str(v).strip().lower() if v is not None and str(v) != "nan" else "" for v in list(df_raw.iloc[i].values)]
            joined = " | ".join(vals)
            has_vendor_col = any(
                "артикул поставщика" in v or v == "артикул продавца"
                for v in vals
            )
            # shk / поставка: артикул + количество
            vc_i = next(
                (
                    j
                    for j, v in enumerate(vals)
                    if "артикул поставщика" in v
                    or v == "артикул продавца"
                    or (
                        v.startswith("артикул")
                        and "wildberries" not in v
                        and "баркод" not in v
                    )
                ),
                None,
            )
            qty_i = next((j for j, v in enumerate(vals) if v.startswith("количество")), None)
            if vc_i is not None and qty_i is not None:
                header_row, col_vc, col_qty, kind = i, vc_i, qty_i, "shk"
                break
            # лист подбора WB-GI
            if "артикул продавца" in joined and ("стикер" in joined or "баркод" in joined):
                vc_i = next((j for j, v in enumerate(vals) if v == "артикул продавца"), None)
                if vc_i is not None:
                    header_row, col_vc, col_qty, kind = i, vc_i, None, "picking"
                    break
            # новый шаблон WB: только баркод + количество
            bc_i = next((j for j, v in enumerate(vals) if v in ("баркод", "barcode", "штрихкод")), None)
            if bc_i is not None and qty_i is not None and not has_vendor_col:
                header_row, col_vc, col_qty, kind = i, bc_i, qty_i, "barcode"
                break
            # fallback: shk headers slightly different
            if "артикул поставщика" in joined and "количество" in joined:
                vc_i = next((j for j, v in enumerate(vals) if "артикул" in v), None)
                qty_i = next((j for j, v in enumerate(vals) if "количество" in v), None)
                if vc_i is not None and qty_i is not None:
                    header_row, col_vc, col_qty, kind = i, vc_i, qty_i, "shk"
                    break

        if header_row is None:
            continue

        agg = {}
        unmapped_barcodes = []
        for i in range(header_row + 1, len(df_raw)):
            row = list(df_raw.iloc[i].values)
            if col_vc >= len(row):
                continue
            raw_key = str(row[col_vc] or "").strip()
            if not raw_key or raw_key.lower() in ("nan", "none", "итого"):
                continue
            if kind == "picking":
                qty = 1
                vc = raw_key
            elif kind == "barcode":
                qty = _parse_int_cell(row[col_qty] if col_qty is not None and col_qty < len(row) else None) or 0
                if qty <= 0:
                    continue
                agg[_normalize_barcode(raw_key)] = agg.get(_normalize_barcode(raw_key), 0) + qty
                continue
            else:
                vc = raw_key
                qty = _parse_int_cell(row[col_qty] if col_qty is not None and col_qty < len(row) else None) or 0
            if qty <= 0:
                continue
            agg[vc] = agg.get(vc, 0) + qty

        if kind == "barcode":
            items, unmapped_barcodes = _resolve_barcode_items(agg)
        else:
            items = [{"vendor_code": vc, "qty": q} for vc, q in sorted(agg.items(), key=lambda x: -x[1])]

        cand = {
            "kind": kind,
            "sheet": sheet,
            "items": items,
            "total_qty": sum(i["qty"] for i in items),
            "articles": len(items),
            "unmapped_barcodes": unmapped_barcodes[:50],
        }
        if best is None or cand["total_qty"] > best["total_qty"]:
            best = cand

    if not best or not best["items"]:
        hint = (
            "Не нашёл артикулы в файле. Ожидаю shk-excel "
            "(Артикул поставщика + Количество), шаблон WB (Баркод + Количество) "
            "или лист подбора WB-GI (Артикул продавца)."
        )
        unmapped = (best or {}).get("unmapped_barcodes") or []
        if unmapped:
            sample = ", ".join(f"{u['barcode']}×{u['qty']}" for u in unmapped[:5])
            hint += f" Не сопоставлены баркоды ({len(unmapped)}): {sample}."
        return {"error": hint}
    best["filename"] = filename or ""
    return best


def _own_wh_join_offer_id(parts: list) -> str:
    """Склеивает артикул Ozon, который в PDF рвётся на строки (046_LK11_Promax_g + rey_O)."""
    return "".join(str(p or "").strip() for p in parts if str(p or "").strip())


def _parse_ozon_picking_words(pages_words: list) -> dict:
    """Лист подбора Ozon FBS из слов с координатами.
    pages_words: [ [(x0, y0, x1, y1, text), ...], ... ]
    Колонки: № | Товар | Артикул | Кол-во. Артикул часто в 1–2 строки."""
    items_agg = {}
    meta = {"date": None, "warehouse": None, "shipments": None, "pages": 0}
    date_re = re.compile(r"(\d{1,2}[./]\d{1,2}[./]\d{2,4})")

    for words in pages_words or []:
        if not words:
            continue
        meta["pages"] += 1
        texts = [(float(w[0]), float(w[1]), str(w[4] or "").strip()) for w in words if len(w) >= 5]
        texts = [t for t in texts if t[2]]
        if not texts:
            continue

        flat = " ".join(t[2] for t in texts)
        if meta["date"] is None:
            m = date_re.search(flat)
            if m:
                meta["date"] = m.group(1)
        if meta["warehouse"] is None:
            wm = re.search(r"Склад:\s*(.+?)(?:\s+Служба|\s+№|$)", flat)
            if wm:
                meta["warehouse"] = wm.group(1).strip()
        if meta["shipments"] is None:
            sm = re.search(r"отправлений:\s*(\d+)", flat, re.I)
            if sm:
                meta["shipments"] = int(sm.group(1))

        art_h = next((t for t in texts if t[2].lower() in ("артикул", "артикул продавца", "offer_id")), None)
        if not art_h:
            continue
        qty_cands = [
            t for t in texts
            if t[2].lower() in ("кол-во", "количество", "qty") and abs(t[1] - art_h[1]) < 10
        ]
        qty_h = next((t for t in qty_cands if t[2].lower() == "кол-во"), None) or (qty_cands[0] if qty_cands else None)
        if not qty_h:
            continue
        article_x = art_h[0]
        qty_x = qty_h[0]
        header_y = min(art_h[1], qty_h[1])
        left_max = min(70.0, article_x * 0.22)

        below = [t for t in texts if t[1] > header_y + 4]
        anchors = sorted(
            ((t[1], int(t[2])) for t in below if t[0] < left_max and t[2].isdigit() and 1 <= int(t[2]) <= 9999),
            key=lambda x: x[0],
        )
        if not anchors:
            continue

        for i, (ay, _n) in enumerate(anchors):
            y0 = (anchors[i - 1][0] + ay) / 2 if i else header_y
            y1 = (ay + anchors[i + 1][0]) / 2 if i + 1 < len(anchors) else 1e9
            row = [t for t in below if y0 < t[1] <= y1]
            art_parts = [t[2] for t in sorted(row, key=lambda t: (t[1], t[0])) if article_x - 12 <= t[0] < qty_x - 12]
            qty_parts = [t[2] for t in row if t[0] >= qty_x - 12]
            vc = _own_wh_join_offer_id(art_parts)
            qty = None
            for p in qty_parts:
                q = _parse_int_cell(p)
                if q is not None and q > 0:
                    qty = q
                    break
            if not vc or not qty:
                continue
            if vc.lower() in ("артикул", "товар", "кол-во", "количество"):
                continue
            items_agg[vc] = items_agg.get(vc, 0) + qty

    items = [{"vendor_code": vc, "qty": q} for vc, q in sorted(items_agg.items(), key=lambda x: -x[1])]
    return {
        "kind": "ozon_picking",
        "channel": "ozon_fbs",
        "items": items,
        "total_qty": sum(i["qty"] for i in items),
        "articles": len(items),
        "date": meta["date"],
        "warehouse": meta["warehouse"],
        "shipments_count": meta["shipments"],
        "pages": meta["pages"],
    }


def parse_own_wh_ozon_picking_pdf(content: bytes, filename: str = "") -> dict:
    """PDF «Лист подбора» Ozon FBS (FPDF): № / Товар / Артикул / Кол-во."""
    if not content:
        return {"error": "Пустой PDF"}
    try:
        import fitz
    except ImportError:
        return {"error": "Сервер не умеет читать PDF (нужен pymupdf). Попроси обновить деплой."}
    try:
        doc = fitz.open(stream=content, filetype="pdf")
    except Exception as e:
        return {"error": f"Не удалось открыть PDF: {e}"}
    pages = []
    head_text = ""
    try:
        for page in doc:
            words = page.get_text("words") or []
            pages.append([(w[0], w[1], w[2], w[3], w[4]) for w in words])
            if not head_text:
                head_text = page.get_text("text") or ""
    finally:
        doc.close()

    parsed = _parse_ozon_picking_words(pages)
    parsed["filename"] = filename or ""
    low = (head_text + " " + (filename or "")).lower()
    looks_ozon = (
        "ozon" in low
        or "озон" in low
        or "лист подбора" in low
        or "отправлений" in low
        or parsed.get("items")
    )
    if not parsed.get("items"):
        return {
            "error": (
                "Не нашёл таблицу «Артикул / Кол-во» в PDF. "
                "Ожидаю лист подбора Ozon (FBS): колонки №, Товар, Артикул, Кол-во."
            )
        }
    if not looks_ozon:
        parsed["kind"] = "ozon_picking"
    return parsed


def parse_own_wh_shipment_file(content: bytes, filename: str = "") -> dict:
    """Excel отгрузки WB/Ozon или PDF листа подбора Ozon."""
    name = (filename or "").lower()
    if name.endswith(".pdf") or (content[:5] == b"%PDF-"):
        return parse_own_wh_ozon_picking_pdf(content, filename)
    return parse_own_wh_shipment_excel(content, filename)


def _own_wh_looks_like_vendor_code(vc: str) -> bool:
    """Грубая проверка: артикул продавца, не число/дата/заголовок."""
    s = str(vc or "").strip()
    if not s or len(s) < 2 or len(s) > 80:
        return False
    low = s.lower()
    if low in ("nan", "none", "итого", "артикул", "количество", "qty", "sku", "offer_id"):
        return False
    if low.startswith("артикул") or low.startswith("количество"):
        return False
    # чистое число без букв — скорее qty/nm, не наш vendor_code
    if re.fullmatch(r"\d+([.,]\d+)?", s):
        return False
    return True


def parse_own_wh_receipt_excel(content: bytes, filename: str = "") -> dict:
    """Поступление / приёмка.
    1) Как отгрузка: заголовки «Артикул…» + «Количество»
    2) Простой файл «Приемка …xlsx»: без шапки, колонка A = артикул, B = кол-во
    """
    # Сначала пробуем общий парсер с заголовками
    headed = parse_own_wh_shipment_excel(content, filename)
    if not headed.get("error") and headed.get("items"):
        headed["kind"] = "receipt"
        return headed

    import io as _io
    try:
        xl = pd.ExcelFile(_io.BytesIO(content))
    except Exception as e:
        return {"error": f"Не удалось прочитать Excel: {e}"}

    best = None
    for sheet in xl.sheet_names:
        try:
            df_raw = pd.read_excel(xl, sheet_name=sheet, header=None, dtype=object)
        except Exception:
            continue
        if df_raw is None or df_raw.empty or df_raw.shape[1] < 2:
            continue

        agg = {}
        for i in range(len(df_raw)):
            row = list(df_raw.iloc[i].values)
            if len(row) < 2:
                continue
            # ищем первую «артикулоподобную» ячейку и ближайшее число справа
            vc = None
            qty = None
            for j, cell in enumerate(row):
                s = str(cell or "").strip()
                if not s or s.lower() in ("nan", "none"):
                    continue
                if _own_wh_looks_like_vendor_code(s):
                    vc = s
                    for k in range(j + 1, len(row)):
                        q = _parse_int_cell(row[k])
                        if q is not None and q > 0:
                            qty = q
                            break
                    break
            if not vc or not qty:
                continue
            agg[vc] = agg.get(vc, 0) + qty

        items = [{"vendor_code": vc, "qty": q} for vc, q in sorted(agg.items(), key=lambda x: -x[1])]
        cand = {
            "kind": "receipt",
            "sheet": sheet,
            "items": items,
            "total_qty": sum(i["qty"] for i in items),
            "articles": len(items),
        }
        if items and (best is None or cand["total_qty"] > best["total_qty"]):
            best = cand

    if not best or not best["items"]:
        err = headed.get("error") if isinstance(headed, dict) else None
        return {
            "error": err or (
                "Не нашёл артикулы в приёмке. Ожидаю файл «Артикул + Количество» "
                "(можно без заголовков, как «Приемка 19.08.xlsx»)."
            )
        }
    best["filename"] = filename or ""
    return best


def _own_wh_shipment_channel(sh: dict) -> str:
    """fbw/fbs = WB; ozon_fbo/ozon_fbs = Ozon. Все списывают один физический склад."""
    ch = str((sh or {}).get("channel") or "").strip().lower()
    if ch in OWN_WH_CHANNELS:
        return ch
    kind = str((sh or {}).get("kind") or "")
    # shk-excel и лист подбора WB-GI — это поставки на склады WB
    if kind in ("shk", "picking"):
        return "fbw"
    if kind == "ozon_picking":
        return "ozon_fbs"
    return "fbw"


def _own_wh_channel_summaries(shipments: list | None = None) -> dict:
    """Сводки по артикулам: WB (fbw/fbs) и Ozon (ozon_fbo/ozon_fbs)."""
    shipments = shipments if shipments is not None else _own_wh_shipments()
    buckets = {ch: {} for ch in OWN_WH_CHANNELS}
    meta = {ch: {"files": 0, "total_qty": 0} for ch in OWN_WH_CHANNELS}
    aliases = _own_wh_sku_aliases()
    for sh in shipments:
        ch = _own_wh_shipment_channel(sh)
        if ch not in meta:
            ch = "fbw"
        meta[ch]["files"] += 1
        meta[ch]["total_qty"] += int(sh.get("total_qty") or 0)
        for it in sh.get("items") or []:
            vc = _own_wh_canonical_vc(str(it.get("vendor_code") or "").strip(), aliases)
            if not vc:
                continue
            try:
                qty = int(it.get("qty") or 0)
            except Exception:
                qty = 0
            if qty <= 0:
                continue
            buckets[ch][vc] = buckets[ch].get(vc, 0) + qty
    out = {}
    for ch in OWN_WH_CHANNELS:
        items = [
            {"vendor_code": vc, "qty": q}
            for vc, q in sorted(buckets[ch].items(), key=lambda x: (-x[1], x[0]))
        ]
        out[ch] = {
            **meta[ch],
            "articles": len(items),
            "items": items,
        }
    return out


def _save_own_wh_shipments(shipments: list) -> bool:
    return save_setting_value(OWN_WH_SHIPMENTS_KEY, shipments)


def _save_own_wh_receipts(receipts: list) -> bool:
    return save_setting_value(OWN_WH_RECEIPTS_KEY, receipts)


def _own_wh_now_stamp():
    """Мск-метка + ISO (UTC) для недель и истории."""
    utc = datetime.now(timezone.utc)
    msk = utc + timedelta(hours=3)
    return msk.strftime("%d.%m.%Y %H:%M"), utc.isoformat()


def _own_wh_response_payload() -> dict:
    ships = OWN_WAREHOUSE_CACHE.get("shipments") or _own_wh_shipments()
    receipts = OWN_WAREHOUSE_CACHE.get("receipts") or _own_wh_receipts()
    archives = OWN_WAREHOUSE_CACHE.get("archives")
    if archives is None:
        archives = _own_wh_archives()
    return {
        "shipments": ships[:40],
        "receipts": receipts[:40],
        "weekly_ledger": _own_wh_weekly_ledger(ships, receipts),
        "channel_summaries": _own_wh_channel_summaries(ships),
        "archives": [
            {
                "id": a.get("id"),
                "archived_at": a.get("archived_at"),
                "archived_at_iso": a.get("archived_at_iso"),
                "reason": a.get("reason"),
                "note": a.get("note"),
                "sheet_as_of": a.get("sheet_as_of"),
                "shipments_qty": a.get("shipments_qty"),
                "receipts_qty": a.get("receipts_qty"),
                "shipments_files": a.get("shipments_files"),
                "receipts_files": a.get("receipts_files"),
            }
            for a in (archives or [])[:30]
        ],
        "sku_aliases": OWN_WAREHOUSE_CACHE.get("sku_aliases") or _own_wh_sku_aliases(),
        "shared_stock": True,
        "marketplaces": ["wb", "ozon"],
        "rows": OWN_WAREHOUSE_CACHE.get("rows") or [],
        "by_vendor": OWN_WAREHOUSE_CACHE.get("by_vendor") or {},
        "as_of": OWN_WAREHOUSE_CACHE.get("as_of"),
        "updated_at": OWN_WAREHOUSE_CACHE.get("updated_at"),
    }


@app.post("/api/own-warehouse-upload-shipment")
async def own_warehouse_upload_shipment(
    files: list[UploadFile] = File(...),
    channel: str = Form("auto"),
):
    """Загрузка Excel отгрузки → списание с общего склада (WB или Ozon).
    channel: auto|fbw|fbs|ozon_fbo|ozon_fbs"""
    if not files:
        raise HTTPException(status_code=400, detail="files required")
    ch_req = str(channel or "auto").strip().lower()
    if ch_req not in ("auto",) + OWN_WH_CHANNELS:
        ch_req = "auto"

    if not OWN_WAREHOUSE_CACHE.get("rows") and not OWN_WAREHOUSE_CACHE.get("personal_sheet"):
        try:
            refresh_own_warehouse_stock()
        except Exception:
            pass

    shipments = _own_wh_shipments()
    applied = []
    errors = []
    created_at, created_iso = _own_wh_now_stamp()
    for f in files:
        try:
            content = await f.read()
        except Exception as e:
            errors.append({"filename": f.filename, "error": str(e)})
            continue
        parsed = parse_own_wh_shipment_file(content, f.filename or "")
        if parsed.get("error"):
            errors.append({"filename": f.filename, "error": parsed["error"]})
            continue
        items = _own_wh_normalize_items(parsed.get("items") or [])
        if not items:
            errors.append({"filename": f.filename, "error": "нет артикулов после нормализации"})
            continue
        total_qty = sum(int(it["qty"]) for it in items)
        if ch_req in OWN_WH_CHANNELS:
            ch = ch_req
        elif parsed.get("channel") in OWN_WH_CHANNELS:
            ch = parsed["channel"]
        elif parsed.get("kind") == "ozon_picking":
            ch = "ozon_fbs"
        else:
            ch = "fbw" if parsed.get("kind") in ("shk", "picking") else "fbw"
        sid = f"sh_{int(time.time())}_{len(shipments)}_{len(applied)}"
        entry = {
            "id": sid,
            "filename": f.filename or parsed.get("filename") or "",
            "kind": parsed.get("kind"),
            "channel": ch,
            "marketplace": "ozon" if ch.startswith("ozon") else "wb",
            "created_at": created_at,
            "created_at_iso": created_iso,
            "items": items,
            "total_qty": total_qty,
            "articles": len(items),
            "unmapped_barcodes": parsed.get("unmapped_barcodes") or [],
            "doc_date": parsed.get("date"),
            "warehouse": parsed.get("warehouse"),
        }
        shipments.insert(0, entry)
        applied.append(entry)

    shipments = shipments[:OWN_WH_DOCS_KEEP]
    if applied and not _save_own_wh_shipments(shipments):
        raise HTTPException(status_code=500, detail="Не удалось сохранить списания в settings")

    OWN_WAREHOUSE_CACHE["shipments"] = shipments
    if not _rebuild_own_wh_from_cache():
        refresh_own_warehouse_stock()

    payload = _own_wh_response_payload()
    return {
        "status": "ok",
        "applied": [
            {
                "id": a["id"],
                "filename": a["filename"],
                "kind": a["kind"],
                "channel": a.get("channel"),
                "marketplace": a.get("marketplace"),
                "total_qty": a["total_qty"],
                "articles": a["articles"],
                "items": a["items"][:40],
                "unmapped_barcodes": (a.get("unmapped_barcodes") or [])[:20],
            }
            for a in applied
        ],
        "errors": errors,
        **payload,
        "note": "Списано с общего склада (WB+Ozon). Google Sheets сам не меняется — в пятницу после пересчёта нажми «Инвентаризация».",
    }


@app.post("/api/own-warehouse-upload-receipt")
async def own_warehouse_upload_receipt(
    files: list[UploadFile] = File(...),
    note: str = Form(""),
):
    """Поступление товара (Excel: Артикул + Количество) — приход на «Наш склад»."""
    file_list = files
    if not file_list:
        raise HTTPException(status_code=400, detail="files required")

    if not OWN_WAREHOUSE_CACHE.get("rows") and not OWN_WAREHOUSE_CACHE.get("personal_sheet"):
        try:
            refresh_own_warehouse_stock()
        except Exception:
            pass

    receipts = _own_wh_receipts()
    applied = []
    errors = []
    created_at, created_iso = _own_wh_now_stamp()
    note_s = str(note or "").strip()
    for f in file_list:
        try:
            content = await f.read()
        except Exception as e:
            errors.append({"filename": f.filename, "error": str(e)})
            continue
        parsed = parse_own_wh_receipt_excel(content, f.filename or "")
        if parsed.get("error"):
            errors.append({"filename": f.filename, "error": parsed["error"]})
            continue
        items = _own_wh_normalize_items(parsed.get("items") or [])
        if not items:
            errors.append({"filename": f.filename, "error": "нет артикулов после нормализации"})
            continue
        total_qty = sum(int(it["qty"]) for it in items)
        rid = f"rc_{int(time.time())}_{len(receipts)}_{len(applied)}"
        entry = {
            "id": rid,
            "filename": f.filename or parsed.get("filename") or "",
            "kind": "receipt",
            "note": note_s,
            "created_at": created_at,
            "created_at_iso": created_iso,
            "items": items,
            "total_qty": total_qty,
            "articles": len(items),
            "unmapped_barcodes": parsed.get("unmapped_barcodes") or [],
        }
        receipts.insert(0, entry)
        applied.append(entry)

    receipts = receipts[:OWN_WH_DOCS_KEEP]
    if applied and not _save_own_wh_receipts(receipts):
        raise HTTPException(status_code=500, detail="Не удалось сохранить поступления")

    OWN_WAREHOUSE_CACHE["receipts"] = receipts
    if not _rebuild_own_wh_from_cache():
        refresh_own_warehouse_stock()

    return {
        "status": "ok",
        "applied": [
            {
                "id": a["id"],
                "filename": a["filename"],
                "total_qty": a["total_qty"],
                "articles": a["articles"],
                "items": a["items"][:40],
                "unmapped_barcodes": (a.get("unmapped_barcodes") or [])[:20],
            }
            for a in applied
        ],
        "errors": errors,
        **_own_wh_response_payload(),
        "note": "Поступление учтено на сайте (+). Google Sheets сам не меняется — когда внесёшь приход в таблицу, сбрось поступления.",
    }


@app.post("/api/own-warehouse-add-receipt")
async def own_warehouse_add_receipt(request: dict):
    """Ручное поступление: {items:[{vendor_code,qty}], note?} или text: 'артикул qty' по строкам."""
    if not OWN_WAREHOUSE_CACHE.get("rows") and not OWN_WAREHOUSE_CACHE.get("personal_sheet"):
        try:
            refresh_own_warehouse_stock()
        except Exception:
            pass

    items_in = request.get("items") or []
    text = str(request.get("text") or "").strip()
    note = str(request.get("note") or "").strip()
    agg = {}
    if items_in:
        for it in items_in:
            vc = str((it or {}).get("vendor_code") or "").strip()
            try:
                qty = int((it or {}).get("qty") or 0)
            except Exception:
                qty = 0
            if vc and qty > 0:
                agg[vc] = agg.get(vc, 0) + qty
    elif text:
        for line in text.splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.replace("\t", " ").replace(";", " ").split()
            if len(parts) < 2:
                continue
            vc = parts[0].strip()
            try:
                qty = int(float(parts[-1].replace(",", ".")))
            except Exception:
                continue
            if vc and qty > 0:
                agg[vc] = agg.get(vc, 0) + qty
    else:
        raise HTTPException(status_code=400, detail="items или text required")

    if not agg:
        raise HTTPException(status_code=400, detail="Не удалось разобрать артикулы")

    items = _own_wh_normalize_items(
        [{"vendor_code": vc, "qty": q} for vc, q in agg.items()]
    )
    created_at, created_iso = _own_wh_now_stamp()
    receipts = _own_wh_receipts()
    rid = f"rc_{int(time.time())}_{len(receipts)}_m"
    entry = {
        "id": rid,
        "filename": note or "ручное поступление",
        "kind": "manual",
        "note": note,
        "created_at": created_at,
        "created_at_iso": created_iso,
        "items": items,
        "total_qty": sum(i["qty"] for i in items),
        "articles": len(items),
    }
    receipts.insert(0, entry)
    receipts = receipts[:OWN_WH_DOCS_KEEP]
    if not _save_own_wh_receipts(receipts):
        raise HTTPException(status_code=500, detail="Не удалось сохранить")
    OWN_WAREHOUSE_CACHE["receipts"] = receipts
    if not _rebuild_own_wh_from_cache():
        refresh_own_warehouse_stock()
    return {
        "status": "ok",
        "applied": [entry],
        **_own_wh_response_payload(),
    }


@app.post("/api/own-warehouse-undo-shipment")
async def own_warehouse_undo_shipment(request: dict):
    """Отменить одно списание по id или все (all=true).
    all=true + archive=true (по умолчанию) — сначала в архив."""
    shipments = _own_wh_shipments()
    if request.get("all"):
        if request.get("archive", True) and shipments:
            try:
                _own_wh_archive_active(
                    reason=str(request.get("reason") or "clear_shipments"),
                    note=str(request.get("note") or "Сброс списаний"),
                    include_shipments=True,
                    include_receipts=False,
                )
            except Exception as e:
                raise HTTPException(status_code=500, detail=str(e))
        shipments = []
    else:
        sid = str(request.get("id") or "").strip()
        if not sid:
            raise HTTPException(status_code=400, detail="id required (или all=true)")
        shipments = [s for s in shipments if str(s.get("id")) != sid]
    if not _save_own_wh_shipments(shipments):
        raise HTTPException(status_code=500, detail="Не удалось сохранить")
    OWN_WAREHOUSE_CACHE["shipments"] = shipments
    if not _rebuild_own_wh_from_cache():
        refresh_own_warehouse_stock()
    return {"status": "ok", **_own_wh_response_payload()}


@app.post("/api/own-warehouse-undo-receipt")
async def own_warehouse_undo_receipt(request: dict):
    """Отменить одно поступление по id или все (all=true)."""
    receipts = _own_wh_receipts()
    if request.get("all"):
        if request.get("archive", True) and receipts:
            try:
                _own_wh_archive_active(
                    reason=str(request.get("reason") or "clear_receipts"),
                    note=str(request.get("note") or "Сброс поступлений"),
                    include_shipments=False,
                    include_receipts=True,
                )
            except Exception as e:
                raise HTTPException(status_code=500, detail=str(e))
        receipts = []
    else:
        rid = str(request.get("id") or "").strip()
        if not rid:
            raise HTTPException(status_code=400, detail="id required (или all=true)")
        receipts = [s for s in receipts if str(s.get("id")) != rid]
    if not _save_own_wh_receipts(receipts):
        raise HTTPException(status_code=500, detail="Не удалось сохранить")
    OWN_WAREHOUSE_CACHE["receipts"] = receipts
    if not _rebuild_own_wh_from_cache():
        refresh_own_warehouse_stock()
    return {"status": "ok", **_own_wh_response_payload()}


@app.post("/api/own-warehouse-friday-reset")
async def own_warehouse_friday_reset(request: dict = None):
    """Пятничная инвентаризация: подтянуть Sheets → архивировать оверлеи → очистить активные.
    Сотрудник сначала правит реальные остатки в Google Sheets, потом жмёт эту кнопку."""
    request = request or {}
    note = str(request.get("note") or "").strip()
    refresh_sheets = request.get("refresh_sheets", True)

    if refresh_sheets:
        try:
            refresh_own_warehouse_stock()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Не удалось обновить Sheets: {e}")

    ships = _own_wh_shipments()
    receipts = _own_wh_receipts()
    archived = None
    if ships or receipts:
        try:
            archived = _own_wh_archive_active(reason="friday", note=note or "Пятничная инвентаризация")
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    if not _save_own_wh_shipments([]):
        raise HTTPException(status_code=500, detail="Не удалось очистить списания")
    if not _save_own_wh_receipts([]):
        raise HTTPException(status_code=500, detail="Не удалось очистить поступления")

    OWN_WAREHOUSE_CACHE["shipments"] = []
    OWN_WAREHOUSE_CACHE["receipts"] = []
    OWN_WAREHOUSE_CACHE["archives"] = _own_wh_archives()
    if not _rebuild_own_wh_from_cache():
        refresh_own_warehouse_stock()

    return {
        "status": "ok",
        "archived": {
            "id": (archived or {}).get("id"),
            "archived_at": (archived or {}).get("archived_at"),
            "shipments_qty": (archived or {}).get("shipments_qty", 0),
            "receipts_qty": (archived or {}).get("receipts_qty", 0),
            "shipments_files": (archived or {}).get("shipments_files", 0),
            "receipts_files": (archived or {}).get("receipts_files", 0),
        } if archived else None,
        "note": "База = Google Sheets. Активные отгрузки/поступления ушли в архив.",
        **_own_wh_response_payload(),
    }


@app.get("/api/own-warehouse-archive")
def own_warehouse_archive_list():
    """Список архивов инвентаризаций / сбросов (без полного состава файлов)."""
    archives = _own_wh_archives()
    return {
        "archives": [
            {
                "id": a.get("id"),
                "archived_at": a.get("archived_at"),
                "archived_at_iso": a.get("archived_at_iso"),
                "reason": a.get("reason"),
                "note": a.get("note"),
                "sheet_as_of": a.get("sheet_as_of"),
                "shipments_qty": a.get("shipments_qty"),
                "receipts_qty": a.get("receipts_qty"),
                "shipments_files": a.get("shipments_files"),
                "receipts_files": a.get("receipts_files"),
            }
            for a in archives
        ]
    }


@app.get("/api/own-warehouse-archive/{archive_id}")
def own_warehouse_archive_detail(archive_id: str):
    """Полный архив за дату (файлы отгрузок/поступлений для аудита)."""
    aid = str(archive_id or "").strip()
    for a in _own_wh_archives():
        if str(a.get("id")) == aid:
            return a
    raise HTTPException(status_code=404, detail="archive not found")


@app.post("/api/own-warehouse-sku-aliases")
async def own_warehouse_sku_aliases(request: dict):
    """Сохранить карту алиасов Ozon/других SKU → канонический артикул WB.
    body: {aliases: {ozon_sku: wb_vendor_code}} или text: 'ozon_sku = wb_article' по строкам.
    replace=true — заменить целиком, иначе merge."""
    aliases = _own_wh_sku_aliases()
    if request.get("replace"):
        aliases = {}
    incoming = request.get("aliases")
    if isinstance(incoming, dict):
        for k, v in incoming.items():
            ak = str(k or "").strip()
            cv = str(v or "").strip()
            if not ak:
                continue
            if not cv:
                aliases.pop(ak, None)
            else:
                aliases[ak] = cv
    text = str(request.get("text") or "").strip()
    if text:
        for line in text.splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                left, right = line.split("=", 1)
            elif "\t" in line:
                left, right = line.split("\t", 1)
            else:
                parts = line.split()
                if len(parts) < 2:
                    continue
                left, right = parts[0], parts[1]
            ak = left.strip()
            cv = right.strip()
            if ak and cv:
                aliases[ak] = cv
    if not save_setting_value(OWN_WH_SKU_ALIASES_KEY, aliases):
        raise HTTPException(status_code=500, detail="Не удалось сохранить алиасы")
    OWN_WAREHOUSE_CACHE["sku_aliases"] = aliases
    if not _rebuild_own_wh_from_cache():
        refresh_own_warehouse_stock()
    return {"status": "ok", "sku_aliases": aliases, **_own_wh_response_payload()}


# ---------- Новые остатки: срок доставки как на витрине WB (по городам) ----------
# Ячейка = остаток на складе, с которого WB везёт в город, + часы витрины (time1+time2).
# >40ч жёлтый, >60ч красный. Это не отчёт warehouse_remains и не MKeeper.
NEW_STOCK_WARN_H = 40
NEW_STOCK_BAD_H = 60
NEW_STOCK_LAYOUT_KEY = "new_stock_layout"
NEW_STOCK_CACHE = {"payload": None, "syncing": False, "error": None, "updated_at": None}
NEW_STOCK_CITIES = [
    {"id": "nsk", "name": "Новосибирск", "group": "Сибирь и ДВ", "dest": -364763, "lat": 55.0302, "lon": 82.9204},
    {"id": "krs", "name": "Красноярск", "group": "Сибирь и ДВ", "dest": 12356481, "lat": 56.0106, "lon": 92.8526},
    {"id": "irk", "name": "Иркутск", "group": "Сибирь и ДВ", "dest": -5827722, "lat": 52.2864, "lon": 104.2807},
    {"id": "omsk", "name": "Омск", "group": "Сибирь и ДВ", "dest": -3902444, "lat": 54.9885, "lon": 73.3242},
    {"id": "krd", "name": "Краснодар", "group": "Юг и СК", "dest": 12358062, "lat": 45.0355, "lon": 38.9753},
    {"id": "rnd", "name": "Ростов-на-Дону", "group": "Юг и СК", "dest": -2228364, "lat": 47.2221, "lon": 39.7203},
    {"id": "vlg", "name": "Волгоград", "group": "Юг и СК", "dest": -4039473, "lat": 48.7080, "lon": 44.5133},
    {"id": "msk", "name": "Москва", "group": "Центр", "dest": -1257786, "lat": 55.7558, "lon": 37.6173},
    {"id": "vrn", "name": "Воронеж", "group": "Центр", "dest": 12358283, "lat": 51.6720, "lon": 39.1843},
    {"id": "ryz", "name": "Рязань", "group": "Центр", "dest": -5817683, "lat": 54.6269, "lon": 39.6916},
    {"id": "spb", "name": "Санкт-Петербург", "group": "Северо-Запад", "dest": -1198055, "lat": 59.9343, "lon": 30.3351},
    {"id": "vld", "name": "Вологда", "group": "Северо-Запад", "dest": 123586880, "lat": 59.2205, "lon": 39.8915},
    {"id": "arh", "name": "Архангельск", "group": "Северо-Запад", "dest": 123589924, "lat": 64.5393, "lon": 40.5187},
    {"id": "ekb", "name": "Екатеринбург", "group": "Урал", "dest": -5818883, "lat": 56.8389, "lon": 60.6057},
    {"id": "chel", "name": "Челябинск", "group": "Урал", "dest": -1581743, "lat": 55.1644, "lon": 61.4368},
    {"id": "tmn", "name": "Тюмень", "group": "Урал", "dest": 12358475, "lat": 57.1522, "lon": 65.5272},
    {"id": "kzn", "name": "Казань", "group": "Волга", "dest": -2133462, "lat": 55.7963, "lon": 49.1088},
    {"id": "nnv", "name": "Нижний Новгород", "group": "Волга", "dest": 12358579, "lat": 56.2965, "lon": 43.9361},
    {"id": "smr", "name": "Самара", "group": "Волга", "dest": -283781, "lat": 53.1959, "lon": 50.1002},
    {"id": "prm", "name": "Пермь", "group": "Волга", "dest": 12358361, "lat": 58.0105, "lon": 56.2502},
]
NEW_STOCK_GROUPS = [
    {"id": "sib", "name": "Сибирь и ДВ", "city_ids": ["nsk", "krs", "irk", "omsk"]},
    {"id": "south", "name": "Юг и СК", "city_ids": ["krd", "rnd", "vlg"]},
    {"id": "center", "name": "Центр", "city_ids": ["msk", "vrn", "ryz"]},
    {"id": "nw", "name": "Северо-Запад", "city_ids": ["spb", "vld", "arh"]},
    {"id": "ural", "name": "Урал", "city_ids": ["ekb", "chel", "tmn"]},
    {"id": "volga", "name": "Волга", "city_ids": ["kzn", "nnv", "smr", "prm"]},
]
# type=128 в stores-data — самовывоз (CC Ковшовой 2с1 и т.п.). Не считаем сроком доставки.
NEW_STOCK_PICKUP_TYPES = {128}
NEW_STOCK_PICKUP_RE = re.compile(r"самовывоз|ковшов", re.I)
NEW_STOCK_STORES_URL = "https://static-basket-01.wbbasket.ru/vol0/data/stores-data.json"
NEW_STOCK_STORES_CACHE = {"by_id": {}, "loaded_at": 0.0}
# Ближайший наш FBS: Центр/СЗ → Москва, Волга → Казань, Юг → Краснодар, Урал/Сибирь → Тюмень.
NEW_STOCK_FBS_HUBS = {
    "msk": {
        "label": "FBS Москва / СЦ Внуково",
        "store_re": re.compile(r"внуков|москв", re.I),
        "stock_re": re.compile(r"внуков|москв", re.I),
    },
    "kzn": {
        "label": "FBS Казань / СЦ Столбище",
        "store_re": re.compile(r"казан", re.I),
        "stock_re": re.compile(r"казан", re.I),
    },
    "krd": {
        "label": "FBS Краснодар / СЦ Тахтамукай",
        "store_re": re.compile(r"краснодар|тахтамукай", re.I),
        "stock_re": re.compile(r"краснодар|тахтамукай|ффиточка", re.I),
    },
    "tmn": {
        "label": "FBS Тюмень / СЦ Харьковская",
        "store_re": re.compile(r"тюмен", re.I),
        "stock_re": re.compile(r"тюмен", re.I),
    },
}
NEW_STOCK_GROUP_HUBS = {
    "center": ["msk"],
    "nw": ["msk"],
    "volga": ["kzn", "msk"],
    "south": ["krd", "msk"],
    "ural": ["tmn", "msk"],
    "sib": ["tmn", "msk"],
}


def _new_stock_hours(product: dict):
    """Клиентский срок витрины, часы: time1 + time2."""
    if not isinstance(product, dict):
        return None
    t1, t2 = product.get("time1"), product.get("time2")
    if t1 is None and t2 is None:
        sizes = product.get("sizes") or []
        if sizes and isinstance(sizes[0], dict):
            t1 = sizes[0].get("time1")
            t2 = sizes[0].get("time2")
    if t1 is None and t2 is None:
        return None
    try:
        return int(t1 or 0) + int(t2 or 0)
    except Exception:
        return None


def _new_stock_qty(product: dict) -> int:
    """Остаток на складах, с которых WB предлагает доставку в этот dest."""
    if not isinstance(product, dict):
        return 0
    qty = 0
    for sz in product.get("sizes") or []:
        if not isinstance(sz, dict):
            continue
        stocks = sz.get("stocks")
        if stocks:
            for st in stocks:
                try:
                    qty += int((st or {}).get("qty") or 0)
                except Exception:
                    pass
        else:
            try:
                qty += int(sz.get("qty") or 0)
            except Exception:
                pass
    if qty > 0:
        return qty
    try:
        return int(product.get("totalQuantity") or product.get("total_quantity") or 0)
    except Exception:
        return 0


def _new_stock_wh(product: dict):
    if not isinstance(product, dict):
        return None
    wh = product.get("wh")
    if wh:
        return wh
    for sz in product.get("sizes") or []:
        if not isinstance(sz, dict):
            continue
        if sz.get("wh"):
            return sz.get("wh")
        for st in sz.get("stocks") or []:
            if st and st.get("wh"):
                return st.get("wh")
    return None


def _fbs_wh_is_ignored(name: str) -> bool:
    """Самовывоз / CC Ковшовой — не склад отгрузки."""
    return bool(NEW_STOCK_PICKUP_RE.search(name or ""))


def _new_stock_city_hubs() -> dict:
    """city_id → список хабов по приоритету (свой региональный, потом Москва)."""
    out = {}
    for g in NEW_STOCK_GROUPS:
        hubs = list(NEW_STOCK_GROUP_HUBS.get(g.get("id"), ["msk"]))
        for cid in g.get("city_ids") or []:
            out[str(cid)] = hubs
    return out


def _new_stock_hub_for_store(name: str, meta: dict):
    """Какой наш FBS-хаб у витринного склада. Самовывоз и склады WB — None."""
    if not isinstance(meta, dict):
        return None
    if meta.get("type") in NEW_STOCK_PICKUP_TYPES:
        return None
    if meta.get("is_wb"):
        return None
    if meta.get("type") not in (None, 2):
        return None
    n = name or meta.get("name") or ""
    if _fbs_wh_is_ignored(n):
        return None
    for hid, hub in NEW_STOCK_FBS_HUBS.items():
        if hub["store_re"].search(n):
            return hid
    return None


def _fbs_hub_qty(fbs_obj, hub_id: str) -> int:
    hub = NEW_STOCK_FBS_HUBS.get(hub_id) or {}
    stock_re = hub.get("stock_re")
    if not isinstance(fbs_obj, dict) or not stock_re:
        return 0
    total = 0
    for w in fbs_obj.get("warehouses") or []:
        if not isinstance(w, dict):
            continue
        name = str(w.get("name") or "")
        if _fbs_wh_is_ignored(name) or not stock_re.search(name):
            continue
        try:
            total += int(w.get("qty") or 0)
        except Exception:
            pass
    return total


def _fbs_pick_hub(fbs_obj, city_id: str, city_hubs: dict):
    for hid in city_hubs.get(str(city_id) or "") or ["msk"]:
        if _fbs_hub_qty(fbs_obj, hid) > 0:
            return hid
    return None


def _new_stock_load_stores() -> dict:
    """id склада витрины → {name, type, is_wb}. Падаем мягко, если CDN недоступен."""
    now = time.time()
    cached = NEW_STOCK_STORES_CACHE.get("by_id") or {}
    ts = float(NEW_STOCK_STORES_CACHE.get("loaded_at") or 0)
    if cached and now - ts < 12 * 3600:
        return cached
    try:
        resp = httpx.get(NEW_STOCK_STORES_URL, timeout=25)
        if not resp.is_success:
            return cached
        arr = resp.json() or []
        by_id = {}
        for s in arr if isinstance(arr, list) else []:
            if not isinstance(s, dict) or s.get("id") is None:
                continue
            try:
                sid = int(s["id"])
            except Exception:
                continue
            by_id[sid] = {
                "name": str(s.get("name") or ""),
                "type": s.get("type"),
                "is_wb": bool(s.get("isWb")),
            }
        NEW_STOCK_STORES_CACHE["by_id"] = by_id
        NEW_STOCK_STORES_CACHE["loaded_at"] = now
        return by_id
    except Exception as e:
        logger.warning(f"new-stock stores-data: {e}")
        return cached


def _new_stock_wh_int(wh):
    if wh is None or wh == "":
        return None
    try:
        return int(wh)
    except Exception:
        return None


def _new_stock_parse_cell(product: dict, stores: dict) -> dict:
    """Ячейка витрины: самовывоз отбрасываем, наш FBS-хаб помечаем."""
    hours = _new_stock_hours(product)
    qty = _new_stock_qty(product)
    wh = _new_stock_wh(product)
    wh_i = _new_stock_wh_int(wh)
    meta = stores.get(wh_i) if (stores and wh_i is not None) else None
    meta = meta if isinstance(meta, dict) else {}
    name = str(meta.get("name") or "")
    pickup = meta.get("type") in NEW_STOCK_PICKUP_TYPES or bool(NEW_STOCK_PICKUP_RE.search(name))
    hub = None if pickup else _new_stock_hub_for_store(name, meta)
    is_fbw = (not pickup) and (hub is None) and (
        bool(meta.get("is_wb")) or meta.get("type") not in (2, 128)
    )
    if pickup:
        return {
            "qty": 0,
            "hours": None,
            "tone": "oos",
            "wh": wh,
            "wh_name": name or None,
            "pickup": True,
            "is_fbw": False,
            "fbs_hub": None,
            "source": "pickup",
        }
    return {
        "qty": qty,
        "hours": hours,
        "tone": _new_stock_tone(hours, qty),
        "wh": wh,
        "wh_name": name or None,
        "pickup": False,
        "is_fbw": is_fbw,
        "fbs_hub": hub,
        "source": "fbs_hub" if hub else ("fbw" if is_fbw else "storefront"),
    }


def _new_stock_median_hours(vals) -> int | None:
    nums = [int(v) for v in (vals or []) if v is not None]
    if not nums:
        return None
    nums.sort()
    n = len(nums)
    if n % 2:
        return nums[n // 2]
    return int(round((nums[n // 2 - 1] + nums[n // 2]) / 2))


def _new_stock_city_fbs_hours(by_nm: dict) -> dict:
    """хаб → город → медиана часов, где витрина уже выбрала этот наш FBS."""
    buckets = {}
    for row in (by_nm or {}).values():
        for cid, cell in ((row or {}).get("cities") or {}).items():
            if not isinstance(cell, dict) or cell.get("pickup"):
                continue
            hub = cell.get("fbs_hub")
            if not hub or cell.get("hours") is None:
                continue
            buckets.setdefault(hub, {}).setdefault(cid, []).append(int(cell["hours"]))
    out = {}
    for hub, cities in buckets.items():
        got = {cid: h for cid, vals in cities.items() if (h := _new_stock_median_hours(vals)) is not None}
        if got:
            out[hub] = got
    return out


def _new_stock_hub_hours(city_fbs_h: dict, hub_id: str, city_id: str):
    if not isinstance(city_fbs_h, dict):
        return None
    inner = city_fbs_h.get(hub_id)
    if not isinstance(inner, dict):
        # старый плоский кэш {city: hours} = только Москва
        if hub_id == "msk" and city_id in city_fbs_h and not isinstance(city_fbs_h.get(city_id), dict):
            try:
                return int(city_fbs_h[city_id])
            except Exception:
                return None
        return None
    if city_id not in inner:
        return None
    try:
        return int(inner[city_id])
    except Exception:
        return None


def _new_stock_city_fbw_hours(by_nm: dict) -> dict:
    """Город → медиана часов, где витрина везёт со склада WB (не наш FBS и не самовывоз)."""
    buckets = {}
    for row in (by_nm or {}).values():
        for cid, cell in ((row or {}).get("cities") or {}).items():
            if not isinstance(cell, dict) or cell.get("pickup") or cell.get("fbs_hub"):
                continue
            if cell.get("hours") is None:
                continue
            if cell.get("is_fbw") or cell.get("source") in ("fbw", "storefront"):
                buckets.setdefault(cid, []).append(int(cell["hours"]))
    return {cid: h for cid, vals in buckets.items() if (h := _new_stock_median_hours(vals)) is not None}


def _new_stock_fbw_qty_by_nm() -> dict:
    """nm_id → штуки на складах WB из каталога (не FBS, не Ковшовая)."""
    out = {}
    for p in WB_PRODUCTS_CACHE.get("products") or []:
        if not isinstance(p, dict):
            continue
        try:
            nm = int(p.get("nm_id"))
        except Exception:
            continue
        total = 0
        for w in p.get("warehouses") or []:
            if not isinstance(w, dict):
                continue
            name = str(w.get("name") or "")
            ch = str(w.get("channel") or "").upper()
            low = name.lower()
            if ch == "FBS" or "fbs" in low or "маркетплейс" in low or _fbs_wh_is_ignored(name):
                continue
            try:
                total += int(w.get("qty") or 0)
            except Exception:
                pass
        if total > 0:
            out[nm] = total
    return out


def _new_stock_apply_fbs_peers(by_nm: dict, city_fbs_h: dict, city_fbw_h: dict | None = None) -> None:
    """Самовывоз: сначала срок склада WB по городу, иначе ближайший наш FBS."""
    city_hubs = _new_stock_city_hubs()
    city_fbw_h = city_fbw_h if isinstance(city_fbw_h, dict) else {}
    for row in (by_nm or {}).values():
        for cid, cell in ((row or {}).get("cities") or {}).items():
            if not isinstance(cell, dict) or not cell.get("pickup"):
                continue
            fbw_h = city_fbw_h.get(cid)
            if fbw_h is not None:
                try:
                    fbw_h = int(fbw_h)
                except Exception:
                    fbw_h = None
            if fbw_h is not None:
                cell["hours"] = fbw_h
                cell["source"] = "fbw_peer"
                cell["is_fbw"] = True
                cell["tone"] = _new_stock_tone(fbw_h, int(cell.get("qty") or 0))
                continue
            peer = None
            hid = None
            for cand in city_hubs.get(str(cid), ["msk"]):
                peer = _new_stock_hub_hours(city_fbs_h, cand, cid)
                if peer is not None:
                    hid = cand
                    break
            if peer is None:
                continue
            cell["hours"] = peer
            cell["fbs_hub"] = hid
            cell["source"] = "fbs_hub_peer"
            cell["tone"] = _new_stock_tone(peer, int(cell.get("qty") or 0))


def _new_stock_tone(hours, qty: int) -> str:
    if not qty:
        return "oos"
    if hours is None:
        return "ok"
    if hours > NEW_STOCK_BAD_H:
        return "bad"
    if hours > NEW_STOCK_WARN_H:
        return "warn"
    return "ok"


def _wb_site_headers() -> dict:
    h = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36",
        "Accept": "application/json",
        "Origin": "https://www.wildberries.ru",
        "Referer": "https://www.wildberries.ru/",
    }
    cookie = (os.getenv("WB_SITE_COOKIE") or "").strip()
    if cookie:
        h["Cookie"] = cookie
    return h


def _new_stock_has_cookie() -> bool:
    return bool((os.getenv("WB_SITE_COOKIE") or "").strip())


def _resolve_wb_dest(city: dict) -> int:
    """Актуальный dest по координатам; при сбое — запасной из NEW_STOCK_CITIES."""
    fallback = int(city.get("dest") or 0)
    lat, lon, name = city.get("lat"), city.get("lon"), city.get("name") or ""
    if lat is None or lon is None:
        return fallback
    try:
        resp = httpx.get(
            "https://user-geo-data.wildberries.ru/get-geo-info",
            params={"latitude": lat, "longitude": lon, "address": name},
            headers={"User-Agent": _wb_site_headers()["User-Agent"]},
            timeout=12,
        )
        if not resp.is_success:
            return fallback
        data = resp.json() or {}
        xinfo = str(data.get("xinfo") or data.get("xInfo") or "")
        m = re.search(r"dest=([^&]+)", xinfo)
        if m:
            return int(m.group(1))
        if data.get("dest") is not None:
            return int(data["dest"])
    except Exception as e:
        logger.warning(f"new-stock geo {name}: {e}")
    return fallback


def _new_stock_articles() -> list:
    raw = (os.getenv("NEW_STOCK_ARTICLES_JSON") or "").strip()
    if raw:
        try:
            arr = json.loads(raw)
        except Exception:
            arr = []
        out = []
        for it in arr if isinstance(arr, list) else []:
            try:
                nm = int((it or {}).get("nm_id") or (it or {}).get("nmId"))
            except Exception:
                continue
            vc = str((it or {}).get("vendor_code") or (it or {}).get("vendorCode") or nm)
            out.append({"nm_id": nm, "vendor_code": vc, "stock_seller": None, "sales_7d": 0})
        if out:
            return out[:80]
    products = WB_PRODUCTS_CACHE.get("products") or []
    if not products:
        try:
            get_wb_products(refresh=False)
            products = WB_PRODUCTS_CACHE.get("products") or []
        except Exception:
            products = []
    out = []
    for p in products or []:
        try:
            nm = int(p.get("nm_id"))
        except Exception:
            continue
        stock = int(p.get("stock") or 0)
        s7 = int(p.get("sales_7d") or 0)
        if stock <= 0 and s7 <= 0:
            continue
        out.append({
            "nm_id": nm,
            "vendor_code": str(p.get("vendor_code") or nm),
            "name": p.get("name") or "",
            "stock_seller": stock,
            "sales_7d": s7,
        })
    out.sort(key=lambda x: (-int(x.get("sales_7d") or 0), -int(x.get("stock_seller") or 0), str(x.get("vendor_code"))))
    return out[:80]


def _fetch_card_detail(nm_ids: list, dest: int) -> dict:
    """card.wb.ru v4. При 403 — ещё раз с cookie, если она задана."""
    if not nm_ids:
        return {"products": [], "status": 0, "error": None}
    ids = ";".join(str(int(n)) for n in nm_ids)
    url = "https://card.wb.ru/cards/v4/detail"
    params = {"appType": 1, "curr": "rub", "dest": dest, "spp": 30, "nm": ids}
    last_status = 0
    last_err = None
    for attempt in range(2):
        try:
            resp = httpx.get(url, params=params, headers=_wb_site_headers(), timeout=25)
        except Exception as e:
            last_err = str(e)
            break
        last_status = resp.status_code
        if resp.status_code == 403 and attempt == 0 and _new_stock_has_cookie():
            time.sleep(0.4)
            continue
        if not resp.is_success:
            last_err = f"HTTP {resp.status_code}"
            break
        try:
            data = resp.json() or {}
        except Exception as e:
            last_err = str(e)
            break
        products = (data.get("data") or {}).get("products") or data.get("products") or []
        return {"products": products if isinstance(products, list) else [], "status": last_status, "error": None}
    return {"products": [], "status": last_status, "error": last_err}


def sync_new_stock():
    if NEW_STOCK_CACHE.get("syncing"):
        return
    NEW_STOCK_CACHE["syncing"] = True
    NEW_STOCK_CACHE["error"] = None
    try:
        articles = _new_stock_articles()
        if not articles:
            NEW_STOCK_CACHE["payload"] = {
                "articles": [],
                "cities": [{"id": c["id"], "name": c["name"], "group": c["group"]} for c in NEW_STOCK_CITIES],
                "groups": NEW_STOCK_GROUPS,
                "as_of": _msk_now().strftime("%d.%m.%Y %H:%M"),
                "cookie_set": _new_stock_has_cookie(),
                "note": "Нет артикулов: обнови «Товары» или задай NEW_STOCK_ARTICLES_JSON.",
            }
            NEW_STOCK_CACHE["updated_at"] = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")
            return
        nm_ids = [a["nm_id"] for a in articles]
        by_nm = {a["nm_id"]: {"nm_id": a["nm_id"], "vendor_code": a["vendor_code"], "name": a.get("name") or "", "sales_7d": a.get("sales_7d") or 0, "stock_seller": a.get("stock_seller"), "cities": {}} for a in articles}
        cities_out = []
        blocked = 0
        dest_fail = 0
        stores = _new_stock_load_stores()
        for city in NEW_STOCK_CITIES:
            dest = _resolve_wb_dest(city)
            cities_out.append({"id": city["id"], "name": city["name"], "group": city["group"], "dest": dest})
            products = []
            for i in range(0, len(nm_ids), 25):
                got = _fetch_card_detail(nm_ids[i:i + 25], dest)
                if got.get("status") == 403:
                    blocked += 1
                if got.get("error") and not got.get("products"):
                    dest_fail += 1
                products.extend(got.get("products") or [])
                time.sleep(0.35)
            found = set()
            for p in products:
                try:
                    nm = int(p.get("id") or p.get("nmId") or p.get("nm_id"))
                except Exception:
                    continue
                if nm not in by_nm:
                    continue
                by_nm[nm]["cities"][city["id"]] = _new_stock_parse_cell(p, stores)
                found.add(nm)
            empty = {"qty": 0, "hours": None, "tone": "oos", "wh": None, "wh_name": None, "pickup": False, "is_fbw": False, "fbs_hub": None, "source": "storefront"}
            for nm, row in by_nm.items():
                if city["id"] not in row["cities"]:
                    row["cities"][city["id"]] = dict(empty)
            logger.info(f"new-stock {city['name']} dest={dest} found={len(found)}/{len(nm_ids)}")
        city_fbs_h = _new_stock_city_fbs_hours(by_nm)
        city_fbw_h = _new_stock_city_fbw_hours(by_nm)
        _new_stock_apply_fbs_peers(by_nm, city_fbs_h, city_fbw_h)

        rows = list(by_nm.values())
        rows.sort(key=lambda r: (
            -max((int(c.get("hours") or 0) for c in (r.get("cities") or {}).values()), default=0),
            -int(r.get("sales_7d") or 0),
            str(r.get("vendor_code")),
        ))
        err = None
        has_data = any(
            (c.get("hours") is not None or int(c.get("qty") or 0) > 0)
            for r in rows
            for c in (r.get("cities") or {}).values()
        )
        if blocked and not has_data:
            err = "WB витрина временно закрыла доступ (403). Подожди и нажми «Обновить» — cookie с сайта не используем."
        elif dest_fail >= len(NEW_STOCK_CITIES) and not has_data:
            err = "Не удалось прочитать витрину WB. Подожди и нажми «Обновить»."
        payload = {
            "articles": rows,
            "cities": cities_out,
            "groups": NEW_STOCK_GROUPS,
            "warn_h": NEW_STOCK_WARN_H,
            "bad_h": NEW_STOCK_BAD_H,
            "as_of": _msk_now().strftime("%d.%m.%Y %H:%M"),
            "cookie_set": _new_stock_has_cookie(),
            "articles_source": "env" if (os.getenv("NEW_STOCK_ARTICLES_JSON") or "").strip() else "catalog",
            "city_fbs_hours": city_fbs_h,
            "city_fbw_hours": city_fbw_h,
            "note": "Срок и штуки — как на витрине WB в этом городе, без самовывоза CC Ковшовой. Если витрина спрятала склад WB за ПВЗ — показываем остаток FBW. Срок с нашего FBS только когда склада WB нет.",
        }
        NEW_STOCK_CACHE["payload"] = payload
        NEW_STOCK_CACHE["error"] = err
        NEW_STOCK_CACHE["updated_at"] = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")
        logger.info(f"new-stock: {len(rows)} arts × {len(cities_out)} cities")
    except Exception as e:
        logger.error(f"sync_new_stock: {e}")
        NEW_STOCK_CACHE["error"] = str(e)
    finally:
        NEW_STOCK_CACHE["syncing"] = False


def _new_stock_stale(max_age_sec: int = 1800) -> bool:
    raw = NEW_STOCK_CACHE.get("updated_at")
    if not raw or not NEW_STOCK_CACHE.get("payload"):
        return True
    try:
        dt = datetime.strptime(str(raw), "%d.%m.%Y %H:%M").replace(tzinfo=timezone.utc)
    except Exception:
        return True
    return (datetime.now(timezone.utc) - dt).total_seconds() > max_age_sec


@app.post("/api/sync-new-stock")
def trigger_new_stock_sync():
    if NEW_STOCK_CACHE.get("syncing"):
        return {"status": "already_running"}
    threading.Thread(target=sync_new_stock, daemon=True, name="new-stock").start()
    return {"status": "started"}


def _new_stock_default_layout() -> dict:
    return {
        "hidden": [],
        "pinned": [],
        "order": [],
        "groups": [],
        "city_order": [],
        "group_order": [],
        "city_group": {},
        "hidden_cities": [],
        "hidden_fbs": [],
        "col_w": 56,
    }


def _uniq_str_list(raw) -> list:
    out, seen = [], set()
    for item in raw or []:
        s = str(item or "").strip()
        if s and s not in seen:
            out.append(s)
            seen.add(s)
    return out


def _normalize_new_stock_layout(raw) -> dict:
    """Общая раскладка для всех: скрытые, закреп, порядок, группы, колонки городов."""
    src = raw if isinstance(raw, dict) else {}
    hidden = _uniq_str_list(src.get("hidden"))
    hidden_set = set(hidden)
    pinned = [s for s in _uniq_str_list(src.get("pinned")) if s not in hidden_set]
    order = _uniq_str_list(src.get("order"))
    groups = []
    used = set()
    for g in src.get("groups") or []:
        if not isinstance(g, dict):
            continue
        name = str(g.get("name") or "").strip()
        gid = str(g.get("id") or "").strip() or f"g_{len(groups)+1}"
        arts = []
        for vc in g.get("articles") or []:
            s = str(vc or "").strip()
            if s and s not in used:
                arts.append(s)
                used.add(s)
        if name:
            groups.append({"id": gid, "name": name, "articles": arts})
    city_group = {}
    raw_cg = src.get("city_group") or {}
    if isinstance(raw_cg, dict):
        for k, v in raw_cg.items():
            ks, vs = str(k or "").strip(), str(v or "").strip()
            if ks and vs:
                city_group[ks] = vs
    return {
        "hidden": hidden,
        "pinned": pinned,
        "order": order,
        "groups": groups,
        "city_order": _uniq_str_list(src.get("city_order")),
        "group_order": _uniq_str_list(src.get("group_order")),
        "city_group": city_group,
        "hidden_cities": _uniq_str_list(src.get("hidden_cities")),
        "hidden_fbs": _uniq_str_list(src.get("hidden_fbs")),
        "col_w": _new_stock_col_w(src.get("col_w")),
    }


def _new_stock_col_w(raw) -> int:
    try:
        n = int(raw)
    except Exception:
        n = 56
    return max(48, min(110, n))


def _fbs_wh_short(name: str) -> str:
    raw = (name or "").strip()
    s = re.sub(r"(?i)^маркетплейс\s*\(fbs\)\s*[·•\-\u2013\u2014:]\s*", "", raw)
    s = re.sub(r"(?i)\s*\(fbs\)\s*$", "", s).strip()
    short = s or raw
    low = short.lower()
    if any(x in low for x in ("dbs", "edbs", "курьер", "мгт")):
        return "Прочие"
    return short


def _new_stock_collect_fbs(rows) -> dict:
    """nm_id -> {short_name: qty} из списка складов {name/warehouse_name, qty/quantity}."""
    by_nm = {}
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        raw_name = (row.get("name") or row.get("warehouse_name") or "").strip()
        if not raw_name:
            continue
        n = raw_name.lower()
        if "fbs" not in n and "маркетплейс" not in n:
            continue
        try:
            nm = int(row.get("nm_id"))
            qty = int(row.get("qty") if row.get("qty") is not None else row.get("quantity") or 0)
        except Exception:
            continue
        if qty <= 0:
            continue
        short = _fbs_wh_short(raw_name)
        if _fbs_wh_is_ignored(raw_name) or _fbs_wh_is_ignored(short):
            continue
        bucket = by_nm.setdefault(nm, {})
        bucket[short] = bucket.get(short, 0) + qty
    return by_nm


def _new_stock_fbs_from_products() -> dict:
    rows = []
    for p in WB_PRODUCTS_CACHE.get("products") or []:
        try:
            nm = int(p.get("nm_id"))
        except Exception:
            continue
        for w in p.get("warehouses") or []:
            if not isinstance(w, dict):
                continue
            rows.append({
                "nm_id": nm,
                "name": w.get("name"),
                "qty": w.get("qty"),
            })
    return _new_stock_collect_fbs(rows)


def _new_stock_fbs_from_supabase() -> dict:
    if not SUPABASE_URL:
        return {}
    try:
        resp = httpx.get(
            f"{SUPABASE_URL}/rest/v1/stock_warehouses"
            "?select=nm_id,warehouse_name,quantity&quantity=gt.0&limit=20000",
            headers=sb_headers(),
            timeout=20,
        )
        rows = resp.json() if resp.is_success else []
        if not isinstance(rows, list):
            rows = []
    except Exception as e:
        logger.warning(f"new-stock FBS warehouses: {e}")
        return {}
    return _new_stock_collect_fbs(rows)


def _new_stock_fbs_bundle() -> tuple:
    """(map nm-> {total, warehouses}, list of {id,name} колонок)."""
    by_nm = _new_stock_fbs_from_products()
    if not by_nm:
        by_nm = _new_stock_fbs_from_supabase()
    totals = {}
    for whs in by_nm.values():
        for name, qty in whs.items():
            totals[name] = totals.get(name, 0) + qty
    fbs_whs = [{"id": n, "name": n} for n, _ in sorted(totals.items(), key=lambda x: (-x[1], x[0].lower()))]
    fbs_map = {}
    for nm, whs in by_nm.items():
        warehouses = [{"name": n, "qty": q} for n, q in sorted(whs.items(), key=lambda x: (-x[1], x[0].lower()))]
        fbs_map[int(nm)] = {"total": sum(whs.values()), "warehouses": warehouses}
    return fbs_map, fbs_whs


def _attach_new_stock_fbs(payload: dict) -> dict:
    src = payload if isinstance(payload, dict) else {}
    fbs_map, fbs_whs = _new_stock_fbs_bundle()
    fbw_map = _new_stock_fbw_qty_by_nm()
    city_fbs_h = src.get("city_fbs_hours") if isinstance(src.get("city_fbs_hours"), dict) else {}
    city_fbw_h = src.get("city_fbw_hours") if isinstance(src.get("city_fbw_hours"), dict) else {}
    city_hubs = _new_stock_city_hubs()
    articles = []
    for a in src.get("articles") or []:
        if not isinstance(a, dict):
            continue
        row = dict(a)
        try:
            nm = int(row.get("nm_id"))
        except Exception:
            nm = None
        row["fbs"] = fbs_map.get(nm) if nm is not None else None
        if not row["fbs"]:
            row["fbs"] = {"total": 0, "warehouses": []}
        fbw_qty = int(fbw_map.get(nm) or 0) if nm is not None else 0
        row["fbw_qty"] = fbw_qty
        cities = {}
        for cid, cell in (row.get("cities") or {}).items():
            if not isinstance(cell, dict):
                continue
            c = dict(cell)
            c["fbw_qty"] = fbw_qty
            if c.get("pickup"):
                fbw_h = city_fbw_h.get(cid)
                try:
                    fbw_h = int(fbw_h) if fbw_h is not None else None
                except Exception:
                    fbw_h = None
                old_src = c.get("source")
                if fbw_qty > 0:
                    c["qty"] = fbw_qty
                    c["is_fbw"] = True
                    c["source"] = "fbw"
                    c["wh_label"] = "Склад WB"
                    if fbw_h is not None:
                        c["hours"] = fbw_h
                    elif old_src in ("fbs_hub_peer", "fbs_msk_peer"):
                        c["hours"] = None
                    c["tone"] = _new_stock_tone(c.get("hours"), fbw_qty)
                else:
                    hid = _fbs_pick_hub(row["fbs"], cid, city_hubs)
                    hub = NEW_STOCK_FBS_HUBS.get(hid) if hid else None
                    primary = (city_hubs.get(str(cid)) or ["msk"])[0]
                    if hid and hub and hid == primary:
                        c["qty"] = _fbs_hub_qty(row["fbs"], hid)
                        peer = _new_stock_hub_hours(city_fbs_h, hid, cid)
                        if peer is not None:
                            c["hours"] = peer
                        c["source"] = "fbs_hub_peer"
                        c["fbs_hub"] = hid
                        c["wh_label"] = hub["label"]
                    else:
                        c["qty"] = 0
                    c["tone"] = _new_stock_tone(c.get("hours"), int(c.get("qty") or 0))
            elif c.get("fbs_hub"):
                hub = NEW_STOCK_FBS_HUBS.get(c.get("fbs_hub"))
                if hub:
                    c["wh_label"] = hub["label"]
                    c["source"] = "fbs_hub"
            elif c.get("is_fbw") or (not c.get("fbs_hub") and fbw_qty > 0):
                if fbw_qty > int(c.get("qty") or 0):
                    c["qty"] = fbw_qty
                    c["is_fbw"] = True
                c["wh_label"] = c.get("wh_label") or "Склад WB"
                c["source"] = "fbw"
            cities[cid] = c
        row["cities"] = cities
        articles.append(row)
    out = dict(src)
    out["articles"] = articles
    out["fbs_warehouses"] = fbs_whs
    return out


def _new_stock_layout() -> dict:
    return _normalize_new_stock_layout(get_setting_json(NEW_STOCK_LAYOUT_KEY, {}))


@app.get("/api/new-stock")
def get_new_stock(refresh: bool = False):
    if refresh or _new_stock_stale():
        if not NEW_STOCK_CACHE.get("syncing"):
            threading.Thread(target=sync_new_stock, daemon=True, name="new-stock").start()
    payload = _attach_new_stock_fbs(NEW_STOCK_CACHE.get("payload") or {})
    return {
        **payload,
        "layout": _new_stock_layout(),
        "syncing": NEW_STOCK_CACHE.get("syncing", False),
        "error": NEW_STOCK_CACHE.get("error") or payload.get("error"),
        "updated_at": NEW_STOCK_CACHE.get("updated_at"),
        "cookie_set": _new_stock_has_cookie(),
    }


@app.post("/api/new-stock-layout")
async def save_new_stock_layout(request: dict):
    """Общая раскладка «Новые остатки»: скрытые / закреп / порядок / группы — на всех одинаково."""
    layout = _normalize_new_stock_layout(request or {})
    if not save_setting_value(NEW_STOCK_LAYOUT_KEY, layout):
        raise HTTPException(status_code=500, detail="Не удалось сохранить раскладку")
    return {"status": "ok", "layout": layout}


# ---------- Рекомендации по поставкам: заказы + продажи по складам (WB Statistics API) ----------
# Заказано — /api/v1/supplier/orders, Выкупили — /api/v1/supplier/sales (только saleID, начинающиеся
# на "S" — это продажи; "R" — возврат, "D" — доплата, их не считаем). Текущий остаток берём из уже
# собранной stock_warehouses (результат sync_stock). Объединяем по nm_id + warehouseName.

def fetch_supplier_feed(endpoint: str, date_from_iso: str, max_pages: int = 5) -> list:
    all_rows = []
    cursor = date_from_iso
    for _ in range(max_pages):
        resp = httpx.get(
            f"{WB_STATISTICS_URL}{endpoint}",
            headers=wb_headers(), params={"dateFrom": cursor}, timeout=60
        )
        if not resp.is_success:
            logger.error(f"WB {endpoint} error {resp.status_code} {resp.text[:200]}")
            break
        batch = resp.json()
        if not batch:
            break
        all_rows.extend(batch)
        cursor = batch[-1].get("lastChangeDate", cursor)
        if len(batch) < 50000:
            break
        time.sleep(61)  # лимит — 1 запрос в минуту на этот метод
    return all_rows

def get_setting_int(key: str, default: int) -> int:
    try:
        resp = httpx.get(f"{SUPABASE_URL}/rest/v1/settings?key=eq.{key}&select=value", headers=sb_headers(), timeout=10)
        if resp.is_success and resp.json():
            return int(resp.json()[0]["value"])
    except Exception:
        pass
    return default

def get_setting_raw(key: str, default=None):
    try:
        resp = httpx.get(
            f"{SUPABASE_URL}/rest/v1/settings?key=eq.{key}&select=value",
            headers=sb_headers(), timeout=10
        )
        if resp.is_success and resp.json():
            return resp.json()[0]["value"]
    except Exception:
        pass
    return default

def get_setting_json(key: str, default=None):
    if default is None:
        default = {}
    raw = get_setting_raw(key, None)
    if raw is None:
        return default
    try:
        import json as _json
        val = _json.loads(raw) if isinstance(raw, str) else raw
        return val if val is not None else default
    except Exception:
        return default

def save_setting_value(key: str, value) -> bool:
    import json as _json
    payload = {
        "key": key,
        "value": value if isinstance(value, str) else _json.dumps(value, ensure_ascii=False),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    try:
        resp = httpx.post(
            f"{SUPABASE_URL}/rest/v1/settings?on_conflict=key",
            json=payload,
            headers=sb_headers(), timeout=10
        )
        return resp.is_success
    except Exception as e:
        logger.error(f"save_setting_value({key}) error: {e}")
        return False

def parse_wb_dt(s: str):
    """WB отдаёт даты в orders/sales без таймзоны (например '2026-06-10T10:00:00').
    Нормализуем всё к naive datetime, чтобы сравнения не падали на tz-aware/naive."""
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is not None:
            dt = dt.replace(tzinfo=None)
        return dt
    except (ValueError, TypeError):
        return None

def fetch_planned_supplies_qty() -> dict:
    """Возвращает {nmId: суммарное количество} по поставкам (FBW), у которых дата
    поставки (supplyDate) попадает в ближайшие 7 дней, и которые ещё не приняты складом
    (factDate пусто) — то есть реально едут/запланированы, а не черновик и не уже приехали."""
    try:
        resp = httpx.post(
            f"{WB_SUPPLIES_URL}/api/v1/supplies",
            headers=wb_headers(), params={"limit": 1000, "offset": 0},
            json={}, timeout=30
        )
        if not resp.is_success:
            logger.error(f"WB supplies list error {resp.status_code} {resp.text[:300]}")
            return {}
        supplies = resp.json()
        if not isinstance(supplies, list):
            logger.error(f"WB supplies list unexpected shape: {str(supplies)[:300]}")
            return {}
    except Exception as e:
        logger.error(f"WB supplies list exception: {e}")
        return {}

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    horizon = now + timedelta(days=7)
    qualifying = []
    for s in supplies:
        if s.get("factDate"):
            continue
        sd = s.get("supplyDate")
        if not sd:
            continue
        try:
            d = datetime.fromisoformat(str(sd).replace("Z", "+00:00")).replace(tzinfo=None)
        except (ValueError, TypeError):
            continue
        if now <= d <= horizon:
            sid, is_preorder = s.get("supplyID"), False
            if not sid:
                sid, is_preorder = s.get("preorderID"), True
            if sid:
                qualifying.append((sid, is_preorder))

    logger.info(f"Planned FBW supplies in next 7 days: {len(qualifying)}")
    planned = {}
    for sid, is_preorder in qualifying:
        try:
            params = {"limit": 1000, "offset": 0}
            if is_preorder:
                params["isPreorderID"] = "true"
            gresp = httpx.get(
                f"{WB_SUPPLIES_URL}/api/v1/supplies/{sid}/goods",
                headers=wb_headers(), params=params, timeout=20
            )
            if not gresp.is_success:
                logger.error(f"WB supply goods error supply={sid} {gresp.status_code} {gresp.text[:200]}")
                continue
            for item in gresp.json():
                nm = item.get("nmID") or item.get("nmId")
                qty = item.get("quantity", 0) or 0
                if nm:
                    planned[nm] = planned.get(nm, 0) + qty
        except Exception as e:
            logger.error(f"WB supply goods exception supply={sid}: {e}")
        time.sleep(0.1)
    return planned

def sync_supply():
    if not WB_TOKEN:
        logger.error("WB_TOKEN not set")
        return
    logger.info("Starting supply (orders/sales) sync...")
    window_days = get_setting_int("sales_window_days", 14)
    cutoff = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=window_days)
    date_from = cutoff.strftime("%Y-%m-%dT00:00:00")

    orders = fetch_supplier_feed("/api/v1/supplier/orders", date_from)
    time.sleep(61)  # отдельный лимит 1 запрос/мин на каждый метод
    sales = fetch_supplier_feed("/api/v1/supplier/sales", date_from)
    logger.info(f"Supply sync: fetched {len(orders)} order rows, {len(sales)} sale rows")

    agg = {}  # (nm_id, warehouseName) -> {"ordered":int,"buyout":int,"vendor_code":str}
    nm_to_barcode = {}
    for o in orders:
        d = parse_wb_dt(o.get("date", ""))
        if d is None or d < cutoff:
            continue
        key = (o.get("nmId"), o.get("warehouseName"))
        a = agg.setdefault(key, {"ordered": 0, "buyout": 0, "vendor_code": ""})
        a["ordered"] += 1
        if o.get("supplierArticle"):
            a["vendor_code"] = o["supplierArticle"]
        if o.get("barcode") and o.get("nmId"):
            nm_to_barcode[o["nmId"]] = o["barcode"]

    for s in sales:
        if not str(s.get("saleID", "")).startswith("S"):
            continue  # пропускаем возвраты (R) и доплаты (D)
        d = parse_wb_dt(s.get("date", ""))
        if d is None or d < cutoff:
            continue
        key = (s.get("nmId"), s.get("warehouseName"))
        a = agg.setdefault(key, {"ordered": 0, "buyout": 0, "vendor_code": ""})
        a["buyout"] += 1
        if s.get("supplierArticle"):
            a["vendor_code"] = s["supplierArticle"]
        if s.get("barcode") and s.get("nmId"):
            nm_to_barcode[s["nmId"]] = s["barcode"]

    # stock_totals часто без vendor_code — берём ratings/feedbacks + артикулы из заказов
    nm_to_vendor = build_nm_to_vendor_map()
    for (nm_id, _), a in agg.items():
        vc = (a.get("vendor_code") or "").strip()
        if not nm_id or not vc or vc == str(nm_id):
            continue
        try:
            nm_int = int(nm_id)
        except (TypeError, ValueError):
            continue
        if nm_int not in nm_to_vendor:
            nm_to_vendor[nm_int] = vc

    try:
        sw = httpx.get(f"{SUPABASE_URL}/rest/v1/stock_warehouses?select=nm_id,warehouse_name,quantity", headers=sb_headers(), timeout=20)
        stock_map = {(r["nm_id"], r["warehouse_name"]): r["quantity"] for r in sw.json()} if sw.is_success else {}
    except Exception as e:
        logger.error(f"sync_supply: stock_warehouses fetch error {e}")
        stock_map = {}

    planned_map = fetch_planned_supplies_qty()

    def resolve_supply_vendor(nm_id, fallback=""):
        try:
            nm_int = int(nm_id)
        except (TypeError, ValueError):
            nm_int = None
        for cand in (nm_to_vendor.get(nm_int) if nm_int is not None else None, fallback):
            vc = (cand or "").strip()
            if vc and vc != str(nm_id):
                return vc
        # любой склад с заказами по этому nm
        for (n, _), aa in agg.items():
            if n != nm_id:
                continue
            vc = (aa.get("vendor_code") or "").strip()
            if vc and vc != str(nm_id):
                return vc
        return str(nm_id)

    keys = set(agg.keys()) | set(stock_map.keys())
    now = datetime.now(timezone.utc).isoformat()
    rows = []
    for nm_id, wh in keys:
        if not nm_id or not wh:
            continue
        a = agg.get((nm_id, wh), {"ordered": 0, "buyout": 0, "vendor_code": ""})
        rows.append({
            "vendor_code": resolve_supply_vendor(nm_id, a.get("vendor_code") or ""),
            "nm_id": nm_id,
            "barcode": nm_to_barcode.get(nm_id),
            "planned_supply_qty": planned_map.get(nm_id, 0),
            "warehouse_name": wh,
            "ordered_qty": a["ordered"],
            "buyout_qty": a["buyout"],
            "current_stock": stock_map.get((nm_id, wh), 0),
            "period_days": window_days,
            "period_start": None,
            "period_end": None,
            "updated_at": now,
        })

    httpx.delete(f"{SUPABASE_URL}/rest/v1/supply_report?id=gte.0", headers=sb_headers(), timeout=15)
    saved = 0
    for i in range(0, len(rows), 300):
        batch = rows[i:i + 300]
        resp = httpx.post(f"{SUPABASE_URL}/rest/v1/supply_report", json=batch, headers=sb_headers(), timeout=30)
        if resp.is_success:
            saved += len(batch)
        else:
            logger.error(f"supply_report insert error {resp.status_code} {resp.text[:300]}")

    httpx.post(
        f"{SUPABASE_URL}/rest/v1/settings",
        json={"key": "last_supply_sync", "value": datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M"),
              "updated_at": now},
        headers=sb_headers(), timeout=10
    )
    logger.info(f"Supply sync complete. Rows saved: {saved}")

# ---------- Продвижение (реклама) ----------

AD_TYPE_NAMES = {
    3: "Карточка товара", 4: "Каталог+поиск", 5: "Карточка",
    6: "Каталог", 8: "Автоматическая", 9: "Поиск"
}

ADS_CACHE = {
    "campaigns": [],
    "updated_at": None,
    "syncing": False,
    "error": None,
    "progress": None,
    "window_days": None,
    "period_begin": None,
    "period_end": None,
}

def _parse_advert_v2_items(payload) -> list:
    if payload is None:
        return []
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("adverts", "data", "items"):
            val = payload.get(key)
            if isinstance(val, list):
                return val
            if isinstance(val, dict) and isinstance(val.get("adverts"), list):
                return val["adverts"]
    return []

def _advert_id_from_item(camp: dict):
    return camp.get("advertId") or camp.get("advert_id") or camp.get("id")

def _advert_name_from_item(camp: dict) -> str:
    """В v2 название лежит в settings.name; в старых ответах — в name."""
    name = (camp.get("name") or "").strip()
    if name:
        return name
    settings = camp.get("settings")
    if isinstance(settings, dict):
        name = (settings.get("name") or "").strip()
        if name:
            return name
    return ""

def _advert_type_from_item(camp: dict) -> int:
    type_id = camp.get("type") or camp.get("type_id")
    if type_id:
        return int(type_id)
    settings = camp.get("settings") if isinstance(camp.get("settings"), dict) else {}
    # иногда тип оплаты/ставки вместо type — оставим 0
    return int(settings.get("type") or 0)

def fetch_campaigns_meta(include_finished: bool = False) -> dict:
    """Активные (9) и на паузе (11). Имена — как в кабинете WB (settings.name)."""
    allowed = (9, 11, 7) if include_finished else (9, 11)
    statuses = "9,11,7" if include_finished else "9,11"
    meta = {}

    # 1) основной путь: /api/advert/v2/adverts
    try:
        param_sets = [
            {"statuses": statuses},
            {"statuses": statuses, "payment_type": "cpm"},
            {"statuses": statuses, "payment_type": "cpc"},
        ]
        for params in param_sets:
            resp = httpx.get(
                f"{WB_PROMOTION_URL}/api/advert/v2/adverts",
                headers=wb_headers(),
                params=params,
                timeout=30,
            )
            if not resp.is_success:
                logger.warning(f"advert/v2/adverts {resp.status_code} params={params}: {resp.text[:200]}")
                continue
            items = _parse_advert_v2_items(resp.json())
            if items and not meta:
                # лог структуры один раз — чтобы видеть поля, если снова сломается
                sample = items[0] if isinstance(items[0], dict) else {}
                logger.info(
                    f"advert/v2 sample keys={list(sample.keys())[:20]} "
                    f"settings_keys={list((sample.get('settings') or {}).keys())[:15] if isinstance(sample.get('settings'), dict) else None}"
                )
            for camp in items:
                if not isinstance(camp, dict):
                    continue
                aid = _advert_id_from_item(camp)
                if not aid:
                    continue
                status = camp.get("status")
                if status is not None and status not in allowed:
                    continue
                type_id = _advert_type_from_item(camp)
                type_name = AD_TYPE_NAMES.get(type_id, f"Тип {type_id}" if type_id else "")
                name = _advert_name_from_item(camp)
                prev = meta.get(aid) or {}
                meta[aid] = {
                    "type_id": type_id or prev.get("type_id") or 0,
                    "type_name": type_name or prev.get("type_name") or "",
                    "status": status if status is not None else prev.get("status"),
                    "name": name or prev.get("name") or "",
                }
            # не break: cpm+cpc могут дополнять друг друга
    except Exception as e:
        logger.error(f"WB advert/v2/adverts exception: {e}")

    # 2) fallback: список id из count
    if not meta:
        try:
            resp = httpx.get(f"{WB_PROMOTION_URL}/adv/v1/promotion/count", headers=wb_headers(), timeout=20)
            if resp.is_success:
                for group in (resp.json() or {}).get("adverts", []) or []:
                    type_id = group.get("type", 0)
                    type_name = AD_TYPE_NAMES.get(type_id, f"Тип {type_id}")
                    status = group.get("status")
                    if status not in allowed:
                        continue
                    for item in group.get("advert_list", []) or []:
                        aid = item.get("advertId")
                        if aid:
                            meta[aid] = {
                                "type_id": type_id,
                                "type_name": type_name,
                                "status": status,
                                "name": "",
                            }
        except Exception as e:
            logger.error(f"WB promotion/count exception: {e}")

    enrich_campaign_names(meta)
    for aid, m in meta.items():
        if not (m.get("name") or "").strip() or is_placeholder_campaign_name(m.get("name") or "", aid):
            # последний шанс — не оставляем голое «Кампания»
            if not (m.get("name") or "").strip() or m.get("name") == "Кампания":
                m["name"] = (m.get("type_name") and f"{m['type_name']} #{aid}") or f"#{aid}"
    return meta

def enrich_campaign_names(meta: dict) -> None:
    """Дотягивает названия через /api/advert/v2/adverts?ids=... и старый promotion/adverts."""
    need = [
        aid for aid, m in meta.items()
        if is_placeholder_campaign_name(m.get("name") or "", aid)
    ]
    if not need:
        return

    for i in range(0, len(need), 50):
        batch = need[i:i + 50]
        got = False
        try:
            resp = httpx.get(
                f"{WB_PROMOTION_URL}/api/advert/v2/adverts",
                headers=wb_headers(),
                params={"ids": ",".join(str(x) for x in batch)},
                timeout=20,
            )
            if resp.is_success:
                for camp in _parse_advert_v2_items(resp.json()):
                    if not isinstance(camp, dict):
                        continue
                    aid = _advert_id_from_item(camp)
                    if aid and aid in meta:
                        name = _advert_name_from_item(camp)
                        if name:
                            meta[aid]["name"] = name
                            got = True
                        type_id = _advert_type_from_item(camp)
                        if type_id:
                            meta[aid]["type_id"] = type_id
                            meta[aid]["type_name"] = AD_TYPE_NAMES.get(type_id, meta[aid].get("type_name"))
            else:
                logger.warning(f"advert/v2 names {resp.status_code}: {resp.text[:200]}")
        except Exception as e:
            logger.warning(f"campaign names v2 skip: {e}")

        # старый endpoint — на случай если v2 без settings.name
        still = [aid for aid in batch if is_placeholder_campaign_name(meta[aid].get("name") or "", aid)]
        if still:
            try:
                resp = httpx.post(
                    f"{WB_PROMOTION_URL}/adv/v1/promotion/adverts",
                    json=still, headers=wb_headers(), timeout=20
                )
                if resp.is_success:
                    for camp in resp.json() or []:
                        aid = camp.get("advertId") or camp.get("advert_id")
                        if aid and aid in meta:
                            name = (camp.get("name") or "").strip()
                            if name:
                                meta[aid]["name"] = name
                                got = True
            except Exception as e:
                logger.warning(f"campaign names v1 skip: {e}")
        if i + 50 < len(need):
            time.sleep(0.3)
        logger.info(f"enrich names batch {i//50+1}: need={len(batch)} got_any={got}")

def is_placeholder_campaign_name(name: str, aid) -> bool:
    """True если имя — заглушка, а не название из кабинета WB."""
    n = (name or "").strip()
    if not n:
        return True
    if n in ("Кампания", "Без названия", "Неизвестно"):
        return True
    if n == f"#{aid}" or n == f"Кампания #{aid}":
        return True
    if n.endswith(f" #{aid}"):
        return True
    return False

def fetch_ad_stats_by_campaign(ids: list, begin_date: str, end_date: str) -> dict:
    """Тянет /adv/v3/fullstats → {campaign_id: метрики}. Пауза 20с только между батчами."""
    agg = {}
    errors = []
    total_batches = max(1, (len(ids) + 49) // 50)
    for i in range(0, len(ids), 50):
        batch = ids[i:i + 50]
        batch_no = i // 50 + 1
        ADS_CACHE["progress"] = f"статистика WB {batch_no}/{total_batches}"
        try:
            resp = httpx.get(
                f"{WB_PROMOTION_URL}/adv/v3/fullstats",
                headers=wb_headers(),
                params={"ids": ",".join(str(x) for x in batch), "beginDate": begin_date, "endDate": end_date},
                timeout=60,
            )
            if resp.status_code == 429:
                logger.warning("WB fullstats 429 — ждём 22с и повторяем")
                ADS_CACHE["progress"] = f"лимит WB, ждём… ({batch_no}/{total_batches})"
                time.sleep(22)
                resp = httpx.get(
                    f"{WB_PROMOTION_URL}/adv/v3/fullstats",
                    headers=wb_headers(),
                    params={"ids": ",".join(str(x) for x in batch), "beginDate": begin_date, "endDate": end_date},
                    timeout=60,
                )
            if not resp.is_success:
                msg = f"fullstats {resp.status_code}: {resp.text[:300]}"
                logger.error(f"WB {msg}")
                errors.append(msg)
                # при 429 на базовом тарифе дальше бессмысленно долбить
                if resp.status_code == 429:
                    break
                continue
            campaigns = resp.json()
        except Exception as e:
            logger.error(f"WB fullstats exception: {e}")
            errors.append(str(e))
            continue

        for camp in campaigns or []:
            campaign_id = camp.get("advertId")
            if not campaign_id:
                continue
            views = int(camp.get("views") or 0)
            clicks = int(camp.get("clicks") or 0)
            atbs = int(camp.get("atbs") or 0)
            orders = int(camp.get("orders") or 0)
            spend = float(camp.get("sum") or 0)
            revenue = float(camp.get("sum_price") or 0)

            if not (views or clicks or orders or spend) and camp.get("days"):
                for day in camp.get("days") or []:
                    views += int(day.get("views") or 0)
                    clicks += int(day.get("clicks") or 0)
                    atbs += int(day.get("atbs") or 0)
                    orders += int(day.get("orders") or 0)
                    spend += float(day.get("sum") or 0)
                    revenue += float(day.get("sum_price") or 0)

            a = agg.setdefault(campaign_id, {
                "views": 0, "clicks": 0, "atbs": 0,
                "orders": 0, "spend": 0.0, "revenue": 0.0,
            })
            a["views"] += views
            a["clicks"] += clicks
            a["atbs"] += atbs
            a["orders"] += orders
            a["spend"] += spend
            a["revenue"] += revenue

        if i + 50 < len(ids):
            time.sleep(20)
    if errors and not agg:
        raise RuntimeError("; ".join(errors[:3]))
    return agg

def _fullstats_nm_day_rows(campaigns) -> list:
    """Разворачивает fullstats → список {nm_id, day, views, spend, clicks, orders}."""
    rows = []
    for camp in campaigns or []:
        for day in camp.get("days") or []:
            day_str = str(day.get("date") or "")[:10]
            if not day_str or len(day_str) < 10:
                continue
            nms = list(day.get("nms") or [])
            for app in day.get("apps") or []:
                nms.extend(app.get("nms") or [])
            for nm in nms:
                if not isinstance(nm, dict):
                    continue
                nm_id = nm.get("nmId") or nm.get("nm_id")
                if not nm_id:
                    continue
                rows.append({
                    "nm_id": int(nm_id),
                    "day": day_str,
                    "views": int(nm.get("views") or 0),
                    "spend": float(nm.get("sum") or 0),
                    "clicks": int(nm.get("clicks") or 0),
                    "orders": int(nm.get("orders") or 0),
                })
    return rows

def fetch_ad_nm_windows(prev_start: datetime, prev_end: datetime, cur_start: datetime, cur_end: datetime):
    """Рекламные показы/затраты по nm за два окна (календарные дни).
    → (cur_by_nm, prev_by_nm) где значение {views, spend, clicks, orders}.
    Один запрос fullstats на весь диапазон (с батчами кампаний)."""
    empty = {}
    if not WB_TOKEN:
        return empty, empty
    begin = min(prev_start.date(), cur_start.date())
    end = max(prev_end.date(), cur_end.date())
    # лимит WB fullstats — 31 день
    if (end - begin).days > 30:
        begin = end - timedelta(days=30)
    try:
        campaigns_meta = fetch_campaigns_meta(include_finished=False)
    except Exception as e:
        logger.error(f"sales-pace ads meta error: {e}")
        return empty, empty
    if not campaigns_meta:
        return empty, empty

    ids = list(campaigns_meta.keys())
    all_rows = []
    for i in range(0, len(ids), 50):
        batch = ids[i:i + 50]
        try:
            resp = httpx.get(
                f"{WB_PROMOTION_URL}/adv/v3/fullstats",
                headers=wb_headers(),
                params={"ids": ",".join(str(x) for x in batch), "beginDate": begin.isoformat(), "endDate": end.isoformat()},
                timeout=60,
            )
            if resp.status_code == 429:
                logger.warning("sales-pace fullstats 429 — ждём 22с")
                time.sleep(22)
                resp = httpx.get(
                    f"{WB_PROMOTION_URL}/adv/v3/fullstats",
                    headers=wb_headers(),
                    params={"ids": ",".join(str(x) for x in batch), "beginDate": begin.isoformat(), "endDate": end.isoformat()},
                    timeout=60,
                )
            if not resp.is_success:
                logger.error(f"sales-pace fullstats {resp.status_code}: {resp.text[:200]}")
                break
            all_rows.extend(_fullstats_nm_day_rows(resp.json()))
        except Exception as e:
            logger.error(f"sales-pace fullstats exception: {e}")
            break
        if i + 50 < len(ids):
            time.sleep(20)

    prev_a0, prev_a1 = prev_start.date(), prev_end.date()
    cur_a0, cur_a1 = cur_start.date(), cur_end.date()
    cur_by, prev_by = {}, {}

    def _add(bucket, nm_id, row):
        a = bucket.setdefault(nm_id, {"views": 0, "spend": 0.0, "clicks": 0, "orders": 0})
        a["views"] += row["views"]
        a["spend"] += row["spend"]
        a["clicks"] += row["clicks"]
        a["orders"] += row["orders"]

    for row in all_rows:
        try:
            d = datetime.strptime(row["day"], "%Y-%m-%d").date()
        except Exception:
            continue
        if cur_a0 <= d <= cur_a1:
            _add(cur_by, row["nm_id"], row)
        if prev_a0 <= d <= prev_a1:
            _add(prev_by, row["nm_id"], row)

    for bucket in (cur_by, prev_by):
        for a in bucket.values():
            a["spend"] = round(a["spend"], 2)
    return cur_by, prev_by

def _cpm(views, spend):
    views = int(views or 0)
    spend = float(spend or 0)
    if views <= 0:
        return None
    return round(spend / views * 1000, 1)

def _ads_period_dates():
    """Период статистики рекламы: ads_date_from/to или окно ads_window_days (макс. 31 день)."""
    today = datetime.now(timezone.utc).date()
    raw_from = get_setting_raw("ads_date_from", None)
    raw_to = get_setting_raw("ads_date_to", None)
    begin = end = None
    try:
        if raw_from:
            begin = datetime.strptime(str(raw_from)[:10], "%Y-%m-%d").date()
        if raw_to:
            end = datetime.strptime(str(raw_to)[:10], "%Y-%m-%d").date()
    except Exception:
        begin = end = None
    if begin and end:
        if end < begin:
            begin, end = end, begin
        # лимит WB fullstats — 31 день
        if (end - begin).days > 30:
            begin = end - timedelta(days=30)
        return begin, end
    window_days = get_setting_int("ads_window_days", 7)
    window_days = min(max(window_days, 1), 31)
    end = today
    begin = end - timedelta(days=window_days - 1)
    return begin, end

def sync_ads():
    """Синк рекламы: список кампаний с затратами/показами/ДРР/заказами."""
    if not WB_TOKEN:
        ADS_CACHE["error"] = "WB_TOKEN не задан"
        logger.error("WB_TOKEN not set")
        return
    if ADS_CACHE.get("syncing"):
        return
    ADS_CACHE["syncing"] = True
    ADS_CACHE["error"] = None
    ADS_CACHE["progress"] = "список кампаний…"
    try:
        logger.info("Starting ads (promotion) sync...")
        begin_date, end_date = _ads_period_dates()
        window_days = (end_date - begin_date).days + 1

        # только активные + пауза — иначе тянем сотни завершённых и ждём по 20с на батч
        campaigns_meta = fetch_campaigns_meta(include_finished=False)
        if not campaigns_meta:
            ADS_CACHE["error"] = "Нет активных или на паузе кампаний"
            logger.info("Ads sync: no eligible campaigns")
            return

        logger.info(f"Ads sync: {len(campaigns_meta)} campaigns, {begin_date}…{end_date}")
        agg = fetch_ad_stats_by_campaign(
            list(campaigns_meta.keys()), begin_date.isoformat(), end_date.isoformat()
        )

        now = datetime.now(timezone.utc)
        now_iso = now.isoformat()
        campaigns = []
        rows = []
        for campaign_id, meta in campaigns_meta.items():
            a = agg.get(campaign_id) or {
                "views": 0, "clicks": 0, "atbs": 0, "orders": 0, "spend": 0.0, "revenue": 0.0,
            }
            views, clicks, atbs, orders = a["views"], a["clicks"], a["atbs"], a["orders"]
            spend, revenue = round(a["spend"], 2), round(a["revenue"], 2)
            drr = round(spend / revenue * 100, 2) if revenue else 0
            ctr = round(clicks / views * 100, 2) if views else 0
            cpc = round(spend / clicks, 2) if clicks else 0
            cr = round(orders / clicks * 100, 2) if clicks else 0
            name = meta.get("name") or f"#{campaign_id}"
            item = {
                "campaign_id": campaign_id,
                "campaign_name": name,
                "campaign_type": meta.get("type_name") or "Неизвестно",
                "views": views,
                "clicks": clicks,
                "atbs": atbs,
                "orders": orders,
                "spend": spend,
                "revenue": revenue,
                "drr": drr,
                "ctr": ctr,
                "cpc": cpc,
                "cr": cr,
                "cv_atb": round(atbs / clicks * 100, 2) if clicks else 0,
                "cv_ord": round(orders / atbs * 100, 2) if atbs else 0,
                "period_days": window_days,
                "updated_at": now_iso,
                "vendor_code": name,
                "nm_id": campaign_id,
            }
            campaigns.append(item)
            rows.append(item)

        campaigns.sort(key=lambda c: (-(c["spend"] or 0), str(c["campaign_name"])))

        ADS_CACHE["campaigns"] = campaigns
        ADS_CACHE["updated_at"] = now.strftime("%d.%m.%Y %H:%M")
        ADS_CACHE["window_days"] = window_days
        ADS_CACHE["period_begin"] = begin_date.isoformat()
        ADS_CACHE["period_end"] = end_date.isoformat()
        ADS_CACHE["error"] = None
        ADS_CACHE["progress"] = None

        if rows:
            try:
                httpx.delete(f"{SUPABASE_URL}/rest/v1/ad_stats?id=gte.0", headers=sb_headers(), timeout=15)
                saved = 0
                for i in range(0, len(rows), 200):
                    batch = rows[i:i + 200]
                    resp = httpx.post(f"{SUPABASE_URL}/rest/v1/ad_stats", json=batch, headers=sb_headers(), timeout=30)
                    if resp.is_success:
                        saved += len(batch)
                    else:
                        logger.error(f"ad_stats insert error {resp.status_code} {resp.text[:300]}")
                logger.info(f"Ads sync: saved {saved}/{len(rows)} rows to supabase")
            except Exception as e:
                logger.error(f"Ads sync supabase write error: {e}")

        httpx.post(
            f"{SUPABASE_URL}/rest/v1/settings",
            json={"key": "last_ads_sync", "value": ADS_CACHE["updated_at"], "updated_at": now_iso},
            headers=sb_headers(), timeout=10,
        )
        logger.info(f"Ads sync complete. Campaigns: {len(campaigns)}")
    except Exception as e:
        logger.error(f"sync_ads error: {e}")
        ADS_CACHE["error"] = str(e)
        ADS_CACHE["progress"] = None
    finally:
        ADS_CACHE["syncing"] = False
        ADS_CACHE["progress"] = None

@app.get("/api/ads")
def get_ads(refresh: bool = False):
    """Список кампаний с метриками. refresh=1 — запустить синк."""
    if refresh and not ADS_CACHE.get("syncing"):
        import threading
        threading.Thread(target=sync_ads, daemon=True).start()
    camps = ADS_CACHE.get("campaigns") or []
    # если кэш пуст — подтянем из supabase (после рестарта)
    if not camps and not ADS_CACHE.get("syncing"):
        try:
            ads = httpx.get(f"{SUPABASE_URL}/rest/v1/ad_stats?select=*", headers=sb_headers(), timeout=20)
            if ads.is_success:
                raw = ads.json() or []
                by = {}
                for r in raw:
                    cid = r.get("campaign_id")
                    if cid is None:
                        continue
                    if cid not in by:
                        by[cid] = {
                            "campaign_id": cid,
                            "campaign_name": r.get("campaign_name") or f"#{cid}",
                            "campaign_type": r.get("campaign_type") or "",
                            "views": 0, "clicks": 0, "atbs": 0, "orders": 0,
                            "spend": 0.0, "revenue": 0.0, "drr": 0,
                            "vendor_code": r.get("campaign_name") or str(cid),
                            "nm_id": cid,
                        }
                    c = by[cid]
                    c["views"] += r.get("views") or 0
                    c["clicks"] += r.get("clicks") or 0
                    c["atbs"] += r.get("atbs") or 0
                    c["orders"] += r.get("orders") or 0
                    c["spend"] += float(r.get("spend") or 0)
                    c["revenue"] += float(r.get("revenue") or 0)
                for c in by.values():
                    c["spend"] = round(c["spend"], 2)
                    c["revenue"] = round(c["revenue"], 2)
                    c["drr"] = round(c["spend"] / c["revenue"] * 100, 2) if c["revenue"] else 0
                camps = sorted(by.values(), key=lambda x: (-x["spend"], str(x["campaign_name"])))
                ADS_CACHE["campaigns"] = camps
        except Exception as e:
            logger.error(f"get_ads supabase fallback: {e}")
    return {
        "campaigns": camps,
        "ad_stats": camps,  # алиас под старый фронт
        "updated_at": ADS_CACHE.get("updated_at"),
        "syncing": ADS_CACHE.get("syncing", False),
        "error": ADS_CACHE.get("error"),
        "progress": ADS_CACHE.get("progress"),
        "window_days": ADS_CACHE.get("window_days"),
        "period_begin": ADS_CACHE.get("period_begin"),
        "period_end": ADS_CACHE.get("period_end"),
    }

# ---------- Календарь акций (Промо) ----------
# Данные акций держим в кэше процесса: WB жёстко лимитирует Календарь акций
# (10 запросов / 6 сек), а на построение матрицы нужно 2 запроса на каждую акцию.
# Обновляется по кнопке, по расписанию и при первом обращении к вкладке.
PROMO_CACHE = {"promotions": [], "articles": [], "updated_at": None, "syncing": False, "error": None}
PROMO_RATE_DELAY = 0.7  # пауза между запросами к Календарю акций (лимит WB: интервал 600 мс)

def fetch_calendar_promotions(start_dt: str, end_dt: str, all_promo: bool = False) -> list:
    """Список акций за период [start_dt, end_dt]. all_promo=False — доступные для участия."""
    promotions, offset, limit = [], 0, 1000
    while True:
        try:
            resp = httpx.get(
                f"{WB_CALENDAR_URL}/api/v1/calendar/promotions",
                headers=wb_headers(),
                params={"startDateTime": start_dt, "endDateTime": end_dt,
                        "allPromo": "true" if all_promo else "false",
                        "limit": limit, "offset": offset},
                timeout=30,
            )
        except Exception as e:
            logger.error(f"calendar promotions fetch error: {e}")
            break
        if not resp.is_success:
            logger.error(f"calendar promotions error {resp.status_code} {resp.text[:200]}")
            break
        batch = (resp.json().get("data") or {}).get("promotions") or []
        promotions.extend(batch)
        if len(batch) < limit:
            break
        offset += limit
        time.sleep(PROMO_RATE_DELAY)
    return promotions

def fetch_promotion_nomenclatures(promotion_id: int, in_action: bool) -> list:
    """Товары акции: in_action=False — можно добавить (есть planPrice для входа), True — уже участвуют."""
    noms, offset, limit = [], 0, 1000
    while True:
        try:
            resp = httpx.get(
                f"{WB_CALENDAR_URL}/api/v1/calendar/promotions/nomenclatures",
                headers=wb_headers(),
                params={"promotionID": promotion_id, "inAction": str(in_action).lower(),
                        "limit": limit, "offset": offset},
                timeout=30,
            )
        except Exception as e:
            logger.error(f"promo nomenclatures fetch error (promo {promotion_id}): {e}")
            break
        if not resp.is_success:
            # 400/404 — у акции просто нет подходящих товаров, это не критично
            if resp.status_code not in (400, 404):
                logger.error(f"promo nomenclatures error {resp.status_code} {resp.text[:200]}")
            break
        batch = (resp.json().get("data") or {}).get("nomenclatures") or []
        noms.extend(batch)
        if len(batch) < limit:
            break
        offset += limit
        time.sleep(PROMO_RATE_DELAY)
    return noms

def fetch_promotions_details(ids: list) -> dict:
    """Детали акций по ID → {promo_id: {...}}. Работает и для автоакций (в отличие от nomenclatures)."""
    out = {}
    for i in range(0, len(ids), 50):
        batch = ids[i:i + 50]
        try:
            resp = httpx.get(
                f"{WB_CALENDAR_URL}/api/v1/calendar/promotions/details",
                headers=wb_headers(),
                params=[("promotionIDs", str(x)) for x in batch],
                timeout=30,
            )
        except Exception as e:
            logger.error(f"promo details fetch error: {e}")
            continue
        if not resp.is_success:
            logger.error(f"promo details error {resp.status_code} {resp.text[:200]}")
            continue
        for d in ((resp.json().get("data") or {}).get("promotions") or []):
            pid = d.get("id")
            if pid is not None:
                out[pid] = d
        if i + 50 < len(ids):
            time.sleep(PROMO_RATE_DELAY)
    return out

def build_promo_nm_to_vendor() -> dict:
    """nm_id → артикул продавца: сначала из stock_totals, добираем из feedbacks."""
    nm_to_vendor = {}
    try:
        st = httpx.get(f"{SUPABASE_URL}/rest/v1/stock_totals?select=nm_id,vendor_code", headers=sb_headers(), timeout=15)
        if st.is_success:
            nm_to_vendor = {r["nm_id"]: r["vendor_code"] for r in st.json() if r.get("nm_id") and r.get("vendor_code")}
    except Exception as e:
        logger.error(f"sync_promotions: stock_totals fetch error {e}")
    try:
        fb = httpx.get(
            f"{SUPABASE_URL}/rest/v1/feedbacks?select=nm_id,article&nm_id=not.is.null&article=not.is.null",
            headers=sb_headers(), timeout=15
        )
        if fb.is_success:
            for r in fb.json():
                nm, art = r.get("nm_id"), r.get("article", "")
                if nm and art and nm not in nm_to_vendor:
                    nm_to_vendor[nm] = art
    except Exception as e:
        logger.error(f"sync_promotions: feedbacks fallback error {e}")
    return nm_to_vendor

def sync_promotions():
    """Строит матрицу: акции × артикулы, с ценой входа и разницей к текущей цене."""
    if not WB_TOKEN:
        logger.error("WB_TOKEN not set")
        PROMO_CACHE["error"] = "WB_TOKEN не задан"
        return
    if PROMO_CACHE.get("syncing"):
        logger.info("Promotions sync already running")
        return
    PROMO_CACHE["syncing"] = True
    PROMO_CACHE["error"] = None
    try:
        logger.info("Starting promotions (calendar) sync...")
        window_days = min(get_setting_int("promo_window_days", 60), 365)
        now = datetime.now(timezone.utc)
        start_dt = now.strftime("%Y-%m-%dT00:00:00Z")
        end_dt = (now + timedelta(days=window_days)).strftime("%Y-%m-%dT23:59:59Z")

        promos = fetch_calendar_promotions(start_dt, end_dt)
        logger.info(f"Promotions sync: {len(promos)} promotions in window")

        nm_to_vendor = build_promo_nm_to_vendor()

        # Детали по всем акциям — работают и для автоакций (участие товаров, охват, буст)
        promo_ids = [p.get("id") for p in promos if p.get("id") is not None]
        details = fetch_promotions_details(promo_ids) if promo_ids else {}

        promotions_out, articles = [], {}
        for p in promos:
            pid = p.get("id")
            if pid is None:
                continue
            start = p.get("startDateTime", "")
            end = p.get("endDateTime", "")
            days_to_start = None
            try:
                sd = datetime.fromisoformat(start.replace("Z", "+00:00"))
                days_to_start = (sd.date() - now.date()).days
            except Exception:
                pass
            ptype = p.get("type", "regular")
            d = details.get(pid, {})
            ranging = d.get("ranging") or []
            max_boost = max((r.get("boost", 0) or 0 for r in ranging), default=0)
            promotions_out.append({
                "id": pid,
                "name": p.get("name", f"#{pid}"),
                "start": start,
                "end": end,
                "type": ptype,
                "days_to_start": days_to_start,
                "in_total": d.get("inPromoActionTotal"),
                "in_leftovers": d.get("inPromoActionLeftovers"),
                "not_in_total": d.get("notInPromoActionTotal"),
                "not_in_leftovers": d.get("notInPromoActionLeftovers"),
                "participation": d.get("participationPercentage"),
                "exceptions": d.get("exceptionProductsCount"),
                "boost": max_boost,
            })
            # Список товаров с ценами входа доступен только для обычных акций.
            # Для автоакций WB не отдаёт номенклатуры — пропускаем, чтобы не ловить 400.
            if ptype != "regular":
                continue
            for in_action in (True, False):
                time.sleep(PROMO_RATE_DELAY)
                for n in fetch_promotion_nomenclatures(pid, in_action):
                    nm = n.get("id")
                    if nm is None:
                        continue
                    price = n.get("price")
                    plan_price = n.get("planPrice")
                    delta = round(price - plan_price) if (price is not None and plan_price is not None) else None
                    entry = articles.setdefault(nm, {
                        "nm_id": nm,
                        "vendor_code": nm_to_vendor.get(nm) or str(nm),
                        "cells": {},
                    })
                    entry["cells"][str(pid)] = {
                        "in_action": bool(n.get("inAction")),
                        "price": price,
                        "plan_price": plan_price,
                        "discount": n.get("discount"),
                        "plan_discount": n.get("planDiscount"),
                        "delta": delta,
                    }

        articles_list = sorted(articles.values(), key=lambda a: str(a["vendor_code"]))
        PROMO_CACHE["promotions"] = promotions_out
        PROMO_CACHE["articles"] = articles_list
        PROMO_CACHE["updated_at"] = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")
        logger.info(f"Promotions sync complete. Promotions: {len(promotions_out)}, articles: {len(articles_list)}")
    except Exception as e:
        logger.error(f"sync_promotions error: {e}")
        PROMO_CACHE["error"] = str(e)
    finally:
        PROMO_CACHE["syncing"] = False

@app.get("/api/promotions")
def get_promotions():
    # первый заход на вкладку — запускаем сбор данных в фоне
    if not PROMO_CACHE["updated_at"] and not PROMO_CACHE["syncing"]:
        import threading
        threading.Thread(target=sync_promotions, daemon=True).start()
    return {
        "promotions": PROMO_CACHE["promotions"],
        "articles": PROMO_CACHE["articles"],
        "updated_at": PROMO_CACHE["updated_at"],
        "syncing": PROMO_CACHE["syncing"],
        "error": PROMO_CACHE["error"],
    }

@app.post("/api/sync-promotions")
def trigger_promotions_sync():
    import threading
    threading.Thread(target=sync_promotions, daemon=True).start()
    return {"status": "started"}


# ---------- Календарь акций (лента как в кабинете WB) ----------

PROMO_CAL_CACHE = {
    "promotions": [],
    "updated_at": None,
    "syncing": False,
    "error": None,
    "from": None,
    "to": None,
}


def _promo_msk_date(iso):
    """Дата акции по Москве: WB отдаёт UTC, 21:00Z это уже следующий день."""
    if not iso:
        return None
    try:
        dt = datetime.fromisoformat(str(iso).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        try:
            from zoneinfo import ZoneInfo
            dt = dt.astimezone(ZoneInfo("Europe/Moscow"))
        except Exception:
            dt = dt.astimezone(timezone(timedelta(hours=3)))
        return dt.date()
    except Exception:
        return None


def _promo_cal_status(start_d, end_d, in_total, today):
    """Участвую / буду участвовать / идёт без нас / доступна / прошла."""
    in_total = int(in_total or 0)
    if not start_d or not end_d:
        return "unknown", "Дата неизвестна"
    if end_d < today:
        return ("ended", "Прошла") if in_total <= 0 else ("ended", "Участвовали")
    if start_d > today:
        if in_total > 0:
            days = (start_d - today).days
            when = "завтра" if days == 1 else f"через {days} дн."
            return "will", f"Буду участвовать · старт {when}"
        days = (start_d - today).days
        when = "завтра" if days == 1 else f"через {days} дн."
        return "soon", f"Старт {when}"
    # идёт сейчас
    left = (end_d - today).days
    left_s = "последний день" if left <= 0 else f"ещё {left} дн."
    if in_total > 0:
        return "in", f"Участвую · {left_s}"
    return "live", f"Идёт · {left_s}"


def build_promo_calendar_items(promos: list, details: dict, today=None) -> list:
    today = today or _msk_now().date()
    out = []
    for p in promos or []:
        pid = p.get("id")
        if pid is None:
            continue
        d = details.get(pid) or {}
        start_iso = d.get("startDateTime") or p.get("startDateTime") or ""
        end_iso = d.get("endDateTime") or p.get("endDateTime") or ""
        start_d = _promo_msk_date(start_iso)
        end_d = _promo_msk_date(end_iso)
        in_total = d.get("inPromoActionTotal")
        if in_total is None:
            in_total = p.get("inPromoActionTotal") or 0
        not_in = d.get("notInPromoActionTotal")
        if not_in is None:
            not_in = p.get("notInPromoActionTotal") or 0
        ranging = d.get("ranging") or []
        max_boost = max((r.get("boost", 0) or 0 for r in ranging), default=0)
        max_rate = max((r.get("participationRate", 0) or 0 for r in ranging), default=0)
        status, status_label = _promo_cal_status(start_d, end_d, in_total, today)
        days_to_start = (start_d - today).days if start_d else None
        ptype = d.get("type") or p.get("type") or "regular"
        out.append({
            "id": pid,
            "name": d.get("name") or p.get("name") or f"#{pid}",
            "type": ptype,
            "start": start_iso,
            "end": end_iso,
            "start_date": start_d.isoformat() if start_d else None,
            "end_date": end_d.isoformat() if end_d else None,
            "days_to_start": days_to_start,
            "status": status,
            "status_label": status_label,
            "in_total": int(in_total or 0),
            "not_in_total": int(not_in or 0),
            "in_leftovers": d.get("inPromoActionLeftovers"),
            "participation": d.get("participationPercentage"),
            "boost": max_boost,
            "plan_discount": max_rate or None,
            "advantages": d.get("advantages") or [],
        })
    out.sort(key=lambda x: (x.get("start_date") or "9999", x.get("name") or ""))
    return out


def sync_promo_calendar():
    """Только список акций и детали — без номенклатуры, чтобы календарь открывался быстро."""
    if not WB_TOKEN:
        PROMO_CAL_CACHE["error"] = "WB_TOKEN не задан"
        return
    if PROMO_CAL_CACHE.get("syncing"):
        return
    PROMO_CAL_CACHE["syncing"] = True
    PROMO_CAL_CACHE["error"] = None
    try:
        today = _msk_now().date()
        start = (today - timedelta(days=14)).strftime("%Y-%m-%dT00:00:00Z")
        end = (today + timedelta(days=90)).strftime("%Y-%m-%dT23:59:59Z")
        logger.info(f"promo calendar sync {start} … {end}")
        promos = fetch_calendar_promotions(start, end, all_promo=False)
        promo_ids = [p.get("id") for p in promos if p.get("id") is not None]
        details = fetch_promotions_details(promo_ids) if promo_ids else {}
        items = build_promo_calendar_items(promos, details, today)
        PROMO_CAL_CACHE["promotions"] = items
        PROMO_CAL_CACHE["updated_at"] = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")
        PROMO_CAL_CACHE["from"] = (today - timedelta(days=14)).isoformat()
        PROMO_CAL_CACHE["to"] = (today + timedelta(days=90)).isoformat()
        logger.info(f"promo calendar: {len(items)} акций")
    except Exception as e:
        logger.error(f"sync_promo_calendar: {e}")
        PROMO_CAL_CACHE["error"] = str(e)
    finally:
        PROMO_CAL_CACHE["syncing"] = False


def _promo_cal_payload():
    items = PROMO_CAL_CACHE.get("promotions") or []
    return {
        "promotions": items,
        "updated_at": PROMO_CAL_CACHE.get("updated_at"),
        "syncing": PROMO_CAL_CACHE.get("syncing", False),
        "error": PROMO_CAL_CACHE.get("error"),
        "from": PROMO_CAL_CACHE.get("from"),
        "to": PROMO_CAL_CACHE.get("to"),
        "counts": {
            "all": len(items),
            "in": sum(1 for x in items if x.get("status") in ("in", "will")),
            "available": sum(1 for x in items if x.get("status") in ("soon", "live")),
            "soon": sum(1 for x in items if x.get("status") in ("soon", "will")),
        },
    }


@app.get("/api/promo-calendar")
def get_promo_calendar(refresh: int = 0):
    if refresh and not PROMO_CAL_CACHE.get("syncing"):
        threading.Thread(target=sync_promo_calendar, daemon=True).start()
    elif not PROMO_CAL_CACHE.get("updated_at") and not PROMO_CAL_CACHE.get("syncing"):
        threading.Thread(target=sync_promo_calendar, daemon=True).start()
    return _promo_cal_payload()


@app.post("/api/sync-promo-calendar")
def trigger_promo_calendar_sync():
    if PROMO_CAL_CACHE.get("syncing"):
        return {"status": "already_running"}
    threading.Thread(target=sync_promo_calendar, daemon=True).start()
    return {"status": "started"}

def _parse_promo_excel_name(filename: str) -> str:
    """Из имени файла WB: «...для акции_<название>_<дата время>.xlsx»."""
    import re as _re
    name = (filename or "").rsplit("/", 1)[-1]
    name = _re.sub(r"\.xlsx?$", "", name, flags=_re.I)
    m = _re.search(r"для акции[_\s]+(.+?)_\d{1,2}\.\d{1,2}\.\d{2,4}", name, flags=_re.I)
    if m:
        return m.group(1).strip(" _-")
    m = _re.search(r"акци[ия][_\s]+(.+)$", name, flags=_re.I)
    if m:
        return m.group(1).strip(" _-")
    return name or "Акция"

PROMO_SESSIONS_KEY = "promo_excel_sessions"


def _promo_sessions_payload(raw=None) -> dict:
    data = raw if isinstance(raw, dict) else {}
    sessions = data.get("sessions") if isinstance(data.get("sessions"), list) else []
    # только словари с id
    clean = []
    for s in sessions:
        if isinstance(s, dict) and s.get("id"):
            clean.append(s)
    active_id = data.get("active_id")
    if active_id and not any(s.get("id") == active_id for s in clean):
        active_id = clean[0]["id"] if clean else None
    if not active_id and clean:
        active_id = clean[0]["id"]
    return {
        "sessions": clean,
        "active_id": active_id,
        "updated_at": data.get("updated_at"),
    }


def _save_promo_sessions(payload: dict) -> bool:
    """Отдельный таймаут — Excel акций может быть большим."""
    import json as _json
    body = {
        "key": PROMO_SESSIONS_KEY,
        "value": _json.dumps(payload, ensure_ascii=False),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    try:
        resp = httpx.post(
            f"{SUPABASE_URL}/rest/v1/settings?on_conflict=key",
            json=body,
            headers=sb_headers(),
            timeout=60,
        )
        return resp.is_success
    except Exception as e:
        logger.error(f"save promo sessions error: {e}")
        return False


@app.get("/api/promo-sessions")
def get_promo_sessions():
    """Общие сессии загруженных Excel акций (видны со всех устройств)."""
    raw = get_setting_json(PROMO_SESSIONS_KEY, {}) or {}
    return _promo_sessions_payload(raw)


@app.put("/api/promo-sessions")
def put_promo_sessions(request: dict):
    """Сохранить список сессий акций + активную."""
    payload = _promo_sessions_payload({
        "sessions": request.get("sessions") if isinstance(request.get("sessions"), list) else [],
        "active_id": request.get("active_id"),
    })
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    if not _save_promo_sessions(payload):
        return {"error": "Не удалось сохранить акции в базу"}
    return {"status": "ok", **payload}


@app.post("/api/upload-promo-excel")
async def upload_promo_excel(file: UploadFile = File(...)):
    """Парсит xlsx «Все товары подходящие для акции_…» из Календаря акций WB.
    Возвращает список артикулов с ценами/участием — фронт сохраняет сессии через /api/promo-sessions."""
    try:
        from openpyxl import load_workbook
        import re as _re

        contents = await file.read()
        if not contents:
            return {"error": "Пустой файл"}

        wb = load_workbook(io.BytesIO(contents), data_only=False)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"error": "Пустой лист"}

        # Ищем строку заголовков
        header_i = None
        header = None
        for i, row in enumerate(rows[:15]):
            vals = [str(c or "").strip().lower() for c in row]
            joined = " | ".join(vals)
            if "артикул" in joined and ("планов" in joined or "участ" in joined or "wb" in joined):
                header_i = i
                header = [str(c or "").strip() for c in row]
                break
        if header_i is None:
            return {"error": "Не найдены заголовки (нужны колонки артикул / плановая цена). Проверь, что это файл из Календаря акций."}

        def find_col(*needles):
            for j, h in enumerate(header):
                hl = h.lower()
                if all(n.lower() in hl for n in needles):
                    return j
            return None

        col_in = find_col("участ")  # «Товар уже участвует в акции»
        col_brand = find_col("бренд")
        col_subject = find_col("предмет")
        col_name = find_col("наименование")
        col_vc = find_col("артикул поставщика") or find_col("артикул продавца")
        col_nm = find_col("артикул wb") or find_col("артикул вб")
        col_turn = find_col("оборачиваемость")
        col_stock_wb = find_col("остаток", "складах") or find_col("остаток товара на складах")
        col_stock_seller = find_col("остаток", "продавца")
        col_plan = find_col("плановая")
        col_price = find_col("текущая розничная") or find_col("розничная цена")
        col_cur_disc = find_col("текущая скидка")
        col_load_disc = find_col("загружаемая скидка")
        col_status = find_col("статус")

        if col_vc is None and col_nm is None:
            return {"error": "Нет колонки артикула поставщика / WB"}

        def cell(row, idx):
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            if v is None:
                return None
            if isinstance(v, str) and not v.strip():
                return None
            return v

        def to_num(v):
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return float(v)
            s = str(v).replace("\xa0", "").replace(" ", "").replace(",", ".").replace("%", "")
            try:
                return float(s)
            except Exception:
                return None

        def to_bool_in_action(v):
            if v is None:
                return False
            if isinstance(v, bool):
                return v
            s = str(v).strip().lower()
            return s in ("да", "yes", "true", "1", "участвует")

        articles = []
        for row in rows[header_i + 1:]:
            if not row or not any(c is not None and str(c).strip() for c in row):
                continue
            vc = cell(row, col_vc)
            nm = cell(row, col_nm)
            if vc is None and nm is None:
                continue
            price = to_num(cell(row, col_price))  # розничная до скидки продавца
            plan = to_num(cell(row, col_plan))
            cur_disc = to_num(cell(row, col_cur_disc))
            # Цена «как на сайте» = розничная минус скидка продавца (6000−18% = 4920)
            price_sale = None
            if price is not None:
                if cur_disc is not None:
                    price_sale = round(price * (100.0 - cur_disc) / 100.0)
                else:
                    price_sale = price
            delta = None
            if price_sale is not None and plan is not None:
                delta = int(round(price_sale - plan))
            turn = to_num(cell(row, col_turn))
            stock_wb = to_num(cell(row, col_stock_wb)) or 0
            stock_seller = to_num(cell(row, col_stock_seller)) or 0
            in_action = to_bool_in_action(cell(row, col_in))
            # слабые/пустые: оборачиваемость 999 у WB = нет продаж
            is_weak = (turn is not None and turn >= 999) or (stock_wb + stock_seller <= 0 and (turn is None or turn >= 100))
            need = (not in_action) and (delta is not None and delta > 0) and not is_weak
            if need:
                priority = "need"
            elif is_weak:
                priority = "weak"
            elif in_action:
                priority = "in"
            else:
                priority = "other"

            articles.append({
                "vendor_code": str(vc).strip() if vc is not None else str(nm),
                "nm_id": int(nm) if isinstance(nm, (int, float)) else (int(to_num(nm)) if to_num(nm) else None),
                "name": str(cell(row, col_name) or "") or None,
                "brand": str(cell(row, col_brand) or "") or None,
                "subject": str(cell(row, col_subject) or "") or None,
                "in_action": in_action,
                "plan_price": plan,
                "price": price,
                "price_sale": price_sale,
                "delta": delta,
                "turnover": turn,
                "stock_wb": stock_wb,
                "stock_seller": stock_seller,
                "stock": stock_wb + stock_seller,
                "cur_discount": cur_disc,
                "load_discount": to_num(cell(row, col_load_disc)),
                "status": str(cell(row, col_status) or "") or None,
                "priority": priority,
            })

        # Сортировка: нужные сверху → в акции → прочие → слабые снизу; внутри по −₽
        order = {"need": 0, "in": 1, "other": 2, "weak": 3}
        articles.sort(key=lambda a: (order.get(a["priority"], 9), a["delta"] if a.get("delta") is not None else 10**12, str(a["vendor_code"])))

        promo_name = _parse_promo_excel_name(file.filename or "")
        need_n = sum(1 for a in articles if a["priority"] == "need")
        in_n = sum(1 for a in articles if a["in_action"])
        weak_n = sum(1 for a in articles if a["priority"] == "weak")

        return {
            "promo_name": promo_name,
            "filename": file.filename,
            "articles": articles,
            "stats": {
                "total": len(articles),
                "need": need_n,
                "in_action": in_n,
                "weak": weak_n,
            },
        }
    except Exception as e:
        logger.error(f"upload-promo-excel error: {e}")
        return {"error": str(e)}

@app.post("/api/upload-ratings")
async def upload_ratings(file: UploadFile = File(...)):
    """
    Принимает xlsx файл "Оценка товара" из WB Partners и сохраняет в Supabase.
    Лист "Товары" содержит правильные данные с учётом исключённых отзывов.
    """
    try:
        contents = await file.read()
        xl = pd.ExcelFile(io.BytesIO(contents))

        # Ищем лист с детальными данными по артикулам, перебирая ВСЕ листы
        # и проверяя где встречается заголовок "Артикул продавца".
        # WB называет этот лист по-разному в зависимости от настроек отчёта:
        # "Товары", "Детализация по артикулам", и т.д.
        sheet_name = None
        header_row = None
        df = None
        for s in xl.sheet_names:
            tmp = pd.read_excel(io.BytesIO(contents), sheet_name=s, header=None)
            for i, row in tmp.iterrows():
                vals = [str(v).strip() for v in row.values]
                if any('артикул продавца' in v.lower() for v in vals):
                    sheet_name = s
                    header_row = i
                    df = tmp
                    break
            if sheet_name:
                break

        logger.info(f"Detected sheet: {sheet_name}, header_row: {header_row} (sheets in file: {xl.sheet_names})")

        if header_row is None:
            return {"error": f"Не найден заголовок 'Артикул продавца' ни на одном листе. Листы в файле: {xl.sheet_names}. Проверь что загружаешь файл 'Оценка товара' из WB Partners → Аналитика."}

        df.columns = df.iloc[header_row].str.strip()
        df = df.iloc[header_row + 1:].reset_index(drop=True)
        df = df.dropna(subset=[df.columns[0]])

        # Маппинг колонок
        col_map = {}
        for c in df.columns:
            cl = str(c).lower().strip()
            if 'артикул продавца' in cl:
                col_map['article'] = c
            elif 'артикул wb' in cl:
                col_map['nm_id'] = c
            elif 'название' in cl:
                col_map['name'] = c
            elif 'рейтинг по отзывам' in cl and 'выше' not in cl:
                col_map['wb_rating'] = c
            elif 'все отзывы за период' in cl or 'всего' in cl:
                col_map['reviews_total'] = c
            elif 'оценки 5' in cl:
                col_map['r5'] = c
            elif 'оценки 4' in cl:
                col_map['r4'] = c
            elif 'оценки 3' in cl:
                col_map['r3'] = c
            elif 'оценки 2' in cl:
                col_map['r2'] = c
            elif 'оценки 1' in cl:
                col_map['r1'] = c
            elif 'исключен' in cl:
                col_map['excluded'] = c

        logger.info(f"Column mapping: {col_map}")

        if 'article' not in col_map:
            return {"error": f"Не найдена колонка 'Артикул продавца'. Найденные колонки: {list(df.columns)}"}

        now = datetime.now(timezone.utc).isoformat()
        rows = []
        for _, row in df.iterrows():
            article = str(row.get(col_map.get('article', ''), '') or '').strip()
            if not article or article == 'nan':
                continue

            def safe_int(key):
                try:
                    v = row.get(col_map.get(key, ''))
                    return int(float(v)) if v and str(v) != 'nan' else 0
                except:
                    return 0

            def safe_int_abs(key):
                return abs(safe_int(key))

            def safe_float(key):
                try:
                    v = row.get(col_map.get(key, ''))
                    return float(v) if v and str(v) != 'nan' else None
                except:
                    return None

            r5v = safe_int('r5') or 0
            r4v = safe_int('r4') or 0
            r3v = safe_int('r3') or 0
            r2v = safe_int('r2') or 0
            r1v = safe_int('r1') or 0
            reviews_total = safe_int('reviews_total') or 0
            star_sum = r5v+r4v+r3v+r2v+r1v

            # Если reviews_total не заполнен но звёзды есть — считаем из них
            if not reviews_total and star_sum > 0:
                reviews_total = star_sum

            # wb_rating берём из колонки "Рейтинг по отзывам"
            # Если там '-' или пусто — считаем из звёзд сами
            wb_rating = safe_float('wb_rating')
            if wb_rating is None and star_sum > 0:
                wb_rating = round((r5v*5+r4v*4+r3v*3+r2v*2+r1v) / star_sum, 2)

            rows.append({
                "article": article,
                "nm_id": safe_int('nm_id') or None,
                "name": str(row.get(col_map.get('name', ''), '') or '').strip() or None,
                "wb_rating": wb_rating,
                "reviews_total": reviews_total,
                "r5": r5v, "r4": r4v, "r3": r3v, "r2": r2v, "r1": r1v,
                "excluded": safe_int_abs('excluded'),
                "updated_at": now
            })

        if not rows:
            return {"error": "Не найдено строк с данными"}

        # Добавляем source='xlsx' каждой строке
        for r in rows:
            r["source"] = "xlsx"

        # Удаляем только НЕ-ручные строки (manual сохраняем)
        httpx.delete(
            f"{SUPABASE_URL}/rest/v1/ratings_official?source=neq.manual",
            headers={**sb_headers(), "Prefer": "return=minimal"}, timeout=15
        )
        # Также удаляем строки без source (старые данные)
        httpx.delete(
            f"{SUPABASE_URL}/rest/v1/ratings_official?source=is.null",
            headers={**sb_headers(), "Prefer": "return=minimal"}, timeout=15
        )

        # Получаем какие артикулы уже заняты ручными записями — их пропускаем
        manual_resp = httpx.get(
            f"{SUPABASE_URL}/rest/v1/ratings_official?source=eq.manual&select=article",
            headers=sb_headers(), timeout=10
        )
        manual_articles = {r["article"] for r in (manual_resp.json() if manual_resp.is_success else [])}
        rows = [r for r in rows if r["article"] not in manual_articles]

        saved = 0
        for i in range(0, len(rows), 100):
            batch = rows[i:i+100]
            resp = httpx.post(
                f"{SUPABASE_URL}/rest/v1/ratings_official",
                json=batch,
                headers=sb_headers(),
                timeout=30
            )
            if resp.is_success:
                saved += len(batch)
            else:
                logger.error(f"Supabase error: {resp.status_code} {resp.text[:300]}")

        # Обновляем время загрузки
        httpx.post(
            f"{SUPABASE_URL}/rest/v1/settings",
            json={"key": "last_ratings_upload", "value": datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M"), "updated_at": now},
            headers=sb_headers(), timeout=10
        )

        return {"status": "ok", "saved": saved, "total_rows": len(rows)}

    except Exception as e:
        logger.error(f"Upload error: {e}")
        return {"error": str(e)}

@app.post("/api/upload-competitor-report")
async def upload_competitor_report(file: UploadFile = File(...)):
    """Принимает xlsx «Сравнение карточек» и парсит только лист Показатели."""
    try:
        import re as _re
        contents = await file.read()

        # ── Период из Общая информация ──
        period_begin = period_end = None
        try:
            df_info = pd.read_excel(io.BytesIO(contents), sheet_name='Общая информация', header=None)
            for _, row in df_info.iterrows():
                if 'Выбранный период' in str(row.iloc[0]):
                    dates = _re.findall(r'\d{4}-\d{2}-\d{2}', str(row.iloc[1]))
                    if len(dates) >= 2:
                        period_begin, period_end = dates[0], dates[1]
                    break
        except Exception:
            pass

        # ── Читаем Показатели ──
        df = pd.read_excel(io.BytesIO(contents), sheet_name='Показатели', header=None)
        headers = [str(v) for v in df.iloc[1].values]

        # Колонки артикулов (не Разница, не предыдущий)
        art_cols = []
        for j, h in enumerate(headers):
            if 'Артикул WB' in h and 'предыдущий' not in h.lower() and 'Разница' not in h:
                m = _re.search(r'(\d{7,10})', h)
                if m:
                    art_cols.append((j, int(m.group(1))))

        if not art_cols:
            return {"error": "Не найдены артикулы в листе Показатели. Проверь формат файла."}

        # Артикулы из файла НЕ помечаем как «мой» — свой артикул добавляется вручную через поиск

        def cell(i, j):
            v = str(df.iloc[i, j]).strip()
            return None if v in ('nan','None','NaT','-','') else v

        def to_num(i, j):
            v = cell(i, j)
            if not v: return None
            try: return float(v.replace('\xa0','').replace(' ','').replace(',','.').replace('%',''))
            except: return None

        def find_row(label):
            for i in range(2, len(df)):
                if str(df.iloc[i, 0]).strip() == label:
                    return i
            return None

        METRICS = {
            'name': (['Название'], True),
            'brand': (['Бренд'], True),
            'card_rating': (['Рейтинг карточки'], False),
            'feedback_rating': (['Рейтинг по отзывам'], False),
            'reviews_count': (['Количество отзывов'], False),
            'price': (['Минимальная цена со скидкой (по размерам), ₽','Цена с учётом скидок, ₽'], False),
            'median_price': (['Медианная цена покупателя, ₽','Медианная цена покупателя'], False),
            'delivery_time': (['Среднее время доставки'], True),
            'avg_position': (['Средняя позиция','Средняя позиция в поиске'], False),
            'views': (['Показы'], False),
            'card_opens': (['Переход в карточку, шт','Переходы в карточку, шт'], False),
            'ctr': (['CTR'], False),
            'cart_adds': (['Добавления в корзину, шт'], False),
            'cart_conv': (['Конверсия в корзину, %'], False),
            'orders': (['Заказы, шт'], False),
            'order_conv': (['Конверсия в заказ, %'], False),
            'buyouts': (['Выкупы, шт'], False),
            'buyout_pct': (['Процент выкупа'], False),
            'cancels': (['Отмены, шт'], False),
        }
        INT_FIELDS = {'reviews_count','views','card_opens','cart_adds','orders','buyouts','cancels'}

        row_idx = {}
        for field, (labels, _) in METRICS.items():
            for label in labels:
                ri = find_row(label)
                if ri is not None:
                    row_idx[field] = ri
                    break

        # ── Создаём сессию ──
        session_row = {"period_begin": period_begin, "period_end": period_end}
        sess_resp = httpx.post(
            f"{SUPABASE_URL}/rest/v1/competitor_sessions",
            json=session_row,
            headers={**sb_headers(), "Prefer": "return=representation"},
            timeout=15
        )
        if not sess_resp.is_success:
            return {"error": f"Ошибка БД ({sess_resp.status_code}): {sess_resp.text[:200]}. Выполни competitor_tables.sql в Supabase."}
        try:
            session_id = sess_resp.json()[0]["id"]
        except Exception as e:
            return {"error": f"Ошибка сессии: {e}. Ответ: {sess_resp.text[:150]}"}

        # ── Сохраняем метрики ──
        now = datetime.now(timezone.utc).isoformat()
        rows = []
        for j, nm_id in art_cols:
            r = {"session_id": session_id, "nm_id": nm_id, "is_own": False, "updated_at": now}
            for field, (_, is_str) in METRICS.items():
                ri = row_idx.get(field)
                if ri is None:
                    r[field] = None
                elif is_str:
                    r[field] = cell(ri, j)
                else:
                    v = to_num(ri, j)
                    r[field] = int(v) if (v is not None and field in INT_FIELDS) else v
            rows.append(r)

        mr = httpx.post(f"{SUPABASE_URL}/rest/v1/competitor_metrics",
                        json=rows, headers=sb_headers(), timeout=20)
        if not mr.is_success:
            return {"error": f"Ошибка сохранения ({mr.status_code}): {mr.text[:200]}"}

        brands = sorted({(r.get("brand") or "").strip() for r in rows if (r.get("brand") or "").strip()})
        period = f"{period_begin or '?'} — {period_end or '?'}"
        logger.info(f"upload-competitor: session={session_id}, {len(rows)} articles, period={period}, brands={brands}")
        return {
            "status": "ok",
            "session_id": session_id,
            "period": period,
            "period_begin": period_begin,
            "period_end": period_end,
            "brands": brands,
            "articles": len(rows),
            "search_queries": 0,
        }

    except Exception as e:
        logger.error(f"upload-competitor error: {e}")
        import traceback; logger.error(traceback.format_exc())
        return {"error": str(e)}

@app.get("/api/own-articles-all")
def own_articles_all(shelf_focus: bool = False):
    """Свои карточки с заказами или выкупами за окно поставок (не весь каталог).

    shelf_focus=1 — только интересные для полок: цена ≥1200₽, не зарядки/кабели/аксессуары,
    с минимальными продажами. Для мониторинга «слабых полок» и задач по ним.
    """
    window_days = get_setting_int("sales_window_days", 14)
    by_nm = {}
    try:
        r = httpx.get(
            f"{SUPABASE_URL}/rest/v1/supply_report"
            f"?select=nm_id,vendor_code,ordered_qty,buyout_qty&limit=20000",
            headers=sb_headers(),
            timeout=25,
        )
        if not r.is_success:
            return {
                "articles": [],
                "days": window_days,
                "error": f"supply_report {r.status_code}",
            }
        for row in r.json() or []:
            nm = row.get("nm_id")
            if nm is None:
                continue
            try:
                nm = int(nm)
            except (TypeError, ValueError):
                continue
            a = by_nm.setdefault(nm, {
                "nm_id": nm,
                "vendor_code": "",
                "ordered_qty": 0,
                "buyout_qty": 0,
            })
            a["ordered_qty"] += int(row.get("ordered_qty") or 0)
            a["buyout_qty"] += int(row.get("buyout_qty") or 0)
            vc = (row.get("vendor_code") or "").strip()
            if vc and vc != str(nm):
                a["vendor_code"] = vc
    except Exception as e:
        logger.error(f"own-articles-all supply_report: {e}")
        return {"articles": [], "days": window_days, "error": str(e)}

    active = [
        a for a in by_nm.values()
        if (a["ordered_qty"] or 0) > 0 or (a["buyout_qty"] or 0) > 0
    ]
    if not active:
        return {
            "articles": [],
            "count": 0,
            "days": window_days,
            "error": "Нет карточек с заказами/выкупами за окно — сначала синхронизируй остатки и поставки",
        }

    nm_to_vendor = build_nm_to_vendor_map()
    for a in active:
        vc = (a.get("vendor_code") or "").strip()
        if not vc or vc == str(a["nm_id"]):
            mapped = nm_to_vendor.get(a["nm_id"])
            if mapped:
                a["vendor_code"] = mapped
        if not a.get("vendor_code"):
            a["vendor_code"] = str(a["nm_id"])

    skipped = []
    if shelf_focus:
        price_map, meta_map = _shelf_focus_enrichment([a["nm_id"] for a in active])
        focused = []
        for a in active:
            nm = a["nm_id"]
            meta = meta_map.get(nm) or {}
            price = price_map.get(nm)
            a["price"] = price
            a["name"] = meta.get("name") or ""
            a["subject"] = meta.get("subject") or ""
            reason = _shelf_focus_skip_reason(
                vendor_code=a.get("vendor_code"),
                name=a.get("name"),
                subject=a.get("subject"),
                price=price,
                ordered_qty=a.get("ordered_qty") or 0,
            )
            if reason:
                skipped.append({
                    "nm_id": nm,
                    "vendor_code": a.get("vendor_code"),
                    "reason": reason,
                    "price": price,
                    "ordered_qty": a.get("ordered_qty") or 0,
                })
                continue
            focused.append(a)
        active = focused

    active.sort(
        key=lambda a: (
            -(a["ordered_qty"] or 0),
            -(a["buyout_qty"] or 0),
            str(a.get("vendor_code") or "").lower(),
        )
    )
    out_articles = []
    for a in active:
        row = {
            "nm_id": a["nm_id"],
            "vendor_code": a["vendor_code"],
            "ordered_qty": a["ordered_qty"],
            "buyout_qty": a["buyout_qty"],
        }
        if shelf_focus:
            row["price"] = a.get("price")
            row["name"] = a.get("name") or ""
            row["subject"] = a.get("subject") or ""
        out_articles.append(row)
    return {
        "articles": out_articles,
        "count": len(out_articles),
        "days": window_days,
        "shelf_focus": bool(shelf_focus),
        "shelf_focus_min_price": SHELF_FOCUS_MIN_PRICE if shelf_focus else None,
        "shelf_focus_min_orders": SHELF_FOCUS_MIN_ORDERS if shelf_focus else None,
        "skipped": skipped[:80] if shelf_focus else [],
        "skipped_count": len(skipped) if shelf_focus else 0,
    }


# Карточки для мониторинга полок: без дешёвых и аксессуаров (зарядки и т.п.)
SHELF_FOCUS_MIN_PRICE = 1200
SHELF_FOCUS_MIN_ORDERS = 5
SHELF_FOCUS_SKIP_RE = re.compile(
    r"(заряд|charger|кабел|cable|адаптер|adapter|power\s*bank|powerbank|"
    r"провод|шнур|usb[\s\-]?[ac]|type[\s\-]?c|док[\s\-]?станц)",
    re.IGNORECASE,
)


def _shelf_focus_skip_reason(
    vendor_code: str = "",
    name: str = "",
    subject: str = "",
    price=None,
    ordered_qty: int = 0,
):
    """Почему карточку не смотрим в задачах по слабым полкам. None = ок."""
    blob = " ".join([str(vendor_code or ""), str(name or ""), str(subject or "")])
    if SHELF_FOCUS_SKIP_RE.search(blob):
        return "accessory"  # зарядки/кабели/чехлы и т.п.
    if price is not None:
        try:
            if float(price) < SHELF_FOCUS_MIN_PRICE:
                return "cheap"
        except (TypeError, ValueError):
            pass
    try:
        if int(ordered_qty or 0) < SHELF_FOCUS_MIN_ORDERS:
            return "low_sales"
    except (TypeError, ValueError):
        pass
    return None


def _shelf_focus_enrichment(nm_ids: list) -> tuple:
    """Цена + название/предмет для фильтра полок."""
    price_map = {}
    meta_map = {}
    # 1) живой кэш СПП
    for a in (SPP_CACHE.get("articles") or []):
        try:
            nm = int(a.get("nm_id"))
        except (TypeError, ValueError):
            continue
        p = a.get("client_price")
        if p is None:
            p = a.get("sale_price")
        if p is None:
            p = a.get("price")
        if p is not None:
            try:
                price_map[nm] = float(p)
            except (TypeError, ValueError):
                pass
        meta_map.setdefault(nm, {})
        if a.get("name"):
            meta_map[nm]["name"] = a.get("name")
    # 2) stock_totals — subject_name
    if nm_ids and SUPABASE_URL and SUPABASE_KEY:
        try:
            ids = ",".join(str(int(x)) for x in nm_ids[:500])
            resp = httpx.get(
                f"{SUPABASE_URL}/rest/v1/stock_totals",
                params={"select": "nm_id,subject_name,vendor_code", "nm_id": f"in.({ids})"},
                headers=sb_headers(),
                timeout=20,
            )
            if resp.is_success:
                for row in resp.json() or []:
                    try:
                        nm = int(row.get("nm_id"))
                    except (TypeError, ValueError):
                        continue
                    meta_map.setdefault(nm, {})
                    if row.get("subject_name"):
                        meta_map[nm]["subject"] = row.get("subject_name")
        except Exception as e:
            logger.warning(f"shelf focus stock_totals: {e}")
        # 3) последний снимок цены, если нет в кэше
        missing = [nm for nm in nm_ids if nm not in price_map]
        if missing:
            try:
                prev = fetch_latest_price_snapshots(missing)
                for nm, row in (prev or {}).items():
                    p = _num_or_none(row.get("client_price"))
                    if p is None:
                        p = _num_or_none(row.get("sale_price"))
                    if p is None:
                        p = _num_or_none(row.get("price"))
                    if p is not None:
                        price_map[int(nm)] = float(p)
            except Exception as e:
                logger.warning(f"shelf focus snapshots: {e}")
    return price_map, meta_map


@app.get("/api/search-own-articles")
def search_own_articles(q: str = ""):
    """Поиск своих артикулов по артикулу продавца (vendorCode).
    Сначала Content API WB (textSearch), иначе — ratings/feedbacks/stock с фильтром
    «не подставляй nmId вместо артикула продавца»."""
    q = (q or "").strip().rstrip(".…").strip()
    for ch in ("\\", "%", ",", "(", ")"):
        q = q.replace(ch, "")
    q = q.strip()

    def is_real_vendor(vc, nm_id=None) -> bool:
        vc = (vc or "").strip()
        if not vc:
            return False
        if nm_id is not None and vc == str(nm_id):
            return False
        return True

    # ── 1) Content API: настоящий vendorCode + поиск по префиксу/тексту ──
    if WB_TOKEN:
        try:
            filt = {"withPhoto": -1}
            if q:
                filt["textSearch"] = q
            resp = httpx.post(
                f"{WB_CONTENT_URL}/content/v2/get/cards/list",
                headers=wb_headers(),
                json={
                    "settings": {
                        "sort": {"ascending": True},
                        "filter": filt,
                        "cursor": {"limit": 80},
                    }
                },
                timeout=20,
            )
            if resp.is_success:
                cards = resp.json().get("cards") or []
                out, seen = [], set()
                ql = q.lower()
                for c in cards:
                    nm = c.get("nmID") or c.get("nmId")
                    vc = (c.get("vendorCode") or "").strip()
                    if not nm or not is_real_vendor(vc, nm) or nm in seen:
                        continue
                    if ql and not vc.lower().startswith(ql) and ql not in vc.lower():
                        continue
                    seen.add(nm)
                    out.append({"nm_id": nm, "vendor_code": vc})
                # Префиксные совпадения выше «содержит»
                if ql:
                    out.sort(key=lambda a: (
                        0 if a["vendor_code"].lower().startswith(ql) else 1,
                        a["vendor_code"].lower(),
                    ))
                else:
                    out.sort(key=lambda a: a["vendor_code"].lower())
                if out:
                    return out[:50]
            else:
                logger.warning(f"search-own content-api {resp.status_code}: {resp.text[:200]}")
        except Exception as e:
            logger.warning(f"search-own content-api error: {e}")

    # ── 2) Fallback из БД: ratings_official → feedbacks → stock_totals ──
    by_nm = {}

    def add_row(nm, vc, prefer=False):
        if not nm or not is_real_vendor(vc, nm):
            return
        vc = vc.strip()
        cur = by_nm.get(nm)
        if cur is None or (prefer and not cur.get("prefer")):
            by_nm[nm] = {"nm_id": nm, "vendor_code": vc, "prefer": prefer}

    try:
        params = {"select": "nm_id,article", "nm_id": "not.is.null", "limit": "80"}
        if q:
            params["article"] = f"ilike.{q}*"
            params["order"] = "article.asc"
        else:
            params["order"] = "article.asc"
            params["limit"] = "50"
        r = httpx.get(f"{SUPABASE_URL}/rest/v1/ratings_official", params=params,
                      headers=sb_headers(), timeout=10)
        if r.is_success:
            for row in r.json() or []:
                add_row(row.get("nm_id"), row.get("article"), prefer=True)
    except Exception as e:
        logger.warning(f"search-own ratings fallback: {e}")

    try:
        params = {"select": "nm_id,article", "nm_id": "not.is.null", "article": "not.is.null", "limit": "120"}
        if q:
            params["article"] = f"ilike.{q}*"
        r = httpx.get(f"{SUPABASE_URL}/rest/v1/feedbacks", params=params,
                      headers=sb_headers(), timeout=10)
        if r.is_success:
            for row in r.json() or []:
                add_row(row.get("nm_id"), row.get("article"), prefer=True)
    except Exception as e:
        logger.warning(f"search-own feedbacks fallback: {e}")

    try:
        params = {"select": "nm_id,vendor_code", "limit": "80"}
        if q:
            params["vendor_code"] = f"ilike.{q}*"
            params["order"] = "vendor_code.asc"
        else:
            params["order"] = "vendor_code.asc"
            params["limit"] = "50"
        r = httpx.get(f"{SUPABASE_URL}/rest/v1/stock_totals", params=params,
                      headers=sb_headers(), timeout=10)
        if r.is_success:
            for row in r.json() or []:
                add_row(row.get("nm_id"), row.get("vendor_code"), prefer=False)
    except Exception as e:
        logger.warning(f"search-own stock fallback: {e}")

    out = [{"nm_id": v["nm_id"], "vendor_code": v["vendor_code"]} for v in by_nm.values()]
    ql = q.lower()
    if ql:
        out = [a for a in out if a["vendor_code"].lower().startswith(ql) or ql in a["vendor_code"].lower()]
        out.sort(key=lambda a: (
            0 if a["vendor_code"].lower().startswith(ql) else 1,
            a["vendor_code"].lower(),
        ))
    else:
        out.sort(key=lambda a: a["vendor_code"].lower())
    return out[:50]

def fetch_own_stats_v3(nm_ids: list[int], date_from: str, date_to: str) -> dict:
    """Метрики своих артикулов за период из WB Sales Funnel v3.
    Возвращает словарь {str(nm_id): {...метрики под таблицу сравнения...}}.
    Одна карточка = один запрос ко всем nm_ids сразу (щадим лимиты WB: 3 req/min)."""
    if not WB_TOKEN or not nm_ids:
        return {}
    try:
        body = {
            "selectedPeriod": {"start": date_from, "end": date_to},
            "nmIds": nm_ids,
            "brandNames": [], "subjectIds": [], "tagIds": [],
            "orderBy": {"field": "orderSum", "mode": "desc"},
            "limit": max(len(nm_ids), 20), "offset": 0,
        }
        resp = httpx.post(
            f"{WB_ANALYTICS_URL}/api/analytics/v3/sales-funnel/products",
            headers=wb_headers(), json=body, timeout=40
        )
        if not resp.is_success:
            logger.error(f"own-stats v3 error {resp.status_code} {resp.text[:250]}")
            return {}
        products = resp.json().get("data", {}).get("products", []) or []
    except Exception as e:
        logger.error(f"own-stats v3 exception: {e}")
        return {}

    # Кол-во отзывов нет в воронке — добираем из ratings_official одним запросом
    reviews = {}
    try:
        ids_csv = ",".join(str(i) for i in nm_ids)
        rq = httpx.get(
            f"{SUPABASE_URL}/rest/v1/ratings_official?nm_id=in.({ids_csv})&select=nm_id,wb_rating,reviews_total",
            headers=sb_headers(), timeout=10
        )
        if rq.is_success:
            for r in rq.json():
                reviews[r["nm_id"]] = r
    except Exception:
        pass

    out = {}
    for p in products:
        prod = p.get("product", {}) or {}
        sel = (p.get("statistic", {}) or {}).get("selected", {}) or {}
        conv = sel.get("conversions", {}) or {}
        nm = prod.get("nmId")
        if nm is None:
            continue
        rev = reviews.get(nm, {})
        fb = prod.get("feedbackRating")
        out[str(nm)] = {
            "nm_id": nm,
            "vendor_code": prod.get("vendorCode") or str(nm),
            "brand": prod.get("brandName") or "",
            "name": prod.get("title") or "",
            "is_own": True,
            # Карточка
            "feedback_rating": fb if fb else rev.get("wb_rating"),
            "card_rating": prod.get("productRating"),
            "reviews_count": rev.get("reviews_total"),
            "price": sel.get("avgPrice"),
            # median_price / avg_position / ctr в воронке WB отсутствуют → остаются "—"
            # Воронка
            "card_opens": sel.get("openCount"),
            "cart_adds": sel.get("cartCount"),
            "orders": sel.get("orderCount"),
            "orders_sum": sel.get("orderSum"),
            "buyouts": sel.get("buyoutCount"),
            "cancels": sel.get("cancelCount"),
            "cart_conv": conv.get("addToCartPercent"),
            "order_conv": conv.get("cartToOrderPercent"),
            "buyout_pct": conv.get("buyoutPercent"),
        }
    return out

@app.get("/api/own-articles-period-stats")
def own_articles_period_stats(nm_ids: str, date_from: str, date_to: str):
    """Метрики своих артикулов за период (WB Sales Funnel v3). nm_ids — через запятую.
    Ответ: {str(nm_id): {...}} — под колонки таблицы сравнения."""
    ids = []
    for x in nm_ids.split(","):
        x = x.strip()
        if x.isdigit():
            ids.append(int(x))
    return fetch_own_stats_v3(ids, date_from, date_to)

@app.get("/api/own-article-period-stats")
def own_article_period_stats(nm_id: int, date_from: str, date_to: str):
    """Статистика одного своего артикула за период (WB Sales Funnel v3)."""
    stats = fetch_own_stats_v3([nm_id], date_from, date_to)
    return stats.get(str(nm_id))

@app.delete("/api/competitor-session/{session_id}")
def delete_competitor_session(session_id: int):
    """Удаляет сессию и все связанные метрики (каскадно через ON DELETE CASCADE)."""
    try:
        resp = httpx.delete(
            f"{SUPABASE_URL}/rest/v1/competitor_sessions?id=eq.{session_id}",
            headers={**sb_headers(), "Prefer": "return=minimal"}, timeout=15
        )
        return {"status": "ok"} if resp.is_success else {"error": resp.text[:200]}
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/competitor-sessions")
def get_competitor_sessions():
    """Список загруженных сессий сравнения (+ уникальные бренды из метрик)."""
    try:
        resp = httpx.get(
            f"{SUPABASE_URL}/rest/v1/competitor_sessions?select=*&order=uploaded_at.desc",
            headers=sb_headers(), timeout=15
        )
        if not resp.is_success:
            return []
        sessions = resp.json() or []
        if not sessions:
            return []

        # Бренды из competitor_metrics — чтобы подпись файла была «14–20 июля · Brand»
        brands_by_sid = {}
        try:
            ids = ",".join(str(s["id"]) for s in sessions if s.get("id") is not None)
            if ids:
                br = httpx.get(
                    f"{SUPABASE_URL}/rest/v1/competitor_metrics?session_id=in.({ids})&select=session_id,brand",
                    headers=sb_headers(), timeout=15
                )
                if br.is_success:
                    for row in br.json() or []:
                        sid = row.get("session_id")
                        brand = (row.get("brand") or "").strip()
                        if sid is None or not brand:
                            continue
                        brands_by_sid.setdefault(sid, [])
                        if brand not in brands_by_sid[sid]:
                            brands_by_sid[sid].append(brand)
        except Exception as e:
            logger.warning(f"competitor-sessions brands enrich: {e}")

        for s in sessions:
            s["brands"] = brands_by_sid.get(s.get("id"), [])
        return sessions
    except Exception:
        return []

@app.get("/api/competitor-data/{session_id}")
def get_competitor_data(session_id: int):
    """Метрики и поисковые запросы по сессии (+ живая цена покупателя / СПП с витрины WB)."""
    try:
        metrics = httpx.get(
            f"{SUPABASE_URL}/rest/v1/competitor_metrics?session_id=eq.{session_id}&select=*",
            headers=sb_headers(), timeout=15
        )
        queries = httpx.get(
            f"{SUPABASE_URL}/rest/v1/competitor_search_queries?session_id=eq.{session_id}&select=*&order=query_count.desc",
            headers=sb_headers(), timeout=15
        )
        rows = metrics.json() if metrics.is_success else []
        if isinstance(rows, list) and rows:
            rows = _enrich_competitor_metrics_prices(rows)
        return {
            "metrics": rows,
            "search_queries": queries.json() if queries.is_success else []
        }
    except Exception as e:
        return {"error": str(e)}


def _enrich_competitor_metrics_prices(metrics: list) -> list:
    """Добавляет client_price / sale_price / spp: живая витрина WB, иначе из отчёта (price + median_price)."""
    if not metrics:
        return metrics
    nms = []
    for m in metrics:
        try:
            nms.append(int(m.get("nm_id")))
        except (TypeError, ValueError):
            continue
    live = {}
    try:
        live, _src = fetch_client_prices(nms)
    except Exception as e:
        logger.warning(f"enrich competitor prices: {e}")
        live = {}

    out = []
    for m in metrics:
        item = dict(m)
        try:
            nm = int(item.get("nm_id"))
        except (TypeError, ValueError):
            out.append(item)
            continue
        info = live.get(nm) or {}
        live_client = _num_or_none(info.get("client_price"))
        live_basic = _num_or_none(info.get("client_basic"))
        report_sale = _num_or_none(item.get("price"))
        report_buyer = _num_or_none(item.get("median_price"))

        client_price = live_client if live_client is not None else report_buyer
        sale_price = live_basic if live_basic is not None else report_sale
        spp_live = _calc_spp(live_basic, live_client)
        spp_report = _calc_spp(report_sale, report_buyer)
        spp = spp_live if spp_live is not None else spp_report

        item["client_price"] = client_price
        item["sale_price"] = sale_price
        item["spp"] = spp
        item["spp_live"] = spp_live
        item["spp_report"] = spp_report
        item["spp_source"] = (
            "live" if spp_live is not None else ("report" if spp_report is not None else None)
        )
        out.append(item)
    return out

@app.get("/api/my-article-stats")
def my_article_stats(begin: str, end: str, nm_ids: str = ""):
    """Тянет метрики своих артикулов с WB Analytics API за указанный период.
    nm_ids — через запятую, пусто = все артикулы продавца."""
    if not WB_TOKEN:
        return {"error": "WB_TOKEN не задан"}
    try:
        body = {
            "period": {"begin": begin, "end": end},
            "brandNames": [], "objectIDs": [], "tagIDs": [],
            "nmIDs": [int(x) for x in nm_ids.split(",") if x.strip()] if nm_ids else [],
            "timezone": "Europe/Moscow",
            "page": 1
        }
        resp = httpx.post(
            f"{WB_ANALYTICS_URL}/api/analytics/v2/nm-report/detail",
            headers=wb_headers(), json=body, timeout=30
        )
        logger.info(f"my-article-stats: {resp.status_code} snippet={resp.text[:300]}")
        if not resp.is_success:
            return {"error": f"WB API {resp.status_code}: {resp.text[:200]}"}
        data = resp.json()
        cards = data.get("data", {}).get("cards") or []

        # Строим маппинг nm_id → vendor_code
        st = httpx.get(f"{SUPABASE_URL}/rest/v1/stock_totals?select=nm_id,vendor_code", headers=sb_headers(), timeout=15)
        nm_to_vc = {r["nm_id"]: r["vendor_code"] for r in st.json()} if st.is_success else {}

        result = []
        for c in cards:
            nm_id = c.get("nmID")
            stats = c.get("statistics", {}).get("selectedPeriod", {})
            result.append({
                "nm_id": nm_id,
                "vendor_code": nm_to_vc.get(nm_id) or str(nm_id),
                "brand": c.get("brandName", ""),
                "name": c.get("objectName", ""),
                "views": stats.get("openCardCount", 0),
                "card_opens": stats.get("openCardCount", 0),
                "cart_adds": stats.get("addToCartCount", 0),
                "orders": stats.get("ordersCount", 0),
                "orders_sum": stats.get("ordersSumRub", 0),
                "buyouts": stats.get("buyoutsCount", 0),
                "buyout_pct": stats.get("buyoutPercent", 0),
                "cancels": stats.get("cancelCount", 0),
                "ctr": round(stats.get("addToCartCount", 0) / stats.get("openCardCount", 1) * 100, 1) if stats.get("openCardCount") else 0,
                "cart_conv": round(stats.get("ordersCount", 0) / stats.get("addToCartCount", 1) * 100, 1) if stats.get("addToCartCount") else 0,
                "order_conv": round(stats.get("buyoutsCount", 0) / stats.get("ordersCount", 1) * 100, 1) if stats.get("ordersCount") else 0,
                "is_own": True,
            })
        return result
    except Exception as e:
        logger.error(f"my-article-stats error: {e}")
        return {"error": str(e)}

def sync_article_daily_stats(days: int = 30):
    """Тянет дневную статистику по своим артикулам с WB Analytics API.
    Без Jam — максимум 7 дней. С Jam — до 365 дней.
    Поля: openCardCount, addToCartCount, ordersCount, buyoutsCount и т.д."""
    if not WB_TOKEN:
        return
    logger.info(f"sync_article_daily_stats: fetching last {days} days...")

    end_dt = datetime.now(timezone.utc).date()
    begin_dt = end_dt - timedelta(days=days)

    # Используем nm-report/detail/history — статистика по дням для nmId
    try:
        resp = httpx.post(
            f"{WB_ANALYTICS_URL}/api/v2/nm-report/detail/history",
            headers=wb_headers(),
            json={
                "nmIDs": [],
                "period": {
                    "begin": begin_dt.isoformat(),
                    "end": end_dt.isoformat()
                },
                "aggregationLevel": "day"
            },
            timeout=60
        )
        if not resp.is_success:
            logger.error(f"sync_daily: WB error {resp.status_code} {resp.text[:300]}")
            return
        data = resp.json()
        logger.info(f"sync_daily: response snippet: {str(data)[:400]}")
        cards = data.get("data", []) or data if isinstance(data, list) else []
        if not cards:
            logger.info("sync_daily: no data returned")
            return
    except Exception as e:
        logger.error(f"sync_daily: exception {e}")
        return

    # Строим nm_id→vendor_code из stock_totals
    try:
        st = httpx.get(f"{SUPABASE_URL}/rest/v1/stock_totals?select=nm_id,vendor_code", headers=sb_headers(), timeout=15)
        nm_to_vc = {r["nm_id"]: r["vendor_code"] for r in st.json()} if st.is_success else {}
    except Exception:
        nm_to_vc = {}

    now = datetime.now(timezone.utc).isoformat()
    rows = []
    for card in cards:
        nm_id = card.get("nmID") or card.get("nmId")
        history = card.get("history") or card.get("days") or []
        for day in history:
            dt = day.get("dt") or day.get("date")
            if not dt or not nm_id:
                continue
            dt_str = str(dt)[:10]
            oc = day.get("openCardCount", 0) or 0
            atc = day.get("addToCartCount", 0) or 0
            ord_ = day.get("ordersCount", 0) or 0
            ord_sum = day.get("ordersSumRub", 0) or 0
            buy = day.get("buyoutsCount", 0) or 0
            buy_sum = day.get("buyoutsSumRub", 0) or 0
            can = day.get("cancelCount", 0) or 0

            rows.append({
                "nm_id": nm_id,
                "vendor_code": nm_to_vc.get(nm_id) or str(nm_id),
                "dt": dt_str,
                "open_card": int(oc),
                "add_to_cart": int(atc),
                "orders": int(ord_),
                "orders_sum": float(ord_sum),
                "buyouts": int(buy),
                "buyouts_sum": float(buy_sum),
                "cancels": int(can),
                "ctr": round(atc / oc * 100, 2) if oc else 0,
                "cart_conv": round(ord_ / atc * 100, 2) if atc else 0,
                "order_conv": round(buy / ord_ * 100, 2) if ord_ else 0,
                "buyout_pct": round(buy / (buy + can) * 100, 2) if (buy + can) else 0,
                "updated_at": now
            })

    logger.info(f"sync_daily: {len(rows)} day-rows to upsert")
    if not rows:
        return

    # Upsert по (nm_id, dt) — обновляем если уже есть
    headers_up = {**sb_headers(), "Prefer": "resolution=merge-duplicates"}
    for i in range(0, len(rows), 500):
        r = httpx.post(
            f"{SUPABASE_URL}/rest/v1/article_daily_stats?on_conflict=nm_id,dt",
            json=rows[i:i+500], headers=headers_up, timeout=30
        )
        if not r.is_success:
            logger.error(f"sync_daily insert error: {r.status_code} {r.text[:200]}")

@app.get("/api/article-daily-stats")
def article_daily_stats(days: int = 30):
    """Дневная статистика по своим артикулам за последние N дней."""
    try:
        dt_from = (datetime.now(timezone.utc).date() - timedelta(days=days)).isoformat()
        resp = httpx.get(
            f"{SUPABASE_URL}/rest/v1/article_daily_stats?dt=gte.{dt_from}&select=*&order=dt.asc",
            headers=sb_headers(), timeout=20
        )
        return resp.json() if resp.is_success else []
    except Exception as e:
        return {"error": str(e)}

@app.post("/api/sync-daily-stats")
def trigger_daily_sync(days: int = 30):
    import threading
    threading.Thread(target=sync_article_daily_stats, args=(days,), daemon=True).start()
    return {"status": "started", "days": days}

# ---------- Рост продаж: темп к прошлому периоду (день/неделя/2 недели/месяц) ----------
# Заказы = orderCount воронки WB («Заказали товаров, шт» в кабинете).
# Вчера «до часа» — из почасового снимка той же воронки. Statistics API — запасной источник.
SALES_PACE_CACHE = {
    "by_period": {},  # period -> payload
    "syncing": False,
    "syncing_period": None,
    "error": None,
}
SALES_PACE_SNAPS_KEY = "sales_pace_funnel_snaps"
SALES_PACE_HIDDEN_KEY = "sales_pace_hidden"
SALES_PACE_PERIODS = ("day", "week", "weeks2", "month")


def _sales_pace_hidden() -> list:
    raw = get_setting_json(SALES_PACE_HIDDEN_KEY, [])
    return _uniq_str_list(raw if isinstance(raw, list) else [])

def _msk_now():
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("Europe/Moscow")).replace(tzinfo=None)
    except Exception:
        return datetime.utcnow() + timedelta(hours=3)

def _parse_ymd(s):
    if not s:
        return None
    try:
        return datetime.strptime(str(s)[:10], "%Y-%m-%d")
    except Exception:
        return None

def _pace_cache_key(period: str, date_cur=None, date_prev=None) -> str:
    if period == "day" and date_cur:
        prev = date_prev or "auto"
        return f"day:{str(date_cur)[:10]}:{str(prev)[:10]}"
    return period

def _pace_windows(period: str, now: datetime, date_cur=None, date_prev=None) -> dict:
    """Окна текущего и прошлого периода.
    Для day + date_cur/date_prev — выбранные календарные дни (полные сутки;
    если выбран сегодняшний — до текущего времени)."""
    today0 = now.replace(hour=0, minute=0, second=0, microsecond=0)
    time_cut = now.strftime("%H:%M")

    def _range_note(equal_cut: bool) -> str:
        if equal_cut:
            return (
                f"Одинаковое окно часов (оба до {time_cut} МСК) — "
                "сравниваем «срез к этому моменту», а не полные сутки"
            )
        return "Полные календарные периоды (или выбранные дни целиком)"

    if period == "day" and date_cur:
        cur_day = _parse_ymd(date_cur)
        if cur_day:
            prev_day = _parse_ymd(date_prev) if date_prev else (cur_day - timedelta(days=1))
            if not prev_day or prev_day.date() >= cur_day.date():
                prev_day = cur_day - timedelta(days=1)
            cur_start = cur_day.replace(hour=0, minute=0, second=0, microsecond=0)
            prev_start = prev_day.replace(hour=0, minute=0, second=0, microsecond=0)
            cur_is_today = cur_start.date() == today0.date()
            prev_is_today = prev_start.date() == today0.date()
            if cur_is_today:
                cur_end = now
                label_cur = f"{cur_start.strftime('%d.%m.%Y')} (сегодня) до {time_cut}"
                col_cur = f"{cur_start.strftime('%d.%m')} до {time_cut}"
            else:
                cur_end = cur_start.replace(hour=23, minute=59, second=59)
                label_cur = f"{cur_start.strftime('%d.%m.%Y')} (полные сутки)"
                col_cur = cur_start.strftime("%d.%m")
            if prev_is_today:
                prev_end = now
                label_prev = f"{prev_start.strftime('%d.%m.%Y')} (сегодня) до {time_cut}"
                col_prev = f"{prev_start.strftime('%d.%m')} до {time_cut}"
            else:
                prev_end = prev_start.replace(hour=23, minute=59, second=59)
                label_prev = f"{prev_start.strftime('%d.%m.%Y')} (полные сутки)"
                col_prev = prev_start.strftime("%d.%m")
            equal_cut = cur_is_today  # выбран «сегодня» vs предыдущий день
            return {
                "cur_start": cur_start, "cur_end": cur_end,
                "prev_start": prev_start, "prev_end": prev_end,
                "label_cur": label_cur,
                "label_prev": label_prev,
                "col_cur": col_cur,
                "col_prev": col_prev,
                "use_snaps": False,
                "custom_dates": True,
                "period_name": "День (выбранный)",
                "mode_hint": _range_note(equal_cut),
                "time_cutoff": time_cut if equal_cut else None,
            }
    if period == "day":
        cur_start = today0
        prev_start = today0 - timedelta(days=1)
        prev_end = prev_start + (now - cur_start)
        return {
            "cur_start": cur_start, "cur_end": now,
            "prev_start": prev_start, "prev_end": prev_end,
            "label_cur": f"сегодня {cur_start.strftime('%d.%m.%Y')} до {time_cut}",
            "label_prev": f"вчера {prev_start.strftime('%d.%m.%Y')} до {time_cut}",
            "col_cur": f"Сегодня {cur_start.strftime('%d.%m')}",
            "col_prev": f"Вчера {prev_start.strftime('%d.%m')}",
            "use_snaps": True,
            "custom_dates": False,
            "period_name": "День",
            "mode_hint": _range_note(True),
            "time_cutoff": time_cut,
        }
    if period == "week":
        # понедельник текущей недели
        cur_start = today0 - timedelta(days=today0.weekday())
        prev_start = cur_start - timedelta(days=7)
        prev_end = prev_start + (now - cur_start)
        return {
            "cur_start": cur_start, "cur_end": now,
            "prev_start": prev_start, "prev_end": prev_end,
            "label_cur": f"эта неделя: {cur_start.strftime('%d.%m')} → сейчас ({now.strftime('%d.%m %H:%M')})",
            "label_prev": f"прошлая неделя: {prev_start.strftime('%d.%m')} → {prev_end.strftime('%d.%m %H:%M')}",
            "col_cur": f"{cur_start.strftime('%d.%m')}–{now.strftime('%d.%m')}",
            "col_prev": f"{prev_start.strftime('%d.%m')}–{prev_end.strftime('%d.%m')}",
            "use_snaps": False,
            "custom_dates": False,
            "period_name": "Неделя (пн → сейчас)",
            "mode_hint": _range_note(True),
            "time_cutoff": time_cut,
        }
    if period == "weeks2":
        cur_start = now - timedelta(days=14)
        prev_start = now - timedelta(days=28)
        prev_end = now - timedelta(days=14)
        return {
            "cur_start": cur_start, "cur_end": now,
            "prev_start": prev_start, "prev_end": prev_end,
            "label_cur": f"последние 14 дн.: {cur_start.strftime('%d.%m.%Y')} → {now.strftime('%d.%m.%Y')}",
            "label_prev": f"предыдущие 14 дн.: {prev_start.strftime('%d.%m.%Y')} → {prev_end.strftime('%d.%m.%Y')}",
            "col_cur": f"{cur_start.strftime('%d.%m')}–{now.strftime('%d.%m')}",
            "col_prev": f"{prev_start.strftime('%d.%m')}–{prev_end.strftime('%d.%m')}",
            "use_snaps": False,
            "custom_dates": False,
            "period_name": "2 недели",
            "mode_hint": "Два окна по 14 дней подряд (без выравнивания по часам)",
            "time_cutoff": None,
        }
    # month — с 1-го числа до сейчас vs прошлый месяц до того же дня/времени
    cur_start = today0.replace(day=1)
    if cur_start.month == 1:
        prev_month_start = cur_start.replace(year=cur_start.year - 1, month=12)
    else:
        prev_month_start = cur_start.replace(month=cur_start.month - 1)
    try:
        prev_end = prev_month_start.replace(day=now.day, hour=now.hour, minute=now.minute, second=now.second)
    except ValueError:
        # 31-е → последний день прошлого месяца
        if prev_month_start.month == 12:
            nxt = prev_month_start.replace(year=prev_month_start.year + 1, month=1, day=1)
        else:
            nxt = prev_month_start.replace(month=prev_month_start.month + 1, day=1)
        prev_end = nxt - timedelta(seconds=1)
        prev_end = prev_end.replace(hour=now.hour, minute=now.minute, second=now.second, microsecond=0)
    return {
        "cur_start": cur_start, "cur_end": now,
        "prev_start": prev_month_start, "prev_end": prev_end,
        "label_cur": f"этот месяц: {cur_start.strftime('%d.%m.%Y')} → {now.strftime('%d.%m.%Y %H:%M')}",
        "label_prev": f"прошлый месяц: {prev_month_start.strftime('%d.%m.%Y')} → {prev_end.strftime('%d.%m.%Y %H:%M')}",
        "col_cur": f"{cur_start.strftime('%d.%m')}–{now.strftime('%d.%m')}",
        "col_prev": f"{prev_month_start.strftime('%d.%m')}–{prev_end.strftime('%d.%m')}",
        "use_snaps": False,
        "custom_dates": False,
        "period_name": "Месяц (1-е → сейчас)",
        "mode_hint": _range_note(True),
        "time_cutoff": time_cut,
    }

def _funnel_products_range(start_str: str, end_str: str, nm_ids: list = None) -> dict:
    """Воронка за период → {nm_id: {opens, cart, orders, vendor_code, name}}."""
    if not WB_TOKEN:
        return {}
    out = {}
    offset = 0
    limit = 1000
    for _ in range(20):
        body = {
            "selectedPeriod": {"start": start_str, "end": end_str},
            "nmIds": nm_ids or [],
            "brandNames": [], "subjectIds": [], "tagIds": [],
            "orderBy": {"field": "orderCount", "mode": "desc"},
            "limit": limit, "offset": offset,
        }
        try:
            resp = httpx.post(
                f"{WB_ANALYTICS_URL}/api/analytics/v3/sales-funnel/products",
                headers=wb_headers(), json=body, timeout=40
            )
            if not resp.is_success:
                logger.error(f"sales-pace funnel error {resp.status_code} {resp.text[:200]}")
                break
            products = resp.json().get("data", {}).get("products", []) or []
        except Exception as e:
            logger.error(f"sales-pace funnel exception: {e}")
            break
        if not products:
            break
        for p in products:
            prod = p.get("product", {}) or {}
            sel = (p.get("statistic", {}) or {}).get("selected", {}) or {}
            nm = prod.get("nmId")
            if nm is None:
                continue
            out[int(nm)] = {
                "nm_id": int(nm),
                "vendor_code": prod.get("vendorCode") or str(nm),
                "name": prod.get("title") or "",
                "opens": int(sel.get("openCount") or 0),
                "cart": int(sel.get("cartCount") or 0),
                "orders": int(sel.get("orderCount") or 0),
            }
        if len(products) < limit:
            break
        offset += limit
        time.sleep(0.7)
    return out

def _funnel_products_day(day_str: str, nm_ids: list = None) -> dict:
    return _funnel_products_range(day_str, day_str, nm_ids)


def _pace_funnel_orders(funnel_map: dict, nm, fallback: int) -> int:
    """orderCount воронки, если артикул в ответе; иначе запасной счётчик Statistics."""
    row = (funnel_map or {}).get(nm)
    if isinstance(row, dict) and "orders" in row:
        return int(row.get("orders") or 0)
    return int(fallback or 0)


def _pace_cache_stale(cached: dict, max_age_sec: int = 600) -> bool:
    """True, если кэш темпа старше max_age_sec (updated_at в UTC)."""
    raw = (cached or {}).get("updated_at")
    if not raw:
        return True
    try:
        dt = datetime.strptime(str(raw), "%d.%m.%Y %H:%M").replace(tzinfo=timezone.utc)
    except Exception:
        return True
    return (datetime.now(timezone.utc) - dt).total_seconds() > max_age_sec

def sync_sales_pace(period: str = "day", date_cur: str = None, date_prev: str = None):
    """Считает темп продаж за выбранный период.
    Для day можно передать date_cur / date_prev (YYYY-MM-DD) — сравнение двух дней."""
    period = period if period in SALES_PACE_PERIODS else "day"
    if period != "day":
        date_cur = date_prev = None
    cache_key = _pace_cache_key(period, date_cur, date_prev)
    if not WB_TOKEN:
        SALES_PACE_CACHE["error"] = "WB_TOKEN не задан"
        return
    if SALES_PACE_CACHE.get("syncing"):
        return
    SALES_PACE_CACHE["syncing"] = True
    SALES_PACE_CACHE["syncing_period"] = cache_key
    SALES_PACE_CACHE["error"] = None
    try:
        now = _msk_now()
        win = _pace_windows(period, now, date_cur, date_prev)
        cur_start, cur_end = win["cur_start"], win["cur_end"]
        prev_start, prev_end = win["prev_start"], win["prev_end"]
        cur_s = cur_start.strftime("%Y-%m-%d")
        cur_e = cur_end.strftime("%Y-%m-%d")
        prev_s = prev_start.strftime("%Y-%m-%d")
        prev_e = prev_end.strftime("%Y-%m-%d")

        # ── Заказы ──
        date_from = prev_start.strftime("%Y-%m-%dT00:00:00")
        orders = fetch_supplier_feed("/api/v1/supplier/orders", date_from, max_pages=5)
        cur_ord, prev_ord, vc_from_orders = {}, {}, {}
        for o in orders:
            nm = o.get("nmId")
            if not nm:
                continue
            d = parse_wb_dt(o.get("date", ""))
            if d is None:
                continue
            if o.get("supplierArticle"):
                vc_from_orders[nm] = o["supplierArticle"]
            qty = int(o.get("quantity") or 1)
            if qty <= 0:
                qty = 1
            if cur_start <= d <= cur_end:
                cur_ord[nm] = cur_ord.get(nm, 0) + qty
            elif prev_start <= d <= prev_end:
                prev_ord[nm] = prev_ord.get(nm, 0) + qty

        funnel_cur, funnel_prev = {}, {}
        compare_as_of = None
        funnel_ready = True
        ads_ready = True
        ads_cur, ads_prev = {}, {}

        # Реклама: fullstats только по дням — для «день» кладём показы в почасовые снимки
        try:
            ads_cur_api, ads_prev_api = fetch_ad_nm_windows(prev_start, prev_end, cur_start, cur_end)
            logger.info(f"sales-pace ads: cur={len(ads_cur_api)} nms, prev_api={len(ads_prev_api)} nms")
        except Exception as e:
            logger.error(f"sales-pace ads error: {e}")
            ads_cur_api, ads_prev_api = {}, {}

        if win.get("use_snaps"):
            # день: снимок воронки + показов (сегодня накопленно; вчера — из снимка на тот же час)
            funnel_cur = _funnel_products_day(cur_s)
            ads_cur = ads_cur_api or {}
            hour_key = now.strftime("%Y-%m-%dT%H")
            snaps = get_setting_json(SALES_PACE_SNAPS_KEY, []) or []
            if not isinstance(snaps, list):
                snaps = []
            products_snap = {}
            for nm, v in (funnel_cur or {}).items():
                products_snap[str(nm)] = {
                    "opens": int(v.get("opens") or 0),
                    "cart": int(v.get("cart") or 0),
                    "orders": int(v.get("orders") or 0),
                    "views": 0,
                    "spend": 0.0,
                }
            for nm, v in (ads_cur or {}).items():
                key = str(nm)
                if key not in products_snap:
                    products_snap[key] = {"opens": 0, "cart": 0, "orders": 0, "views": 0, "spend": 0.0}
                products_snap[key]["views"] = int(v.get("views") or 0)
                products_snap[key]["spend"] = float(v.get("spend") or 0)
            snap_payload = {
                "hour_key": hour_key,
                "as_of": now.strftime("%Y-%m-%d %H:%M"),
                "day": cur_s,
                "products": products_snap,
            }
            snaps = [s for s in snaps if s.get("hour_key") != hour_key]
            snaps.append(snap_payload)
            cutoff_day = (cur_start - timedelta(days=3)).strftime("%Y-%m-%d")
            snaps = [s for s in snaps if (s.get("day") or "") >= cutoff_day]
            snaps.sort(key=lambda s: s.get("hour_key") or "")
            save_setting_value(SALES_PACE_SNAPS_KEY, snaps)

            yest_str = prev_s
            target_yest_hour = prev_end.strftime("%Y-%m-%dT%H")
            yest_snap = None
            for s in snaps:
                if s.get("day") == yest_str and (s.get("hour_key") or "") <= target_yest_hour:
                    yest_snap = s
            if yest_snap is None:
                yest_candidates = [s for s in snaps if s.get("day") == yest_str]
                if yest_candidates:
                    yest_snap = min(
                        yest_candidates,
                        key=lambda s: abs(
                            (datetime.strptime(s["hour_key"], "%Y-%m-%dT%H") - prev_end.replace(minute=0, second=0, microsecond=0)).total_seconds()
                        ) if s.get("hour_key") else 10**9
                    )
            funnel_prev_raw = (yest_snap or {}).get("products") or {}
            funnel_prev = {}
            ads_prev = {}
            ads_from_snap = False
            for k, v in funnel_prev_raw.items():
                try:
                    nm_i = int(k)
                except Exception:
                    continue
                funnel_prev[nm_i] = v
                if isinstance(v, dict) and ("views" in v or "spend" in v):
                    ads_from_snap = True
                    ads_prev[nm_i] = {
                        "views": int(v.get("views") or 0),
                        "spend": float(v.get("spend") or 0),
                        "clicks": int(v.get("clicks") or 0),
                        "orders": int(v.get("orders") or 0),
                    }
            compare_as_of = (yest_snap or {}).get("as_of")
            funnel_ready = bool(yest_snap)
            # показы «вчера до этого часа» только из снимка (API отдаёт весь день)
            ads_ready = bool(yest_snap) and ads_from_snap
            if not ads_ready:
                ads_prev = {}
        else:
            # неделя / 2 недели / месяц — два запроса воронки по диапазонам дат
            funnel_cur = _funnel_products_range(cur_s, cur_e)
            time.sleep(0.7)
            funnel_prev = _funnel_products_range(prev_s, prev_e)
            compare_as_of = f"{prev_s}–{prev_e}"
            funnel_ready = True
            ads_cur = ads_cur_api or {}
            ads_prev = ads_prev_api or {}
            ads_ready = True

        try:
            st = httpx.get(
                f"{SUPABASE_URL}/rest/v1/stock_totals?select=nm_id,vendor_code,quantity_warehouses_full,in_way_to_client,in_way_from_client",
                headers=sb_headers(), timeout=15
            )
            stock_rows = st.json() if st.is_success else []
            nm_to_vendor = {r["nm_id"]: r["vendor_code"] for r in (stock_rows or []) if r.get("nm_id") is not None}
            stock_by_nm = {
                int(r["nm_id"]): {
                    "stock": int(r.get("quantity_warehouses_full") or 0),
                    "in_way": int(r.get("in_way_to_client") or 0) + int(r.get("in_way_from_client") or 0),
                }
                for r in (stock_rows or [])
                if r.get("nm_id") is not None
            }
        except Exception:
            nm_to_vendor = {}
            stock_by_nm = {}

        # Текущая география складов + вчерашний снимок (для «обнулился склад»)
        stock_wh_by_nm = _fetch_stock_wh_by_nm()
        disabled_wh = get_disabled_warehouses()
        try:
            # в снимках храним все склады; отключённые фильтруем при сравнении
            save_stock_warehouse_snapshot_by_nm(stock_wh_by_nm)
        except Exception as e:
            logger.warning(f"sales-pace stock WH snapshot: {e}")
        stock_wh_prev_snap = get_stock_warehouse_snap_for_day(prev_s)

        period_days = {"day": 1, "week": 7, "weeks2": 14, "month": 30}.get(period, 1)

        def _cr_pct(num, den):
            if not den:
                return None
            return round(100.0 * float(num) / float(den), 1)

        # артикулы с заказами в воронке или (запасной) Statistics
        all_nms = set(cur_ord) | set(prev_ord) | set(funnel_cur) | set(funnel_prev)
        articles = []
        for nm in all_nms:
            ft = funnel_cur.get(nm) or {}
            fy = funnel_prev.get(nm) or {}
            # как в кабинете WB: «Заказали товаров, шт» = orderCount воронки
            o_t = _pace_funnel_orders(funnel_cur, nm, cur_ord.get(nm, 0))
            o_y = _pace_funnel_orders(funnel_prev, nm, prev_ord.get(nm, 0)) if funnel_ready else prev_ord.get(nm, 0)
            if o_t <= 0 and o_y <= 0:
                continue
            opens_t = int(ft.get("opens") or 0)
            opens_y = int(fy.get("opens") or 0)
            cart_t = int(ft.get("cart") or 0)
            cart_y = int(fy.get("cart") or 0)
            ad_t = ads_cur.get(nm) or {}
            ad_y = ads_prev.get(nm) or {}
            # полный вчерашний день из API (даже в режиме снимков «до часа»)
            ad_y_full = (ads_prev_api or {}).get(nm) or {}
            views_t = int(ad_t.get("views") or 0)
            views_y = int(ad_y.get("views") or 0) if ads_ready else None
            views_y_full = int(ad_y_full.get("views") or 0)
            spend_t = float(ad_t.get("spend") or 0)
            spend_y = float(ad_y.get("spend") or 0) if ads_ready else None
            spend_y_full = float(ad_y_full.get("spend") or 0)
            cpm_t = _cpm(views_t, spend_t)
            cpm_y = _cpm(views_y, spend_y) if ads_ready else None
            cpm_y_full = _cpm(views_y_full, spend_y_full)
            cpm_delta = None
            if ads_ready and cpm_t is not None and cpm_y is not None:
                cpm_delta = round(cpm_t - cpm_y, 1)
            views_delta = (views_t - views_y) if ads_ready and views_y is not None else None
            cart_cr_t = _cr_pct(cart_t, opens_t)
            cart_cr_y = _cr_pct(cart_y, opens_y)
            cart_cr_delta = (
                round(cart_cr_t - cart_cr_y, 1)
                if cart_cr_t is not None and cart_cr_y is not None else None
            )
            st_info = stock_by_nm.get(int(nm)) or {}
            stock_qty_raw = int(st_info.get("stock") or 0)
            in_way = int(st_info.get("in_way") or 0)
            geo = stock_wh_geo_compare(nm, stock_wh_by_nm, stock_wh_prev_snap, disabled_wh)
            # остаток для темпа = сумма по включённым складам, если есть детализация
            nm_slot = (stock_wh_by_nm or {}).get(str(int(nm))) or {}
            if nm_slot.get("w"):
                stock_qty = int(geo.get("stock_qty_enabled") or 0)
            else:
                stock_qty = stock_qty_raw
            # дней запаса ≈ остаток / среднесут. заказам в окне
            daily_orders = max(o_t, o_y, 0) / float(period_days or 1)
            days_left = round(stock_qty / daily_orders, 1) if daily_orders > 0 else None
            if stock_qty <= 0:
                stock_flag = "oos"
            elif days_left is not None and days_left < 5:
                stock_flag = "low"
            else:
                stock_flag = "ok"
            opens_delta = opens_t - opens_y if funnel_ready else None
            cart_delta = cart_t - cart_y if funnel_ready else None
            funnel_down = (
                (opens_delta is not None and opens_delta < 0)
                or (cart_delta is not None and cart_delta < 0)
                or (cart_cr_delta is not None and cart_cr_delta <= -1)
            )
            orders_down = (o_t - o_y) < 0
            geo_bad = geo.get("stock_geo_flag") in ("oos", "narrow", "emptied")
            stock_linked = bool(
                (funnel_down or orders_down)
                and (stock_flag in ("oos", "low") or geo_bad)
            )
            articles.append({
                "nm_id": nm,
                "vendor_code": ft.get("vendor_code") or fy.get("vendor_code") or nm_to_vendor.get(nm) or vc_from_orders.get(nm) or str(nm),
                "name": ft.get("name") or fy.get("name") or "",
                "orders_today": o_t,
                "orders_yesterday": o_y,
                "orders_delta": o_t - o_y,
                "opens_today": opens_t,
                "opens_yesterday": opens_y,
                "opens_delta": opens_delta,
                "clicks_today": opens_t,
                "clicks_yesterday": opens_y,
                "clicks_delta": opens_delta,
                "cart_today": cart_t,
                "cart_yesterday": cart_y,
                "cart_delta": cart_delta,
                "cart_cr_today": cart_cr_t,
                "cart_cr_yesterday": cart_cr_y,
                "cart_cr_delta": cart_cr_delta,
                "views_today": views_t,
                "views_yesterday": views_y,
                "views_yesterday_full": views_y_full,
                "views_delta": views_delta,
                "spend_today": spend_t,
                "spend_yesterday": spend_y,
                "spend_yesterday_full": spend_y_full,
                "cpm_today": cpm_t,
                "cpm_yesterday": cpm_y,
                "cpm_yesterday_full": cpm_y_full,
                "cpm_delta": cpm_delta,
                "stock": stock_qty,
                "in_way": in_way,
                "days_left": days_left,
                "stock_flag": stock_flag,
                "wh_live": geo.get("wh_live"),
                "wh_live_prev": geo.get("wh_live_prev"),
                "wh_emptied": geo.get("wh_emptied") or [],
                "wh_names": geo.get("wh_names") or [],
                "stock_geo_flag": geo.get("stock_geo_flag") or "ok",
                "funnel_down": funnel_down,
                "stock_linked": stock_linked,
                "funnel_compare_ready": funnel_ready,
                "ads_compare_ready": ads_ready,
            })

        articles.sort(key=lambda a: (a["orders_delta"], a["orders_today"], str(a["vendor_code"])))

        payload = {
            "period": period,
            "cache_key": cache_key,
            "articles": articles,
            "as_of": now.strftime("%d.%m.%Y %H:%M"),
            "compare_as_of": compare_as_of,
            "label_cur": win["label_cur"],
            "label_prev": win["label_prev"],
            "col_cur": win.get("col_cur") or ("Сегодня" if period == "day" else "Текущий"),
            "col_prev": win.get("col_prev") or ("Вчера" if period == "day" else "Прошлый"),
            "custom_dates": bool(win.get("custom_dates")),
            "period_name": win.get("period_name") or period,
            "mode_hint": win.get("mode_hint") or "",
            "time_cutoff": win.get("time_cutoff"),
            "date_cur": cur_s,
            "date_prev": prev_s,
            "today": cur_s,
            "yesterday": prev_s,
            "now_time": now.strftime("%H:%M"),
            "updated_at": datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M"),
            "funnel_ready": funnel_ready,
            "ads_ready": ads_ready,
            "disabled_warehouses": sorted(disabled_wh),
            "error": None,
        }
        SALES_PACE_CACHE.setdefault("by_period", {})[cache_key] = payload
        SALES_PACE_CACHE["syncing"] = False
        SALES_PACE_CACHE["syncing_period"] = None
        logger.info(f"sales-pace[{cache_key}]: {len(articles)} arts, {win['label_cur']} vs {win['label_prev']}")
    except Exception as e:
        logger.error(f"sync_sales_pace({cache_key}) error: {e}")
        SALES_PACE_CACHE["error"] = str(e)
        SALES_PACE_CACHE["syncing"] = False
        SALES_PACE_CACHE["syncing_period"] = None
    finally:
        SALES_PACE_CACHE["syncing"] = False
        SALES_PACE_CACHE["syncing_period"] = None

def _enrich_pace_articles_stock(articles: list, period: str = "day") -> list:
    """Докидывает текущий остаток/флаги/географию складов к кэшу темпа (без полного пересчёта WB)."""
    if not articles:
        return articles
    need_stock = not all(isinstance(a, dict) and a.get("stock") is not None for a in articles)
    need_geo = not all(isinstance(a, dict) and a.get("wh_live") is not None for a in articles)
    if not need_stock and not need_geo:
        return articles
    stock_by_nm = {}
    try:
        st = httpx.get(
            f"{SUPABASE_URL}/rest/v1/stock_totals?select=nm_id,quantity_warehouses_full,in_way_to_client,in_way_from_client",
            headers=sb_headers(), timeout=15,
        )
        if st.is_success:
            for r in st.json() or []:
                if r.get("nm_id") is None:
                    continue
                stock_by_nm[int(r["nm_id"])] = {
                    "stock": int(r.get("quantity_warehouses_full") or 0),
                    "in_way": int(r.get("in_way_to_client") or 0) + int(r.get("in_way_from_client") or 0),
                }
    except Exception as e:
        logger.warning(f"enrich pace stock: {e}")
        if need_stock:
            return articles
    stock_wh_by_nm = _fetch_stock_wh_by_nm() if need_geo else {}
    disabled_wh = get_disabled_warehouses() if (need_geo or need_stock) else set()
    prev_day = (_msk_now() - timedelta(days=1)).strftime("%Y-%m-%d")
    prev_snap = get_stock_warehouse_snap_for_day(prev_day) if need_geo else None
    period_days = {"day": 1, "week": 7, "weeks2": 14, "month": 30}.get(period, 1)
    out = []
    for a in articles:
        item = dict(a)
        try:
            nm = int(item.get("nm_id"))
        except Exception:
            out.append(item)
            continue
        st_info = stock_by_nm.get(nm) or {}
        if need_stock:
            stock_qty = int(st_info.get("stock") or 0)
            in_way = int(st_info.get("in_way") or 0)
        else:
            stock_qty = int(item.get("stock") if item.get("stock") is not None else (st_info.get("stock") or 0))
            in_way = int(item.get("in_way") if item.get("in_way") is not None else (st_info.get("in_way") or 0))
        geo = stock_wh_geo_compare(nm, stock_wh_by_nm, prev_snap, disabled_wh) if need_geo else {
            "wh_live": item.get("wh_live"),
            "wh_live_prev": item.get("wh_live_prev"),
            "wh_emptied": item.get("wh_emptied") or [],
            "wh_names": item.get("wh_names") or [],
            "stock_geo_flag": item.get("stock_geo_flag") or "ok",
            "stock_qty_enabled": None,
        }
        nm_slot = (stock_wh_by_nm or {}).get(str(nm)) or {}
        if need_stock and nm_slot.get("w"):
            stock_qty = int(geo.get("stock_qty_enabled") or 0)
        o_t = int(item.get("orders_today") or 0)
        o_y = int(item.get("orders_yesterday") or 0)
        daily_orders = max(o_t, o_y, 0) / float(period_days or 1)
        days_left = round(stock_qty / daily_orders, 1) if daily_orders > 0 else None
        if stock_qty <= 0:
            stock_flag = "oos"
        elif days_left is not None and days_left < 5:
            stock_flag = "low"
        else:
            stock_flag = "ok"
        opens_d = item.get("opens_delta")
        cart_d = item.get("cart_delta")
        cr_t = item.get("cart_cr_today")
        cr_y = item.get("cart_cr_yesterday")
        if cr_t is None and item.get("opens_today"):
            try:
                cr_t = round(100.0 * float(item.get("cart_today") or 0) / float(item["opens_today"]), 1)
            except Exception:
                cr_t = None
        if cr_y is None and item.get("opens_yesterday"):
            try:
                cr_y = round(100.0 * float(item.get("cart_yesterday") or 0) / float(item["opens_yesterday"]), 1)
            except Exception:
                cr_y = None
        cr_d = item.get("cart_cr_delta")
        if cr_d is None and cr_t is not None and cr_y is not None:
            cr_d = round(cr_t - cr_y, 1)
        funnel_down = (
            (opens_d is not None and opens_d < 0)
            or (cart_d is not None and cart_d < 0)
            or (cr_d is not None and cr_d <= -1)
        )
        orders_down = (item.get("orders_delta") or 0) < 0
        geo_bad = geo.get("stock_geo_flag") in ("oos", "narrow", "emptied")
        item.update({
            "stock": stock_qty,
            "in_way": in_way,
            "days_left": days_left,
            "stock_flag": stock_flag,
            "cart_cr_today": cr_t,
            "cart_cr_yesterday": cr_y,
            "cart_cr_delta": cr_d,
            "wh_live": geo.get("wh_live"),
            "wh_live_prev": geo.get("wh_live_prev"),
            "wh_emptied": geo.get("wh_emptied") or [],
            "wh_names": geo.get("wh_names") or [],
            "stock_geo_flag": geo.get("stock_geo_flag") or "ok",
            "funnel_down": funnel_down,
            "stock_linked": bool(
                (funnel_down or orders_down)
                and (stock_flag in ("oos", "low") or geo_bad)
            ),
        })
        out.append(item)
    return out


def _pace_parse_snap_day(captured_at) -> str:
    if not captured_at:
        return ""
    s = str(captured_at)
    # 2026-08-11T... or 11.08.2026
    if len(s) >= 10 and s[4] == "-":
        return s[:10]
    return ""


def fetch_pace_price_compare(nm_ids: list, date_cur: str = None, date_prev: str = None) -> dict:
    """Сравнение цены покупателя и СПП: сейчас vs день базы (date_prev).

    Источник: price_snapshots (+ живой SPP_CACHE для «сейчас»).
    """
    if not nm_ids or not SUPABASE_URL or not SUPABASE_KEY:
        return {}
    ids = []
    for x in nm_ids:
        try:
            ids.append(int(x))
        except Exception:
            pass
    ids = list(dict.fromkeys(ids))
    if not ids:
        return {}

    live = {}
    for a in (SPP_CACHE.get("articles") or []):
        try:
            nm = int(a.get("nm_id"))
        except Exception:
            continue
        live[nm] = a

    since = (datetime.now(timezone.utc) - timedelta(days=45)).isoformat()
    by_nm = {}
    for i in range(0, len(ids), 80):
        chunk = ids[i:i + 80]
        try:
            resp = httpx.get(
                f"{SUPABASE_URL}/rest/v1/price_snapshots",
                params={
                    "select": "nm_id,sale_price,client_price,spp,captured_at",
                    "nm_id": f"in.({','.join(str(x) for x in chunk)})",
                    "captured_at": f"gte.{since}",
                    "order": "captured_at.asc",
                    "limit": "12000",
                },
                headers={
                    "apikey": SUPABASE_KEY,
                    "Authorization": f"Bearer {SUPABASE_KEY}",
                },
                timeout=30,
            )
            if not resp.is_success:
                if resp.status_code == 404:
                    logger.warning("price_snapshots missing — run supabase/price_snapshots.sql")
                continue
            for row in resp.json() or []:
                try:
                    nm = int(row.get("nm_id"))
                except Exception:
                    continue
                by_nm.setdefault(nm, []).append(row)
        except Exception as e:
            logger.warning(f"fetch_pace_price_compare: {e}")

    msk = _msk_now()
    cur_day = (date_cur or msk.strftime("%Y-%m-%d"))[:10]
    prev_day = (date_prev or (msk - timedelta(days=1)).strftime("%Y-%m-%d"))[:10]

    out = {}
    for nm in ids:
        snaps = by_nm.get(nm) or []
        lv = live.get(nm) or {}

        cur_client = _num_or_none(lv.get("client_price"))
        cur_spp = _num_or_none(lv.get("spp"))
        cur_sale = _num_or_none(lv.get("sale_price"))
        if snaps:
            last = snaps[-1]
            if cur_client is None:
                cur_client = _num_or_none(last.get("client_price"))
            if cur_spp is None:
                cur_spp = _num_or_none(last.get("spp"))
            if cur_sale is None:
                cur_sale = _num_or_none(last.get("sale_price"))

        prev = None
        for s in reversed(snaps):
            day = _pace_parse_snap_day(s.get("captured_at"))
            if day and day <= prev_day:
                prev = s
                break
        # запасной: предпоследний снимок, если база = «вчера», а снимков за вчера нет
        if prev is None and len(snaps) >= 2:
            prev = snaps[-2]
            # не сравнивать с самим собой
            if prev is snaps[-1]:
                prev = None

        prev_client = _num_or_none(prev.get("client_price")) if prev else None
        prev_spp = _num_or_none(prev.get("spp")) if prev else None
        prev_sale = _num_or_none(prev.get("sale_price")) if prev else None

        client_delta = (
            round(cur_client - prev_client, 2)
            if cur_client is not None and prev_client is not None else None
        )
        spp_delta = (
            round(cur_spp - prev_spp, 1)
            if cur_spp is not None and prev_spp is not None else None
        )
        sale_delta = (
            round(cur_sale - prev_sale, 2)
            if cur_sale is not None and prev_sale is not None else None
        )
        out[nm] = {
            "client_price": cur_client,
            "prev_client_price": prev_client,
            "client_delta": client_delta,
            "spp": round(cur_spp, 1) if cur_spp is not None else None,
            "prev_spp": round(prev_spp, 1) if prev_spp is not None else None,
            "spp_delta": spp_delta,
            "sale_price": cur_sale,
            "prev_sale_price": prev_sale,
            "sale_delta": sale_delta,
            "price_compare_day": prev_day,
            "price_as_of": cur_day,
        }
    return out


def _enrich_pace_articles_prices(articles: list, date_cur: str = None, date_prev: str = None) -> list:
    """Цена на сайте (для покупателя) и СПП: сейчас vs день сравнения темпа."""
    if not articles:
        return articles
    nms = [a.get("nm_id") for a in articles if a.get("nm_id") is not None]
    try:
        price_map = fetch_pace_price_compare(nms, date_cur=date_cur, date_prev=date_prev)
    except Exception as e:
        logger.warning(f"enrich pace prices: {e}")
        return articles
    if not price_map:
        return articles
    out = []
    for a in articles:
        item = dict(a)
        try:
            nm = int(item.get("nm_id"))
        except Exception:
            out.append(item)
            continue
        p = price_map.get(nm) or {}
        item.update(p)
        # флаг: подорожало для покупателя при падении заказов
        od = item.get("orders_delta")
        cd = item.get("client_delta")
        sd = item.get("spp_delta")
        price_up = cd is not None and cd >= 30  # +30₽ и выше на сайте
        spp_down = sd is not None and sd <= -1.0
        item["price_linked"] = bool(
            od is not None and od < 0 and (price_up or spp_down)
        )
        out.append(item)
    return out


@app.get("/api/sales-pace")
def get_sales_pace(period: str = "day", refresh: bool = False, date_cur: str = None, date_prev: str = None):
    period = period if period in SALES_PACE_PERIODS else "day"
    if period != "day":
        date_cur = date_prev = None
    cache_key = _pace_cache_key(period, date_cur, date_prev)
    by = SALES_PACE_CACHE.get("by_period") or {}
    cached = by.get(cache_key)
    if refresh or not cached or _pace_cache_stale(cached):
        if not SALES_PACE_CACHE.get("syncing"):
            import threading
            threading.Thread(
                target=sync_sales_pace,
                kwargs={"period": period, "date_cur": date_cur, "date_prev": date_prev},
                daemon=True,
            ).start()
    cached = (SALES_PACE_CACHE.get("by_period") or {}).get(cache_key) or {}
    articles = _enrich_pace_articles_stock(cached.get("articles") or [], period)
    articles = _enrich_pace_articles_prices(
        articles,
        date_cur=cached.get("date_cur") or date_cur,
        date_prev=cached.get("date_prev") or date_prev,
    )
    return {
        "period": period,
        "cache_key": cache_key,
        "articles": articles,
        "as_of": cached.get("as_of"),
        "compare_as_of": cached.get("compare_as_of"),
        "label_cur": cached.get("label_cur"),
        "label_prev": cached.get("label_prev"),
        "col_cur": cached.get("col_cur"),
        "col_prev": cached.get("col_prev"),
        "custom_dates": cached.get("custom_dates"),
        "period_name": cached.get("period_name"),
        "mode_hint": cached.get("mode_hint"),
        "time_cutoff": cached.get("time_cutoff"),
        "date_cur": cached.get("date_cur") or date_cur,
        "date_prev": cached.get("date_prev") or date_prev,
        "today": cached.get("today"),
        "yesterday": cached.get("yesterday"),
        "now_time": cached.get("now_time"),
        "updated_at": cached.get("updated_at"),
        "funnel_ready": cached.get("funnel_ready"),
        "ads_ready": cached.get("ads_ready"),
        "syncing": SALES_PACE_CACHE.get("syncing", False) and SALES_PACE_CACHE.get("syncing_period") == cache_key,
        "error": SALES_PACE_CACHE.get("error") or cached.get("error"),
        "hidden": _sales_pace_hidden(),
    }


@app.post("/api/sales-pace-hidden")
async def save_sales_pace_hidden(request: dict):
    hidden = _uniq_str_list((request or {}).get("hidden"))
    if not save_setting_value(SALES_PACE_HIDDEN_KEY, hidden):
        raise HTTPException(status_code=500, detail="Не удалось сохранить скрытые артикулы")
    return {"status": "ok", "hidden": hidden}

@app.post("/api/sync-sales-pace")
async def trigger_sales_pace_sync(period: str = "day", date_cur: str = None, date_prev: str = None):
    import threading
    period = period if period in SALES_PACE_PERIODS else "day"
    if period != "day":
        date_cur = date_prev = None
    cache_key = _pace_cache_key(period, date_cur, date_prev)
    if SALES_PACE_CACHE.get("syncing"):
        return {"status": "already_running", "period": SALES_PACE_CACHE.get("syncing_period")}
    threading.Thread(
        target=sync_sales_pace,
        kwargs={"period": period, "date_cur": date_cur, "date_prev": date_prev},
        daemon=True,
    ).start()
    return {"status": "started", "period": period, "cache_key": cache_key, "date_cur": date_cur, "date_prev": date_prev}

# ─────────── Цены и СПП ───────────
SPP_CACHE = {
    "articles": [],
    "updated_at": None,
    "syncing": False,
    "error": None,
    "client_source": None,
}

def _money_to_rub(v, force_kopecks: bool = False):
    """WB: priceU/basic/product часто в копейках. Кабинет — обычно уже в рублях."""
    if v is None:
        return None
    try:
        x = float(v)
    except Exception:
        return None
    if x <= 0:
        return None
    if force_kopecks or x >= 10000:
        return round(x / 100.0, 2)
    return round(x, 2)

def fetch_cabinet_prices() -> list:
    """Цены из кабинета продавца: /api/v2/list/goods/filter."""
    if not WB_TOKEN:
        return []
    out = []
    offset = 0
    limit = 1000
    for _ in range(50):
        try:
            resp = httpx.get(
                f"{WB_PRICES_URL}/api/v2/list/goods/filter",
                headers=wb_headers(),
                params={"limit": limit, "offset": offset},
                timeout=40,
            )
        except Exception as e:
            logger.error(f"cabinet prices exception: {e}")
            break
        if not resp.is_success:
            logger.error(f"cabinet prices {resp.status_code}: {resp.text[:300]}")
            break
        goods = (resp.json().get("data") or {}).get("listGoods") or []
        if not goods:
            break
        for g in goods:
            nm = g.get("nmID") or g.get("nmId")
            if not nm:
                continue
            sizes = g.get("sizes") or []
            # берём мин. цену после скидки по размерам (то, что «цена продавца» до СПП)
            price = None
            sale = None
            for s in sizes:
                p = s.get("price")
                dp = s.get("discountedPrice")
                if p is not None:
                    price = min(price, float(p)) if price is not None else float(p)
                if dp is not None:
                    sale = min(sale, float(dp)) if sale is not None else float(dp)
            if sale is None and g.get("discountedPrice") is not None:
                sale = float(g.get("discountedPrice"))
            if price is None and g.get("price") is not None:
                price = float(g.get("price"))
            if sale is None and price is not None and g.get("discount") is not None:
                try:
                    sale = round(price * (100 - float(g.get("discount"))) / 100.0, 2)
                except Exception:
                    sale = price
            out.append({
                "nm_id": int(nm),
                "vendor_code": (g.get("vendorCode") or "").strip() or str(nm),
                "price": round(price, 2) if price is not None else None,
                "discount": g.get("discount"),
                "sale_price": round(sale, 2) if sale is not None else None,
                "club_discount": g.get("clubDiscount"),
            })
        if len(goods) < limit:
            break
        offset += limit
        time.sleep(0.7)
    return out

def _parse_client_product(p: dict) -> dict:
    """Достаёт цену покупателя и «базу до СПП» из продукта card/search API."""
    if not isinstance(p, dict):
        return {}
    nm = p.get("id") or p.get("nmId") or p.get("nmID")
    basic = product = None
    cashback = None
    sizes = p.get("sizes") or []
    for s in sizes:
        pr = s.get("price") if isinstance(s, dict) else None
        if not isinstance(pr, dict):
            continue
        b = pr.get("basic")
        prod = pr.get("product")
        cb = pr.get("cashback")
        if b is not None:
            b_rub = _money_to_rub(b)
            basic = min(basic, b_rub) if basic is not None and b_rub is not None else (b_rub if basic is None else basic)
        if prod is not None:
            p_rub = _money_to_rub(prod)
            product = min(product, p_rub) if product is not None and p_rub is not None else (p_rub if product is None else product)
        if cb is not None:
            try:
                cb_n = float(cb)
                cashback = max(cashback, cb_n) if cashback is not None else cb_n
            except Exception:
                pass
    # legacy fields (копейки)
    if product is None and p.get("salePriceU") is not None:
        product = _money_to_rub(p.get("salePriceU"), force_kopecks=True)
    if basic is None and p.get("priceU") is not None:
        basic = _money_to_rub(p.get("priceU"), force_kopecks=True)
    if product is None and p.get("salePrice") is not None:
        product = _money_to_rub(p.get("salePrice"))
    spp_hint = p.get("spp")
    try:
        spp_hint = float(spp_hint) if spp_hint is not None else None
    except Exception:
        spp_hint = None
    if nm is None:
        return {}
    return {
        "nm_id": int(nm),
        "client_price": product,
        "client_basic": basic,
        "spp_hint": spp_hint,
        "cashback_pct": cashback,
        "name": p.get("name") or p.get("title") or "",
    }

def fetch_client_prices(nm_ids: list) -> tuple:
    """Цены с витрины WB. Возвращает ({nm_id: info}, source_name)."""
    ids = [int(x) for x in nm_ids if x]
    if not ids:
        return {}, None
    by_nm = {}
    source = None
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
        "Accept": "application/json",
    }
    # батчами по 50
    for i in range(0, len(ids), 50):
        batch = ids[i:i + 50]
        nm_param = ";".join(str(x) for x in batch)
        # spp не передаём: иначе WB подставит «виртуальную» скидку покупателя и СПП будет фейковым
        urls = [
            f"https://card.wb.ru/cards/v4/detail?appType=1&curr=rub&dest=-1257786&nm={nm_param}",
            f"https://card.wb.ru/cards/v2/detail?appType=1&curr=rub&dest=-1257786&nm={nm_param}",
        ]
        hdrs = {
            **headers,
            "Origin": "https://www.wildberries.ru",
            "Referer": "https://www.wildberries.ru/",
        }
        got = False
        for url in urls:
            try:
                resp = httpx.get(url, headers=hdrs, timeout=30)
            except Exception as e:
                logger.warning(f"client prices fetch error: {e}")
                continue
            if not resp.is_success:
                continue
            try:
                data = resp.json()
            except Exception:
                continue
            products = data.get("products") or (data.get("data") or {}).get("products") or []
            if not products:
                continue
            for p in products:
                info = _parse_client_product(p)
                if info.get("nm_id"):
                    by_nm[info["nm_id"]] = info
            got = True
            source = "card.wb.ru"
            break
        if not got:
            # fallback: search по одному nm (медленнее)
            for nm in batch:
                try:
                    resp = httpx.get(
                        "https://search.wb.ru/exactmatch/ru/common/v18/search",
                        headers=hdrs,
                        params={
                            "appType": 1, "curr": "rub", "dest": "-1257786",
                            "query": str(nm), "resultset": "catalog", "lang": "ru",
                        },
                        timeout=25,
                    )
                    if resp.status_code == 429:
                        time.sleep(2)
                        continue
                    if not resp.is_success:
                        continue
                    data = resp.json()
                    products = data.get("products") or (data.get("data") or {}).get("products") or []
                    for p in products:
                        info = _parse_client_product(p)
                        if info.get("nm_id") == int(nm):
                            by_nm[int(nm)] = info
                            source = "search.wb.ru"
                            break
                except Exception as e:
                    logger.warning(f"search client price nm={nm}: {e}")
                time.sleep(0.35)
        time.sleep(0.25)
    return by_nm, source

def _calc_spp(sale_price, client_price):
    if not sale_price or not client_price or sale_price <= 0 or client_price <= 0:
        return None
    # СПП = насколько витрина дешевле цены продавца после его скидки
    spp = (1.0 - float(client_price) / float(sale_price)) * 100.0
    if spp < -1:
        return round(spp, 1)
    return round(max(0.0, spp), 1)

def _fmt_snap_dt(raw):
    if not raw:
        return None
    try:
        s = str(raw).replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc).strftime("%d.%m.%Y %H:%M")
    except Exception:
        return str(raw)[:16]

def _num_or_none(v):
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return None

def fetch_latest_price_snapshots(nm_ids: list) -> dict:
    """Последний снимок по каждому nm_id (за 30 дней)."""
    if not nm_ids or not SUPABASE_URL or not SUPABASE_KEY:
        return {}
    out = {}
    since = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    for i in range(0, len(nm_ids), 80):
        chunk = [int(x) for x in nm_ids[i:i + 80] if x is not None]
        if not chunk:
            continue
        ids = ",".join(str(x) for x in chunk)
        try:
            resp = httpx.get(
                f"{SUPABASE_URL}/rest/v1/price_snapshots",
                params={
                    "select": "nm_id,sale_price,client_price,price,spp,captured_at",
                    "nm_id": f"in.({ids})",
                    "captured_at": f"gte.{since}",
                    "order": "captured_at.desc",
                    "limit": "8000",
                },
                headers={
                    "apikey": SUPABASE_KEY,
                    "Authorization": f"Bearer {SUPABASE_KEY}",
                },
                timeout=30,
            )
            if not resp.is_success:
                if resp.status_code == 404:
                    logger.warning("price_snapshots table missing — run supabase/price_snapshots.sql")
                else:
                    logger.warning(f"fetch_latest_price_snapshots: {resp.status_code} {resp.text[:160]}")
                continue
            for row in resp.json() or []:
                nm = row.get("nm_id")
                if nm is not None and nm not in out:
                    out[int(nm)] = row
        except Exception as e:
            logger.warning(f"fetch_latest_price_snapshots error: {e}")
    return out

def attach_price_deltas(articles: list, prev_map: dict) -> list:
    """Добавляет prev_* и дельты относительно прошлого снимка + пояснение смены цены для клиента."""
    for a in articles:
        nm = a.get("nm_id")
        prev = prev_map.get(int(nm)) if nm is not None else None
        if not prev:
            a["prev_sale_price"] = None
            a["prev_client_price"] = None
            a["prev_spp"] = None
            a["prev_captured_at"] = None
            a["sale_delta"] = None
            a["client_delta"] = None
            a["spp_delta"] = None
            a["client_change_reason"] = None
            a["client_change_tip"] = None
            continue
        prev_sale = _num_or_none(prev.get("sale_price"))
        prev_client = _num_or_none(prev.get("client_price"))
        prev_spp = _num_or_none(prev.get("spp"))
        cur_sale = _num_or_none(a.get("sale_price"))
        cur_client = _num_or_none(a.get("client_price"))
        cur_spp = _num_or_none(a.get("spp"))
        a["prev_sale_price"] = prev_sale
        a["prev_client_price"] = prev_client
        a["prev_spp"] = round(prev_spp, 1) if prev_spp is not None else None
        a["prev_captured_at"] = _fmt_snap_dt(prev.get("captured_at"))
        a["sale_delta"] = (
            round(cur_sale - prev_sale, 2)
            if cur_sale is not None and prev_sale is not None else None
        )
        a["client_delta"] = (
            round(cur_client - prev_client, 2)
            if cur_client is not None and prev_client is not None else None
        )
        a["spp_delta"] = (
            round(cur_spp - prev_spp, 1)
            if cur_spp is not None and prev_spp is not None else None
        )
        reason, tip = _explain_client_price_change(a)
        a["client_change_reason"] = reason
        a["client_change_tip"] = tip
    return articles


def _explain_client_price_change(a: dict):
    """Поясняет, почему изменилась цена для клиента: наша цена и/или СПП."""
    client_delta = a.get("client_delta")
    if client_delta is None or abs(float(client_delta)) < 0.5:
        return None, None
    sale_delta = a.get("sale_delta")
    spp_delta = a.get("spp_delta")
    prev_client = a.get("prev_client_price")
    cur_client = a.get("client_price")
    prev_at = a.get("prev_captured_at") or ""

    parts = []
    # что сделали мы с ценой продавца
    if sale_delta is not None and abs(float(sale_delta)) >= 0.5:
        sd = float(sale_delta)
        if sd < 0:
            parts.append(f"мы снизили цену продавца на {abs(int(round(sd)))} ₽")
        else:
            parts.append(f"мы подняли цену продавца на {int(round(sd))} ₽")

    # изменение СПП (скидка WB для покупателя)
    if spp_delta is not None and abs(float(spp_delta)) >= 0.3:
        sp = float(spp_delta)
        prev_spp = a.get("prev_spp")
        cur_spp = a.get("spp")
        spp_bit = ""
        if prev_spp is not None and cur_spp is not None:
            spp_bit = f" ({prev_spp}% → {cur_spp}%)"
        if sp > 0:
            parts.append(f"вырос СПП на {sp:g} п.п.{spp_bit} — WB дал больше скидки")
        else:
            parts.append(f"упал СПП на {abs(sp):g} п.п.{spp_bit} — покупателю дороже")

    if not parts:
        parts.append("причина не по цене продавца и не по СПП — открой график")

    # доминирующая причина для короткого бейджа
    sale_abs = abs(float(sale_delta)) if sale_delta is not None else 0
    spp_abs = abs(float(spp_delta)) if spp_delta is not None else 0
    if sale_abs >= 0.5 and spp_abs < 0.3:
        reason = "our_price_down" if float(sale_delta) < 0 else "our_price_up"
    elif spp_abs >= 0.3 and sale_abs < 0.5:
        reason = "spp_up" if float(spp_delta) > 0 else "spp_down"
    elif sale_abs >= 0.5 and spp_abs >= 0.3:
        reason = "both"
    else:
        reason = "unknown"

    direction = "снизилась" if float(client_delta) < 0 else "выросла"
    prev_s = f"{int(round(prev_client))} ₽" if prev_client is not None else "—"
    cur_s = f"{int(round(cur_client))} ₽" if cur_client is not None else "—"
    tip = (
        f"Цена для клиента {direction}: было {prev_s} → стало {cur_s}"
        + (f" (снимок {prev_at})" if prev_at else "")
        + ". "
        + "Причина: "
        + "; ".join(parts)
        + "."
    )
    return reason, tip

def save_price_snapshots(articles: list) -> int:
    """Пишет снимок цен после sync. Возвращает число строк."""
    if not articles or not SUPABASE_URL or not SUPABASE_KEY:
        return 0
    now = datetime.now(timezone.utc).isoformat()
    rows = []
    for a in articles:
        nm = a.get("nm_id")
        if nm is None:
            continue
        rows.append({
            "nm_id": int(nm),
            "vendor_code": a.get("vendor_code") or "",
            "price": a.get("price"),
            "sale_price": a.get("sale_price"),
            "client_price": a.get("client_price"),
            "spp": a.get("spp"),
            "captured_at": now,
        })
    if not rows:
        return 0
    saved = 0
    for i in range(0, len(rows), 200):
        batch = rows[i:i + 200]
        try:
            resp = httpx.post(
                f"{SUPABASE_URL}/rest/v1/price_snapshots",
                json=batch,
                headers={
                    "apikey": SUPABASE_KEY,
                    "Authorization": f"Bearer {SUPABASE_KEY}",
                    "Content-Type": "application/json",
                    "Prefer": "return=minimal",
                },
                timeout=40,
            )
            if resp.is_success:
                saved += len(batch)
            else:
                logger.error(f"save_price_snapshots: {resp.status_code} {resp.text[:200]}")
                if resp.status_code == 404:
                    logger.error("Создай таблицу: supabase/price_snapshots.sql")
                break
        except Exception as e:
            logger.error(f"save_price_snapshots error: {e}")
            break
    return saved

def sync_spp_prices():
    if SPP_CACHE.get("syncing"):
        return
    SPP_CACHE["syncing"] = True
    SPP_CACHE["error"] = None
    try:
        if not WB_TOKEN:
            SPP_CACHE["error"] = "WB_TOKEN не задан"
            return
        cabinet = fetch_cabinet_prices()
        if not cabinet:
            SPP_CACHE["error"] = "Не удалось получить цены из кабинета (проверь токен категории «Цены и скидки»)"
            SPP_CACHE["articles"] = []
            return
        nm_ids = [a["nm_id"] for a in cabinet]
        prev_map = fetch_latest_price_snapshots(nm_ids)
        client_map, source = fetch_client_prices(nm_ids)
        articles = []
        missing_client = 0
        for a in cabinet:
            c = client_map.get(a["nm_id"]) or {}
            client_price = c.get("client_price")
            sale = a.get("sale_price")
            spp = _calc_spp(sale, client_price)
            # если product нет, но WB отдал поле spp — считаем клиентскую цену от кабинета
            if spp is None and sale and c.get("spp_hint") is not None:
                try:
                    hint = float(c["spp_hint"])
                    if 0 <= hint <= 95:
                        spp = round(hint, 1)
                        if client_price is None:
                            client_price = round(float(sale) * (1.0 - hint / 100.0), 2)
                except Exception:
                    pass
            # запасной путь: basic на витрине ≈ цена до СПП, product — после
            if spp is None and c.get("client_basic") and c.get("client_price"):
                spp = _calc_spp(c.get("client_basic"), c.get("client_price"))
                if client_price is None:
                    client_price = c.get("client_price")
            if client_price is None:
                missing_client += 1
            cashback_pct = c.get("cashback_pct")
            cashback_rub = None
            if cashback_pct is not None and client_price is not None:
                try:
                    cashback_rub = round(float(client_price) * float(cashback_pct) / 100.0, 0)
                except Exception:
                    cashback_rub = None
            articles.append({
                **a,
                "client_price": client_price,
                "client_basic": c.get("client_basic"),
                "spp": spp,
                "name": c.get("name") or "",
                "cashback_pct": cashback_pct,
                "cashback_rub": cashback_rub,
            })
        attach_price_deltas(articles, prev_map)
        articles.sort(key=lambda x: (-(x.get("spp") or -1), str(x.get("vendor_code") or "")))
        SPP_CACHE["articles"] = articles
        SPP_CACHE["updated_at"] = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")
        SPP_CACHE["client_source"] = source
        snap_n = save_price_snapshots(articles)
        if missing_client == len(articles):
            SPP_CACHE["error"] = "Цены кабинета загружены, но витрину WB не удалось прочитать (блокировка). СПП пока пустой — нажми Обновить позже."
        elif missing_client:
            SPP_CACHE["error"] = None
            logger.warning(f"SPP: no client price for {missing_client}/{len(articles)}")
        logger.info(
            f"SPP sync: {len(articles)} arts, client_source={source}, "
            f"missing={missing_client}, snapshots={snap_n}"
        )
    except Exception as e:
        logger.error(f"sync_spp_prices error: {e}")
        SPP_CACHE["error"] = str(e)
    finally:
        SPP_CACHE["syncing"] = False

@app.get("/api/spp-prices")
def get_spp_prices(refresh: bool = False):
    if refresh or not SPP_CACHE.get("articles"):
        if not SPP_CACHE.get("syncing"):
            threading.Thread(target=sync_spp_prices, daemon=True).start()
    return {
        "articles": SPP_CACHE.get("articles") or [],
        "updated_at": SPP_CACHE.get("updated_at"),
        "syncing": SPP_CACHE.get("syncing", False),
        "error": SPP_CACHE.get("error"),
        "client_source": SPP_CACHE.get("client_source"),
    }

@app.get("/api/price-history")
def get_price_history(nm_id: int, days: int = 30):
    """История снимков цен для графика (с момента включения, без бэкфилла WB)."""
    days = max(1, min(int(days or 30), 90))
    if not SUPABASE_URL or not SUPABASE_KEY:
        return {"nm_id": nm_id, "days": days, "points": [], "error": "Supabase не настроен"}
    since = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    try:
        resp = httpx.get(
            f"{SUPABASE_URL}/rest/v1/price_snapshots",
            params={
                "select": "nm_id,vendor_code,price,sale_price,client_price,spp,captured_at",
                "nm_id": f"eq.{int(nm_id)}",
                "captured_at": f"gte.{since}",
                "order": "captured_at.asc",
                "limit": "2000",
            },
            headers={
                "apikey": SUPABASE_KEY,
                "Authorization": f"Bearer {SUPABASE_KEY}",
            },
            timeout=30,
        )
        if not resp.is_success:
            err = "Таблица price_snapshots не создана — выполни supabase/price_snapshots.sql" if resp.status_code == 404 else resp.text[:200]
            return {"nm_id": nm_id, "days": days, "points": [], "error": err}
        rows = resp.json() or []
        points = []
        for r in rows:
            points.append({
                "captured_at": r.get("captured_at"),
                "label": _fmt_snap_dt(r.get("captured_at")),
                "price": _num_or_none(r.get("price")),
                "sale_price": _num_or_none(r.get("sale_price")),
                "client_price": _num_or_none(r.get("client_price")),
                "spp": _num_or_none(r.get("spp")),
                "vendor_code": r.get("vendor_code") or "",
            })
        vendor = points[-1]["vendor_code"] if points else ""
        return {"nm_id": nm_id, "days": days, "vendor_code": vendor, "points": points}
    except Exception as e:
        return {"nm_id": nm_id, "days": days, "points": [], "error": str(e)}

@app.post("/api/sync-spp-prices")
async def trigger_spp_prices_sync():
    if SPP_CACHE.get("syncing"):
        return {"status": "already_running"}
    threading.Thread(target=sync_spp_prices, daemon=True).start()
    return {"status": "started"}


def _wb_chat_now_ms() -> int:
    return int(time.time() * 1000)


def _wb_chat_load_state() -> dict:
    raw = get_setting_json(WB_CHAT_AUTOREPLY_KEY, {}) or {}
    if not isinstance(raw, dict):
        raw = {}
    text = str(raw.get("text") or WB_CHAT_DEFAULT_TEXT).strip()[:1000]
    replied = raw.get("replied_chats")
    if not isinstance(replied, dict):
        replied = {}
    events_next = raw.get("events_next")
    try:
        events_next = int(events_next) if events_next not in (None, "", 0, "0") else None
    except (TypeError, ValueError):
        events_next = None
    try:
        since_ts = int(raw.get("since_ts") or 0)
    except (TypeError, ValueError):
        since_ts = 0
    try:
        sent_total = int(raw.get("sent_total") or 0)
    except (TypeError, ValueError):
        sent_total = 0
    try:
        sent_last_run = int(raw.get("sent_last_run") or 0)
    except (TypeError, ValueError):
        sent_last_run = 0
    return {
        "enabled": bool(raw.get("enabled")),
        "text": text or WB_CHAT_DEFAULT_TEXT,
        "once_per_chat": raw.get("once_per_chat", True) is not False,
        "since_ts": since_ts,
        "events_next": events_next,
        "replied_chats": replied,
        "last_run": raw.get("last_run"),
        "last_error": str(raw.get("last_error") or ""),
        "last_result": str(raw.get("last_result") or ""),
        "sent_total": sent_total,
        "sent_last_run": sent_last_run,
    }


def _wb_chat_save_state(state: dict) -> bool:
    payload = dict(state)
    replied = payload.get("replied_chats") or {}
    if isinstance(replied, dict) and len(replied) > WB_CHAT_REPLIED_KEEP:
        items = sorted(
            replied.items(),
            key=lambda kv: str((kv[1] or {}).get("at") if isinstance(kv[1], dict) else kv[1] or ""),
            reverse=True,
        )
        payload["replied_chats"] = dict(items[:WB_CHAT_REPLIED_KEEP])
    return save_setting_value(WB_CHAT_AUTOREPLY_KEY, payload)


def _wb_chat_public(state: dict) -> dict:
    replied = state.get("replied_chats") or {}
    return {
        "enabled": bool(state.get("enabled")),
        "text": state.get("text") or WB_CHAT_DEFAULT_TEXT,
        "once_per_chat": state.get("once_per_chat", True) is not False,
        "last_run": state.get("last_run"),
        "last_error": state.get("last_error") or "",
        "last_result": state.get("last_result") or "",
        "sent_total": int(state.get("sent_total") or 0),
        "sent_last_run": int(state.get("sent_last_run") or 0),
        "replied_chats": len(replied) if isinstance(replied, dict) else 0,
        "running": _WB_CHAT_RUNNING,
        "default_text": WB_CHAT_DEFAULT_TEXT,
    }


def _wb_chat_auth_hint(status_code: int) -> str:
    if status_code in (401, 403):
        return (
            "WB_TOKEN без категории «Чат с покупателями». "
            "В кабинете WB → Настройки → Доступ к API перевыпусти токен с этой категорией "
            "и обнови WB_TOKEN в Railway."
        )
    if status_code == 402:
        return "WB API: не оплачен доступ к категории «Чат с покупателями»."
    return ""


def _wb_chat_get(path: str, params: dict = None, timeout: float = 20):
    return httpx.get(
        f"{WB_CHAT_URL}{path}",
        headers=wb_headers(),
        params=params or {},
        timeout=timeout,
    )


def _wb_chat_fetch_chats() -> tuple:
    """→ (chatID → replySign, error)."""
    try:
        resp = _wb_chat_get("/api/v1/seller/chats")
    except Exception as e:
        return {}, f"чаты: {e}"
    if not resp.is_success:
        hint = _wb_chat_auth_hint(resp.status_code)
        return {}, hint or f"чаты HTTP {resp.status_code}: {resp.text[:180]}"
    body = resp.json() if resp.content else {}
    rows = body.get("result") if isinstance(body, dict) else body
    if not isinstance(rows, list):
        rows = []
    signs = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        cid = str(row.get("chatID") or "").strip()
        sign = str(row.get("replySign") or "").strip()
        if cid and sign:
            signs[cid] = sign
    return signs, None


def _wb_chat_fetch_events(next_cursor=None) -> tuple:
    """→ (result_dict, error)."""
    params = {}
    if next_cursor:
        params["next"] = int(next_cursor)
    try:
        resp = _wb_chat_get("/api/v1/seller/events", params=params)
    except Exception as e:
        return {}, f"события: {e}"
    if resp.status_code == 400 and next_cursor:
        try:
            resp = _wb_chat_get("/api/v1/seller/events")
        except Exception as e:
            return {}, f"события: {e}"
    if not resp.is_success:
        hint = _wb_chat_auth_hint(resp.status_code)
        return {}, hint or f"события HTTP {resp.status_code}: {resp.text[:180]}"
    body = resp.json() if resp.content else {}
    result = body.get("result") if isinstance(body, dict) else {}
    if not isinstance(result, dict):
        result = {}
    return result, None


def _wb_chat_event_ts(ev: dict) -> int:
    ts = ev.get("addTimestamp")
    try:
        return int(ts)
    except (TypeError, ValueError):
        pass
    add_time = ev.get("addTime") or ""
    if add_time:
        try:
            dt = datetime.fromisoformat(str(add_time).replace("Z", "+00:00"))
            return int(dt.timestamp() * 1000)
        except (ValueError, TypeError):
            return 0
    return 0


def _wb_chat_is_client(ev: dict) -> bool:
    sender = str(ev.get("sender") or "").strip().lower()
    if sender in ("seller", "wb", "support", "employee"):
        return False
    if sender in ("client", "buyer", "customer", "user", "клиент"):
        return True
    source = str(ev.get("source") or "").strip().lower()
    if source in ("seller-public-api", "seller"):
        return False
    if ev.get("isNewChat"):
        return True
    if source in ("rusite", "site", "android", "ios", "mobile"):
        return True
    return False


def _wb_chat_send(reply_sign: str, text: str) -> tuple:
    """→ (ok, error)."""
    try:
        resp = httpx.post(
            f"{WB_CHAT_URL}/api/v1/seller/message",
            headers={"Authorization": WB_TOKEN},
            data={"replySign": reply_sign, "message": text},
            timeout=30,
        )
    except Exception as e:
        return False, str(e)
    if resp.is_success:
        return True, None
    hint = _wb_chat_auth_hint(resp.status_code)
    return False, hint or f"отправка HTTP {resp.status_code}: {resp.text[:180]}"


def sync_wb_chat_autoreply(force: bool = False):
    """Автоответ на входящие сообщения в чатах WB. Один шаблон, по умолчанию один раз на чат."""
    global _WB_CHAT_RUNNING
    if not WB_TOKEN:
        return {"status": "error", "error": "WB_TOKEN не задан"}
    with _WB_CHAT_LOCK:
        if _WB_CHAT_RUNNING:
            return {"status": "already_running"}
        _WB_CHAT_RUNNING = True
    try:
        state = _wb_chat_load_state()
        if not state.get("enabled") and not force:
            return {"status": "disabled"}
        text = (state.get("text") or WB_CHAT_DEFAULT_TEXT).strip()[:1000]
        if not text:
            state["last_run"] = datetime.now(timezone.utc).isoformat()
            state["last_error"] = "пустой текст автоответа"
            state["last_result"] = ""
            state["sent_last_run"] = 0
            _wb_chat_save_state(state)
            return {"status": "error", "error": state["last_error"]}

        since_ts = int(state.get("since_ts") or 0)
        if not since_ts:
            since_ts = _wb_chat_now_ms()
            state["since_ts"] = since_ts

        cursor = state.get("events_next") or since_ts
        events = []
        last_next = cursor
        pages = 0
        err = None
        while pages < 6:
            pages += 1
            result, err = _wb_chat_fetch_events(cursor)
            if err:
                break
            batch = result.get("events") or []
            if not isinstance(batch, list):
                batch = []
            events.extend(ev for ev in batch if isinstance(ev, dict))
            nxt = result.get("next")
            total = result.get("totalEvents")
            try:
                total = int(total) if total is not None else len(batch)
            except (TypeError, ValueError):
                total = len(batch)
            if nxt not in (None, "", 0, "0"):
                try:
                    last_next = int(nxt)
                    cursor = last_next
                except (TypeError, ValueError):
                    pass
            if total == 0 or not batch:
                break
            time.sleep(1.1)

        if err and not events:
            state["last_run"] = datetime.now(timezone.utc).isoformat()
            state["last_error"] = err
            state["last_result"] = ""
            state["sent_last_run"] = 0
            _wb_chat_save_state(state)
            logger.error(f"wb chat autoreply: {err}")
            return {"status": "error", "error": err}

        if last_next:
            state["events_next"] = last_next

        latest = {}
        for ev in events:
            if str(ev.get("eventType") or "message").lower() not in ("message", ""):
                continue
            ts = _wb_chat_event_ts(ev)
            if ts and ts < since_ts:
                continue
            cid = str(ev.get("chatID") or "").strip()
            if not cid:
                continue
            prev = latest.get(cid)
            if prev is None or ts >= _wb_chat_event_ts(prev):
                latest[cid] = ev

        need = []
        replied = state.get("replied_chats") or {}
        once = state.get("once_per_chat", True) is not False
        for cid, ev in latest.items():
            if not _wb_chat_is_client(ev):
                continue
            prev = replied.get(cid)
            if once and prev:
                continue
            if isinstance(prev, dict) and prev.get("event_id") and prev.get("event_id") == ev.get("eventID"):
                continue
            need.append((cid, ev))

        sent = 0
        errors = []
        signs = {}
        if need:
            time.sleep(1.1)
            signs, chat_err = _wb_chat_fetch_chats()
            if chat_err:
                errors.append(chat_err)

        for cid, ev in need[:8]:
            sign = signs.get(cid) or ""
            if not sign and ev.get("isNewChat"):
                sign = str(ev.get("replySign") or "").strip()
            if not sign:
                errors.append(f"{cid}: нет replySign — чат ещё не появился в списке")
                continue
            ok, send_err = _wb_chat_send(sign, text)
            time.sleep(1.1)
            if not ok:
                errors.append(f"{cid}: {send_err}")
                continue
            sent += 1
            replied[cid] = {
                "at": datetime.now(timezone.utc).isoformat(),
                "event_id": ev.get("eventID") or "",
                "client": ev.get("clientName") or "",
            }

        state["replied_chats"] = replied
        state["sent_total"] = int(state.get("sent_total") or 0) + sent
        state["sent_last_run"] = sent
        state["last_run"] = datetime.now(timezone.utc).isoformat()
        state["last_error"] = "; ".join(errors[:4]) if errors else ""
        parts = [f"проверено чатов: {len(latest)}", f"отправленных: {sent}"]
        if err:
            parts.append(f"события: {err}")
        state["last_result"] = ", ".join(parts)
        _wb_chat_save_state(state)
        if sent:
            logger.info(f"wb chat autoreply sent={sent} checked={len(latest)}")
        return {
            "status": "ok",
            "sent": sent,
            "checked": len(latest),
            "error": state["last_error"] or None,
        }
    except Exception as e:
        logger.exception("wb chat autoreply failed")
        try:
            state = _wb_chat_load_state()
            state["last_run"] = datetime.now(timezone.utc).isoformat()
            state["last_error"] = str(e)
            _wb_chat_save_state(state)
        except Exception:
            pass
        return {"status": "error", "error": str(e)}
    finally:
        _WB_CHAT_RUNNING = False


@app.get("/api/wb-chat-autoreply")
def get_wb_chat_autoreply():
    return _wb_chat_public(_wb_chat_load_state())


@app.put("/api/wb-chat-autoreply")
async def put_wb_chat_autoreply(request: dict):
    state = _wb_chat_load_state()
    was_enabled = bool(state.get("enabled"))
    if "enabled" in request:
        state["enabled"] = bool(request.get("enabled"))
    if "text" in request:
        text = str(request.get("text") or "").strip()[:1000]
        state["text"] = text or WB_CHAT_DEFAULT_TEXT
    if "once_per_chat" in request:
        state["once_per_chat"] = bool(request.get("once_per_chat"))
    if state["enabled"] and not was_enabled:
        state["since_ts"] = _wb_chat_now_ms()
        state["events_next"] = state["since_ts"]
        state["last_error"] = ""
        state["last_result"] = "включено — старые чаты не трогаем, отвечаем только на новые"
    if not _wb_chat_save_state(state):
        return {"error": "не удалось сохранить настройки"}
    return _wb_chat_public(state)


@app.post("/api/wb-chat-autoreply/run")
def trigger_wb_chat_autoreply():
    state = _wb_chat_load_state()
    if not state.get("enabled"):
        return {"status": "disabled", "error": "сначала включи автоответы"}
    if _WB_CHAT_RUNNING:
        return {"status": "already_running"}
    threading.Thread(target=sync_wb_chat_autoreply, daemon=True).start()
    return {"status": "started"}


@app.post("/api/wb-chat-autoreply/test")
def test_wb_chat_access():
    if not WB_TOKEN:
        return {"ok": False, "error": "WB_TOKEN не задан"}
    signs, err = _wb_chat_fetch_chats()
    if err:
        return {"ok": False, "error": err, "chats": 0}
    return {"ok": True, "chats": len(signs), "error": None}


scheduler = BackgroundScheduler()
scheduler.add_job(sync_all, "interval", minutes=30, id="sync")
scheduler.add_job(sync_stock, "interval", hours=3, id="sync_stock")
scheduler.add_job(sync_supply, "interval", hours=4, id="sync_supply")
scheduler.add_job(sync_ads, "interval", hours=4, id="sync_ads")
scheduler.add_job(lambda: sync_article_daily_stats(30), "interval", hours=6, id="sync_daily")
scheduler.add_job(sync_promotions, "interval", hours=6, id="sync_promotions")
scheduler.add_job(sync_promo_calendar, "interval", hours=6, id="sync_promo_calendar")
scheduler.add_job(lambda: sync_sales_pace("day"), "interval", minutes=15, id="sync_sales_pace")
scheduler.add_job(sync_new_stock, "interval", hours=2, id="sync_new_stock")
scheduler.add_job(sync_spp_prices, "interval", hours=3, id="sync_spp_prices")
# Каталог товаров держим тёплым: он живёт только в памяти и обнуляется при редеплое,
# а без него «что заканчивается» отвечает пустотой.
scheduler.add_job(
    lambda: refresh_wb_products_catalog(sync_sources=True),
    "interval", hours=3, id="sync_wb_products",
)
scheduler.add_job(sync_wb_chat_autoreply, "interval", minutes=1, id="wb_chat_autoreply")
scheduler.start()
# Разово чистим ошибочные api-рейтинги после деплоя (item-rating ломал склейки).
threading.Thread(target=sync_ratings_official, daemon=True).start()

FRONTEND_CANDIDATES = [
    Path(__file__).resolve().parent.parent / "frontend",  # repo/frontend
    Path(__file__).resolve().parent / "frontend",         # backend/frontend
    Path.cwd() / "frontend",
    Path.cwd().parent / "frontend",
]

def _resolve_frontend_dir():
    for p in FRONTEND_CANDIDATES:
        if (p / "index.html").exists():
            return p
    return FRONTEND_CANDIDATES[0]

FRONTEND_DIR = _resolve_frontend_dir()
logger.info(f"FRONTEND_DIR={FRONTEND_DIR} exists={FRONTEND_DIR.exists()} index={(FRONTEND_DIR / 'index.html').exists()}")

@app.get("/")
def root():
    index = FRONTEND_DIR / "index.html"
    if index.exists():
        return FileResponse(index, media_type="text/html; charset=utf-8")
    tried = [str(p) for p in FRONTEND_CANDIDATES]
    return {"status": "ok", "hint": "frontend/index.html not found", "tried": tried}

@app.get("/index.html")
def root_index():
    index = FRONTEND_DIR / "index.html"
    if index.exists():
        return FileResponse(index, media_type="text/html; charset=utf-8")
    return HTMLResponse("<h1>frontend missing</h1>", status_code=404)

if (FRONTEND_DIR / "index.html").exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")

@app.get("/api/status")
def status():
    try:
        resp = httpx.get(
            f"{SUPABASE_URL}/rest/v1/settings?key=eq.last_sync&select=value",
            headers=sb_headers(), timeout=5
        )
        last_sync = resp.json()[0]["value"] if resp.is_success and resp.json() else None
        count_resp = httpx.get(
            f"{SUPABASE_URL}/rest/v1/feedbacks?select=id",
            headers={**sb_headers(), "Prefer": "count=exact"}, timeout=5
        )
        total = int(count_resp.headers.get("content-range", "0/0").split("/")[-1])
    except:
        last_sync = None
        total = 0
    return {"status": "ok", "last_sync": last_sync, "total_feedbacks": total}

@app.post("/api/sync")
def trigger_sync():
    import threading
    threading.Thread(target=sync_all, daemon=True).start()
    return {"status": "started"}

@app.post("/api/save-manual-rating")
async def save_manual_rating(request: dict):
    """Сохраняет ручной рейтинг (разбивку по звёздам) для артикула без данных."""
    article = request.get("article")
    nm_id = request.get("nm_id")
    r5 = int(request.get("r5") or 0)
    r4 = int(request.get("r4") or 0)
    r3 = int(request.get("r3") or 0)
    r2 = int(request.get("r2") or 0)
    r1 = int(request.get("r1") or 0)
    if not article:
        return {"error": "article required"}
    total = r5 + r4 + r3 + r2 + r1
    wb_rating = round((r5*5 + r4*4 + r3*3 + r2*2 + r1*1) / total, 2) if total else 0
    row = {
        "article": article, "nm_id": nm_id,
        "wb_rating": wb_rating, "reviews_total": total,
        "r5": r5, "r4": r4, "r3": r3, "r2": r2, "r1": r1,
        "excluded": 0, "source": "manual",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    try:
        # Удаляем старую запись если есть, вставляем новую
        httpx.delete(
            f"{SUPABASE_URL}/rest/v1/ratings_official?article=eq.{article}",
            headers={**sb_headers(), "Prefer": "return=minimal"}, timeout=10
        )
        resp = httpx.post(
            f"{SUPABASE_URL}/rest/v1/ratings_official",
            json=[row], headers=sb_headers(), timeout=15
        )
        if not resp.is_success:
            return {"error": f"DB error: {resp.status_code} {resp.text[:200]}"}
        return {"status": "ok", "wb_rating": wb_rating, "total": total}
    except Exception as e:
        return {"error": str(e)}



@app.post("/api/sync-ratings")
def trigger_ratings_sync():
    """Совместимость: больше не тянем item-rating; чистим битые api-рейтинги."""
    return sync_ratings_official()

@app.post("/api/sync-stock")
def trigger_stock_sync():
    import threading
    threading.Thread(target=sync_stock, daemon=True).start()
    return {"status": "started"}

@app.get("/api/fbs-stocks")
def get_fbs_stocks(limit: int = 15):
    """Живой запрос остатков FBS (Маркетплейс) — для проверки и превью."""
    limit = max(1, min(int(limit or 15), 50))
    data = fetch_fbs_stocks()
    samples = (data.get("samples") or [])[:limit]
    # дополним vendor_code из stock_totals, если пусто
    try:
        st = httpx.get(
            f"{SUPABASE_URL}/rest/v1/stock_totals?select=nm_id,vendor_code",
            headers=sb_headers(), timeout=15,
        )
        vc_map = {
            int(r["nm_id"]): r.get("vendor_code") or ""
            for r in (st.json() or [])
            if r.get("nm_id") is not None
        } if st.is_success else {}
        for s in samples:
            if not s.get("vendor_code") and s.get("nm_id") in vc_map:
                s["vendor_code"] = vc_map[int(s["nm_id"])]
    except Exception:
        pass
    return {
        "samples": samples,
        "fbs_warehouses": data.get("fbs_warehouses") or [],
        "skus_count": data.get("skus_count"),
        "nms_with_stock": len(data.get("by_nm") or {}),
        "warehouse_rows": len(data.get("warehouses") or []),
        "errors": data.get("errors") or [],
        "error": data.get("error"),
    }


def _warehouse_channel(name: str) -> str:
    n = (name or "").lower()
    if "fbs" in n or "маркетплейс" in n:
        return "FBS"
    return "FBW"


WB_PRODUCTS_CACHE = {
    "products": [],
    "updated_at": None,
    "stock_updated_at": None,
    "prices_updated_at": None,
    "sales_updated_at": None,
    "syncing": False,
    "error": None,
    "sales_by_nm": {},  # nm -> {yesterday, d7, d28}
}


def _fetch_orders_sales_periods_fast() -> dict:
    """Только Supabase article_daily_stats (без медленного Statistics API)."""
    now = _msk_now()
    today = now.date()
    yest = today - timedelta(days=1)
    start_7 = today - timedelta(days=6)
    start_28 = today - timedelta(days=27)
    out = {}
    try:
        dt_from = start_28.isoformat()
        resp = httpx.get(
            f"{SUPABASE_URL}/rest/v1/article_daily_stats"
            f"?dt=gte.{dt_from}&select=nm_id,dt,orders&limit=50000",
            headers=sb_headers(),
            timeout=30,
        )
        rows = resp.json() if resp.is_success else []
        if not isinstance(rows, list):
            return out
        for row in rows:
            try:
                nm = int(row.get("nm_id"))
            except (TypeError, ValueError):
                continue
            try:
                d = datetime.strptime(str(row.get("dt"))[:10], "%Y-%m-%d").date()
            except Exception:
                continue
            qty = int(row.get("orders") or 0)
            if qty <= 0:
                continue
            slot = out.setdefault(nm, {"yesterday": 0, "d7": 0, "d28": 0})
            if d == yest:
                slot["yesterday"] += qty
            if start_7 <= d <= today:
                slot["d7"] += qty
            if start_28 <= d <= today:
                slot["d28"] += qty
    except Exception as e:
        logger.error(f"wb-products daily_stats sales: {e}")
    return out


def _fetch_orders_sales_periods() -> dict:
    """Заказы: сначала daily_stats, иначе Statistics API supplier/orders. Даты по Москве."""
    out = _fetch_orders_sales_periods_fast()
    if out:
        logger.info(f"wb-products sales from daily_stats: {len(out)} nms")
        return out

    now = _msk_now()
    today = now.date()
    yest = today - timedelta(days=1)
    start_7 = today - timedelta(days=6)
    start_28 = today - timedelta(days=27)
    if not WB_TOKEN:
        return out
    try:
        date_from = start_28.strftime("%Y-%m-%dT00:00:00")
        orders = fetch_supplier_feed("/api/v1/supplier/orders", date_from, max_pages=5)
        for o in orders or []:
            nm = o.get("nmId")
            if not nm:
                continue
            try:
                nm = int(nm)
            except (TypeError, ValueError):
                continue
            d = parse_wb_dt(o.get("date", ""))
            if d is None:
                continue
            day = d.date() if hasattr(d, "date") else d
            if day < start_28 or day > today:
                continue
            slot = out.setdefault(nm, {"yesterday": 0, "d7": 0, "d28": 0})
            if day == yest:
                slot["yesterday"] += 1
            if start_7 <= day <= today:
                slot["d7"] += 1
            if start_28 <= day <= today:
                slot["d28"] += 1
        logger.info(f"wb-products sales from orders API: {len(out)} nms, rows={len(orders or [])}")
    except Exception as e:
        logger.error(f"wb-products orders sales: {e}")
    return out


def build_wb_products_catalog(sales_by_nm: dict | None = None) -> dict:
    """Каталог товаров: цена покупателя, остаток по складам, канал FBW/FBS, продажи."""
    totals = []
    warehouses = []
    try:
        r = httpx.get(
            f"{SUPABASE_URL}/rest/v1/stock_totals?select=nm_id,vendor_code,quantity_warehouses_full,updated_at",
            headers=sb_headers(),
            timeout=30,
        )
        totals = r.json() if r.is_success else []
        if not isinstance(totals, list):
            totals = []
    except Exception as e:
        logger.error(f"wb-products stock_totals: {e}")
        totals = []
    try:
        r = httpx.get(
            f"{SUPABASE_URL}/rest/v1/stock_warehouses?select=nm_id,warehouse_name,quantity,updated_at",
            headers=sb_headers(),
            timeout=30,
        )
        warehouses = r.json() if r.is_success else []
        if not isinstance(warehouses, list):
            warehouses = []
    except Exception as e:
        logger.error(f"wb-products stock_warehouses: {e}")
        warehouses = []

    by_nm_wh = {}  # nm -> {name: qty}
    stock_updated = None
    for row in warehouses:
        nm = row.get("nm_id")
        if nm is None:
            continue
        try:
            nm = int(nm)
        except (TypeError, ValueError):
            continue
        name = (row.get("warehouse_name") or "").strip()
        qty = int(row.get("quantity") or 0)
        if not name or qty <= 0:
            continue
        by_nm_wh.setdefault(nm, {})
        by_nm_wh[nm][name] = by_nm_wh[nm].get(name, 0) + qty
        ua = row.get("updated_at")
        if ua and (not stock_updated or str(ua) > str(stock_updated)):
            stock_updated = ua

    totals_map = {}
    for row in totals:
        nm = row.get("nm_id")
        if nm is None:
            continue
        try:
            nm = int(nm)
        except (TypeError, ValueError):
            continue
        totals_map[nm] = row
        ua = row.get("updated_at")
        if ua and (not stock_updated or str(ua) > str(stock_updated)):
            stock_updated = ua

    price_map = {}
    for a in (SPP_CACHE.get("articles") or []):
        try:
            nm = int(a.get("nm_id"))
        except (TypeError, ValueError):
            continue
        price_map[nm] = a

    if sales_by_nm is None:
        sales_by_nm = WB_PRODUCTS_CACHE.get("sales_by_nm") or {}
    # ключи могли прийти строками из JSON-кэша
    sales_norm = {}
    for k, v in (sales_by_nm or {}).items():
        try:
            sales_norm[int(k)] = v if isinstance(v, dict) else {}
        except (TypeError, ValueError):
            continue

    vc_map = {}
    try:
        vc_map = build_nm_to_vendor_map() or {}
    except Exception:
        vc_map = {}

    nm_ids = set(totals_map.keys()) | set(by_nm_wh.keys()) | set(price_map.keys()) | set(sales_norm.keys())
    products = []
    for nm in nm_ids:
        t = totals_map.get(nm) or {}
        p = price_map.get(nm) or {}
        wh_map = by_nm_wh.get(nm) or {}
        wh_list = [
            {"name": name, "qty": qty, "channel": _warehouse_channel(name)}
            for name, qty in sorted(wh_map.items(), key=lambda x: (-x[1], x[0].lower()))
        ]
        stock = sum(w["qty"] for w in wh_list)
        if not stock:
            stock = int(t.get("quantity_warehouses_full") or 0)
        channels = []
        for ch in ("FBW", "FBS"):
            if any(w["channel"] == ch for w in wh_list):
                channels.append(ch)
        if not channels and stock > 0:
            channels = ["FBW"]
        vc = (
            (p.get("vendor_code") or "").strip()
            or (t.get("vendor_code") or "").strip()
            or (vc_map.get(nm) or "").strip()
            or str(nm)
        )
        if vc == str(nm) and vc_map.get(nm):
            vc = str(vc_map.get(nm)).strip()
        client_price = p.get("client_price")
        sale_price = p.get("sale_price")
        sales = sales_norm.get(nm) or {}
        products.append({
            "nm_id": nm,
            "vendor_code": vc,
            "name": (p.get("name") or "").strip() or None,
            "client_price": client_price,
            "sale_price": sale_price,
            "spp": p.get("spp"),
            "stock": int(stock or 0),
            "warehouse_count": len(wh_list),
            "warehouses": wh_list,
            "channels": channels,
            "sales_yesterday": int(sales.get("yesterday") or 0),
            "sales_7d": int(sales.get("d7") or 0),
            "sales_28d": int(sales.get("d28") or 0),
            "url": f"https://www.wildberries.ru/catalog/{nm}/detail.aspx",
        })

    products.sort(key=lambda x: (
        -(x.get("sales_7d") or 0),
        -(x.get("stock") or 0),
        str(x.get("vendor_code") or "").lower(),
    ))
    stock_updated_fmt = None
    if stock_updated:
        try:
            from zoneinfo import ZoneInfo
            dt = datetime.fromisoformat(str(stock_updated).replace("Z", "+00:00"))
            stock_updated_fmt = dt.astimezone(ZoneInfo("Europe/Moscow")).strftime("%d.%m.%Y %H:%M")
        except Exception:
            stock_updated_fmt = str(stock_updated)[:16]

    return {
        "products": products,
        "count": len(products),
        "with_stock": sum(1 for x in products if (x.get("stock") or 0) > 0),
        "updated_at": _msk_now().strftime("%d.%m.%Y %H:%M"),
        "stock_updated_at": stock_updated_fmt,
        "prices_updated_at": SPP_CACHE.get("updated_at"),
        "sales_updated_at": WB_PRODUCTS_CACHE.get("sales_updated_at"),
        "sales_by_nm": sales_norm,
        "error": None,
    }


def refresh_wb_products_catalog(sync_sources: bool = False):
    if WB_PRODUCTS_CACHE.get("syncing"):
        return
    WB_PRODUCTS_CACHE["syncing"] = True
    WB_PRODUCTS_CACHE["error"] = None
    try:
        if sync_sources:
            try:
                sync_stock()
            except Exception as e:
                logger.error(f"wb-products sync_stock: {e}")
            if not SPP_CACHE.get("articles") and not SPP_CACHE.get("syncing"):
                try:
                    sync_spp_prices()
                except Exception as e:
                    logger.error(f"wb-products sync_spp: {e}")
            try:
                # подтянуть дневную статистику (если таблица есть) — для быстрых периодов
                sync_article_daily_stats(30)
            except Exception as e:
                logger.warning(f"wb-products sync_daily: {e}")
            try:
                sales = _fetch_orders_sales_periods()
                WB_PRODUCTS_CACHE["sales_by_nm"] = sales
                WB_PRODUCTS_CACHE["sales_updated_at"] = _msk_now().strftime("%d.%m.%Y %H:%M")
            except Exception as e:
                logger.error(f"wb-products sales: {e}")
        elif not WB_PRODUCTS_CACHE.get("sales_by_nm"):
            try:
                sales = _fetch_orders_sales_periods_fast()
                WB_PRODUCTS_CACHE["sales_by_nm"] = sales
                if sales:
                    WB_PRODUCTS_CACHE["sales_updated_at"] = _msk_now().strftime("%d.%m.%Y %H:%M")
            except Exception as e:
                logger.error(f"wb-products sales soft: {e}")
        data = build_wb_products_catalog(WB_PRODUCTS_CACHE.get("sales_by_nm") or {})
        WB_PRODUCTS_CACHE.update(data)
        WB_PRODUCTS_CACHE["syncing"] = False
    except Exception as e:
        logger.error(f"wb-products refresh: {e}")
        WB_PRODUCTS_CACHE["syncing"] = False
        WB_PRODUCTS_CACHE["error"] = str(e)


@app.get("/api/wb-products")
def get_wb_products(refresh: bool = False):
    """Товары WB: цена покупателя, остаток (склады), канал FBW/FBS, продажи."""
    need = refresh or not WB_PRODUCTS_CACHE.get("products")
    if need and not WB_PRODUCTS_CACHE.get("syncing"):
        try:
            # продажи: только быстрый daily_stats; полный orders — через sync-wb-products
            if not WB_PRODUCTS_CACHE.get("sales_by_nm"):
                try:
                    sales = _fetch_orders_sales_periods_fast()
                    if sales:
                        WB_PRODUCTS_CACHE["sales_by_nm"] = sales
                        WB_PRODUCTS_CACHE["sales_updated_at"] = _msk_now().strftime("%d.%m.%Y %H:%M")
                except Exception:
                    pass
            data = build_wb_products_catalog(WB_PRODUCTS_CACHE.get("sales_by_nm") or {})
            WB_PRODUCTS_CACHE.update({**data, "syncing": False, "error": None})
        except Exception as e:
            WB_PRODUCTS_CACHE["error"] = str(e)
    return {
        "products": WB_PRODUCTS_CACHE.get("products") or [],
        "count": WB_PRODUCTS_CACHE.get("count") or len(WB_PRODUCTS_CACHE.get("products") or []),
        "with_stock": WB_PRODUCTS_CACHE.get("with_stock"),
        "updated_at": WB_PRODUCTS_CACHE.get("updated_at"),
        "stock_updated_at": WB_PRODUCTS_CACHE.get("stock_updated_at"),
        "prices_updated_at": WB_PRODUCTS_CACHE.get("prices_updated_at"),
        "sales_updated_at": WB_PRODUCTS_CACHE.get("sales_updated_at"),
        "syncing": WB_PRODUCTS_CACHE.get("syncing", False),
        "error": WB_PRODUCTS_CACHE.get("error"),
    }


@app.post("/api/sync-wb-products")
def sync_wb_products():
    if WB_PRODUCTS_CACHE.get("syncing"):
        return {"status": "already_running"}
    threading.Thread(
        target=refresh_wb_products_catalog,
        kwargs={"sync_sources": True},
        daemon=True,
    ).start()
    return {"status": "started"}


@app.post("/api/sync-supply")
def trigger_supply_sync():
    import threading
    threading.Thread(target=sync_supply, daemon=True).start()
    return {"status": "started"}

def sync_stock_then_supply():
    sync_stock()
    sync_supply()

@app.post("/api/sync-supply-full")
def trigger_supply_full_sync():
    """Обновляет остатки (для текущих остатков по складам), затем заказы/продажи (для рекомендаций) — одной кнопкой."""
    import threading
    threading.Thread(target=sync_stock_then_supply, daemon=True).start()
    return {"status": "started"}

@app.post("/api/sync-ads")
def trigger_ads_sync():
    import threading
    if ADS_CACHE.get("syncing"):
        return {"status": "already_running"}
    threading.Thread(target=sync_ads, daemon=True).start()
    return {"status": "started"}

@app.post("/api/save-setting")
async def save_setting(request: dict):
    """Сохраняет произвольную настройку (например target_coverage_days) в таблицу settings."""
    key = request.get("key")
    value = request.get("value")
    if not key:
        return {"error": "key required"}
    try:
        if key == SUPPLY_WH_DISABLED_KEY:
            value = _normalize_disabled_warehouses(value)
        # списки/объекты — как JSON-строка (иначе str(list) сломает get_setting_json)
        if isinstance(value, (list, dict)):
            import json as _json
            store_val = _json.dumps(value, ensure_ascii=False)
        else:
            store_val = str(value)
        resp = httpx.post(
            f"{SUPABASE_URL}/rest/v1/settings?on_conflict=key",
            json={"key": key, "value": store_val, "updated_at": datetime.now(timezone.utc).isoformat()},
            headers=sb_headers(), timeout=10
        )
        if not resp.is_success:
            return {"error": f"Supabase error: {resp.status_code} {resp.text[:200]}"}
        # сброс кэшей, где настройка влияет на UI/расчёты
        if key == SUPPLY_WH_DISABLED_KEY:
            SALES_PACE_CACHE["by_period"] = {}
            _invalidate_dash_cache()
        elif key in (
            "target_coverage_days", "sales_window_days",
            "last_supply_sync", "last_ads_sync", "last_sync",
        ):
            _invalidate_dash_cache()
        return {"status": "ok", "key": key, "value": value if key == SUPPLY_WH_DISABLED_KEY else store_val}
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/supply-wh-disabled")
def get_supply_wh_disabled():
    """Актуальный список отключённых складов (без кэша dashboard-data)."""
    names = sorted(get_disabled_warehouses())
    return {"status": "ok", "disabled": names}


@app.put("/api/supply-wh-disabled")
async def put_supply_wh_disabled(request: dict):
    """Сохранить список отключённых складов (общий для всех устройств)."""
    names = _normalize_disabled_warehouses(request.get("disabled", request.get("value")))
    if not _save_disabled_warehouses(names):
        return {"error": "не удалось сохранить в settings"}
    SALES_PACE_CACHE["by_period"] = {}
    _invalidate_dash_cache()
    return {"status": "ok", "disabled": names}

# ---------- Proxy endpoints: фронтенд обращается только к Railway, ----------
# ---------- никогда напрямую к Supabase (для пользователей у которых ----------
# ---------- Supabase плохо доступен напрямую). Railway сам ходит в Supabase. ----------

_DASH_CACHE = {"ts": 0.0, "data": None}
_DASH_CACHE_LOCK = threading.Lock()
_DASH_CACHE_TTL = float(os.getenv("DASHBOARD_CACHE_TTL", "45"))


def _dash_get(path: str, timeout: float = 15):
    return httpx.get(f"{SUPABASE_URL}/rest/v1/{path}", headers=sb_headers(), timeout=timeout)


def _dash_rpc(name: str, payload: dict, timeout: float = 20):
    return httpx.post(
        f"{SUPABASE_URL}/rest/v1/rpc/{name}",
        json=payload, headers=sb_headers(), timeout=timeout,
    )


@app.get("/api/dashboard-data")
def dashboard_data():
    """Все данные дашборда одним ответом. Запросы к Supabase — параллельно + короткий кэш."""
    now = time.time()
    with _DASH_CACHE_LOCK:
        cached = _DASH_CACHE["data"]
        if cached is not None and (now - _DASH_CACHE["ts"]) < _DASH_CACHE_TTL:
            return cached

    result = {
        "groups": [], "ratings": [], "feedback_stats": [], "negative_counts": {},
        "settings": {}, "stock_totals": [], "stock_warehouses": [],
        "supply_report": [], "ad_stats": [],
    }
    neg_days = [1, 2, 3, 4, 5, 7, 14, 30]

    def load_groups():
        r = _dash_get("groups_config?select=name,articles,sort_order&order=sort_order", 15)
        return ("groups", r.json() if r.is_success else [])

    def load_ratings():
        # Как неделю назад: только xlsx/manual. source=api — прирост за период, ломает склейки.
        r = _dash_get("ratings_official?select=*&source=neq.api", 15)
        if not r.is_success:
            r = _dash_get("ratings_official?select=*", 15)
            rows = r.json() if r.is_success else []
            rows = [x for x in rows if (x.get("source") or "") != "api"]
            return ("ratings", rows)
        return ("ratings", r.json() if r.is_success else [])

    def load_feedback_stats():
        r = _dash_rpc("get_article_stats", {}, 20)
        return ("feedback_stats", r.json() if r.is_success else [])

    def load_neg(days: int):
        # ★1–2 — «жесткий» негатив (совпадает с фильтром NEG_STARS на фронте)
        r = _dash_rpc("get_negative_counts", {"days_back": days, "max_stars": 2}, 20)
        return ("neg", days, r.json() if r.is_success else [])

    def load_settings():
        r = _dash_get("settings?select=key,value", 10)
        out = {}
        if r.is_success:
            for row in r.json():
                out[row["key"]] = row["value"]
        return ("settings", out)

    def load_stock_totals():
        r = _dash_get("stock_totals?select=*", 15)
        return ("stock_totals", r.json() if r.is_success else [])

    def load_stock_warehouses():
        r = _dash_get("stock_warehouses?select=*", 15)
        return ("stock_warehouses", r.json() if r.is_success else [])

    def load_supply():
        r = _dash_get("supply_report?select=*", 20)
        return ("supply_report", r.json() if r.is_success else [])

    def load_ads():
        if ADS_CACHE.get("campaigns"):
            return ("ad_stats", ADS_CACHE["campaigns"])
        r = _dash_get("ad_stats?select=*", 20)
        return ("ad_stats", r.json() if r.is_success else [])

    jobs = [
        load_groups, load_ratings, load_feedback_stats, load_settings,
        load_stock_totals, load_stock_warehouses, load_supply, load_ads,
    ]
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=12) as pool:
        futs = [pool.submit(fn) for fn in jobs]
        futs += [pool.submit(load_neg, d) for d in neg_days]
        for fut in as_completed(futs):
            try:
                item = fut.result()
                if item[0] == "neg":
                    _, days, rows = item
                    result["negative_counts"][str(days)] = rows
                else:
                    key, val = item
                    result[key] = val
            except Exception as e:
                logger.error(f"dashboard-data parallel error: {e}")

    logger.info(f"dashboard-data built in {time.time() - t0:.2f}s")
    with _DASH_CACHE_LOCK:
        _DASH_CACHE["ts"] = time.time()
        _DASH_CACHE["data"] = result
    return result


@app.get("/api/article-feedbacks")
def article_feedbacks(article: str, days: int = 30, max_stars: int = 3, limit: int = 50):
    """
    Возвращает тексты отзывов по конкретному артикулу за последние N дней,
    с оценкой <= max_stars, отсортированные по дате (новые сверху).
    Используется при раскрытии артикула в таблице товаров.
    """
    try:
        cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
        resp = httpx.get(
            f"{SUPABASE_URL}/rest/v1/feedbacks",
            params={
                "article": f"eq.{article}",
                "stars": f"lte.{max_stars}",
                "created_date": f"gte.{cutoff}",
                "select": "id,stars,created_date,text,is_answered",
                "order": "created_date.desc",
                "limit": str(limit),
            },
            headers=sb_headers(), timeout=15
        )
        if not resp.is_success:
            return {"error": f"Supabase error: {resp.status_code} {resp.text[:200]}"}
        return {"feedbacks": resp.json()}
    except Exception as e:
        logger.error(f"article-feedbacks error: {e}")
        return {"error": str(e)}

@app.post("/api/save-groups")
async def save_groups(request: dict):
    """
    Сохраняет конфигурацию склеек. Ожидает {"groups": {"Название": ["арт1","арт2"], ...}}
    """
    groups = request.get("groups", {})
    try:
        del_resp = httpx.delete(
            f"{SUPABASE_URL}/rest/v1/groups_config?id=gte.1",
            headers=sb_headers(), timeout=15
        )
        rows = [{"name": name, "articles": articles, "sort_order": i + 1}
                for i, (name, articles) in enumerate(groups.items())]
        if rows:
            ins_resp = httpx.post(
                f"{SUPABASE_URL}/rest/v1/groups_config",
                json=rows,
                headers={**sb_headers(), "Prefer": "return=minimal"},
                timeout=15
            )
            if not ins_resp.is_success:
                return {"error": f"Insert failed: {ins_resp.status_code} {ins_resp.text[:200]}"}
        return {"status": "ok"}
    except Exception as e:
        logger.error(f"save-groups error: {e}")
        return {"error": str(e)}

# ---------- Финансы: себестоимость остатков ----------
COST_PRICES_KEY = "cost_prices"
COST_META_KEY = "cost_prices_meta"

def _parse_cost_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v) if float(v) >= 0 else None
    s = str(v).strip().replace("\xa0", " ").replace(" ", "").replace(",", ".")
    s = s.replace("₽", "").replace("руб.", "").replace("руб", "")
    if not s or s.lower() in ("nan", "none", "-", "—"):
        return None
    try:
        n = float(s)
        return n if n >= 0 else None
    except Exception:
        return None

def _norm_vendor_key(v):
    """Артикул продавца: trim + латиница O вместо кириллической О (частая путаница в Excel)."""
    s = str(v or "").strip()
    if not s:
        return ""
    return s.replace("\u041e", "O").replace("\u043e", "o")


def build_nm_to_vendor_map() -> dict:
    """nm_id → артикул продавца (033_…). stock_totals часто без vendor_code — берём из рейтингов/отзывов."""
    m = {}
    try:
        r = httpx.get(
            f"{SUPABASE_URL}/rest/v1/ratings_official?select=nm_id,article&nm_id=not.is.null&article=not.is.null&limit=5000",
            headers=sb_headers(), timeout=20,
        )
        if r.is_success:
            for row in r.json() or []:
                nm, art = row.get("nm_id"), _norm_vendor_key(row.get("article"))
                if nm is not None and art and art != str(nm):
                    m[int(nm)] = art
    except Exception as e:
        logger.warning(f"build_nm_to_vendor_map ratings: {e}")
    try:
        r = httpx.get(
            f"{SUPABASE_URL}/rest/v1/feedbacks?select=nm_id,article&nm_id=not.is.null&article=not.is.null&limit=5000",
            headers=sb_headers(), timeout=20,
        )
        if r.is_success:
            for row in r.json() or []:
                nm, art = row.get("nm_id"), _norm_vendor_key(row.get("article"))
                if nm is None or not art or art == str(nm):
                    continue
                nm = int(nm)
                if nm not in m:
                    m[nm] = art
    except Exception as e:
        logger.warning(f"build_nm_to_vendor_map feedbacks: {e}")
    try:
        r = httpx.get(
            f"{SUPABASE_URL}/rest/v1/stock_totals?select=nm_id,vendor_code&limit=5000",
            headers=sb_headers(), timeout=15,
        )
        if r.is_success:
            for row in r.json() or []:
                nm, art = row.get("nm_id"), _norm_vendor_key(row.get("vendor_code"))
                if nm is None or not art or art == str(nm):
                    continue
                nm = int(nm)
                if nm not in m:
                    m[nm] = art
    except Exception as e:
        logger.warning(f"build_nm_to_vendor_map stock: {e}")
    return m

def _parse_header_date(v):
    """Парсит дату из заголовка колонки (datetime / '2026-03-30' / '30.03.2026')."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or s.lower() in ("по умолчанию", "default", "sku", "артикул", "наименование", "размер"):
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            pass
    try:
        # excel serial sometimes comes as number string
        n = float(s)
        if 30000 < n < 60000:
            from datetime import date as _date
            return (_date(1899, 12, 30) + timedelta(days=int(n)))
    except Exception:
        pass
    return None

def _effective_cost_from_history(default_cost, dated_costs, as_of=None):
    """
    dated_costs: [(date, cost), ...] — только даты, где цена явно задана.
    Берём последнюю дату <= as_of, иначе default.
    """
    as_of = as_of or datetime.now(timezone.utc).date()
    applicable = [(d, c) for d, c in dated_costs if d is not None and d <= as_of and c is not None]
    if applicable:
        d, c = max(applicable, key=lambda x: x[0])
        return c, d.isoformat()
    if default_cost is not None:
        return default_cost, None
    return None, None

def _cost_entry_value(entry):
    """Достаёт актуальную себестоимость из float или объекта."""
    if entry is None:
        return None
    if isinstance(entry, dict):
        return _parse_cost_number(entry.get("cost"))
    return _parse_cost_number(entry)

def _cost_entry_meta(entry):
    if isinstance(entry, dict):
        return {
            "cost": _parse_cost_number(entry.get("cost")),
            "default": _parse_cost_number(entry.get("default")),
            "as_of": entry.get("as_of"),
        }
    c = _parse_cost_number(entry)
    return {"cost": c, "default": c, "as_of": None}

def parse_cost_price_workbook(contents: bytes, as_of=None):
    """
    Формат листа «Себестоимость»:
      row1: SKU | Артикул | … | По умолчанию | По умолчанию | 2026-03-30 | 2026-03-30 | …
      row2:          …        | Себестоимость | Фулфилмент | Себестоимость | Фулфилмент | …
    Для остатков берём только «Себестоимость»: default + последняя дата <= сегодня.
    """
    from openpyxl import load_workbook
    as_of = as_of or datetime.now(timezone.utc).date()
    wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    # предпочитаем лист с «себестоим» в названии
    ws = None
    for name in wb.sheetnames:
        if "себестоим" in name.lower() or "cost" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        h1 = next(rows_iter)
        h2 = next(rows_iter)
    except StopIteration:
        wb.close()
        return {"by_vendor": {}, "by_nm": {}}, {"error": "Пустой файл"}

    # колонки себестоимости: (col_idx, date_or_None_for_default)
    cost_cols = []
    for i, (top, sub) in enumerate(zip(h1, h2)):
        sub_l = str(sub or "").strip().lower()
        top_s = str(top or "").strip().lower()
        if "себестоим" not in sub_l and "себестоим" not in top_s and "cost" not in sub_l:
            # default pair sometimes has sub only
            if top_s == "по умолчанию" and ("себестоим" in sub_l or sub_l == ""):
                # только если сосед/этот — себестоимость; skip fulfillment
                if "фулфил" in sub_l or "fulfill" in sub_l:
                    continue
            else:
                continue
        if "фулфил" in sub_l or "fulfill" in sub_l:
            continue
        d = _parse_header_date(top)
        is_default = d is None and ("умолчан" in top_s or top_s in ("", "none", "nan"))
        if d is None and not is_default and "умолчан" not in top_s:
            # заголовок не дата и не default — пропускаем
            continue
        cost_cols.append((i, d))  # d=None → default

    # если по sub-заголовку не нашли — ищем пары «По умолчанию»/даты где чётные = себес
    if not cost_cols:
        for i, top in enumerate(h1):
            top_s = str(top or "").strip().lower()
            d = _parse_header_date(top)
            if "умолчан" in top_s:
                # первая из пары default = себес (col 4), вторая фулфилмент
                # определяем: если следующий top такой же — это пара, берём только первый
                prev_same = i > 0 and str(h1[i - 1] or "").strip().lower() == top_s
                if prev_same:
                    continue  # вторая колонка пары
                cost_cols.append((i, None))
            elif d is not None:
                prev_d = _parse_header_date(h1[i - 1]) if i > 0 else None
                if prev_d == d:
                    continue  # fulfillment twin
                cost_cols.append((i, d))

    # артикул продавца + SKU (nm_id WB)
    vc_col = 1
    sku_col = 0
    for i, top in enumerate(h1):
        t = str(top or "").strip().lower()
        if t == "артикул" or "артикул продавца" in t or t == "vendorcode" or t == "vendor_code":
            vc_col = i
        if t in ("sku", "nm_id", "nmid", "код нм", "номенклатура", "нм", "nm"):
            sku_col = i
    # если заголовки пустые из‑за merge — эвристика по данным первых строк
    peek = []
    rows_list = list(rows_iter)
    for row in rows_list[:30]:
        if row:
            peek.append(row)
    if peek and (not h1[0] or str(h1[0]).strip() == ""):
        # col0 выглядит как nm_id (длинное число), col1 — артикул продавца
        c0 = peek[0][0] if peek[0] else None
        c1 = peek[0][1] if len(peek[0]) > 1 else None
        try:
            if c0 is not None and float(str(c0)) > 10000:
                sku_col = 0
        except Exception:
            pass
        if c1 is not None and not str(c1).replace(".", "").isdigit():
            vc_col = 1

    default_idxs = [i for i, d in cost_cols if d is None]
    dated_idxs = [(i, d) for i, d in cost_cols if d is not None]

    by_vendor = {}
    by_nm = {}
    for row in rows_list:
        if not row or vc_col >= len(row):
            continue
        vc = _norm_vendor_key(row[vc_col])
        if not vc or vc.lower() in ("артикул", "nan", "none"):
            continue
        nm_id = None
        if sku_col is not None and sku_col < len(row) and row[sku_col] not in (None, ""):
            try:
                raw_sku = str(row[sku_col]).strip()
                # nm_id WB — обычно 6–12 цифр; не путать с артикулом продавца
                if raw_sku.replace(".", "", 1).isdigit():
                    nm_id = int(float(raw_sku))
                    if nm_id < 10000:
                        nm_id = None
            except Exception:
                nm_id = None
        default_cost = None
        for i in default_idxs:
            if i < len(row):
                c = _parse_cost_number(row[i])
                if c is not None:
                    default_cost = c
                    break
        dated = []
        for i, d in dated_idxs:
            if i < len(row) and row[i] not in (None, ""):
                c = _parse_cost_number(row[i])
                if c is not None:
                    dated.append((d, c))
        eff, as_of_used = _effective_cost_from_history(default_cost, dated, as_of)
        if eff is None:
            continue
        entry = {
            "cost": round(eff, 4),
            "default": round(default_cost, 4) if default_cost is not None else None,
            "as_of": as_of_used,
            "vendor_code": vc,
            "nm_id": nm_id,
            "history": (
                ([{"date": None, "cost": round(default_cost, 4)}] if default_cost is not None else [])
                + [{"date": d.isoformat(), "cost": round(c, 4)} for d, c in sorted(dated, key=lambda x: x[0])]
            ),
        }
        by_vendor[vc] = entry
        if nm_id is not None:
            by_nm[str(nm_id)] = entry
    wb.close()
    return {"by_vendor": by_vendor, "by_nm": by_nm}, {
        "format": "dated_cost_matrix",
        "default_cols": len(default_idxs),
        "date_cols": len(dated_idxs),
        "as_of": as_of.isoformat(),
        "vendors": len(by_vendor),
        "nms": len(by_nm),
    }

@app.post("/api/upload-costs")
async def upload_costs(file: UploadFile = File(...)):
    """
    Excel себестоимости:
    - формат с датами (По умолчанию + колонки дат Себестоимость/Фулфилмент)
    - или простой файл Артикул + Себестоимость
    Актуальная цена остатков = последняя себестоимость с датой <= сегодня, иначе «По умолчанию».
    Матчинг остатков WB: по артикулу продавца и по SKU (nm_id).
    """
    try:
        contents = await file.read()
        name = (file.filename or "").lower()
        by_vendor, by_nm = {}, {}
        parse_meta = {}

        if name.endswith(".xlsx") or name.endswith(".xls") or not name.endswith(".csv"):
            try:
                parsed, parse_meta = parse_cost_price_workbook(contents)
                by_vendor = (parsed or {}).get("by_vendor") or {}
                by_nm = (parsed or {}).get("by_nm") or {}
            except Exception as e:
                logger.warning(f"dated cost parse failed, fallback: {e}")
                by_vendor, by_nm, parse_meta = {}, {}, {"dated_error": str(e)}

        if not by_vendor:
            if name.endswith(".csv"):
                df = pd.read_csv(io.BytesIO(contents), dtype=str, sep=None, engine="python")
            else:
                xl = pd.ExcelFile(io.BytesIO(contents))
                best = None
                for s in xl.sheet_names:
                    tmp = pd.read_excel(io.BytesIO(contents), sheet_name=s, header=None, dtype=str)
                    for i, row in tmp.iterrows():
                        vals = [str(v).strip().lower() for v in row.values]
                        joined = " | ".join(vals)
                        if "артикул" in joined and ("себестоим" in joined or "cost" in joined or "умолчан" in joined):
                            best = (i, tmp)
                            break
                    if best:
                        break
                if not best:
                    tmp = pd.read_excel(io.BytesIO(contents), sheet_name=0, header=None, dtype=str)
                    best = (0, tmp)
                header_row, tmp = best
                tmp.columns = [str(c).strip() for c in tmp.iloc[header_row].tolist()]
                df = tmp.iloc[header_row + 1:].reset_index(drop=True)

            cols = {str(c).strip().lower(): c for c in df.columns}
            def find_col(*needles):
                for low, orig in cols.items():
                    for n in needles:
                        if n in low:
                            return orig
                return None
            col_vc = find_col("артикул продавца") or find_col("артикул") or find_col("vendor")
            col_sku = find_col("sku", "nm_id", "nmid")
            col_cost = find_col("по умолчанию") or find_col("себестоим", "cost", "закуп") or find_col("цена")
            if not col_vc or not col_cost:
                return {
                    "error": "Не удалось прочитать файл. Нужен Excel как cost_price: SKU + Артикул + По умолчанию + даты.",
                    "columns": list(df.columns.astype(str)),
                    "parse_meta": parse_meta,
                }
            for _, row in df.iterrows():
                vc = _norm_vendor_key(row.get(col_vc))
                if not vc or vc.lower() in ("nan", "none", "артикул"):
                    continue
                cost = _parse_cost_number(row.get(col_cost))
                if cost is None:
                    continue
                nm_id = None
                if col_sku is not None:
                    try:
                        nm_id = int(float(str(row.get(col_sku)).strip()))
                    except Exception:
                        nm_id = None
                entry = {
                    "cost": round(cost, 4),
                    "default": round(cost, 4),
                    "as_of": None,
                    "vendor_code": vc,
                    "nm_id": nm_id,
                    "history": [],
                }
                by_vendor[vc] = entry
                if nm_id is not None:
                    by_nm[str(nm_id)] = entry
            parse_meta["format"] = "simple"

        if not by_vendor:
            return {"error": "В файле не найдено артикулов с себестоимостью", "parse_meta": parse_meta}

        payload = {"_v": 2, "by_vendor": by_vendor, "by_nm": by_nm}
        save_setting_value(COST_PRICES_KEY, payload)
        meta = {
            "filename": file.filename,
            "uploaded_at": datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M"),
            "rows_in_file": len(by_vendor),
            "total_articles": len(by_vendor),
            "nm_mapped": len(by_nm),
            "format": parse_meta.get("format"),
            "as_of": parse_meta.get("as_of"),
            "date_cols": parse_meta.get("date_cols"),
        }
        save_setting_value(COST_META_KEY, meta)

        sample_vc = "039_DT10_mini_gold_O"
        sample = by_vendor.get(sample_vc)
        return {
            "status": "ok",
            "loaded": len(by_vendor),
            "total": len(by_vendor),
            "nm_mapped": len(by_nm),
            "meta": meta,
            "sample": {sample_vc: sample} if sample else None,
        }
    except Exception as e:
        logger.error(f"upload-costs error: {e}")
        return {"error": str(e)}

def _load_cost_indexes():
    """Возвращает (by_vendor, by_nm) из settings — поддерживает старый и новый формат."""
    raw = get_setting_json(COST_PRICES_KEY, {}) or {}
    if not isinstance(raw, dict):
        return {}, {}
    if raw.get("_v") == 2 or ("by_vendor" in raw or "by_nm" in raw):
        by_vendor = raw.get("by_vendor") or {}
        by_nm = raw.get("by_nm") or {}
    else:
        # старый формат: {vendor: entry|float}
        by_vendor, by_nm = {}, {}
        for k, v in raw.items():
            if str(k).startswith("_"):
                continue
            vc = _norm_vendor_key(k)
            if not vc:
                continue
            meta = _cost_entry_meta(v)
            if meta["cost"] is None:
                continue
            entry = {**meta, "vendor_code": vc, "nm_id": None}
            if isinstance(v, dict):
                entry["nm_id"] = v.get("nm_id")
                entry["history"] = v.get("history") or []
                if entry["nm_id"] is not None:
                    by_nm[str(entry["nm_id"])] = entry
            by_vendor[vc] = entry
    # нормализуем meta
    out_v, out_n = {}, {}
    for vc, v in by_vendor.items():
        m = _cost_entry_meta(v)
        if m["cost"] is None:
            continue
        entry = {
            **m,
            "vendor_code": (v.get("vendor_code") if isinstance(v, dict) else None) or vc,
            "nm_id": v.get("nm_id") if isinstance(v, dict) else None,
        }
        out_v[_norm_vendor_key(vc)] = entry
        if entry.get("nm_id") is not None:
            out_n[str(entry["nm_id"])] = entry
    for nm, v in by_nm.items():
        m = _cost_entry_meta(v)
        if m["cost"] is None:
            continue
        entry = {
            **m,
            "vendor_code": (v.get("vendor_code") if isinstance(v, dict) else None) or "",
            "nm_id": int(nm) if str(nm).isdigit() else (v.get("nm_id") if isinstance(v, dict) else None),
        }
        out_n[str(nm)] = entry
        if entry["vendor_code"]:
            out_v.setdefault(_norm_vendor_key(entry["vendor_code"]), entry)
    return out_v, out_n

@app.get("/api/finance")
def get_finance():
    """Себестоимость остатков: WB + наш склад (актуальная цена на сегодня)."""
    by_vendor, by_nm = _load_cost_indexes()
    meta = get_setting_json(COST_META_KEY, {}) or {}
    nm_to_vendor = build_nm_to_vendor_map()
    # дополняем карту из файла себестоимости (если SKU был в Excel)
    for nm_s, entry in by_nm.items():
        try:
            nm = int(nm_s)
        except Exception:
            continue
        vc = _norm_vendor_key(entry.get("vendor_code"))
        if vc and vc != str(nm) and nm not in nm_to_vendor:
            nm_to_vendor[nm] = vc

    def resolve_cost(vendor_code=None, nm_id=None):
        vc = _norm_vendor_key(vendor_code)
        if vc and vc in by_vendor:
            return by_vendor[vc]
        if nm_id is not None and str(nm_id) in by_nm:
            return by_nm[str(nm_id)]
        # иногда в by_vendor ключ с другим регистром
        if vc:
            low = vc.lower()
            for k, e in by_vendor.items():
                if k.lower() == low:
                    return e
        return {}

    # WB остатки — vendor_code в stock_totals часто пустой, матчим через ratings/nm_id
    wb_rows = []
    try:
        st = httpx.get(
            f"{SUPABASE_URL}/rest/v1/stock_totals?select=nm_id,vendor_code,quantity_warehouses_full,in_way_to_client,in_way_from_client,subject_name",
            headers=sb_headers(), timeout=20,
        )
        if st.is_success:
            for r in st.json() or []:
                nm_id = r.get("nm_id")
                qty = int(r.get("quantity_warehouses_full") or 0)
                if qty <= 0:
                    continue
                try:
                    nm_int = int(nm_id) if nm_id is not None else None
                except Exception:
                    nm_int = None
                stock_vc = _norm_vendor_key(r.get("vendor_code"))
                if stock_vc and nm_int is not None and stock_vc == str(nm_int):
                    stock_vc = ""
                seller = (
                    stock_vc
                    or (nm_to_vendor.get(nm_int) if nm_int is not None else "")
                    or ""
                )
                seller = _norm_vendor_key(seller)
                cm = resolve_cost(seller, nm_id)
                # если себес нашли по nm — подтянем артикул из файла
                if not seller:
                    seller = _norm_vendor_key(cm.get("vendor_code")) or ""
                cost = cm.get("cost")
                value = round(qty * cost, 2) if cost is not None else None
                wb_rows.append({
                    "vendor_code": seller or (str(nm_id) if nm_id else ""),
                    "nm_id": nm_id,
                    "name": r.get("subject_name") or "",
                    "qty": qty,
                    "cost": cost,
                    "cost_default": cm.get("default"),
                    "cost_as_of": cm.get("as_of"),
                    "value": value,
                    "in_way": int(r.get("in_way_to_client") or 0) + int(r.get("in_way_from_client") or 0),
                })
    except Exception as e:
        logger.error(f"finance stock_totals: {e}")

    # Наш склад
    own = OWN_WAREHOUSE_CACHE.get("rows") or []
    if not own and not OWN_WAREHOUSE_CACHE.get("syncing"):
        try:
            refresh_own_warehouse_stock()
            own = OWN_WAREHOUSE_CACHE.get("rows") or []
        except Exception as e:
            logger.error(f"finance own-wh refresh: {e}")

    own_rows = []
    seen_own = set()
    for r in own:
        vc = _norm_vendor_key(r.get("vendor_code"))
        if not vc or vc in seen_own:
            continue
        seen_own.add(vc)
        qty = int(r.get("stock") or 0)
        if qty <= 0:
            continue
        cm = resolve_cost(vc, None)
        cost = cm.get("cost")
        value = round(qty * cost, 2) if cost is not None else None
        own_rows.append({
            "vendor_code": vc,
            "name": r.get("name") or "",
            "qty": qty,
            "cost": cost,
            "cost_default": cm.get("default"),
            "cost_as_of": cm.get("as_of"),
            "value": value,
            "family_stock": r.get("family_stock"),
        })

    def summarize(rows):
        with_cost = [x for x in rows if x.get("value") is not None]
        without = [x for x in rows if x.get("value") is None]
        return {
            "total_value": round(sum(x["value"] for x in with_cost), 2),
            "total_qty": sum(x["qty"] for x in rows),
            "qty_with_cost": sum(x["qty"] for x in with_cost),
            "qty_without_cost": sum(x["qty"] for x in without),
            "articles": len(rows),
            "articles_without_cost": len(without),
        }

    wb_sum = summarize(wb_rows)
    own_sum = summarize(own_rows)
    wb_rows.sort(key=lambda x: (-(x["value"] or 0), str(x["vendor_code"])))
    own_rows.sort(key=lambda x: (-(x["value"] or 0), str(x["vendor_code"])))

    return {
        "costs_count": len(by_vendor),
        "nm_mapped": len(by_nm),
        "meta": meta,
        "wb": {**wb_sum, "rows": wb_rows},
        "own": {
            **own_sum,
            "rows": own_rows,
            "as_of": OWN_WAREHOUSE_CACHE.get("as_of"),
            "updated_at": OWN_WAREHOUSE_CACHE.get("updated_at"),
        },
        "grand_total": round(wb_sum["total_value"] + own_sum["total_value"], 2),
        "costs": sorted(
            [
                {
                    "vendor_code": e.get("vendor_code") or vc,
                    "nm_id": e.get("nm_id"),
                    "cost": e.get("cost"),
                    "default": e.get("default"),
                    "as_of": e.get("as_of"),
                    "manual": bool(e.get("manual")),
                }
                for vc, e in by_vendor.items()
            ],
            key=lambda x: str(x.get("vendor_code") or ""),
        ),
    }

@app.post("/api/finance/cost")
async def save_finance_cost(request: dict):
    """Ручное изменение себестоимости по артикулу продавца и/или nm_id."""
    vc = _norm_vendor_key(request.get("vendor_code"))
    nm_raw = request.get("nm_id")
    nm_id = None
    if nm_raw not in (None, ""):
        try:
            nm_id = int(float(str(nm_raw).strip()))
        except Exception:
            return {"error": "Некорректный nm_id"}
    cost = _parse_cost_number(request.get("cost"))
    if cost is None:
        return {"error": "Укажи себестоимость числом ≥ 0"}
    if not vc and nm_id is None:
        return {"error": "Нужен артикул продавца или nm_id"}

    raw = get_setting_json(COST_PRICES_KEY, {}) or {}
    if not isinstance(raw, dict):
        raw = {}
    if raw.get("_v") == 2 or "by_vendor" in raw or "by_nm" in raw:
        by_vendor = dict(raw.get("by_vendor") or {})
        by_nm = dict(raw.get("by_nm") or {})
    else:
        by_vendor, by_nm = {}, {}
        for k, v in raw.items():
            if str(k).startswith("_"):
                continue
            key = _norm_vendor_key(k)
            if not key:
                continue
            m = _cost_entry_meta(v)
            entry = {
                **m,
                "vendor_code": key,
                "nm_id": v.get("nm_id") if isinstance(v, dict) else None,
                "history": v.get("history") if isinstance(v, dict) else [],
            }
            by_vendor[key] = entry
            if entry.get("nm_id") is not None:
                by_nm[str(entry["nm_id"])] = entry

    # найти существующую запись
    prev = None
    if vc and vc in by_vendor:
        prev = by_vendor[vc]
    elif nm_id is not None and str(nm_id) in by_nm:
        prev = by_nm[str(nm_id)]
        if not vc:
            vc = _norm_vendor_key(prev.get("vendor_code"))

    today = datetime.now(timezone.utc).date().isoformat()
    prev_default = None
    prev_history = []
    if isinstance(prev, dict):
        prev_default = _parse_cost_number(prev.get("default"))
        prev_history = list(prev.get("history") or [])
        if not vc:
            vc = _norm_vendor_key(prev.get("vendor_code"))
        if nm_id is None and prev.get("nm_id") is not None:
            nm_id = prev.get("nm_id")

    if not vc:
        vc = f"nm_{nm_id}" if nm_id is not None else ""
    if prev_default is None:
        prev_default = cost

    entry = {
        "cost": round(cost, 4),
        "default": round(prev_default, 4) if prev_default is not None else round(cost, 4),
        "as_of": today,
        "vendor_code": vc,
        "nm_id": nm_id,
        "manual": True,
        "history": prev_history + [{"date": today, "cost": round(cost, 4), "manual": True}],
    }
    by_vendor[vc] = entry
    if nm_id is not None:
        by_nm[str(nm_id)] = entry

    save_setting_value(COST_PRICES_KEY, {"_v": 2, "by_vendor": by_vendor, "by_nm": by_nm})
    meta = get_setting_json(COST_META_KEY, {}) or {}
    if isinstance(meta, dict):
        meta["last_manual_edit"] = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")
        meta["total_articles"] = len(by_vendor)
        save_setting_value(COST_META_KEY, meta)

    return {"status": "ok", "entry": entry}


# ---------- Финансы: CFO баланс / кредиты ----------
CFO_SNAPSHOT_KEY = "cfo_snapshot"

DEFAULT_CFO_SNAPSHOT = {
    "as_of": "",
    "cash": 0,
    "suppliers": 0,
    "inventory_wb_own": 0,
    "inventory_transit": 0,
    "inventory_ozon": 0,
    "salary_month": 0,
    "realization_month": 0,
    "target_margin": 0.15,
    "cash_floor": 0,
    "wb_compensation_pending": 0,
    "wb_receivables": [],
    "ozon_receivables": [],
    "pnl_wb": 0,
    "pnl_ozon": 0,
    "loans": [],
}


def _cfo_num(v, default=0.0):
    try:
        if v is None or v == "":
            return float(default)
        return float(v)
    except (TypeError, ValueError):
        return float(default)


def _cfo_interest_month(loan: dict) -> float:
    if loan.get("interest_only"):
        return _cfo_num(loan.get("payment"))
    # факт по ВТБ крупному из июня
    if loan.get("id") == "vtb_big":
        return 88000.0 + _cfo_num(loan.get("fee_month"))
    rate = _cfo_num(loan.get("rate"))
    bal = _cfo_num(loan.get("balance"))
    fee = _cfo_num(loan.get("fee_month"))
    interest = (bal * rate / 12.0) if (rate > 0 and bal > 0) else 0.0
    return interest + fee


def _migrate_cfo_loans_jul30(data: dict) -> tuple:
    """Разово: закрыт Сбер#1, Сбер#2 −2.8млн/платёж 40к, линия 6 млн,
    новая дебиторка WB и долг поставщикам 21 344 400."""
    changed = False
    loans = data.get("loans")
    if isinstance(loans, list) and loans:
        ids = {l.get("id") for l in loans if isinstance(l, dict)}
        sber1 = next((l for l in loans if isinstance(l, dict) and l.get("id") == "sber1"), None)
        need_loans = ("line6m" not in ids) and sber1 is not None and _cfo_num(sber1.get("payment")) >= 150000
        if need_loans:
            new_loans = []
            for l in loans:
                if not isinstance(l, dict):
                    continue
                if l.get("id") == "sber1":
                    continue
                item = dict(l)
                if item.get("id") == "sber2":
                    bal = _cfo_num(item.get("balance"))
                    if bal >= 2_500_000:
                        item["balance"] = round(bal - 2_800_000, 2)
                    item["payment"] = 40000
                    item["notes"] = "после досрочки −2.8 млн, платёж 40к"
                item["fee_month"] = _cfo_num(item.get("fee_month"))
                new_loans.append(item)
            if not any(l.get("id") == "line6m" for l in new_loans):
                line = {
                    "id": "line6m",
                    "name": "Кредитная линия 6 млн",
                    "contract": "",
                    "balance": 6_000_000,
                    "rate": 0.189,
                    "payment": 243600,
                    "fee_month": 24000,
                    "close": "2029-07",
                    "early_repay": "",
                    "interest_only": False,
                    "notes": "18.9%/год + 0.4% лимита (24к). 36 мес. Переплата 2.77 млн (46.2%)",
                }
                idx = next((i for i, l in enumerate(new_loans) if l.get("id") == "sber2"), -1)
                new_loans.insert(idx + 1, line)
            data["loans"] = new_loans
            changed = True

    # дебиторка/поставщики — только для уже существующего среза Ярослава (старые значения)
    new_wb = [5719189.43, 30441.68, 5963238.65, 6034075.99, 21484.81, 4382878.38, 31802.27]
    new_suppliers = 21344400
    old_suppliers = _cfo_num(data.get("suppliers"))
    old_wb = data.get("wb_receivables") or []
    looks_like_legacy = (
        data.get("cfo_recv_jul30") != True
        and (
            abs(old_suppliers - 22_680_000) < 1
            or (isinstance(old_wb, list) and any(abs(_cfo_num(x) - 5_400_759.27) < 1 for x in old_wb))
            or (isinstance(loans, list) and any(isinstance(l, dict) and l.get("id") == "sber1" for l in loans))
        )
    )
    if looks_like_legacy:
        data["wb_receivables"] = list(new_wb)
        data["suppliers"] = new_suppliers
        data["cfo_recv_jul30"] = True
        changed = True

    if changed:
        data["as_of"] = data.get("as_of") or "2026-07-30"
    return data, changed


def enrich_cfo_snapshot(raw: dict) -> dict:
    data = {**DEFAULT_CFO_SNAPSHOT, **(raw or {})}
    data.pop("personal", None)
    if not isinstance(data.get("loans"), list) or not data["loans"]:
        data["loans"] = [dict(x) for x in DEFAULT_CFO_SNAPSHOT["loans"]]
    else:
        data, _ = _migrate_cfo_loans_jul30(data)

    loans = []
    for i, loan in enumerate(data["loans"]):
        if not isinstance(loan, dict):
            continue
        item = dict(loan)
        item["id"] = item.get("id") or f"loan_{i}"
        item["balance"] = _cfo_num(item.get("balance"))
        item["rate"] = _cfo_num(item.get("rate"))
        item["payment"] = _cfo_num(item.get("payment"))
        item["fee_month"] = _cfo_num(item.get("fee_month"))
        item["interest_month"] = round(_cfo_interest_month(item), 2)
        item["principal_month"] = round(max(0.0, item["payment"] - item["interest_month"]), 2)
        loans.append(item)
    data["loans"] = loans

    wb_recv = [_cfo_num(x) for x in (data.get("wb_receivables") or [])]
    oz_recv = [_cfo_num(x) for x in (data.get("ozon_receivables") or [])]
    data["wb_receivables"] = wb_recv
    data["ozon_receivables"] = oz_recv

    cash = _cfo_num(data.get("cash"))
    suppliers = _cfo_num(data.get("suppliers"))
    inv_wb = _cfo_num(data.get("inventory_wb_own"))
    inv_tr = _cfo_num(data.get("inventory_transit"))
    inv_oz = _cfo_num(data.get("inventory_ozon"))
    stock = inv_wb + inv_tr + inv_oz
    bank = sum(_cfo_num(l.get("balance")) for l in loans)
    bank_pay = sum(_cfo_num(l.get("payment")) for l in loans)
    bank_int = sum(_cfo_num(l.get("interest_month")) for l in loans)
    mp_recv = sum(wb_recv) + sum(oz_recv)
    assets = cash + mp_recv + stock
    liabilities = suppliers + bank
    pnl_wb = _cfo_num(data.get("pnl_wb"))
    pnl_oz = _cfo_num(data.get("pnl_ozon"))
    salary = _cfo_num(data.get("salary_month"))
    realization = _cfo_num(data.get("realization_month"))
    target_margin = _cfo_num(data.get("target_margin"), 0.15) or 0.15
    cash_floor = _cfo_num(data.get("cash_floor"), 4000000)
    wb_comp = _cfo_num(data.get("wb_compensation_pending"))
    pnl_channels = pnl_wb + pnl_oz
    pnl_real = pnl_channels - salary - bank_int
    # прибыль до процентов (после ЗП) — для coverage
    ebit_like = pnl_channels - salary
    avg_rate = (sum(_cfo_num(l.get("balance")) * _cfo_num(l.get("rate")) for l in loans) / bank) if bank else 0.0
    margin_channels = (pnl_channels / realization) if realization else None
    margin_real = (pnl_real / realization) if realization else None
    interest_coverage = (ebit_like / bank_int) if bank_int > 0 else None
    # рычаги
    leverage_bank_months = (bank / pnl_real) if pnl_real > 0 else None
    debt_to_assets = (liabilities / assets) if assets else None
    equity = assets - liabilities
    # красные линии при целевой марже
    min_realiz_cover_pct = (salary + bank_int) / target_margin if target_margin else None
    min_realiz_cover_cash = (salary + bank_pay) / target_margin if target_margin else None
    min_pnl_to_interest = salary + bank_int
    min_pnl_to_debt_service = salary + bank_pay

    def _status(ok: bool, warn: bool = False) -> str:
        if ok:
            return "ok"
        if warn:
            return "warn"
        return "bad"

    health = {
        "avg_rate": round(avg_rate, 4),
        "avg_rate_pct": round(avg_rate * 100, 2),
        "interest_coverage": round(interest_coverage, 2) if interest_coverage is not None else None,
        "margin_channels": round(margin_channels, 4) if margin_channels is not None else None,
        "margin_channels_pct": round(margin_channels * 100, 2) if margin_channels is not None else None,
        "margin_real": round(margin_real, 4) if margin_real is not None else None,
        "margin_real_pct": round(margin_real * 100, 2) if margin_real is not None else None,
        "leverage_bank_months": round(leverage_bank_months, 2) if leverage_bank_months is not None else None,
        "debt_to_assets": round(debt_to_assets, 3) if debt_to_assets is not None else None,
        "target_margin": target_margin,
        "target_margin_pct": round(target_margin * 100, 2),
        "cash_floor": cash_floor,
        "min_realization_for_salary_interest": round(min_realiz_cover_pct, 0) if min_realiz_cover_pct else None,
        "min_realization_for_debt_service": round(min_realiz_cover_cash, 0) if min_realiz_cover_cash else None,
        "min_pnl_channels_salary_interest": round(min_pnl_to_interest, 0),
        "min_pnl_channels_debt_service": round(min_pnl_to_debt_service, 0),
        "floors": {
            "cash": {"value": cash, "floor": cash_floor, "status": _status(cash >= cash_floor, cash >= cash_floor * 0.75)},
            "coverage": {
                "value": interest_coverage,
                "floor": 3.0,
                "status": _status(
                    interest_coverage is not None and interest_coverage >= 3,
                    interest_coverage is not None and interest_coverage >= 1.5,
                ),
            },
            "current_ratio": {
                "value": round(assets / liabilities, 3) if liabilities else None,
                "floor": 1.1,
                "status": _status(
                    liabilities > 0 and assets / liabilities >= 1.1,
                    liabilities > 0 and assets / liabilities >= 1.0,
                ),
            },
            "margin": {
                "value": margin_channels,
                "floor": target_margin,
                "status": _status(
                    margin_channels is not None and margin_channels >= target_margin,
                    margin_channels is not None and margin_channels >= target_margin * 0.85,
                ),
            },
            "avg_rate": {
                "value": avg_rate,
                "floor": 0.25,
                "status": _status(avg_rate <= 0.25, avg_rate <= 0.30),
            },
            "realization": {
                "value": realization,
                "floor": min_realiz_cover_pct,
                "status": _status(
                    min_realiz_cover_pct is not None and realization >= min_realiz_cover_pct,
                    min_realiz_cover_pct is not None and realization >= min_realiz_cover_pct * 0.85,
                ),
            },
        },
    }

    data["totals"] = {
        "cash": cash,
        "wb_receivable": round(sum(wb_recv), 2),
        "ozon_receivable": round(sum(oz_recv), 2),
        "mp_receivable": round(mp_recv, 2),
        "inventory": round(stock, 2),
        "assets": round(assets, 2),
        "suppliers": suppliers,
        "bank_debt": round(bank, 2),
        "liabilities": round(liabilities, 2),
        "equity": round(equity, 2),
        "bank_payment": round(bank_pay, 2),
        "bank_interest": round(bank_int, 2),
        "bank_principal": round(bank_pay - bank_int, 2),
        "pnl_channels": round(pnl_channels, 2),
        "pnl_real": round(pnl_real, 2),
        "ebit_like": round(ebit_like, 2),
        "realization": round(realization, 2),
        "wb_compensation_pending": round(wb_comp, 2),
        "current_ratio": round(assets / liabilities, 3) if liabilities else None,
        "health": health,
    }
    return data


WB_MONEY_STORE_KEY = "wb_money_store"


def _wb_money(v):
    """WB finance API часто отдаёт деньги строкой."""
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace("\xa0", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def _money_store() -> dict:
    raw = get_setting_json(WB_MONEY_STORE_KEY, None)
    if not isinstance(raw, dict):
        return {"reports": [], "payments": [], "marks": {}, "updated_at": None}
    return {
        "reports": list(raw.get("reports") or []),
        "payments": list(raw.get("payments") or []),
        "marks": dict(raw.get("marks") or {}),
        "updated_at": raw.get("updated_at"),
        "balance": raw.get("balance"),
    }


def _save_money_store(store: dict) -> bool:
    store = dict(store or {})
    store["updated_at"] = datetime.now(timezone.utc).isoformat()
    return save_setting_value(WB_MONEY_STORE_KEY, store)


def _report_key(r: dict) -> str:
    rid = r.get("report_id") or r.get("id") or ""
    return "|".join([
        str(r.get("date_from") or "")[:10],
        str(r.get("date_to") or "")[:10],
        str(r.get("type") or ""),
        str(rid),
    ])


def _pick_report_amount(money: dict, fallback=None):
    """Сумма «Итого к оплате» из отчёта реализации.

    В кабинете ВБ это bankPaymentSum (не retailAmountSum = продажа
    и не forPaySum = к перечислению до удержаний).
    """
    if not isinstance(money, dict):
        return fallback
    preferred = (
        "bankPaymentSum", "bank_payment_sum",
        "totalToPay", "total_to_pay", "Итого к оплате", "итог_к_оплате",
        "forPaySum", "for_pay_sum", "toPay", "forPay", "ppvz_for_pay",
        "paid_sum", "transferAmount",
    )
    for k in preferred:
        if k in money and money[k] is not None:
            try:
                return float(money[k])
            except (TypeError, ValueError):
                continue
    return fallback


def _report_type_label(typ) -> str:
    if typ is None or typ == "":
        return "Отчёт"
    # finance-api: 1 = основной, 2 = по выкупам
    try:
        n = int(typ)
        if n == 1:
            return "Основной"
        if n == 2:
            return "По выкупам"
    except (TypeError, ValueError):
        pass
    s = str(typ).strip()
    return s or "Отчёт"


def _normalize_report_row(row: dict) -> dict:
    """Пересчитывает amount из money (bankPaymentSum) и подписи типа."""
    if not isinstance(row, dict):
        return row
    row = dict(row)
    money = row.get("money") if isinstance(row.get("money"), dict) else {}
    sale = _wb_money(money.get("retailAmountSum"))
    for_pay = _wb_money(money.get("forPaySum"))
    to_pay = _pick_report_amount(money, row.get("amount"))
    if sale is not None:
        row["sale_amount"] = float(sale)
    if for_pay is not None:
        row["for_pay_amount"] = float(for_pay)
    if to_pay is not None:
        row["amount"] = float(to_pay)
    row["type"] = _report_type_label(row.get("type"))
    row["key"] = row.get("key") or _report_key(row)
    return row


def _match_payment_status(report: dict, payments: list, marks: dict) -> dict:
    """Сверяет отчёт с историей платежей и ручными отметками."""
    key = report.get("key") or _report_key(report)
    amount = report.get("amount")
    manual = (marks or {}).get(key)
    if manual in ("paid", "unpaid", "processing", "partial"):
        return {
            "payment_status": manual,
            "payment_source": "manual",
            "matched_payment": None,
            "key": key,
        }

    if amount is None or not payments:
        return {
            "payment_status": "unknown",
            "payment_source": None,
            "matched_payment": None,
            "key": key,
        }

    best = None
    best_diff = None
    for p in payments:
        try:
            pam = float(p.get("amount") or 0)
        except (TypeError, ValueError):
            continue
        diff = abs(pam - float(amount))
        if best_diff is None or diff < best_diff:
            best_diff = diff
            best = p

    if best is None or best_diff is None:
        status = "unpaid"
        matched = None
    elif best_diff <= 1.0:
        st = (best.get("status") or "").lower()
        if best.get("paid_at") or st in ("paid", "done", "success", "оплачено"):
            status = "paid"
        elif st in ("processing", "queue", "pending", "обрабатывается", "очередь"):
            status = "processing"
        else:
            status = "processing" if not best.get("paid_at") else "paid"
        matched = best
    elif best_diff <= max(5000.0, abs(float(amount)) * 0.01):
        # близко — считаем near-match, статус по платежу
        st = (best.get("status") or "").lower()
        if best.get("paid_at") or st in ("paid", "done", "success", "оплачено"):
            status = "paid"
        else:
            status = "processing"
        matched = {**best, "near_match_diff": round(best_diff, 2)}
    else:
        status = "unpaid"
        matched = None

    return {
        "payment_status": status,
        "payment_source": "payment_match" if matched else None,
        "matched_payment": matched,
        "key": key,
    }


def fetch_wb_account_balance() -> dict:
    """Виджет баланса с главной seller.wildberries.ru."""
    if not WB_TOKEN:
        return {"error": "WB_TOKEN не задан"}
    try:
        r = httpx.get(
            f"{WB_FINANCE_URL}/api/v1/account/balance",
            headers=wb_headers(),
            timeout=30,
        )
    except Exception as e:
        return {"error": str(e)}
    if r.status_code != 200:
        return {"error": f"HTTP {r.status_code}", "body": r.text[:400]}
    data = r.json() if r.content else {}
    if isinstance(data, dict) and isinstance(data.get("data"), dict):
        data = data["data"]
    return {
        "currency": data.get("currency") or "RUB",
        "current": _wb_money(data.get("current")),
        "for_withdraw": _wb_money(data.get("for_withdraw")),
        "raw_keys": sorted(data.keys()) if isinstance(data, dict) else [],
    }


def fetch_wb_sales_reports(date_from: str, date_to: str) -> dict:
    """Список еженедельных отчётов реализации (Финансы → отчёты)."""
    if not WB_TOKEN:
        return {"error": "WB_TOKEN не задан", "reports": []}
    payload = {"dateFrom": date_from, "dateTo": date_to}
    try:
        r = httpx.post(
            f"{WB_FINANCE_URL}/api/finance/v1/sales-reports/list",
            headers={**wb_headers(), "Content-Type": "application/json"},
            json=payload,
            timeout=60,
        )
    except Exception as e:
        return {"error": str(e), "reports": []}
    if r.status_code == 204:
        return {"reports": [], "note": "пусто за период"}
    if r.status_code != 200:
        return {"error": f"HTTP {r.status_code}", "body": r.text[:500], "reports": []}
    raw = r.json() if r.content else []
    items = raw
    if isinstance(raw, dict):
        items = raw.get("data") or raw.get("reports") or raw.get("list") or []
    if not isinstance(items, list):
        return {"error": "неожиданный формат", "raw_type": str(type(raw)), "reports": []}

    reports = []
    for it in items:
        if not isinstance(it, dict):
            continue
        money_fields = {}
        for k, v in it.items():
            mv = _wb_money(v)
            if mv is not None and any(
                x in str(k).lower()
                for x in ("pay", "sum", "total", "amount", "transfer", "оплат", "перечисл", "sale", "retail")
            ):
                money_fields[str(k)] = mv
        date_from_v = it.get("dateFrom") or it.get("date_from") or it.get("begin")
        date_to_v = it.get("dateTo") or it.get("date_to") or it.get("end")
        typ = it.get("reportType") or it.get("type") or it.get("category") or "Отчёт"
        rid = it.get("reportId") or it.get("report_id") or it.get("id")
        row = {
            "report_id": rid,
            "date_from": str(date_from_v)[:10] if date_from_v else None,
            "date_to": str(date_to_v)[:10] if date_to_v else None,
            "created": it.get("createDate") or it.get("createdAt") or it.get("created"),
            "type": typ,
            "api_status": it.get("status") or it.get("paymentStatus") or it.get("state"),
            "money": money_fields,
            "source": "wb_api",
        }
        # key до смены type→лейбл, чтобы стабильно матчить кэш
        row["key"] = _report_key(row)
        row = _normalize_report_row(row)
        reports.append(row)
    return {"reports": reports, "count": len(reports)}


def parse_wb_weekly_pay_excel(content: bytes) -> list:
    """Excel «Еженедельный отчет …» с колонкой «Итого к оплате»."""
    import io
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl не установлен")
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = {}
    for c in range(1, (ws.max_column or 1) + 1):
        h = ws.cell(1, c).value
        if h is None:
            continue
        headers[str(h).strip().lower()] = c

    def col(*names):
        for n in names:
            if n.lower() in headers:
                return headers[n.lower()]
        return None

    c_id = col("№ отчета", "номер отчета", "report id", "id")
    c_from = col("дата начала", "date from", "from")
    c_to = col("дата конца", "дата окончания", "date to", "to")
    c_type = col("тип отчета", "тип", "type")
    c_pay = col("итого к оплате", "к оплате", "total to pay")
    if not c_from or not c_to or not c_pay:
        raise RuntimeError("Не нашёл колонки «Дата начала/конца» и «Итого к оплате»")

    def cell_date(v):
        if v is None:
            return None
        if hasattr(v, "date"):
            try:
                return v.date().isoformat()
            except Exception:
                pass
        s = str(v)
        if "T" in s:
            return s[:10]
        if len(s) >= 10 and s[4] == "-":
            return s[:10]
        # 01.06.2026
        try:
            return datetime.strptime(s[:10], "%d.%m.%Y").date().isoformat()
        except Exception:
            return s[:10]

    out = []
    for r in range(2, (ws.max_row or 1) + 1):
        amount = _wb_money(ws.cell(r, c_pay).value)
        if amount is None:
            continue
        dfrom = cell_date(ws.cell(r, c_from).value)
        dto = cell_date(ws.cell(r, c_to).value)
        typ = ws.cell(r, c_type).value if c_type else "Основной"
        rid = ws.cell(r, c_id).value if c_id else None
        row = {
            "report_id": str(rid) if rid is not None else None,
            "date_from": dfrom,
            "date_to": dto,
            "type": str(typ or "Основной"),
            "amount": float(amount),
            "source": "excel",
        }
        row["key"] = _report_key(row)
        out.append(row)
    wb.close()
    return out


def _merge_reports(existing: list, incoming: list) -> list:
    by = {}
    for r in existing or []:
        if isinstance(r, dict) and r.get("key"):
            by[r["key"]] = r
    for r in incoming or []:
        if not isinstance(r, dict):
            continue
        key = r.get("key") or _report_key(r)
        r = {**r, "key": key}
        prev = by.get(key) or {}
        # excel amount предпочтительнее, если api без суммы
        money = r.get("money") if isinstance(r.get("money"), dict) else None
        if not money and isinstance(prev.get("money"), dict):
            money = prev.get("money")
        amount = r.get("amount")
        if money:
            picked = _pick_report_amount(money)
            if picked is not None:
                amount = picked
        if amount is None:
            amount = prev.get("amount")
        elif prev.get("source") == "excel" and r.get("source") == "wb_api" and prev.get("amount") is not None:
            # Excel «Итого к оплате» важнее, если API ещё без bankPaymentSum
            if not money or _pick_report_amount(money) is None:
                amount = prev.get("amount")
                r = {**r, "source": "excel+api"}
            else:
                # оба есть — API bankPaymentSum точнее для сверки с платежами
                r = {**r, "source": "excel+api"}
        merged = {**prev, **r, "amount": amount, "key": key}
        if money:
            merged["money"] = money
        by[key] = _normalize_report_row(merged)
    rows = list(by.values())
    rows.sort(key=lambda x: (str(x.get("date_from") or ""), str(x.get("type") or "")), reverse=True)
    return rows


def _enrich_reports_with_payments(store: dict) -> dict:
    reports = []
    payments = list(store.get("payments") or [])
    marks = store.get("marks") or {}
    # жадное сопоставление: каждый платёж — максимум к одному отчёту
    used_pay_idx = set()
    pending = []
    for r in store.get("reports") or []:
        if not isinstance(r, dict):
            continue
        row = _normalize_report_row(r)
        manual = marks.get(row["key"])
        if manual in ("paid", "unpaid", "processing", "partial"):
            row.update({
                "payment_status": manual,
                "payment_source": "manual",
                "matched_payment": None,
            })
            reports.append(row)
        else:
            pending.append(row)

    # сначала точные совпадения (±1₽), потом near-match
    def try_match(rows, max_diff_fn):
        for row in rows:
            amount = row.get("amount")
            if amount is None:
                continue
            best_i = None
            best_diff = None
            for i, p in enumerate(payments):
                if i in used_pay_idx:
                    continue
                try:
                    pam = float(p.get("amount") or 0)
                except (TypeError, ValueError):
                    continue
                diff = abs(pam - float(amount))
                limit = max_diff_fn(float(amount))
                if diff > limit:
                    continue
                if best_diff is None or diff < best_diff:
                    best_diff = diff
                    best_i = i
            if best_i is None:
                continue
            p = payments[best_i]
            used_pay_idx.add(best_i)
            st = (p.get("status") or "").lower()
            if p.get("paid_at") or st in ("paid", "done", "success", "оплачено"):
                status = "paid"
            elif st in ("processing", "queue", "pending", "обрабатывается", "очередь"):
                status = "processing"
            else:
                status = "processing" if not p.get("paid_at") else "paid"
            matched = dict(p)
            if best_diff is not None and best_diff > 1.0:
                matched["near_match_diff"] = round(best_diff, 2)
            row.update({
                "payment_status": status,
                "payment_source": "payment_match",
                "matched_payment": matched,
            })

    # exact
    try_match(pending, lambda a: 1.0)
    # near
    remaining = [r for r in pending if "payment_status" not in r]
    try_match(remaining, lambda a: max(5000.0, abs(a) * 0.01))

    for row in pending:
        if "payment_status" not in row:
            if row.get("amount") is None:
                row.update({"payment_status": "unknown", "payment_source": None, "matched_payment": None})
            else:
                row.update({"payment_status": "unpaid", "payment_source": None, "matched_payment": None})
        reports.append(row)

    reports.sort(key=lambda x: (str(x.get("date_from") or ""), str(x.get("type") or "")), reverse=True)

    def _sum(status):
        return round(sum(float(r.get("amount") or 0) for r in reports if r.get("payment_status") == status), 2)

    return {
        "reports": reports,
        "summary": {
            "count": len(reports),
            "paid": _sum("paid"),
            "processing": _sum("processing"),
            "partial": _sum("partial"),
            "unpaid": _sum("unpaid"),
            "unknown": _sum("unknown"),
            "total_amount": round(sum(float(r.get("amount") or 0) for r in reports), 2),
        },
        "payments": payments,
        "marks": marks,
        "updated_at": store.get("updated_at"),
        "balance": store.get("balance"),
    }


def _wb_money_default_period():
    """По умолчанию — с 1 января текущего (МСК) года по сегодня."""
    today = _msk_now().date()
    return today.replace(month=1, day=1).isoformat(), today.isoformat()


def fetch_wb_sales_reports_range(date_from: str, date_to: str, chunk_days: int = 100) -> dict:
    """Тянет отчёты кусками — у list-API иногда режется длинный период."""
    try:
        d0 = datetime.strptime(str(date_from)[:10], "%Y-%m-%d").date()
        d1 = datetime.strptime(str(date_to)[:10], "%Y-%m-%d").date()
    except Exception:
        return fetch_wb_sales_reports(date_from, date_to)

    if d1 < d0:
        d0, d1 = d1, d0

    all_rows = []
    errors = []
    cur = d0
    while cur <= d1:
        end = min(cur + timedelta(days=chunk_days - 1), d1)
        part = fetch_wb_sales_reports(cur.isoformat(), end.isoformat())
        if part.get("error"):
            errors.append(f"{cur}…{end}: {part.get('error')}")
        else:
            all_rows.extend(part.get("reports") or [])
        cur = end + timedelta(days=1)

    merged = _merge_reports([], all_rows)
    out = {"reports": merged, "count": len(merged)}
    if errors and not merged:
        out["error"] = "; ".join(errors[:3])
    elif errors:
        out["note"] = "; ".join(errors[:3])
    return out


@app.get("/api/wb-money")
def get_wb_money(date_from: str = None, date_to: str = None, refresh: bool = False):
    """Отчёты ВБ + статус оплаты (сверка с историей платежей)."""
    store = _money_store()
    def_from, def_to = _wb_money_default_period()
    if not date_to:
        date_to = def_to
    if not date_from:
        date_from = def_from

    balance = store.get("balance")
    api_err = None
    if refresh or not store.get("reports"):
        balance = fetch_wb_account_balance()
        api = fetch_wb_sales_reports_range(date_from, date_to)
        if api.get("error"):
            api_err = api.get("error")
            if api.get("body"):
                api_err = f"{api_err}: {api.get('body')}"
        else:
            store["reports"] = _merge_reports(store.get("reports") or [], api.get("reports") or [])
        store["balance"] = balance
        _save_money_store(store)

    enriched = _enrich_reports_with_payments(store)
    # период в ответе — фактический охват отчётов, если шире запроса
    reps = enriched.get("reports") or []
    if reps:
        real_from = min((r.get("date_from") or date_from) for r in reps)
        real_to = max((r.get("date_to") or r.get("date_from") or date_to) for r in reps)
    else:
        real_from, real_to = date_from, date_to
    return {
        "as_of": _msk_now().strftime("%d.%m.%Y %H:%M"),
        "period": {"from": real_from, "to": real_to, "requested_from": date_from, "requested_to": date_to},
        "balance": balance or {},
        "api_error": api_err,
        **enriched,
        "payment_history_note": (
            "Историю платежей из кабинета API не отдаёт — добавь заявки вручную "
            "(сумма + статус), отчёты сверятся автоматически. "
            "Или загрузи Excel «Еженедельный отчет» с колонкой «Итого к оплате»."
        ),
    }


@app.post("/api/wb-money/sync")
def sync_wb_money(date_from: str = None, date_to: str = None):
    return get_wb_money(date_from=date_from, date_to=date_to, refresh=True)


@app.post("/api/wb-money/upload-reports")
async def upload_wb_money_reports(file: UploadFile = File(...)):
    content = await file.read()
    try:
        rows = parse_wb_weekly_pay_excel(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    store = _money_store()
    store["reports"] = _merge_reports(store.get("reports") or [], rows)
    _save_money_store(store)
    return {
        "status": "ok",
        "imported": len(rows),
        "total_reports": len(store["reports"]),
        "summary": _enrich_reports_with_payments(store)["summary"],
    }


@app.post("/api/wb-money/payments")
async def save_wb_money_payments(request: dict):
    """Сохранить заявки из «Истории платежей».

    body: {payments:[{id, amount, created, paid_at, status}], replace?: bool}
    status: paid | processing | queue
    """
    items = request.get("payments") if isinstance(request, dict) else None
    if not isinstance(items, list):
        raise HTTPException(status_code=400, detail="нужен payments: []")
    cleaned = []
    for p in items:
        if not isinstance(p, dict):
            continue
        amount = _wb_money(p.get("amount"))
        if amount is None:
            continue
        status = str(p.get("status") or "processing").strip().lower()
        if status in ("оплачено", "done", "success"):
            status = "paid"
        elif status in ("очередь", "в очереди", "поручение в очереди"):
            status = "queue"
        elif status in ("обрабатывается", "оплата обрабатывается"):
            status = "processing"
        cleaned.append({
            "id": str(p.get("id") or p.get("payment_id") or ""),
            "amount": float(amount),
            "created": str(p.get("created") or p.get("created_at") or "")[:32] or None,
            "paid_at": str(p.get("paid_at") or p.get("paid") or "")[:32] or None,
            "status": status,
        })
    store = _money_store()
    replace = bool(request.get("replace", True))
    if replace:
        store["payments"] = cleaned
    else:
        by_id = {p.get("id"): p for p in (store.get("payments") or []) if p.get("id")}
        for p in cleaned:
            if p.get("id"):
                by_id[p["id"]] = p
            else:
                store.setdefault("payments", []).append(p)
        store["payments"] = list(by_id.values())
    _save_money_store(store)
    return {
        "status": "ok",
        "payments": len(store["payments"]),
        "summary": _enrich_reports_with_payments(store)["summary"],
    }


@app.post("/api/wb-money/mark")
async def mark_wb_money_report(request: dict):
    """Ручная отметка отчёта: paid | unpaid | processing | partial | clear."""
    key = str((request or {}).get("key") or "").strip()
    status = str((request or {}).get("status") or "").strip().lower()
    if not key:
        raise HTTPException(status_code=400, detail="нужен key")
    store = _money_store()
    marks = dict(store.get("marks") or {})
    if status in ("", "clear", "none", "auto"):
        marks.pop(key, None)
    elif status in ("paid", "unpaid", "processing", "partial"):
        marks[key] = status
    else:
        raise HTTPException(status_code=400, detail="status: paid|unpaid|processing|partial|clear")
    store["marks"] = marks
    _save_money_store(store)
    return {"status": "ok", "key": key, "mark": marks.get(key), "summary": _enrich_reports_with_payments(store)["summary"]}


@app.get("/api/finance/cfo")
def get_finance_cfo():
    raw = get_setting_json(CFO_SNAPSHOT_KEY, None)
    if not raw or not isinstance(raw, dict):
        raw = dict(DEFAULT_CFO_SNAPSHOT)
        save_setting_value(CFO_SNAPSHOT_KEY, raw)
        return enrich_cfo_snapshot(raw)
    migrated, changed = _migrate_cfo_loans_jul30(dict(raw))
    if changed:
        # сохраняем очищенные loans без computed полей
        clean_loans = []
        for i, loan in enumerate(migrated.get("loans") or []):
            if not isinstance(loan, dict):
                continue
            clean_loans.append({
                "id": loan.get("id") or f"loan_{i}",
                "name": str(loan.get("name") or f"Кредит {i+1}"),
                "contract": str(loan.get("contract") or ""),
                "balance": _cfo_num(loan.get("balance")),
                "rate": _cfo_num(loan.get("rate")),
                "payment": _cfo_num(loan.get("payment")),
                "fee_month": _cfo_num(loan.get("fee_month")),
                "close": str(loan.get("close") or ""),
                "early_repay": str(loan.get("early_repay") or ""),
                "interest_only": bool(loan.get("interest_only")),
                "notes": str(loan.get("notes") or ""),
            })
        to_save = {**raw, "loans": clean_loans, "as_of": migrated.get("as_of") or raw.get("as_of")}
        if "wb_receivables" in migrated:
            to_save["wb_receivables"] = [_cfo_num(x) for x in (migrated.get("wb_receivables") or [])]
        if "suppliers" in migrated:
            to_save["suppliers"] = _cfo_num(migrated.get("suppliers"))
        if migrated.get("cfo_recv_jul30"):
            to_save["cfo_recv_jul30"] = True
        to_save["updated_at"] = datetime.now(timezone.utc).isoformat()
        to_save.pop("totals", None)
        to_save.pop("personal", None)
        save_setting_value(CFO_SNAPSHOT_KEY, to_save)
        raw = to_save
    return enrich_cfo_snapshot(raw)


@app.post("/api/finance/cfo")
async def save_finance_cfo(request: dict):
    if not isinstance(request, dict):
        raise HTTPException(status_code=400, detail="invalid body")
    # не сохраняем computed totals и личные платежи
    payload = {k: v for k, v in request.items() if k not in ("totals", "personal")}
    # нормализация чисел в loans / receivables
    loans_in = payload.get("loans")
    if isinstance(loans_in, list):
        clean_loans = []
        for i, loan in enumerate(loans_in):
            if not isinstance(loan, dict):
                continue
            clean_loans.append({
                "id": loan.get("id") or f"loan_{i}",
                "name": str(loan.get("name") or f"Кредит {i+1}"),
                "contract": str(loan.get("contract") or ""),
                "balance": _cfo_num(loan.get("balance")),
                "rate": _cfo_num(loan.get("rate")),
                "payment": _cfo_num(loan.get("payment")),
                "fee_month": _cfo_num(loan.get("fee_month")),
                "close": str(loan.get("close") or ""),
                "early_repay": str(loan.get("early_repay") or ""),
                "interest_only": bool(loan.get("interest_only")),
                "notes": str(loan.get("notes") or ""),
            })
        payload["loans"] = clean_loans
    for key in ("wb_receivables", "ozon_receivables"):
        if key in payload and isinstance(payload[key], list):
            payload[key] = [_cfo_num(x) for x in payload[key]]
    for key in ("cash", "suppliers", "inventory_wb_own", "inventory_transit", "inventory_ozon",
                "salary_month", "pnl_wb", "pnl_ozon", "realization_month", "target_margin",
                "cash_floor", "wb_compensation_pending"):
        if key in payload:
            payload[key] = _cfo_num(payload[key])
    payload["as_of"] = str(payload.get("as_of") or datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    if not save_setting_value(CFO_SNAPSHOT_KEY, payload):
        raise HTTPException(status_code=500, detail="save failed")
    return {"status": "ok", "snapshot": enrich_cfo_snapshot(payload)}


# ── Позиции в клиентском поиске WB ──────────────────────────────────────────
SEARCH_KEYWORDS_PATH = Path(__file__).resolve().parent / "data" / "search_keywords.json"

# dest — регион выдачи витрины WB (как у покупателя). Москва по умолчанию.
WB_SEARCH_CITIES = [
    {"id": "moscow", "name": "Москва", "dest": -1257786},
    {"id": "spb", "name": "Санкт-Петербург", "dest": -1124448},
    {"id": "kazan", "name": "Казань", "dest": -2133462},
    {"id": "ekb", "name": "Екатеринбург", "dest": -1113276},
    {"id": "nsk", "name": "Новосибирск", "dest": -140294},
    {"id": "nn", "name": "Нижний Новгород", "dest": -1190344},
    {"id": "krasnodar", "name": "Краснодар", "dest": -1221148},
    {"id": "rostov", "name": "Ростов-на-Дону", "dest": -1197210},
    {"id": "samara", "name": "Самара", "dest": -1235864},
    {"id": "chelyabinsk", "name": "Челябинск", "dest": -1382589},
]

_wb_search_lock = threading.Lock()
_wb_search_last_ts = 0.0
_wb_search_host_i = 0

WB_SEARCH_HOSTS = [
    "https://search.wb.ru/exactmatch/ru/common/v9/search",
    "https://u-search.wb.ru/exactmatch/ru/common/v9/search",
    "https://search.wb.ru/exactmatch/ru/common/v14/search",
    "https://u-search.wb.ru/exactmatch/ru/common/v18/search",
]


def _load_search_keywords():
    try:
        with open(SEARCH_KEYWORDS_PATH, encoding="utf-8") as f:
            data = json.load(f)
        keys = data.get("keywords") or []
        return {
            "source": data.get("source") or "",
            "updated": data.get("updated") or "",
            "keywords": keys,
            "count": len(keys),
        }
    except Exception as e:
        logger.error(f"search_keywords load: {e}")
        return {"source": "", "updated": "", "keywords": [], "count": 0, "error": str(e)}


def _wb_search_throttle(min_interval: float = 0.9):
    """Не долбим search.wb.ru — иначе 429."""
    global _wb_search_last_ts
    with _wb_search_lock:
        now = time.time()
        wait = min_interval - (now - _wb_search_last_ts)
        if wait > 0:
            time.sleep(wait)
        _wb_search_last_ts = time.time()


def _wb_search_next_host():
    global _wb_search_host_i
    with _wb_search_lock:
        host = WB_SEARCH_HOSTS[_wb_search_host_i % len(WB_SEARCH_HOSTS)]
        _wb_search_host_i += 1
        return host


def find_nm_in_wb_search(nm_id: int, query: str, dest: int, max_pages: int = 3):
    """Ищет nm_id в клиентской выдаче WB по запросу. Позиция с 1, None = не в топ max_pages*100."""
    query = (query or "").strip()
    if not query or not nm_id:
        return {"query": query, "position": None, "page": None, "total": None, "error": "bad input"}

    # Важно: значения заголовков должны быть latin-1/ascii — кириллица в Referer валит httpx
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
        ),
        "Accept": "*/*",
        "Accept-Language": "ru-RU,ru;q=0.9",
        "Origin": "https://www.wildberries.ru",
        "Referer": "https://www.wildberries.ru/",
    }
    last_total = None
    last_err = None
    max_pages = max(1, min(int(max_pages or 3), 10))

    try:
        with httpx.Client(timeout=30, headers=headers, follow_redirects=True) as client:
            for page in range(1, max_pages + 1):
                page_ok = False
                for attempt in range(5):
                    base = _wb_search_next_host()
                    _wb_search_throttle(0.85 + 0.15 * attempt)
                    try:
                        resp = client.get(
                            base,
                            params={
                                "appType": 1,
                                "curr": "rub",
                                "dest": dest,
                                "query": query,
                                "resultset": "catalog",
                                "sort": "popular",
                                "spp": 30,
                                "page": page,
                            },
                        )
                    except Exception as e:
                        last_err = str(e)[:120]
                        time.sleep(0.8 * (attempt + 1))
                        continue

                    if resp.status_code == 429:
                        last_err = "429"
                        time.sleep(2.0 * (attempt + 1))
                        continue
                    if not resp.is_success:
                        last_err = f"http {resp.status_code}"
                        time.sleep(0.5 * (attempt + 1))
                        continue

                    try:
                        data = resp.json()
                    except Exception as e:
                        last_err = f"json {e}"
                        time.sleep(0.4)
                        continue

                    products = data.get("products") or (data.get("data") or {}).get("products") or []
                    last_total = data.get("total")
                    if last_total is None:
                        last_total = (data.get("data") or {}).get("total")
                    page_ok = True
                    last_err = None

                    for i, p in enumerate(products):
                        pid = p.get("id") or p.get("nmId") or p.get("nmID")
                        if pid == nm_id or str(pid) == str(nm_id):
                            return {
                                "query": query,
                                "position": (page - 1) * 100 + i + 1,
                                "page": page,
                                "total": last_total,
                                "error": None,
                            }

                    if not products:
                        return {
                            "query": query,
                            "position": None,
                            "page": None,
                            "total": last_total,
                            "error": None,
                            "not_found": True,
                        }
                    break

                if not page_ok:
                    # не смогли получить страницу — дальше нет смысла
                    break

        return {
            "query": query,
            "position": None,
            "page": None,
            "total": last_total,
            "error": last_err,
            "not_found": last_err is None,
        }
    except Exception as e:
        logger.exception(f"find_nm_in_wb_search: {e}")
        return {
            "query": query,
            "position": None,
            "page": None,
            "total": None,
            "error": str(e)[:160],
        }


@app.get("/api/search-keywords")
def get_search_keywords():
    """Кураторский список ключей (аналитика) + города для dest."""
    data = _load_search_keywords()
    return {
        **data,
        "cities": WB_SEARCH_CITIES,
        "default_city": "moscow",
        "default_dest": -1257786,
    }


@app.post("/api/search-positions")
def search_positions(request: dict):
    """Позиции nm_id в клиентском поиске WB по списку запросов.
    Body: {nm_id, dest?, queries: [str], max_pages?}
    dest по умолчанию Москва (-1257786).
    Пачки маленькие — иначе WB отвечает 429.
    """
    if not isinstance(request, dict):
        raise HTTPException(status_code=400, detail="invalid body")
    try:
        nm_id = int(request.get("nm_id"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="nm_id required")

    try:
        dest = int(request.get("dest") if request.get("dest") is not None else -1257786)
    except (TypeError, ValueError):
        dest = -1257786

    max_pages = request.get("max_pages", 3)
    try:
        max_pages = int(max_pages)
    except (TypeError, ValueError):
        max_pages = 3

    queries = request.get("queries")
    if not isinstance(queries, list) or not queries:
        raise HTTPException(status_code=400, detail="queries required (non-empty list)")
    clean = []
    seen = set()
    for q in queries:
        s = str(q or "").strip()
        if not s:
            continue
        k = s.lower()
        if k in seen:
            continue
        seen.add(k)
        clean.append(s)
        if len(clean) >= 8:
            break

    results = []
    for q in clean:
        try:
            results.append(find_nm_in_wb_search(nm_id, q, dest, max_pages=max_pages))
        except Exception as e:
            logger.exception(f"search-positions query={q}: {e}")
            results.append({"query": q, "position": None, "page": None, "total": None, "error": str(e)[:160]})

    city_name = next((c["name"] for c in WB_SEARCH_CITIES if c["dest"] == dest), str(dest))
    return {
        "nm_id": nm_id,
        "dest": dest,
        "city": city_name,
        "max_pages": max_pages,
        "checked": len(results),
        "results": results,
    }


def wb_product_thumb_url(nm_id: int) -> str:
    """Публичный CDN превью карточки WB."""
    try:
        nm_id = int(nm_id)
    except (TypeError, ValueError):
        return ""
    vol = nm_id // 100000
    part = nm_id // 1000
    ranges = [
        143, 287, 431, 719, 1007, 1061, 1115, 1169, 1313, 1601, 1655, 1919,
        2045, 2189, 2405, 2621, 2837, 3053, 3269, 3485, 3701, 3917, 4133,
        4349, 4565, 4877, 5189, 5501, 5813, 6125, 6437, 6749, 7061, 7373,
        7685, 7997, 8309, 8741, 9173, 9605, 10373, 11141, 11909, 12677,
        13445, 14213,
    ]
    basket = 47
    for i, r in enumerate(ranges):
        if vol <= r:
            basket = i + 1
            break
    return f"https://basket-{basket:02d}.wbbasket.ru/vol{vol}/part{part}/{nm_id}/images/tm/1.webp"


def fetch_wb_card_brief(nm_id: int, dest: int = -1257786):
    """Краткие данные карточки (бренд/название) через клиентский card API."""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
        ),
        "Accept": "*/*",
        "Origin": "https://www.wildberries.ru",
        "Referer": f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
    }
    try:
        with httpx.Client(timeout=20, headers=headers, follow_redirects=True) as client:
            resp = client.get(
                "https://card.wb.ru/cards/v4/detail",
                params={"appType": 1, "curr": "rub", "dest": dest, "nm": nm_id},
            )
            if not resp.is_success:
                return None
            products = resp.json().get("products") or []
            if not products:
                return None
            p = products[0]
            price_info = _parse_client_product(p)
            return {
                "nm_id": p.get("id") or nm_id,
                "brand": p.get("brand") or "",
                "name": p.get("name") or "",
                "supplier": p.get("supplier") or "",
                "thumb": wb_product_thumb_url(p.get("id") or nm_id),
                "url": f"https://www.wildberries.ru/catalog/{p.get('id') or nm_id}/detail.aspx",
                "client_price": price_info.get("client_price"),
                "sale_price": price_info.get("client_basic"),
                "spp": _calc_spp(price_info.get("client_basic"), price_info.get("client_price")),
            }
    except Exception as e:
        logger.warning(f"fetch_wb_card_brief {nm_id}: {e}")
        return None


def fetch_wb_see_also_shelf(nm_id: int, dest: int = -1257786, limit: int = 15):
    """Полка «Смотрите также» у карточки (клиентский recom.wb.ru)."""
    limit = max(1, min(int(limit or 15), 30))
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
        ),
        "Accept": "*/*",
        "Origin": "https://www.wildberries.ru",
        "Referer": f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
    }
    # query=<nm> даёт полку see-also для этой карточки
    # spp не передаём — иначе WB подставит виртуальную скидку
    url = "https://recom.wb.ru/recom/ru/common/v8/search"
    params = {
        "appType": 1,
        "curr": "rub",
        "dest": dest,
        "resultset": "catalog",
        "query": str(nm_id),
        "suppressSpellcheck": "false",
    }
    last_err = None
    for attempt in range(4):
        try:
            with httpx.Client(timeout=25, headers=headers, follow_redirects=True) as client:
                resp = client.get(url, params=params)
            if resp.status_code == 429:
                last_err = "429"
                time.sleep(1.5 * (attempt + 1))
                continue
            if not resp.is_success:
                last_err = f"http {resp.status_code}"
                time.sleep(0.5)
                continue
            data = resp.json()
            products = data.get("products") or (data.get("data") or {}).get("products") or []
            items = []
            for i, p in enumerate(products[:limit], 1):
                pid = p.get("id") or p.get("nmId") or p.get("nmID")
                if not pid:
                    continue
                price_info = _parse_client_product(p)
                client_price = price_info.get("client_price")
                sale_price = price_info.get("client_basic")
                items.append({
                    "position": i,
                    "nm_id": pid,
                    "brand": p.get("brand") or "",
                    "name": p.get("name") or "",
                    "supplier": p.get("supplier") or "",
                    "rating": p.get("reviewRating") or p.get("rating"),
                    "feedbacks": p.get("feedbacks"),
                    "thumb": wb_product_thumb_url(pid),
                    "url": f"https://www.wildberries.ru/catalog/{pid}/detail.aspx",
                    "client_price": client_price,
                    "sale_price": sale_price,
                    "spp": _calc_spp(sale_price, client_price),
                })
            return {"items": items, "total": len(products), "error": None}
        except Exception as e:
            last_err = str(e)[:160]
            time.sleep(0.6 * (attempt + 1))
    return {"items": [], "total": 0, "error": last_err or "unknown"}


def _watch_shape(vendor_code: str = "", name: str = "", brand: str = "") -> str:
    """круглые / квадратные / неясно / skip (не часы)."""
    vc = str(vendor_code or "").strip()
    blob = f"{vc} {name or ''} {brand or ''}".lower().replace("ё", "е")
    vc_l = vc.lower()
    if any(k in blob for k in (
        "заряд", "charger", "ремеш", "браслет", "бланк", "переходник",
        "кабел", "adapter", "powerbank", "pods", "наушник",
    )):
        return "skip"
    if any(k in blob for k in ("кругл", "round")):
        return "round"
    if any(k in blob for k in ("квадрат", "square")):
        return "square"
    # квадратные: линейка 11 / S11 / LK11 / HK11 / Pro Max / mini
    if re.search(r"(^|[_/])(031|034|035|038|039|040|042|046)([_/]|$)", vc_l):
        return "square"
    if any(k in vc_l for k in ("lk11", "hk11", "s11", "promax", "dt11")):
        return "square"
    if any(k in blob for k in ("11 series", "11 серия", "11 сери", "pro max", "promax", "s11", "lk11", "hk11")):
        return "square"
    # круглые: GT5 / Ultra / Watch 6 Pro / X8–X10
    if re.search(r"(^|[_/])(026|033|036|037|041|044|045)([_/]|$)", vc_l):
        return "round"
    if any(k in vc_l for k in ("gt5", "g7pro", "ultra", "x8", "x10", "watch_6", "6pro", "х10", "х8")):
        return "round"
    if any(k in blob for k in ("gt5", "gt 5", "ultra", "x8", "x10", "6pro", "6 pro", "х10", "х8")):
        return "round"
    return "unknown"


_WATCH_SHAPE_LABEL = {
    "round": "круглые",
    "square": "квадратные",
    "unknown": "неясно",
    "skip": "не часы",
}


def _own_nm_vendor_map() -> dict:
    """nm_id(int) → vendor_code для наших карточек."""
    out = {}
    try:
        st = httpx.get(
            f"{SUPABASE_URL}/rest/v1/stock_totals?select=nm_id,vendor_code&limit=5000",
            headers=sb_headers(), timeout=15,
        )
        if st.is_success:
            for r in st.json() or []:
                nm = r.get("nm_id")
                if nm is None:
                    continue
                try:
                    nm = int(nm)
                except (TypeError, ValueError):
                    continue
                vc = (r.get("vendor_code") or "").strip()
                if nm and (nm not in out or (vc and not out[nm])):
                    out[nm] = vc or str(nm)
    except Exception as e:
        logger.warning(f"own nm map stock_totals: {e}")
    try:
        rt = httpx.get(
            f"{SUPABASE_URL}/rest/v1/ratings_official?select=nm_id,article&nm_id=not.is.null&limit=5000",
            headers=sb_headers(), timeout=15,
        )
        if rt.is_success:
            for r in rt.json() or []:
                nm = r.get("nm_id")
                art = (r.get("article") or "").strip()
                if nm is None:
                    continue
                try:
                    nm = int(nm)
                except (TypeError, ValueError):
                    continue
                if nm and art and (nm not in out or not out.get(nm) or out[nm] == str(nm)):
                    out[nm] = art
    except Exception as e:
        logger.warning(f"own nm map ratings: {e}")
    return out


_SHELF_TOP_CACHE = {"ts": 0.0, "days": 0, "items": []}


def _own_top_sellers_week(top_n: int = 20, days: int = 7) -> list:
    """Топ наших nm по заказам за последние days дней.
    Источники: article_daily_stats → кэш sales-pace → живая воронка WB."""
    top_n = max(1, min(int(top_n or 20), 40))
    days = max(1, min(int(days or 7), 30))
    now_ts = time.time()
    if (
        _SHELF_TOP_CACHE.get("items")
        and _SHELF_TOP_CACHE.get("days") == days
        and now_ts - float(_SHELF_TOP_CACHE.get("ts") or 0) < 600
    ):
        return list(_SHELF_TOP_CACHE["items"])[:top_n]

    own_map = _own_nm_vendor_map()
    dt_from = (datetime.now(timezone.utc).date() - timedelta(days=days)).isoformat()
    agg = {}

    # 1) дневная статистика в Supabase
    if own_map:
        try:
            resp = httpx.get(
                f"{SUPABASE_URL}/rest/v1/article_daily_stats"
                f"?dt=gte.{dt_from}&select=nm_id,vendor_code,orders,open_card,add_to_cart,dt&order=dt.desc",
                headers=sb_headers(), timeout=30,
            )
            rows = resp.json() if resp.is_success else []
        except Exception as e:
            logger.warning(f"top sellers daily: {e}")
            rows = []
        for r in rows or []:
            try:
                nm = int(r.get("nm_id"))
            except (TypeError, ValueError):
                continue
            if nm not in own_map:
                continue
            slot = agg.setdefault(nm, {
                "nm_id": nm,
                "vendor_code": (r.get("vendor_code") or own_map.get(nm) or str(nm)).strip(),
                "orders": 0, "opens": 0, "cart": 0, "name": "",
            })
            vc = (r.get("vendor_code") or "").strip()
            if vc and (not slot["vendor_code"] or slot["vendor_code"] == str(nm)):
                slot["vendor_code"] = vc
            try:
                slot["orders"] += int(r.get("orders") or 0)
                slot["opens"] += int(r.get("open_card") or 0)
                slot["cart"] += int(r.get("add_to_cart") or 0)
            except (TypeError, ValueError):
                pass

    # 2) кэш sales-pace (day/week)
    if not agg:
        by = SALES_PACE_CACHE.get("by_period") or {}
        for key in ("week", "day"):
            cached = by.get(key) or {}
            arts = cached.get("articles") or []
            if not arts:
                continue
            for a in arts:
                try:
                    nm = int(a.get("nm_id"))
                except (TypeError, ValueError):
                    continue
                if own_map and nm not in own_map:
                    continue
                agg[nm] = {
                    "nm_id": nm,
                    "vendor_code": (a.get("vendor_code") or (own_map or {}).get(nm) or str(nm)).strip(),
                    "orders": int(a.get("orders_today") or 0),
                    "opens": int(a.get("opens_today") or a.get("clicks_today") or 0),
                    "cart": int(a.get("cart_today") or 0),
                    "name": a.get("name") or "",
                }
            if agg:
                break

    # 3) живая воронка WB за период (nm-report detail — все карточки продавца)
    if not agg and WB_TOKEN:
        end_d = datetime.now(timezone.utc).date()
        begin_d = end_d - timedelta(days=days)
        try:
            resp = httpx.post(
                f"{WB_ANALYTICS_URL}/api/analytics/v2/nm-report/detail",
                headers=wb_headers(),
                json={
                    "period": {"begin": begin_d.isoformat(), "end": end_d.isoformat()},
                    "brandNames": [], "objectIDs": [], "tagIDs": [],
                    "nmIDs": [],
                    "timezone": "Europe/Moscow",
                    "page": 1,
                },
                timeout=45,
            )
            if resp.is_success:
                cards = (resp.json() or {}).get("data", {}).get("cards") or []
                for c in cards:
                    nm = c.get("nmID") or c.get("nmId")
                    try:
                        nm = int(nm)
                    except (TypeError, ValueError):
                        continue
                    if own_map and nm not in own_map:
                        # всё равно берём — это карточки кабинета
                        pass
                    stats = (c.get("statistics") or {}).get("selectedPeriod") or {}
                    vc = (own_map or {}).get(nm) or c.get("vendorCode") or str(nm)
                    agg[nm] = {
                        "nm_id": nm,
                        "vendor_code": str(vc).strip(),
                        "orders": int(stats.get("ordersCount") or 0),
                        "opens": int(stats.get("openCardCount") or 0),
                        "cart": int(stats.get("addToCartCount") or 0),
                        "name": c.get("objectName") or c.get("brandName") or "",
                    }
            else:
                logger.warning(f"top sellers live nm-report: {resp.status_code} {resp.text[:180]}")
        except Exception as e:
            logger.warning(f"top sellers live: {e}")

    # 4) fallback: sales-funnel v3 по списку own nm (батчами)
    if not agg and WB_TOKEN and own_map:
        end_d = datetime.now(timezone.utc).date()
        begin_d = end_d - timedelta(days=days)
        ids = list(own_map.keys())
        for i in range(0, len(ids), 100):
            chunk = ids[i:i + 100]
            stats = fetch_own_stats_v3(chunk, begin_d.isoformat(), end_d.isoformat())
            for s in (stats or {}).values():
                try:
                    nm = int(s.get("nm_id"))
                except (TypeError, ValueError):
                    continue
                agg[nm] = {
                    "nm_id": nm,
                    "vendor_code": (s.get("vendor_code") or own_map.get(nm) or str(nm)).strip(),
                    "orders": int(s.get("orders") or 0),
                    "opens": int(s.get("card_opens") or 0),
                    "cart": int(s.get("cart_adds") or 0),
                    "name": s.get("name") or "",
                }
            if i + 100 < len(ids):
                time.sleep(1.1)

    items = []
    for nm, slot in agg.items():
        vc = slot.get("vendor_code") or (own_map or {}).get(nm) or str(nm)
        name = slot.get("name") or ""
        shape = _watch_shape(vc, name)
        if shape == "skip":
            continue
        if int(slot.get("orders") or 0) <= 0 and int(slot.get("cart") or 0) <= 0:
            continue
        items.append({
            "nm_id": nm,
            "vendor_code": vc,
            "name": name,
            "orders": int(slot.get("orders") or 0),
            "opens": int(slot.get("opens") or 0),
            "cart": int(slot.get("cart") or 0),
            "shape": shape,
            "shape_label": _WATCH_SHAPE_LABEL.get(shape, shape),
            "thumb": wb_product_thumb_url(nm),
            "url": f"https://www.wildberries.ru/catalog/{nm}/detail.aspx",
        })
    items.sort(key=lambda x: (-x["orders"], -x["cart"], -x["opens"], x["vendor_code"]))
    items = items[:top_n]
    _SHELF_TOP_CACHE["ts"] = now_ts
    _SHELF_TOP_CACHE["days"] = days
    _SHELF_TOP_CACHE["items"] = items
    return items


def _shelf_suggest_add(competitor: dict, shelf_items: list, top_n: int = 20) -> dict:
    """Какие наши топ-продажи ещё не стоят в полке конкурента (с учётом формы)."""
    own_map = _own_nm_vendor_map()
    own_set = set(own_map.keys())
    shelf_nms = set()
    already = []
    for it in shelf_items or []:
        try:
            nid = int(it.get("nm_id"))
        except (TypeError, ValueError):
            continue
        shelf_nms.add(nid)
        if nid in own_set:
            vc = own_map.get(nid) or ""
            already.append({
                "nm_id": nid,
                "vendor_code": vc,
                "position": it.get("position"),
                "brand": it.get("brand") or "",
                "name": it.get("name") or "",
                "thumb": it.get("thumb") or wb_product_thumb_url(nid),
            })

    comp_shape = _watch_shape("", competitor.get("name") or "", competitor.get("brand") or "")
    if comp_shape == "unknown" and already:
        shapes = [_watch_shape(a.get("vendor_code") or "", a.get("name") or "") for a in already]
        sq = sum(1 for s in shapes if s == "square")
        rd = sum(1 for s in shapes if s == "round")
        if sq >= 2 and sq > rd:
            comp_shape = "square"
        elif rd >= 2 and rd > sq:
            comp_shape = "round"

    top = _own_top_sellers_week(top_n=top_n, days=7)
    missing = [t for t in top if t["nm_id"] not in shelf_nms]

    def sort_key(t):
        sh = t.get("shape") or "unknown"
        if comp_shape in ("round", "square"):
            if sh == comp_shape:
                bucket = 0
            elif sh == "unknown":
                bucket = 1
            else:
                bucket = 2
        else:
            bucket = 0
        return (bucket, -int(t.get("orders") or 0), -int(t.get("cart") or 0))

    missing.sort(key=sort_key)
    same = [t for t in missing if t.get("shape") == comp_shape] if comp_shape in ("round", "square") else list(missing)
    other = [t for t in missing if t not in same]

    return {
        "own_nm_ids": sorted(own_set),
        "comp_shape": comp_shape,
        "comp_shape_label": _WATCH_SHAPE_LABEL.get(comp_shape, comp_shape),
        "top_period_days": 7,
        "top_n": top_n,
        "already_in_shelf": already,
        "already_count": len(already),
        "top_in_shelf_count": sum(1 for t in top if t["nm_id"] in shelf_nms),
        "suggest_add": missing,
        "suggest_same_shape": same,
        "suggest_other_shape": other,
        "top_sellers": top,
    }


@app.get("/api/competitor-shelf")
def get_competitor_shelf(nm_id: int, dest: int = -1257786, limit: int = 15, top: int = 20):
    """Топ полки «Смотрите также» у конкурента + предложения из нашего топ-20 продаж за неделю."""
    if not nm_id or nm_id < 1:
        raise HTTPException(status_code=400, detail="nm_id required")
    limit = max(1, min(int(limit or 15), 30))
    try:
        top = max(5, min(int(top or 20), 40))
    except (TypeError, ValueError):
        top = 20
    try:
        dest = int(dest)
    except (TypeError, ValueError):
        dest = -1257786

    card = fetch_wb_card_brief(nm_id, dest=dest)
    shelf = fetch_wb_see_also_shelf(nm_id, dest=dest, limit=limit)
    city_name = next((c["name"] for c in WB_SEARCH_CITIES if c["dest"] == dest), str(dest))
    competitor = card or {
        "nm_id": nm_id,
        "brand": "",
        "name": "",
        "thumb": wb_product_thumb_url(nm_id),
        "url": f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
    }
    items = shelf.get("items") or []
    suggest = _shelf_suggest_add(competitor=competitor, shelf_items=items, top_n=top)
    return {
        "nm_id": nm_id,
        "dest": dest,
        "city": city_name,
        "limit": limit,
        "competitor": competitor,
        "items": items,
        "shelf_total": shelf.get("total") or 0,
        "error": shelf.get("error"),
        **suggest,
    }


def _crm_headers() -> dict:
    h = {"Content-Type": "application/json"}
    if CRM_PASSWORD:
        h["x-crm-password"] = CRM_PASSWORD
    return h


def _crm_find_employee(employees: list, aliases: tuple) -> dict | None:
    for emp in employees or []:
        if not emp.get("active", True):
            continue
        name = str(emp.get("name") or "").strip().lower().replace("ё", "е")
        if not name:
            continue
        for a in aliases:
            if a in name:
                return emp
    return None


@app.post("/api/crm-shelf-boost-task")
async def crm_shelf_boost_task(request: dict):
    """Создать в Team CRM задачу «прокачать полки» на Афину или Заиру.

    Body: {
      manager: "afina"|"zaira",
      own_vendor_code, own_nm_id,
      competitor_nm_id, competitor_brand?, competitor_name?
    }
    """
    if not CRM_API_URL:
        return {
            "ok": False,
            "error": "CRM_API_URL не задан в Railway (URL team-crm без слэша в конце)",
        }
    if not isinstance(request, dict):
        raise HTTPException(status_code=400, detail="invalid body")

    manager_key = str(request.get("manager") or "").strip().lower()
    aliases = CRM_MANAGER_ALIASES.get(manager_key)
    if not aliases:
        return {"ok": False, "error": "manager: укажи afina или zaira"}

    own_vc = str(request.get("own_vendor_code") or "").strip()
    try:
        own_nm = int(request.get("own_nm_id"))
        competitor_nm = int(request.get("competitor_nm_id"))
    except (TypeError, ValueError):
        return {"ok": False, "error": "нужны own_nm_id и competitor_nm_id"}
    if not own_vc:
        own_vc = str(own_nm)
    if own_nm < 1 or competitor_nm < 1:
        return {"ok": False, "error": "некорректные nm_id"}

    comp_brand = str(request.get("competitor_brand") or "").strip()
    comp_name = str(request.get("competitor_name") or "").strip()

    title = (
        f'Раздача "{own_vc}" и {own_nm}, '
        f'через этого конкурента "{competitor_nm}".'
    )
    marker = f"[dash:shelf-boost:{own_nm}:{competitor_nm}]"
    comp_bits = [str(competitor_nm)]
    if comp_brand:
        comp_bits.insert(0, comp_brand)
    if comp_name:
        comp_bits.append(comp_name)
    description = (
        f"{marker}\n"
        f"Прокачать полки.\n"
        f"Наш: {own_vc} · https://www.wildberries.ru/catalog/{own_nm}/detail.aspx\n"
        f"Конкурент: {' · '.join(comp_bits)} · "
        f"https://www.wildberries.ru/catalog/{competitor_nm}/detail.aspx"
    )

    try:
        board_resp = httpx.get(
            f"{CRM_API_URL}/api/board",
            headers=_crm_headers(),
            timeout=25,
        )
    except Exception as e:
        return {"ok": False, "error": f"CRM недоступен: {e}"}
    if board_resp.status_code == 401:
        return {"ok": False, "error": "CRM: неверный CRM_PASSWORD (x-crm-password)"}
    if not board_resp.is_success:
        return {"ok": False, "error": f"CRM board HTTP {board_resp.status_code}: {board_resp.text[:180]}"}

    board = board_resp.json() or {}
    employees = board.get("employees") or []
    assignee = _crm_find_employee(employees, aliases)
    if not assignee:
        names = ", ".join(str(e.get("name") or "") for e in employees[:20])
        return {
            "ok": False,
            "error": f"В CRM не найден менеджер «{manager_key}». Есть: {names}",
        }

    owner = next((e for e in employees if str(e.get("role") or "") == "owner"), None)
    created_by_id = (owner or assignee).get("id")

    payload = {
        "title": title[:500],
        "description": description[:2000],
        "articles": f"{own_vc} {own_nm} / {competitor_nm}"[:500],
        "assignee_id": assignee.get("id"),
        "assignee_ids": [assignee.get("id")],
        "created_by_id": created_by_id,
        "status": "todo",
        "kind": "once",
        "priority": "normal",
        "notify_now": True,
    }
    try:
        create_resp = httpx.post(
            f"{CRM_API_URL}/api/tasks",
            headers=_crm_headers(),
            json=payload,
            timeout=30,
        )
    except Exception as e:
        return {"ok": False, "error": f"CRM create: {e}"}
    if not create_resp.is_success:
        return {
            "ok": False,
            "error": f"CRM tasks HTTP {create_resp.status_code}: {create_resp.text[:220]}",
        }
    task = create_resp.json() or {}
    return {
        "ok": True,
        "task_id": task.get("id"),
        "title": task.get("title") or title,
        "assignee_name": task.get("assignee_name") or assignee.get("name"),
        "notified": task.get("notified"),
        "notify_error": task.get("notify_error"),
        "manager": manager_key,
    }


@app.post("/api/shelf-presence")
def shelf_presence(request: dict):
    """Где наша карточка стоит в полках «Смотрите также» у списка конкурентов.

    Body: {own_nm_id, competitor_nm_ids: [int], dest?, limit?}
    limit — глубина полки (1–30, по умолчанию 15).
    """
    if not isinstance(request, dict):
        raise HTTPException(status_code=400, detail="invalid body")
    try:
        own_nm_id = int(request.get("own_nm_id"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="own_nm_id required")
    if own_nm_id < 1:
        raise HTTPException(status_code=400, detail="own_nm_id required")

    try:
        dest = int(request.get("dest") if request.get("dest") is not None else -1257786)
    except (TypeError, ValueError):
        dest = -1257786

    limit = request.get("limit", 15)
    try:
        limit = max(1, min(int(limit or 15), 30))
    except (TypeError, ValueError):
        limit = 15

    raw_ids = request.get("competitor_nm_ids")
    if not isinstance(raw_ids, list) or not raw_ids:
        raise HTTPException(status_code=400, detail="competitor_nm_ids required (non-empty list)")

    competitor_ids = []
    seen = set()
    for x in raw_ids:
        try:
            nid = int(x)
        except (TypeError, ValueError):
            continue
        if nid < 1 or nid == own_nm_id or nid in seen:
            continue
        seen.add(nid)
        competitor_ids.append(nid)
        if len(competitor_ids) >= 80:
            break

    if not competitor_ids:
        raise HTTPException(status_code=400, detail="no valid competitor_nm_ids")

    results = []
    for i, nm in enumerate(competitor_ids):
        if i:
            time.sleep(0.35)
        shelf = fetch_wb_see_also_shelf(nm, dest=dest, limit=limit)
        items = shelf.get("items") or []
        position = None
        for it in items:
            try:
                if int(it.get("nm_id")) == own_nm_id:
                    position = it.get("position")
                    break
            except (TypeError, ValueError):
                continue
        results.append({
            "competitor_nm_id": nm,
            "found": position is not None,
            "position": position,
            "shelf_total": shelf.get("total") or len(items),
            "shelf_checked": len(items),
            "error": shelf.get("error"),
            "thumb": wb_product_thumb_url(nm),
            "url": f"https://www.wildberries.ru/catalog/{nm}/detail.aspx",
        })

    city_name = next((c["name"] for c in WB_SEARCH_CITIES if c["dest"] == dest), str(dest))
    found = [r for r in results if r["found"]]
    found.sort(key=lambda r: (r["position"] is None, r["position"] or 999))
    not_found = [r for r in results if not r["found"]]
    return {
        "own_nm_id": own_nm_id,
        "dest": dest,
        "city": city_name,
        "limit": limit,
        "checked": len(results),
        "found_count": len(found),
        "own_thumb": wb_product_thumb_url(own_nm_id),
        "own_url": f"https://www.wildberries.ru/catalog/{own_nm_id}/detail.aspx",
        "results": found + not_found,
    }


# ---------- «Продавец рекомендует» (Content API, настраиваемый блок на КТ) ----------

SELLER_RECS_PATHS = (
    f"{WB_CONTENT_URL}/content/v1/recommendations/list",
    f"{WB_CONTENT_URL}/api/content/v1/recommendations/list",
)

# Короткий кэш полного агрегата (источник — Content API)
SELLER_RECS_AGG_CACHE = {"ts": 0.0, "data": None}


def fetch_seller_recommendations_raw(limit: int = 1000):
    """
    Весь список настроек «Продавец рекомендует» из Content API.
    → (hosts, error)
      hosts: [{nm_id, vendor_code, brand, name, thumb, recom_nms, recom_pics, recom_count, updated_at}]
    """
    if not WB_TOKEN:
        return [], "WB_TOKEN не задан"

    limit = max(1, min(int(limit or 1000), 5000))
    hosts = []
    next_cur = 0
    used_path = None
    last_err = None

    for page in range(80):
        body = {"limit": limit, "next": next_cur}
        resp = None
        for path in (used_path,) if used_path else SELLER_RECS_PATHS:
            if not path:
                continue
            try:
                resp = httpx.post(path, headers=wb_headers(), json=body, timeout=45)
            except Exception as e:
                last_err = str(e)[:200]
                resp = None
                continue
            if resp.status_code == 404 and not used_path:
                last_err = f"404 {path}"
                continue
            used_path = path
            break

        if resp is None:
            return hosts, last_err or "Content API недоступен"

        if resp.status_code == 429:
            time.sleep(2.0)
            try:
                resp = httpx.post(used_path, headers=wb_headers(), json=body, timeout=45)
            except Exception as e:
                return hosts, str(e)[:200]

        if not resp.is_success:
            text = (resp.text or "")[:280]
            # 401/403 — нет категории Контент / опции рекомендаций
            hint = ""
            if resp.status_code in (401, 403):
                hint = " — проверь WB_TOKEN (категория «Контент») и доступ к блоку «Продавец рекомендует»"
            elif resp.status_code == 402:
                hint = " — рекомендациями управляет тариф/опция продавца (Джем и т.п.)"
            return hosts, f"Content API {resp.status_code}: {text}{hint}"

        payload = resp.json() or {}
        # схемы: {data: [...], next: int} или обёртка data
        rows = payload.get("data")
        if rows is None and isinstance(payload.get("data"), dict):
            rows = (payload.get("data") or {}).get("data")
        if not isinstance(rows, list):
            rows = []

        if not rows:
            break

        for row in rows:
            if not isinstance(row, dict):
                continue
            nm = row.get("nmId") or row.get("nmID") or row.get("nm_id")
            if not nm:
                continue
            recom_nms = row.get("recomNms") or row.get("recom_nms") or []
            recom_pics = row.get("recomPics") or row.get("recom_pics") or []
            try:
                recom_nms = [int(x) for x in recom_nms if x is not None]
            except (TypeError, ValueError):
                recom_nms = []
            hosts.append({
                "nm_id": int(nm),
                "vendor_code": (row.get("vendorCode") or row.get("vendor_code") or "").strip(),
                "brand": (row.get("brandName") or row.get("brand_name") or row.get("brand") or "").strip(),
                "name": (row.get("title") or row.get("name") or "").strip(),
                "subject": (row.get("subjectName") or row.get("subject_name") or "").strip(),
                "thumb": row.get("pic") or wb_product_thumb_url(int(nm)),
                "recom_count": int(row.get("recomCount") or row.get("recom_count") or len(recom_nms) or 0),
                "recom_nms": recom_nms,
                "recom_pics": list(recom_pics) if isinstance(recom_pics, list) else [],
                "updated_at": row.get("updatedAt") or row.get("updated_at"),
            })

        new_next = payload.get("next")
        try:
            new_next = int(new_next) if new_next is not None else 0
        except (TypeError, ValueError):
            new_next = 0

        # курсор = последний nmId; если не сдвинулся — стоп
        if not new_next or new_next == next_cur or len(rows) < limit:
            # если next == 0 после страницы — конец
            if not new_next:
                break
            if new_next == next_cur:
                break
        next_cur = new_next
        # лимит 100 req/min — берём с запасом
        time.sleep(0.65)

    logger.info(f"seller recommendations: {len(hosts)} host cards via {used_path}")
    return hosts, None


def aggregate_seller_recommendations(hosts: list, catalog: list | None = None):
    """
    Инверт: для каждого рекомендуемого nm — у скольких карточек он в топ-5 / ниже.
    Порядок в recom_nms = место (1-based).
    missing — свои карточки, которых нет ни в одном recomNms.
    """
    own = {}
    for h in hosts:
        nm = int(h.get("nm_id") or 0)
        if nm:
            own[nm] = h
    for c in catalog or []:
        nm = int(c.get("nm_id") or 0)
        if not nm:
            continue
        prev = own.get(nm)
        if not prev:
            own[nm] = c
        else:
            # дополняем метаданные из полного каталога
            for k in ("vendor_code", "brand", "name", "thumb"):
                if not prev.get(k) and c.get(k):
                    prev[k] = c[k]

    by_nm = {}
    hosts_with = 0
    for host in hosts:
        recom_nms = host.get("recom_nms") or []
        if not recom_nms:
            continue
        hosts_with += 1
        host_nm = int(host["nm_id"])
        host_vc = host.get("vendor_code") or str(host_nm)
        pics = host.get("recom_pics") or []
        for pos, rnm in enumerate(recom_nms, 1):
            try:
                rnm = int(rnm)
            except (TypeError, ValueError):
                continue
            if not rnm or rnm == host_nm:
                continue
            row = by_nm.get(rnm)
            if not row:
                meta = own.get(rnm) or {}
                pic = ""
                if pos - 1 < len(pics) and pics[pos - 1]:
                    pic = pics[pos - 1]
                row = {
                    "nm_id": rnm,
                    "vendor_code": meta.get("vendor_code") or "",
                    "brand": meta.get("brand") or "",
                    "name": meta.get("name") or "",
                    "thumb": pic or meta.get("thumb") or wb_product_thumb_url(rnm),
                    "top5": 0,
                    "below": 0,
                    "hosts_top5": [],
                    "hosts_below": [],
                    "is_mine": rnm in own,
                }
                by_nm[rnm] = row
            else:
                meta = own.get(rnm)
                if meta:
                    row["is_mine"] = True
                    if not row.get("vendor_code") and meta.get("vendor_code"):
                        row["vendor_code"] = meta["vendor_code"]
                    if not row.get("brand") and meta.get("brand"):
                        row["brand"] = meta["brand"]
                    if not row.get("name") and meta.get("name"):
                        row["name"] = meta["name"]
                    if not row.get("thumb") and meta.get("thumb"):
                        row["thumb"] = meta["thumb"]

            host_info = {
                "host_nm": host_nm,
                "host_vc": host_vc,
                "position": pos,
            }
            if pos <= 5:
                if not any(x["host_nm"] == host_nm for x in row["hosts_top5"]):
                    row["top5"] += 1
                    row["hosts_top5"].append(host_info)
            else:
                if not any(x["host_nm"] == host_nm for x in row["hosts_below"]):
                    row["below"] += 1
                    row["hosts_below"].append(host_info)

    rows = list(by_nm.values())
    rows.sort(key=lambda r: (-r["top5"], -r["below"], r["nm_id"]))

    recommended = set(by_nm.keys())
    missing = []
    # каталог — все свои nm; fallback own keys if catalog empty
    all_own = catalog if catalog else list(own.values())
    seen_miss = set()
    for c in all_own:
        try:
            nm = int(c.get("nm_id") or 0)
        except (TypeError, ValueError):
            continue
        if not nm or nm in recommended or nm in seen_miss:
            continue
        seen_miss.add(nm)
        missing.append({
            "nm_id": nm,
            "vendor_code": c.get("vendor_code") or "",
            "brand": c.get("brand") or "",
            "name": c.get("name") or "",
            "thumb": c.get("thumb") or wb_product_thumb_url(nm),
            "top5": 0,
            "below": 0,
            "hosts_top5": [],
            "hosts_below": [],
            "is_mine": True,
        })
    missing.sort(key=lambda r: (
        (r.get("vendor_code") or "").lower(),
        r["nm_id"],
    ))

    return {
        "hosts_total": len(hosts),
        "hosts_with_recs": hosts_with,
        "catalog_total": len(all_own) if catalog else len(own),
        "rows": rows,
        "missing": missing,
    }


def fetch_all_own_content_cards() -> list:
    """Все свои nm-карточки из Content API: nm_id, vendor_code, brand, name, thumb."""
    if not WB_TOKEN:
        return []
    out, seen = [], set()
    cursor = {"limit": 100}
    for _ in range(200):
        try:
            resp = httpx.post(
                f"{WB_CONTENT_URL}/content/v2/get/cards/list",
                headers=wb_headers(),
                json={
                    "settings": {
                        "sort": {"ascending": True},
                        "filter": {"withPhoto": -1},
                        "cursor": cursor,
                    }
                },
                timeout=40,
            )
        except Exception as e:
            logger.error(f"cards/list for seller recs catalog: {e}")
            break
        if not resp.is_success:
            logger.error(f"cards/list seller recs {resp.status_code}: {resp.text[:200]}")
            break
        payload = resp.json() or {}
        cards = payload.get("cards") or []
        if not cards:
            break
        for c in cards:
            nm = c.get("nmID") or c.get("nmId")
            if not nm:
                continue
            try:
                nm = int(nm)
            except (TypeError, ValueError):
                continue
            if nm in seen:
                continue
            seen.add(nm)
            vc = (c.get("vendorCode") or "").strip()
            brand = (c.get("brand") or "").strip()
            name = (c.get("title") or c.get("subjectName") or "").strip()
            thumb = wb_product_thumb_url(nm)
            photos = c.get("photos") or c.get("mediaFiles") or []
            if isinstance(photos, list) and photos:
                p0 = photos[0]
                if isinstance(p0, dict):
                    thumb = (
                        p0.get("c516x688")
                        or p0.get("big")
                        or p0.get("square")
                        or p0.get("tm")
                        or thumb
                    )
                elif isinstance(p0, str) and p0.startswith("http"):
                    thumb = p0
            out.append({
                "nm_id": nm,
                "vendor_code": vc,
                "brand": brand,
                "name": name,
                "thumb": thumb,
            })
        curs = payload.get("cursor") or {}
        updated = curs.get("updatedAt")
        nm_cur = curs.get("nmID") or curs.get("nmId")
        if len(cards) < 100 or not updated or nm_cur is None:
            break
        cursor = {"limit": 100, "updatedAt": updated, "nmID": nm_cur}
        time.sleep(0.35)
    logger.info(f"seller recs catalog cards: {len(out)}")
    return out


def fetch_fbs_speed_report_data(days: int = 14) -> dict:
    """
    Выгружает сборочные задания FBS, поставки и склады из WB Marketplace API v3,
    рассчитывает время сдачи (от создания заказа до скана/закрытия поставки),
    коэффициент kC (правила с 07.08.2026) и финансовый эффект (скидка/штраф к комиссии).
    """
    if not WB_TOKEN:
        return {"error": "WB_TOKEN не задан"}

    now = datetime.now(timezone.utc)
    date_from_dt = now - timedelta(days=days)
    date_from_ts = int(date_from_dt.timestamp())

    headers = wb_headers()

    # 1. Склады FBS
    wh_map = {}
    try:
        r_wh = httpx.get(f"{WB_MARKETPLACE_URL}/api/v3/warehouses", headers=headers, timeout=20)
        if r_wh.is_success:
            for w in r_wh.json():
                wh_map[w.get("id")] = w.get("name")
    except Exception as e:
        logger.warning(f"fetch_fbs_speed_report warehouses error: {e}")

    # 2. Поставки FBS (supplies)
    supplies_map = {}
    try:
        next_val = 0
        while True:
            r_sup = httpx.get(
                f"{WB_MARKETPLACE_URL}/api/v3/supplies",
                headers=headers,
                params={"limit": 1000, "next": next_val},
                timeout=30,
            )
            if not r_sup.is_success:
                logger.warning(f"fetch_fbs_speed_report supplies error: {r_sup.status_code} {r_sup.text[:200]}")
                break
            data = r_sup.json()
            items = data.get("supplies") if isinstance(data, dict) else (data if isinstance(data, list) else [])
            if not items:
                break
            for sup in items:
                sid = sup.get("id")
                if sid:
                    supplies_map[sid] = sup
            next_val = data.get("next", 0) if isinstance(data, dict) else 0
            if not next_val or len(items) < 1000:
                break
    except Exception as e:
        logger.warning(f"fetch_fbs_speed_report supplies fetch error: {e}")

    # 3. Заказы FBS (сборочные задания)
    orders = []
    try:
        next_val = 0
        for _ in range(20):  # до 20,000 заказов
            r_ord = httpx.get(
                f"{WB_MARKETPLACE_URL}/api/v3/orders",
                headers=headers,
                params={"limit": 1000, "next": next_val, "dateFrom": date_from_ts},
                timeout=35,
            )
            if not r_ord.is_success:
                logger.warning(f"fetch_fbs_speed_report orders error: {r_ord.status_code} {r_ord.text[:200]}")
                break
            data = r_ord.json()
            batch = data.get("orders") if isinstance(data, dict) else (data if isinstance(data, list) else [])
            if not batch:
                break
            orders.extend(batch)
            next_val = data.get("next", 0) if isinstance(data, dict) else 0
            if not next_val or len(batch) < 1000:
                break
    except Exception as e:
        logger.warning(f"fetch_fbs_speed_report orders fetch error: {e}")

    # 4. Анализ каждого заказа и расчет kC
    orders_analyzed = []
    bracket_counts = {
        "bonus_13": 0,    # <= 13ч (-5.0%)
        "norm_42": 0,     # 13-42ч (-3.5%)
        "base_48": 0,     # 42-48ч (0%)
        "fine_54": 0,     # 48-54ч (+0.30%/ч)
        "fine_60": 0,     # 54-60ч (+0.35%/ч)
        "fine_over": 0,   # > 60ч (+0.45%/ч)
        "pending": 0,     # еще не сданы
    }
    wh_stats = {}  # wh_name -> {total, <=13, 13-42, >48, bonus_rub, fine_rub, hours_list}

    total_bonus_rub = 0.0
    total_fine_rub = 0.0
    all_hours = []

    for o in orders:
        created_str = o.get("createdAt") or ""
        created_dt = parse_wb_dt(created_str)
        if not created_dt:
            continue

        sup_id = o.get("supplyId")
        sup = supplies_map.get(sup_id, {}) if sup_id else {}
        
        # Определяем время сканирования/сдачи
        scan_str = sup.get("scanDt") or sup.get("closedAt") or o.get("scanDt") or ""
        scan_dt = parse_wb_dt(scan_str)

        wh_id = o.get("warehouseId")
        wh_name = wh_map.get(wh_id) or f"Склад #{wh_id}"

        if wh_name not in wh_stats:
            wh_stats[wh_name] = {
                "warehouse_id": wh_id,
                "warehouse_name": wh_name,
                "total_orders": 0,
                "bonus_13_count": 0,
                "norm_42_count": 0,
                "base_48_count": 0,
                "fine_48_count": 0,
                "pending_count": 0,
                "bonus_rub": 0.0,
                "fine_rub": 0.0,
                "hours_list": [],
            }
        wstat = wh_stats[wh_name]
        wstat["total_orders"] += 1

        price_raw = o.get("convertedPrice") or o.get("price") or 0
        price_rub = (price_raw / 100.0) if price_raw > 100000 else float(price_raw)

        if not scan_dt:
            bracket_counts["pending"] += 1
            wstat["pending_count"] += 1
            orders_analyzed.append({
                "order_id": o.get("id"),
                "created_at": created_str,
                "scan_at": None,
                "hours": None,
                "warehouse_name": wh_name,
                "article": o.get("article"),
                "nm_id": o.get("nmId"),
                "price_rub": price_rub,
                "status": "pending",
                "kc_pct": 0.0,
                "impact_rub": 0.0,
            })
            continue

        diff_seconds = (scan_dt - created_dt).total_seconds()
        hours = max(0.0, round(diff_seconds / 3600.0, 2))
        all_hours.append(hours)
        wstat["hours_list"].append(hours)

        # Расчет kC
        kc_pct = 0.0
        bracket = ""
        if hours <= 13.0:
            kc_pct = -5.0
            bracket = "bonus_13"
            bracket_counts["bonus_13"] += 1
            wstat["bonus_13_count"] += 1
            b_rub = abs(kc_pct) / 100.0 * price_rub
            total_bonus_rub += b_rub
            wstat["bonus_rub"] += b_rub
            impact_rub = b_rub
        elif hours <= 42.0:
            kc_pct = -3.5
            bracket = "norm_42"
            bracket_counts["norm_42"] += 1
            wstat["norm_42_count"] += 1
            b_rub = abs(kc_pct) / 100.0 * price_rub
            total_bonus_rub += b_rub
            wstat["bonus_rub"] += b_rub
            impact_rub = b_rub
        elif hours <= 48.0:
            kc_pct = 0.0
            bracket = "base_48"
            bracket_counts["base_48"] += 1
            wstat["base_48_count"] += 1
            impact_rub = 0.0
        elif hours <= 54.0:
            kc_pct = (hours - 48.0) * 0.30
            bracket = "fine_54"
            bracket_counts["fine_54"] += 1
            wstat["fine_48_count"] += 1
            f_rub = kc_pct / 100.0 * price_rub
            total_fine_rub += f_rub
            wstat["fine_rub"] += f_rub
            impact_rub = -f_rub
        elif hours <= 60.0:
            kc_pct = (6.0 * 0.30) + ((hours - 54.0) * 0.35)
            bracket = "fine_60"
            bracket_counts["fine_60"] += 1
            wstat["fine_48_count"] += 1
            f_rub = kc_pct / 100.0 * price_rub
            total_fine_rub += f_rub
            wstat["fine_rub"] += f_rub
            impact_rub = -f_rub
        else:
            kc_pct = (6.0 * 0.30) + (6.0 * 0.35) + ((hours - 60.0) * 0.45)
            bracket = "fine_over"
            bracket_counts["fine_over"] += 1
            wstat["fine_48_count"] += 1
            f_rub = kc_pct / 100.0 * price_rub
            total_fine_rub += f_rub
            wstat["fine_rub"] += f_rub
            impact_rub = -f_rub

        orders_analyzed.append({
            "order_id": o.get("id"),
            "created_at": created_str,
            "scan_at": scan_str,
            "hours": hours,
            "warehouse_name": wh_name,
            "article": o.get("article"),
            "nm_id": o.get("nmId"),
            "price_rub": price_rub,
            "bracket": bracket,
            "kc_pct": round(kc_pct, 2),
            "impact_rub": round(impact_rub, 2),
        })

    import statistics
    median_hours = round(float(statistics.median(all_hours)), 1) if all_hours else 0.0
    delivered_count = len(all_hours)

    # Формируем сводку по складам
    wh_summary = []
    for w in wh_stats.values():
        h_list = w.pop("hours_list")
        w["median_hours"] = round(float(statistics.median(h_list)), 1) if h_list else 0.0
        w["bonus_rub"] = round(w["bonus_rub"], 2)
        w["fine_rub"] = round(w["fine_rub"], 2)
        w["net_rub"] = round(w["bonus_rub"] - w["fine_rub"], 2)
        wh_summary.append(w)
    wh_summary.sort(key=lambda x: -x["total_orders"])

    return {
        "period_days": days,
        "total_orders": len(orders_analyzed),
        "delivered_orders": delivered_count,
        "pending_orders": bracket_counts["pending"],
        "median_hours": median_hours,
        "bracket_counts": bracket_counts,
        "total_bonus_rub": round(total_bonus_rub, 2),
        "total_fine_rub": round(total_fine_rub, 2),
        "net_profit_rub": round(total_bonus_rub - total_fine_rub, 2),
        "warehouses": wh_summary,
        "orders_sample": orders_analyzed[:500],
    }


@app.get("/api/fbs-speed-report")
def get_fbs_speed_report(days: int = 14):
    """Отчет по скорости отгрузки FBS, порогам kC и экономии на комиссии."""
    try:
        data = fetch_fbs_speed_report_data(days=days)
        return data
    except Exception as e:
        logger.error(f"get_fbs_speed_report error: {e}")
        return {"error": str(e)}


# ---------- География заказов FBS / FBW (лента заказов + Statistics API) ----------

ORDERS_GEO_CACHE = {
    "orders": [],
    "updated_at": None,
    "source": None,
    "filename": None,
    "syncing": False,
    "error": None,
}
_ORDERS_GEO_LOCK = threading.Lock()
_ORDERS_GEO_FILE = Path(__file__).resolve().parent.parent / "data" / "orders_geo_cache.json"
# Диск контейнера на Railway эфемерный, поэтому долговременно храним в Supabase.
ORDERS_GEO_SETTING_KEY = "orders_geo_cache"


def _orders_geo_ensure_dir():
    try:
        _ORDERS_GEO_FILE.parent.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass


def _orders_geo_payload() -> dict:
    return {
        "orders": ORDERS_GEO_CACHE.get("orders") or [],
        "updated_at": ORDERS_GEO_CACHE.get("updated_at"),
        "source": ORDERS_GEO_CACHE.get("source"),
        "filename": ORDERS_GEO_CACHE.get("filename"),
    }


def _orders_geo_apply_payload(payload, source_fallback: str) -> bool:
    orders = payload.get("orders") if isinstance(payload, dict) else payload
    if not isinstance(orders, list) or not orders:
        return False
    ORDERS_GEO_CACHE["orders"] = orders
    if isinstance(payload, dict):
        ORDERS_GEO_CACHE["updated_at"] = payload.get("updated_at")
        ORDERS_GEO_CACHE["source"] = payload.get("source") or source_fallback
        ORDERS_GEO_CACHE["filename"] = payload.get("filename")
    else:
        ORDERS_GEO_CACHE["updated_at"] = None
        ORDERS_GEO_CACHE["source"] = source_fallback
        ORDERS_GEO_CACHE["filename"] = None
    return True


def _orders_geo_encode(payload: dict) -> str:
    """Гзипуем: 12 тысяч заказов это около 3 МБ JSON, в settings столько лить незачем."""
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    return base64.b64encode(gzip.compress(raw, 6)).decode("ascii")


def _orders_geo_decode(blob):
    if not blob:
        return None
    if isinstance(blob, (dict, list)):
        return blob
    s = str(blob).strip()
    if not s:
        return None
    if s.startswith("{") or s.startswith("["):
        return json.loads(s)
    return json.loads(gzip.decompress(base64.b64decode(s)).decode("utf-8"))


def _orders_geo_save_supabase() -> bool:
    if not SUPABASE_URL or not SUPABASE_KEY:
        return False
    try:
        blob = _orders_geo_encode(_orders_geo_payload())
    except Exception as e:
        logger.error(f"orders_geo encode error: {e}")
        return False
    ok = save_setting_value(ORDERS_GEO_SETTING_KEY, blob)
    if ok:
        logger.info(f"orders_geo: сохранено в Supabase, {len(blob)} символов")
    return ok


def _orders_geo_load_supabase() -> bool:
    if not SUPABASE_URL or not SUPABASE_KEY:
        return False
    try:
        raw = get_setting_raw(ORDERS_GEO_SETTING_KEY, None)
        payload = _orders_geo_decode(raw)
    except Exception as e:
        logger.warning(f"orders_geo load supabase error: {e}")
        return False
    if not _orders_geo_apply_payload(payload, "supabase"):
        return False
    logger.info(f"orders_geo: поднято из Supabase, {len(ORDERS_GEO_CACHE['orders'])} заказов")
    return True


def _orders_geo_save_file():
    _orders_geo_ensure_dir()
    try:
        with open(_ORDERS_GEO_FILE, "w", encoding="utf-8") as f:
            json.dump(_orders_geo_payload(), f, ensure_ascii=False)
    except Exception as e:
        logger.warning(f"orders_geo save file error: {e}")


def _orders_geo_persist():
    """Локальный файл — быстрый кэш, Supabase — то, что переживает редеплой."""
    _orders_geo_save_file()
    _orders_geo_save_supabase()


def _orders_geo_load_file() -> bool:
    try:
        if not _ORDERS_GEO_FILE.exists():
            return False
        with open(_ORDERS_GEO_FILE, "r", encoding="utf-8") as f:
            payload = json.load(f)
        return _orders_geo_apply_payload(payload, "file")
    except Exception as e:
        logger.warning(f"orders_geo load file error: {e}")
        return False


def _orders_geo_load_seed() -> bool:
    seed = Path(__file__).resolve().parent.parent / "tmp" / "orders_ribbon_preprocessed.json"
    try:
        if not seed.exists():
            return False
        with open(seed, "r", encoding="utf-8") as f:
            orders = json.load(f)
        if not isinstance(orders, list) or not orders:
            return False
        ORDERS_GEO_CACHE["orders"] = orders
        ORDERS_GEO_CACHE["updated_at"] = datetime.now(timezone.utc).isoformat()
        ORDERS_GEO_CACHE["source"] = "seed"
        ORDERS_GEO_CACHE["filename"] = seed.name
        _orders_geo_persist()
        return True
    except Exception as e:
        logger.warning(f"orders_geo seed error: {e}")
        return False


def _orders_geo_ensure_loaded() -> list:
    with _ORDERS_GEO_LOCK:
        if ORDERS_GEO_CACHE.get("orders"):
            return ORDERS_GEO_CACHE["orders"]
        # Supabase первым: после редеплоя контейнера локального файла уже нет.
        if _orders_geo_load_supabase():
            _orders_geo_persist()
        elif _orders_geo_load_file():
            _orders_geo_save_supabase()
        else:
            _orders_geo_load_seed()
        return ORDERS_GEO_CACHE.get("orders") or []


def _orders_geo_normalize_channel(tipo_sklada: str, warehouse: str = "") -> str:
    t = (tipo_sklada or "").strip().lower()
    w = (warehouse or "").strip().lower()
    if "сво" in t or "продавц" in t or "fbs" in t:
        return "FBS"
    if "склад wb" in t or "склады wb" in t or "fbw" in t or "fbo" in t:
        return "FBW"
    if "продавц" in w or "склад продавца" in w:
        return "FBS"
    return "FBW"


def _orders_geo_parse_ribbon_df(df: "pd.DataFrame") -> list:
    """Парсит лист «Все заказы» / «Активные» из отчёта «Лента заказов» WB."""
    if df is None or df.empty:
        return []

    cols = [str(c).strip() if c is not None else "" for c in df.columns]
    # Иногда регион/город разбиты на две колонки (название + Unnamed)
    rename = {}
    for i, c in enumerate(cols):
        cl = c.lower()
        if "артикул продавца" in cl:
            rename[df.columns[i]] = "article"
        elif c == "Артикул WB" or "артикул wb" in cl:
            rename[df.columns[i]] = "nm_id"
        elif c == "Название" or cl == "название":
            rename[df.columns[i]] = "name"
        elif "дата оформления" in cl:
            rename[df.columns[i]] = "order_dt"
        elif "статус заказа" in cl:
            rename[df.columns[i]] = "status"
        elif "регион отправки" in cl:
            rename[df.columns[i]] = "src_region"
        elif "регион прибытия" in cl:
            rename[df.columns[i]] = "dest_region"
        elif "цена со скидкой" in cl:
            rename[df.columns[i]] = "price"
        elif "тип склада" in cl:
            rename[df.columns[i]] = "warehouse_type"
        elif "id заказа" in cl:
            rename[df.columns[i]] = "order_id"

    df = df.rename(columns=rename)

    # Unnamed колонки сразу после региона — склад отправки / город прибытия
    cols_now = list(df.columns)
    for i, c in enumerate(cols_now):
        if c == "src_region" and i + 1 < len(cols_now):
            nxt = cols_now[i + 1]
            if str(nxt).startswith("Unnamed") or nxt not in ("article", "nm_id", "name", "order_dt", "status", "dest_region", "price", "warehouse_type", "order_id", "warehouse", "dest_city"):
                df = df.rename(columns={nxt: "warehouse"})
        if c == "dest_region" and i + 1 < len(cols_now):
            nxt = cols_now[i + 1]
            if str(nxt).startswith("Unnamed") or nxt not in ("article", "nm_id", "name", "order_dt", "status", "src_region", "price", "warehouse_type", "order_id", "warehouse", "dest_city"):
                df = df.rename(columns={nxt: "dest_city"})

    # если склад/город всё ещё unnamed — эвристика по позиции
    if "warehouse" not in df.columns:
        for c in df.columns:
            if str(c).startswith("Unnamed"):
                sample = df[c].dropna().astype(str).head(20).str.lower()
                if sample.str.contains("склад|сц ").any():
                    df = df.rename(columns={c: "warehouse"})
                    break
    if "dest_city" not in df.columns:
        for c in df.columns:
            if str(c).startswith("Unnamed"):
                df = df.rename(columns={c: "dest_city"})
                break

    records = []
    for _, r in df.iterrows():
        try:
            dt_raw = r.get("order_dt")
            if pd.isna(dt_raw):
                continue
            dt = pd.to_datetime(dt_raw, errors="coerce")
            if pd.isna(dt):
                continue
            date_str = dt.strftime("%Y-%m-%d")
            wh = str(r.get("warehouse") or "").strip() or "Не указан"
            wtype = str(r.get("warehouse_type") or "").strip()
            channel = _orders_geo_normalize_channel(wtype, wh)
            try:
                price = float(r.get("price") or 0)
            except (TypeError, ValueError):
                price = 0.0
            if price != price:  # NaN
                price = 0.0
            try:
                nm = int(r.get("nm_id") or 0)
            except (TypeError, ValueError):
                nm = 0
            records.append({
                "order_id": str(r.get("order_id") or ""),
                "date": date_str,
                "dt": dt.strftime("%Y-%m-%d %H:%M:%S"),
                "channel": channel,
                "warehouse": wh,
                "dest_region": str(r.get("dest_region") or "").strip() or "Не указан",
                "dest_city": str(r.get("dest_city") or "").strip() or "Не указан",
                "article": str(r.get("article") or "").strip(),
                "nm_id": nm,
                "name": str(r.get("name") or "").strip(),
                "price": round(price, 2),
                "status": str(r.get("status") or "").strip(),
            })
        except Exception:
            continue
    return records


def parse_orders_geo_excel(content: bytes, filename: str = "") -> list:
    """Читает xlsx ленты заказов WB и возвращает нормализованный список заказов."""
    bio = io.BytesIO(content)
    xl = pd.ExcelFile(bio)
    sheet = None
    for name in xl.sheet_names:
        low = str(name).lower()
        if "все заказ" in low or "активн" in low:
            sheet = name
            break
    if sheet is None:
        sheet = xl.sheet_names[-1] if xl.sheet_names else 0

    # пробуем header=1 (типичный формат ленты), иначе header=0
    df = pd.read_excel(xl, sheet_name=sheet, header=1)
    # если колонки Unnamed и мало смысла — перечитать с header=0
    named = [c for c in df.columns if not str(c).startswith("Unnamed")]
    if len(named) < 5:
        df = pd.read_excel(xl, sheet_name=sheet, header=0)

    # если первая ячейка — «Все заказы», сдвигаем заголовок
    if df.shape[0] > 0 and str(df.iloc[0, 0]).strip().lower().startswith("артикул"):
        df.columns = [str(x).strip() for x in df.iloc[0].tolist()]
        df = df.iloc[1:].reset_index(drop=True)

    records = _orders_geo_parse_ribbon_df(df)
    if not records:
        # fallback: raw без заголовка
        raw = pd.read_excel(xl, sheet_name=sheet, header=None)
        if raw.shape[0] > 2:
            header_row = None
            for i in range(min(5, len(raw))):
                row_vals = [str(x).lower() for x in raw.iloc[i].tolist()]
                if any("артикул продавца" in v for v in row_vals):
                    header_row = i
                    break
            if header_row is not None:
                df2 = raw.iloc[header_row + 1:].copy()
                df2.columns = [str(x).strip() for x in raw.iloc[header_row].tolist()]
                records = _orders_geo_parse_ribbon_df(df2)
    logger.info(f"orders_geo parse {filename or sheet}: {len(records)} orders")
    return records


def aggregate_orders_geo(
    orders: list,
    date_from: str = None,
    date_to: str = None,
    channel: str = "all",
    warehouse: str = "all",
    region: str = "all",
    city: str = "all",
    search: str = "",
) -> dict:
    filtered = []
    search = (search or "").strip().lower()
    ch_filter = (channel or "all").upper()
    wh_filter = (warehouse or "all")
    reg_filter = (region or "all")
    city_filter = (city or "all")

    # опции фильтров — по датам (чтобы селекты не схлопывались при выборе канала)
    date_scoped = []
    for o in orders or []:
        o_date = o.get("date") or ""
        if date_from and o_date < date_from:
            continue
        if date_to and o_date > date_to:
            continue
        date_scoped.append(o)

    for o in date_scoped:
        o_ch = (o.get("channel") or "FBW").upper()
        if ch_filter not in ("ALL", "") and o_ch != ch_filter:
            continue
        o_wh = o.get("warehouse") or ""
        if wh_filter not in ("all", "", None) and o_wh != wh_filter:
            continue
        o_reg = o.get("dest_region") or ""
        if reg_filter not in ("all", "", None) and o_reg != reg_filter:
            continue
        o_city = o.get("dest_city") or ""
        if city_filter not in ("all", "", None) and o_city != city_filter:
            continue
        if search:
            blob = " ".join([
                str(o.get("article") or ""),
                str(o.get("nm_id") or ""),
                str(o.get("name") or ""),
                str(o_wh),
                str(o_city),
                str(o_reg),
            ]).lower()
            if search not in blob:
                continue
        filtered.append(o)

    total_orders = len(filtered)
    total_rev = sum(float(o.get("price") or 0) for o in filtered)
    fbs_orders = sum(1 for o in filtered if (o.get("channel") or "").upper() == "FBS")
    fbs_rev = sum(float(o.get("price") or 0) for o in filtered if (o.get("channel") or "").upper() == "FBS")
    fbw_orders = total_orders - fbs_orders
    fbw_rev = total_rev - fbs_rev

    by_day_dict = {}
    by_wh_dict = {}
    by_reg_dict = {}
    by_city_dict = {}
    by_art_dict = {}
    warehouses_set = {(o.get("warehouse") or "Не указан") for o in date_scoped}
    regions_set = {(o.get("dest_region") or "Не указан") for o in date_scoped}
    cities_set = {(o.get("dest_city") or "Не указан") for o in date_scoped}

    for o in filtered:
        d = o.get("date") or ""
        ch = (o.get("channel") or "FBW").upper()
        p = float(o.get("price") or 0)
        wh = o.get("warehouse") or "Не указан"
        reg = o.get("dest_region") or "Не указан"
        city_name = o.get("dest_city") or "Не указан"
        art = (o.get("article") or "").strip() or str(o.get("nm_id") or "—")

        if d:
            slot = by_day_dict.setdefault(d, {
                "date": d, "total": 0, "fbs": 0, "fbw": 0,
                "revenue": 0.0, "fbs_revenue": 0.0, "fbw_revenue": 0.0,
            })
            slot["total"] += 1
            slot["revenue"] += p
            if ch == "FBS":
                slot["fbs"] += 1
                slot["fbs_revenue"] += p
            else:
                slot["fbw"] += 1
                slot["fbw_revenue"] += p

        wslot = by_wh_dict.setdefault(wh, {
            "warehouse": wh, "channel": ch, "orders": 0, "revenue": 0.0, "regions": {},
        })
        wslot["orders"] += 1
        wslot["revenue"] += p
        wslot["regions"][reg] = wslot["regions"].get(reg, 0) + 1

        rslot = by_reg_dict.setdefault(reg, {
            "region": reg, "fbs_orders": 0, "fbw_orders": 0,
            "total_orders": 0, "revenue": 0.0, "cities": {},
        })
        rslot["total_orders"] += 1
        rslot["revenue"] += p
        if ch == "FBS":
            rslot["fbs_orders"] += 1
        else:
            rslot["fbw_orders"] += 1
        rslot["cities"][city_name] = rslot["cities"].get(city_name, 0) + 1

        cslot = by_city_dict.setdefault(city_name, {
            "city": city_name, "region": reg, "fbs_orders": 0, "fbw_orders": 0,
            "total_orders": 0, "revenue": 0.0,
        })
        cslot["total_orders"] += 1
        cslot["revenue"] += p
        if ch == "FBS":
            cslot["fbs_orders"] += 1
        else:
            cslot["fbw_orders"] += 1

        aslot = by_art_dict.setdefault(art, {
            "article": art, "nm_id": o.get("nm_id"), "name": o.get("name") or "",
            "fbs_orders": 0, "fbw_orders": 0, "total_orders": 0, "revenue": 0.0, "cities": {},
        })
        aslot["total_orders"] += 1
        aslot["revenue"] += p
        if ch == "FBS":
            aslot["fbs_orders"] += 1
        else:
            aslot["fbw_orders"] += 1
        aslot["cities"][city_name] = aslot["cities"].get(city_name, 0) + 1

    by_day = sorted(by_day_dict.values(), key=lambda x: x["date"])
    for s in by_day:
        s["revenue"] = round(s["revenue"], 2)
        s["fbs_revenue"] = round(s["fbs_revenue"], 2)
        s["fbw_revenue"] = round(s["fbw_revenue"], 2)

    by_wh = []
    for item in by_wh_dict.values():
        top_regs = [r[0] for r in sorted(item["regions"].items(), key=lambda x: -x[1])[:3]]
        by_wh.append({
            "warehouse": item["warehouse"],
            "channel": item["channel"],
            "orders": item["orders"],
            "revenue": round(item["revenue"], 2),
            "share_pct": round(item["orders"] / max(1, total_orders) * 100, 1),
            "top_regions": top_regs,
        })
    by_wh.sort(key=lambda x: -x["orders"])

    by_reg = []
    for item in by_reg_dict.values():
        top_cities = [c[0] for c in sorted(item["cities"].items(), key=lambda x: -x[1])[:4]]
        by_reg.append({
            "region": item["region"],
            "fbs_orders": item["fbs_orders"],
            "fbw_orders": item["fbw_orders"],
            "total_orders": item["total_orders"],
            "revenue": round(item["revenue"], 2),
            "share_pct": round(item["total_orders"] / max(1, total_orders) * 100, 1),
            "top_cities": top_cities,
        })
    by_reg.sort(key=lambda x: -x["total_orders"])

    by_city = []
    for item in by_city_dict.values():
        by_city.append({
            "city": item["city"],
            "region": item["region"],
            "fbs_orders": item["fbs_orders"],
            "fbw_orders": item["fbw_orders"],
            "total_orders": item["total_orders"],
            "revenue": round(item["revenue"], 2),
            "share_pct": round(item["total_orders"] / max(1, total_orders) * 100, 1),
        })
    by_city.sort(key=lambda x: -x["total_orders"])

    by_art = []
    for item in by_art_dict.values():
        top_city = sorted(item["cities"].items(), key=lambda x: -x[1])[0][0] if item["cities"] else ""
        by_art.append({
            "article": item["article"],
            "nm_id": item["nm_id"],
            "name": item["name"],
            "fbs_orders": item["fbs_orders"],
            "fbw_orders": item["fbw_orders"],
            "total_orders": item["total_orders"],
            "revenue": round(item["revenue"], 2),
            "avg_price": round(item["revenue"] / max(1, item["total_orders"]), 2),
            "top_city": top_city,
        })
    by_art.sort(key=lambda x: -x["total_orders"])

    all_orders = _orders_geo_ensure_loaded()
    dates = sorted({o.get("date") for o in all_orders if o.get("date")})
    return {
        "summary": {
            "total_orders": total_orders,
            "total_revenue": round(total_rev, 2),
            "avg_order_price": round(total_rev / max(1, total_orders), 2),
            "fbs_orders": fbs_orders,
            "fbs_revenue": round(fbs_rev, 2),
            "fbs_share_pct": round(fbs_orders / max(1, total_orders) * 100, 1),
            "fbw_orders": fbw_orders,
            "fbw_revenue": round(fbw_rev, 2),
            "fbw_share_pct": round(fbw_orders / max(1, total_orders) * 100, 1),
        },
        "by_day": by_day,
        "by_warehouse": by_wh[:40],
        "by_region": by_reg[:30],
        "by_city": by_city[:80],
        "by_article": by_art[:80],
        "filters": {
            "warehouses": sorted(warehouses_set),
            "regions": sorted(regions_set),
            "cities": sorted(cities_set)[:300],
            "date_min": dates[0] if dates else None,
            "date_max": dates[-1] if dates else None,
            "total_cached": len(all_orders),
        },
        "meta": {
            "updated_at": ORDERS_GEO_CACHE.get("updated_at"),
            "source": ORDERS_GEO_CACHE.get("source"),
            "filename": ORDERS_GEO_CACHE.get("filename"),
            "filtered": total_orders,
            "city_note": _orders_geo_city_note(ORDERS_GEO_CACHE.get("source")),
        },
    }


def _orders_geo_city_note(source: str) -> str:
    if source == "ribbon":
        return "Точные города — из отчёта «Лента заказов»."
    if source == "statistics_api":
        return "Данные с API: это область/край, а не город. Точные города — только в отчёте «Лента заказов»."
    if source == "ribbon+api":
        return "Из ленты — точные города, из API — область/край."
    return ""


def fetch_fbs_office_cities(days: int = 30) -> dict:
    """Город сдачи заказа FBS: Marketplace API v3 отдаёт offices — куда везём заказ.
    Statistics API для FBS пишет обезличенный «Склад WB РФ», поэтому этим уточняем склад
    отгрузки. Ключ — rid (он же srid в Statistics API)."""
    if not WB_TOKEN:
        return {}
    date_from_ts = int((datetime.now(timezone.utc) - timedelta(days=max(1, min(int(days), 90)))).timestamp())
    out = {}
    try:
        next_val = 0
        for _ in range(30):
            r = httpx.get(
                f"{WB_MARKETPLACE_URL}/api/v3/orders",
                headers=wb_headers(),
                params={"limit": 1000, "next": next_val, "dateFrom": date_from_ts},
                timeout=35,
            )
            if not r.is_success:
                logger.warning(f"orders_geo fbs offices error {r.status_code} {r.text[:200]}")
                break
            data = r.json()
            batch = data.get("orders") if isinstance(data, dict) else (data if isinstance(data, list) else [])
            if not batch:
                break
            for o in batch:
                offices = o.get("offices") or []
                city = str(offices[0]).strip() if offices else ""
                if not city:
                    continue
                for key in (o.get("rid"), o.get("orderUid"), o.get("id")):
                    if key:
                        out[str(key)] = city
            next_val = data.get("next", 0) if isinstance(data, dict) else 0
            if not next_val or len(batch) < 1000:
                break
    except Exception as e:
        logger.warning(f"orders_geo fbs offices exception: {e}")
    logger.info(f"orders_geo fbs offices: {len(out)} rids with city")
    return out


def sync_orders_geo_from_statistics(days: int = 30) -> dict:
    """Подтягивает заказы из Statistics API (склад + регион).
    Города для FBS дотягиваем из Marketplace API v3 (offices)."""
    if not WB_TOKEN:
        return {"error": "WB_TOKEN не задан"}
    if ORDERS_GEO_CACHE.get("syncing"):
        return {"status": "already_running"}
    ORDERS_GEO_CACHE["syncing"] = True
    ORDERS_GEO_CACHE["error"] = None
    try:
        cutoff = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=max(1, min(int(days), 90)))
        date_from = cutoff.strftime("%Y-%m-%dT00:00:00")
        raw = fetch_supplier_feed("/api/v1/supplier/orders", date_from, max_pages=5)
        fbs_cities = fetch_fbs_office_cities(days=days)
        records = []
        for o in raw or []:
            d = parse_wb_dt(o.get("date") or "")
            if d is None or d < cutoff:
                continue
            wh = str(o.get("warehouseName") or "").strip() or "Не указан"
            # Statistics API: warehouseType / склад продавца ≈ FBS
            wtype = str(o.get("warehouseType") or o.get("orderType") or "")
            channel = _orders_geo_normalize_channel(wtype, wh)
            if "seller" in wh.lower() or "продав" in wh.lower():
                channel = "FBS"
            price = float(o.get("finishedPrice") or o.get("priceWithDisc") or o.get("totalPrice") or 0)
            srid = str(o.get("srid") or "")
            # Для FBS Statistics API пишет обезличенный склад — берём город сдачи из Marketplace.
            if channel == "FBS" and srid:
                office = fbs_cities.get(srid)
                if office:
                    wh = f"Сдача: {office}"
            records.append({
                "order_id": srid or str(o.get("gNumber") or ""),
                "date": d.strftime("%Y-%m-%d"),
                "dt": d.strftime("%Y-%m-%d %H:%M:%S"),
                "channel": channel,
                "warehouse": wh,
                "dest_region": str(o.get("oblastOkrugName") or o.get("regionName") or "").strip() or "Не указан",
                "dest_city": str(o.get("regionName") or "").strip() or "Не указан",
                "article": str(o.get("supplierArticle") or "").strip(),
                "nm_id": int(o.get("nmId") or 0),
                "name": str(o.get("subject") or "").strip(),
                "price": round(price, 2),
                "status": "Отменён" if o.get("isCancel") else "Заказ",
            })
        with _ORDERS_GEO_LOCK:
            # если уже есть лента с городами — не затираем, а дополняем только новые даты API
            existing = ORDERS_GEO_CACHE.get("orders") or []
            if existing and ORDERS_GEO_CACHE.get("source") == "ribbon":
                exist_ids = {x.get("order_id") for x in existing if x.get("order_id")}
                exist_keys = {(x.get("date"), x.get("nm_id"), x.get("warehouse"), x.get("article")) for x in existing}
                added = 0
                for r in records:
                    key = (r.get("date"), r.get("nm_id"), r.get("warehouse"), r.get("article"))
                    if r.get("order_id") and r["order_id"] in exist_ids:
                        continue
                    if key in exist_keys:
                        continue
                    existing.append(r)
                    added += 1
                ORDERS_GEO_CACHE["orders"] = existing
                ORDERS_GEO_CACHE["updated_at"] = datetime.now(timezone.utc).isoformat()
                ORDERS_GEO_CACHE["source"] = "ribbon+api"
                _orders_geo_persist()
                return {"status": "ok", "added": added, "total": len(existing), "source": "ribbon+api"}
            ORDERS_GEO_CACHE["orders"] = records
            ORDERS_GEO_CACHE["updated_at"] = datetime.now(timezone.utc).isoformat()
            ORDERS_GEO_CACHE["source"] = "statistics_api"
            ORDERS_GEO_CACHE["filename"] = None
            _orders_geo_persist()
        return {"status": "ok", "total": len(records), "source": "statistics_api"}
    except Exception as e:
        ORDERS_GEO_CACHE["error"] = str(e)
        logger.error(f"sync_orders_geo_from_statistics: {e}")
        return {"error": str(e)}
    finally:
        ORDERS_GEO_CACHE["syncing"] = False


@app.get("/api/orders-geo")
def get_orders_geo(
    date_from: str = None,
    date_to: str = None,
    channel: str = "all",
    warehouse: str = "all",
    region: str = "all",
    city: str = "all",
    search: str = "",
):
    """Сводка географии заказов FBS/FBW за период (из кэша ленты или Statistics API)."""
    orders = _orders_geo_ensure_loaded()
    if not orders:
        return {
            "summary": {
                "total_orders": 0, "total_revenue": 0, "avg_order_price": 0,
                "fbs_orders": 0, "fbs_revenue": 0, "fbs_share_pct": 0,
                "fbw_orders": 0, "fbw_revenue": 0, "fbw_share_pct": 0,
            },
            "by_day": [], "by_warehouse": [], "by_region": [], "by_city": [], "by_article": [],
            "filters": {"warehouses": [], "regions": [], "cities": [], "date_min": None, "date_max": None, "total_cached": 0},
            "meta": {"updated_at": None, "source": None, "filename": None, "filtered": 0, "empty": True},
            "hint": "Загрузите Excel «Лента заказов» (вкладка Все заказы) или нажмите «Подтянуть с WB».",
        }
    # дефолтный период — последние 28 дней от max даты в кэше
    if not date_from and not date_to:
        dates = sorted({o.get("date") for o in orders if o.get("date")})
        if dates:
            date_to = dates[-1]
            try:
                d_to = datetime.strptime(date_to, "%Y-%m-%d").date()
                date_from = (d_to - timedelta(days=27)).isoformat()
            except Exception:
                date_from = dates[0]
    return aggregate_orders_geo(
        orders,
        date_from=date_from,
        date_to=date_to,
        channel=channel,
        warehouse=warehouse,
        region=region,
        city=city,
        search=search,
    )


@app.post("/api/orders-geo/upload")
async def upload_orders_geo(file: UploadFile = File(...)):
    """Загрузка Excel «Лента заказов» WB для раздела географии."""
    try:
        content = await file.read()
        if not content:
            return {"error": "Пустой файл"}
        records = parse_orders_geo_excel(content, filename=file.filename or "")
        if not records:
            return {"error": "Не удалось разобрать файл. Нужна вкладка «Все заказы» из ленты заказов WB."}
        with _ORDERS_GEO_LOCK:
            ORDERS_GEO_CACHE["orders"] = records
            ORDERS_GEO_CACHE["updated_at"] = datetime.now(timezone.utc).isoformat()
            ORDERS_GEO_CACHE["source"] = "ribbon"
            ORDERS_GEO_CACHE["filename"] = file.filename
            ORDERS_GEO_CACHE["error"] = None
            _orders_geo_persist()
        dates = sorted({o.get("date") for o in records if o.get("date")})
        return {
            "status": "ok",
            "total": len(records),
            "fbs": sum(1 for o in records if o.get("channel") == "FBS"),
            "fbw": sum(1 for o in records if o.get("channel") == "FBW"),
            "date_min": dates[0] if dates else None,
            "date_max": dates[-1] if dates else None,
            "filename": file.filename,
        }
    except Exception as e:
        logger.error(f"upload_orders_geo: {e}")
        return {"error": str(e)}


@app.post("/api/orders-geo/sync")
def sync_orders_geo(days: int = 30):
    """Подтянуть заказы из Statistics API WB (склад + регион)."""
    import threading
    if ORDERS_GEO_CACHE.get("syncing"):
        return {"status": "already_running"}
    threading.Thread(target=sync_orders_geo_from_statistics, args=(days,), daemon=True).start()
    return {"status": "started", "days": days}


DELIVERY_TIME_CACHE = {"data": None, "ts": 0, "days": 0}


def compute_delivery_times(days: int = 30, region: str = "") -> dict:
    """
    Сколько идёт заказ до покупателя. Считаем от оформления до выкупа, сопоставляя
    supplier/orders и supplier/sales по srid.

    Важно: это не чистая логистика. В срок входит время, пока покупатель забирает
    посылку из пункта выдачи, а невыкупленные заказы сюда вообще не попадают.
    Поэтому число всегда чуть больше реального срока доставки, но сравнивать
    склады и регионы между собой оно позволяет.
    """
    import statistics

    date_from = (_msk_now() - timedelta(days=max(1, int(days or 30)))).strftime("%Y-%m-%dT00:00:00")
    orders = fetch_supplier_feed("/api/v1/supplier/orders", date_from, max_pages=5)
    sales = fetch_supplier_feed("/api/v1/supplier/sales", date_from, max_pages=5)

    ord_by_srid = {}
    for o in orders:
        srid = str(o.get("srid") or "")
        if srid:
            ord_by_srid[srid] = o

    reg_filter = (region or "").strip().lower()
    pairs = []
    for s in sales:
        if not str(s.get("saleID") or "").startswith("S"):
            continue  # возвраты и корректировки пропускаем
        o = ord_by_srid.get(str(s.get("srid") or ""))
        if not o:
            continue
        od, sd = parse_wb_dt(o.get("date")), parse_wb_dt(s.get("date"))
        if not od or not sd:
            continue
        hours = (sd - od).total_seconds() / 3600.0
        if hours <= 0 or hours > 24 * 60:
            continue
        reg = str(o.get("regionName") or "").strip() or "Не указан"
        if reg_filter and reg_filter not in reg.lower():
            continue
        wh = str(o.get("warehouseName") or "").strip() or "Не указан"
        pairs.append({
            "days": hours / 24.0,
            "region": reg,
            "district": str(o.get("oblastOkrugName") or "").strip(),
            "warehouse": wh,
            "channel": _orders_geo_normalize_channel("", wh),
        })

    def summarize(items):
        d = sorted(x["days"] for x in items)
        return {
            "orders": len(d),
            "median_days": round(statistics.median(d), 1) if d else None,
            "avg_days": round(sum(d) / len(d), 1) if d else None,
            "p90_days": round(d[int(len(d) * 0.9)], 1) if len(d) >= 10 else None,
            "fastest_days": round(d[0], 1) if d else None,
        }

    by_wh, by_reg = {}, {}
    for p in pairs:
        by_wh.setdefault(p["warehouse"], []).append(p)
        by_reg.setdefault(p["region"], []).append(p)

    wh_rows = [{"warehouse": k, "channel": v[0]["channel"], **summarize(v)} for k, v in by_wh.items()]
    wh_rows.sort(key=lambda x: -x["orders"])
    reg_rows = [{"region": k, **summarize(v)} for k, v in by_reg.items()]
    reg_rows.sort(key=lambda x: -x["orders"])

    return {
        "period_days": days,
        "region_filter": region or None,
        "matched_pairs": len(pairs),
        "orders_seen": len(orders),
        "sales_seen": len(sales),
        "overall": summarize(pairs),
        "by_warehouse": wh_rows[:30],
        "by_region": reg_rows[:40],
        "note": ("Срок считается от оформления заказа до выкупа, поэтому включает время, "
                 "пока покупатель забирает посылку. Невыкупленные заказы не учитываются."),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }


@app.get("/api/delivery-time")
def get_delivery_time(days: int = 30, region: str = "", refresh: int = 0):
    """Сроки доставки до покупателя по складам отгрузки и регионам."""
    key_days = int(days or 30)
    fresh = (
        DELIVERY_TIME_CACHE.get("data")
        and DELIVERY_TIME_CACHE.get("days") == key_days
        and not region
        and time.time() - float(DELIVERY_TIME_CACHE.get("ts") or 0) < 3600
    )
    if fresh and not refresh:
        return DELIVERY_TIME_CACHE["data"]
    try:
        data = compute_delivery_times(days=key_days, region=region)
    except Exception as e:
        logger.error(f"delivery-time: {e}")
        return {"error": str(e)}
    if not region:
        DELIVERY_TIME_CACHE.update({"data": data, "ts": time.time(), "days": key_days})
    return data


@app.get("/api/seller-recommendations-agg")
def get_seller_recommendations_agg(refresh: int = 0):
    """
    Сводка «Продавец рекомендует»: какой nm в скольких карточках в топ-5 / ниже
    + missing — свои карточки, которых нигде нет в рекомендациях.
    """
    now = time.time()
    if (
        not refresh
        and SELLER_RECS_AGG_CACHE.get("data")
        and now - float(SELLER_RECS_AGG_CACHE.get("ts") or 0) < 300
    ):
        return SELLER_RECS_AGG_CACHE["data"]

    hosts, err = fetch_seller_recommendations_raw()
    if err and not hosts:
        raise HTTPException(status_code=502, detail=err)

    catalog = fetch_all_own_content_cards()
    agg = aggregate_seller_recommendations(hosts, catalog=catalog)
    out = {
        **agg,
        "error": err,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    SELLER_RECS_AGG_CACHE["ts"] = now
    SELLER_RECS_AGG_CACHE["data"] = out
    return out


def _warm_caches_after_start():
    """Через пару минут после старта прогреваем то, что не переживает редеплой."""
    time.sleep(120)
    try:
        _orders_geo_ensure_loaded()
    except Exception as e:
        logger.warning(f"warmup orders_geo: {e}")
    try:
        if not WB_PRODUCTS_CACHE.get("sales_by_nm"):
            refresh_wb_products_catalog(sync_sources=True)
    except Exception as e:
        logger.warning(f"warmup wb_products: {e}")
    try:
        sync_new_stock()
    except Exception as e:
        logger.warning(f"warmup new_stock: {e}")


threading.Thread(target=_warm_caches_after_start, daemon=True, name="warmup").start()
threading.Thread(target=sync_promo_calendar, daemon=True, name="promo-cal-boot").start()


# ---------- Телеграм-бот (только чтение) ----------
# Живёт в этом же процессе. Отдельный сервер не нужен: Railway уже за рубежом.
try:
    import telegram_bot

    telegram_bot.start_bot({
        "wb_products": lambda: get_wb_products(),
        "orders_geo": lambda **kw: aggregate_orders_geo(_orders_geo_ensure_loaded(), **kw),
        "fbs_speed": lambda days=14: fetch_fbs_speed_report_data(days=days),
        "sales_pace": lambda period="day": get_sales_pace(period=period),
        "own_warehouse": lambda: get_own_warehouse_stock(),
    })
    @app.get("/api/telegram-status")
    def telegram_status():
        """Что настроено у бота и жив ли он. Секреты не отдаёт."""
        return telegram_bot.bot_status()

    @app.get("/api/llm-models")
    def llm_models(contains: str = ""):
        """Какие модели доступны настроенному ключу — чтобы не гадать с точным id."""
        return telegram_bot.list_models(contains)

    @app.get("/api/llm-ping")
    def llm_ping(q: str = "Ответь одним словом: работает"):
        """Прогоняет ту же цепочку, что и бот, — чтобы проверять её без телеграма."""
        started = time.time()
        answer = telegram_bot.ask_llm(0, q)
        return {
            "question": q,
            "answer": answer,
            "provider": telegram_bot.bot_status().get("llm_provider"),
            "model": telegram_bot.bot_status().get("model"),
            "seconds": round(time.time() - started, 1),
        }

except Exception as e:
    logger.warning(f"telegram bot не поднялся: {e}")


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8080))
    # Один процесс (scheduler), но больше потоков под sync-эндпоинты —
    # иначе долгий /api/dashboard-data блокирует весь сайт.
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=port,
        timeout_keep_alive=30,
        limit_concurrency=40,
    )
